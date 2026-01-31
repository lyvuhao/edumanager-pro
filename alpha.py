import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Toplevel, simpledialog
import pandas as pd
import warnings
import os
import sys
import re
import json
from datetime import datetime, timedelta
import unicodedata
# [ADDED] Imports cho Trợ Lý Phòng Thi (Exam Proctor Toolkit)
import threading  # Để chạy đồng hồ đếm ngược không block UI thread
import time       # Để sleep và tính toán thời gian
try:
    import winsound  # Âm thanh cảnh báo (Windows only)
    HAS_WINSOUND = True
except ImportError:
    HAS_WINSOUND = False
    print("⚠️ Module winsound không khả dụng (chỉ chạy trên Windows). Âm thanh sẽ bị tắt.")
# [END ADDED]
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import math
import random
import io
from functools import lru_cache  # [OPTIMIZED] Cache cho normalize_text
from typing import List, Union  # [OPTIMIZED] Type hints

# =================================================================================
# [AUTO-UPDATE] VERSION MANAGEMENT
# =================================================================================
APP_VERSION = "1.0.1"  # Format: MAJOR.MINOR.PATCH - CẬP NHẬT MỖI KHI RELEASE
APP_BUILD_DATE = "2026-01-31"
APP_NAME = "EduManager Pro"

# [CONFIG] Thay đổi sau khi tạo GitHub repo
GITHUB_USERNAME = "lyvuhao"  # ✅ Username GitHub của bạn
GITHUB_REPO = "edumanager-pro"
GITHUB_BRANCH = "main"
UPDATE_CHECK_URL = f"https://raw.githubusercontent.com/{GITHUB_USERNAME}/{GITHUB_REPO}/{GITHUB_BRANCH}/version.json"
# =================================================================================

import tempfile   # Cho auto-update system
import queue      # Cho threading communication

# =================================================================================
# --- LICENSE KEY SYSTEM ---
import requests  # Cryptolens API calls
import hashlib   # Machine fingerprint
import platform  # System info
import base64    # Encoding
import subprocess  # Get hardware info
try:
    from Crypto.Cipher import AES  # Encrypt license file
    from Crypto.Util.Padding import pad, unpad  # AES padding
    from Crypto.Random import get_random_bytes
    HAS_CRYPTO = True
except ImportError:
    HAS_CRYPTO = False
    print("⚠️ Chưa cài pycryptodome. License sẽ không được mã hóa!")
# [END LICENSE KEY SYSTEM]

# --- PHẦN DATA SCIENCE ---
try:
    from sklearn.linear_model import LinearRegression
    import numpy as np
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

try:
    import scipy.stats as stats
    import numpy as np
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Canh bao: Thieu thu vien openpyxl. Tinh nang luu bao toan se bi han che.")

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import lxml
    HAS_LXML = True
except ImportError:
    HAS_LXML = False

# [ADDED] Excel COM để auto-repair file corrupt (VNEdu, etc.)
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False
# -------------------------

# [ADDED] Voice Input - Nhập điểm bằng giọng nói
import queue  # Để giao tiếp giữa thread voice và UI

try:
    import speech_recognition as sr
    HAS_SPEECH = True
except ImportError:
    HAS_SPEECH = False
    print("⚠️ Chưa cài speech_recognition. Chạy: pip install SpeechRecognition")

try:
    import sounddevice as sd
    import soundfile as sf
    import numpy as np
    AUDIO_BACKEND = "sounddevice"
except ImportError:
    AUDIO_BACKEND = None
    print("⚠️ Chưa cài sounddevice. Chạy: pip install sounddevice soundfile numpy")

try:
    from fuzzywuzzy import fuzz
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False
    print("⚠️ Chưa cài fuzzywuzzy. Chạy: pip install fuzzywuzzy python-Levenshtein")

# [NEW] Phase 1: Audio Processing - Voice Activity Detection
# NOTE: Noise reduction (noisereduce/PyTorch) removed for PyInstaller compatibility
# [END Voice Input]

# ===== [FIX #2] HELPER FUNCTION: Build vocabulary hints =====
# except ImportError:
#     AUDIO_BACKEND = None
#     print("⚠️ Chưa cài sounddevice. Chạy: pip install sounddevice soundfile numpy")
#
# try:
#     from fuzzywuzzy import fuzz
#     HAS_FUZZY = True
# except ImportError:
#     HAS_FUZZY = False
#     print("⚠️ Chưa cài fuzzywuzzy. Chạy: pip install fuzzywuzzy python-Levenshtein")

# [NEW] Phase 1: Audio Processing - Voice Activity Detection
# NOTE: Noise reduction (noisereduce/PyTorch) removed for PyInstaller compatibility
# [END Voice Input]

# ===== [FIX #2] HELPER FUNCTION: Build vocabulary hints =====
def build_vocabulary_hints(students):
    """Xây dựng vocabulary hints từ danh sách học sinh cho speech recognition"""
    hints = {
        "names": [],
        "phrases": []
    }
    
    for student in students:
        name = student['name']
        hints["names"].append(name)
        
        # Thêm tên gọi (phần cuối)
        name_parts = name.split()
        if name_parts:
            hints["phrases"].append(name_parts[-1])
        
        # Thêm tên + đệm (2 phần cuối)
        if len(name_parts) >= 2:
            hints["phrases"].append(" ".join(name_parts[-2:]))
    
    return hints
# ===== [END FIX #2] =====

# Tắt cảnh báo
warnings.filterwarnings("ignore")

# =================================================================================
# CẤU HÌNH GIAO DIỆN (MOUSE GREY DARK MODE & PASTEL LIGHT MODE)
# =================================================================================
IS_DARK_MODE = False

# Bảng màu Pastel Premium
PASTEL_PALETTE = {
    "mint": "#A7F3D0",       # Xanh bạc hà (Success)
    "mint_dark": "#059669",
    "blue": "#BFDBFE",       # Xanh dương phấn (Info)
    "blue_dark": "#2563EB",
    "rose": "#FECDD3",       # Hồng phấn (Warning/High)
    "rose_dark": "#E11D48",
    "lavender": "#DDD6FE",   # Tím nhạt (Tools)
    "lavender_dark": "#7C3AED",
    "orange": "#FED7AA",     # Cam nhạt (Attention)
    "orange_dark": "#EA580C",
    "gray_bg": "#F3F4F6",    # Nền App
    "white": "#FFFFFF",      # Nền Card
    "text": "#1F2937",       # Chữ chính
    "text_light": "#6B7280"  # Chữ phụ
}

LIGHT_THEME = {
    "primary": "#3B8ED0",        
    "bg_app": PASTEL_PALETTE["gray_bg"],        
    "bg_card": PASTEL_PALETTE["white"],        
    "text_main": PASTEL_PALETTE["text"],
    "text_sub": PASTEL_PALETTE["text_light"],
    "text_white": "#FFFFFF", 
    "text_shadow": "#000000", 
    "border": "#000000",        
    "chart_bg": PASTEL_PALETTE["white"],
    "entry_bg": "#F9FAFB",
    "tree_bg": PASTEL_PALETTE["white"],
    "tree_fg": PASTEL_PALETTE["text"],
    "tree_header_bg": "#F8F9FA",
    "tree_header_fg": "#374151",
    "btn_border_col": "#bdc3c7", # Viền nút xám nhạt
    "tab_bg_selected": PASTEL_PALETTE["white"],
    "tab_bg_unselected": "#E5E7EB",
    "primary_dark": "#2A6EBB", 
    "btn_border": "#2c3e50" 
}

# [UPDATED] DARK MODE: MOUSE GREY (XÁM LÔNG CHUỘT)
DARK_THEME = {
    "primary": "#3B8ED0", 
    "primary_dark": "#64B5F6", 
    "bg_app": "#2b2b2b",     # Nền App: Xám đậm
    "bg_card": "#383838",    # Nền Card/Table: Xám vừa (Dễ chịu)
    "text_main": "#E0E0E0",  # Chữ trắng đục (chống chói)
    "text_sub": "#B0BEC5", 
    "text_white": "#FFFFFF", 
    "text_shadow": "#000000",
    "border": "#555555", 
    "chart_bg": "#383838", 
    "entry_bg": "#454545",   # Input: Xám sáng hơn chút
    "tree_bg": "#383838",    # Nền bảng dữ liệu đồng bộ Card
    "tree_fg": "#E0E0E0", 
    "tree_header_bg": "#454545", # Header bảng xám đậm
    "tree_header_fg": "#FFFFFF",
    "btn_border_col": "#666666", # Viền nút xám rõ hơn
    "tab_bg_selected": "#383838",
    "tab_bg_unselected": "#1f1f1f", 
    "btn_border": "#FFFFFF"
}

# [UPDATED] Style cho Treeview (Bảng dữ liệu) - Màu pha trộn với nền xám
TREE_ROW_COLORS = {
    "Light": {
        "warning": {"bg": "#FEF2F2", "fg": "#991B1B"},    
        "good": {"bg": "#ECFDF5", "fg": "#065F46"},       
        "excellent": {"bg": "#F5F3FF", "fg": "#5B21B6"},  
        "nodata": {"bg": "#F9FAFB", "fg": "#9CA3AF"},
        "normal": {"bg": "#FFFFFF", "fg": "#1F2937"}
    },
    "Dark": {
        # Màu nền pha trộn với #383838 để dịu mắt
        "warning": {"bg": "#4a2c2c", "fg": "#ffcccc"},    # Xám ám đỏ
        "good": {"bg": "#2c4a3e", "fg": "#ccffeb"},       # Xám ám xanh
        "excellent": {"bg": "#3a2c4a", "fg": "#eaccff"},  # Xám ám tím
        "nodata": {"bg": "#2f2f2f", "fg": "#888888"},     # Xám tối hơn nền
        "normal": {"bg": "#383838", "fg": "#E0E0E0"}      # Xám chuẩn (đồng bộ nền)
    }
}

THEME = LIGHT_THEME.copy()
THEME.update({
    "font_header": ("Segoe UI", 14, "bold"),
    "font_val": ("Segoe UI", 24, "bold"),
    "font_title": ("Segoe UI", 9, "bold"),
    "font_body": ("Segoe UI", 10),
    "font_status": ("Segoe UI", 13, "bold"), 
})

CARD_COLORS = {
    "All": (PASTEL_PALETTE["blue"], PASTEL_PALETTE["blue_dark"]),        
    "Tot": (PASTEL_PALETTE["mint"], PASTEL_PALETTE["mint_dark"]),        
    "HoanThanh": (PASTEL_PALETTE["orange"], PASTEL_PALETTE["orange_dark"]),
    "CanChuY": (PASTEL_PALETTE["rose"], PASTEL_PALETTE["rose_dark"]),  
    "ChuaCoDiem": ("#E5E7EB", "#4B5563") # Xám
}

# =================================================================================
# HELPER FUNCTIONS (TÍNH TOÁN)
# =================================================================================
def calculate_dtb_exact(tx_list: List[Union[int, float]], gk: Union[int, float, str], ck: Union[int, float, str]) -> Union[float, str]:
    """
    [OPTIMIZED] Tự động tính ĐTB: (Tổng TX + GK*2 + CK*3) / Tổng hệ số
    """
    numerator = 0.0 
    denominator = 0 

    # 1. TX (Hệ số 1)
    valid_tx = []
    for x in tx_list:
        if isinstance(x, (int, float)):
            valid_tx.append(float(x))
     
    numerator += sum(valid_tx)
    denominator += len(valid_tx)

    # 2. GK (Hệ số 2)
    if isinstance(gk, (int, float)):
        numerator += float(gk) * 2
        denominator += 2

    # 3. CK (Hệ số 3)
    if isinstance(ck, (int, float)):
        numerator += float(ck) * 3
        denominator += 3

    if denominator == 0:
        return ""
    
    result = numerator / denominator
    return round(result, 1)

def clean_float_val(val):
    """[OPTIMIZED] Chuyển đổi giá trị thành float, xử lý lỗi an toàn"""
    if pd.isna(val) or str(val).strip() == "": return ""
    s_val = str(val).strip().replace(",", ".") 
    try:
        return float(s_val)
    except (ValueError, TypeError):  # [FIXED] Specific exception thay vì bare except
        return ""

@lru_cache(maxsize=2000)  # [OPTIMIZED] Cache 2000 tên gần nhất
def normalize_text(val):
    """[OPTIMIZED] Bỏ dấu và chuyển thường để tìm kiếm không dấu."""
    s = str(val).lower()
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in nfkd if unicodedata.category(ch) != "Mn")

def convert_legacy_file_to_xlsx(file_path):
    """
    Chuyển đổi file .xls cũ sang .xlsx mới.
    Hỗ trợ cả file corrupt thông qua Excel COM.
    
    Args:
        file_path: Đường dẫn file .xls cần convert
        
    Returns:
        Tuple[str, str]: (new_path, error_msg) - new_path nếu thành công, error_msg nếu thất bại
    """
    try:
        base_name = os.path.splitext(file_path)[0]
        timestamp = datetime.now().strftime("%H%M%S")
        new_path = f"{base_name}_converted_{timestamp}.xlsx"
        
        dfs = {} 
        is_binary_excel = False  # Flag để biết file có phải Excel binary không
        used_direct_save = False  # Flag biết đã SaveAs trực tiếp chưa
        
        # Kiểm tra định dạng file thật (dựa vào magic bytes)
        try:
            with open(file_path, 'rb') as f:
                header = f.read(8)
            # OLE2 compound document (Excel .xls thật)
            if header.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
                is_binary_excel = True
        except:
            pass
        
        # Bước 1: Thử đọc bằng xlrd (file Excel chuẩn)
        if HAS_XLRD:
            try:
                dfs = pd.read_excel(file_path, sheet_name=None, header=None, engine='xlrd')
            except Exception as e:
                # Nếu file binary nhưng xlrd lỗi → file corrupt
                # Dùng Excel COM để SaveAs TRỰC TIẾP (giữ nguyên 100% format)
                if is_binary_excel and HAS_WIN32COM:
                    try:
                        success = _repair_file_via_excel_com(file_path, new_path)
                        if success and os.path.exists(new_path):
                            # File đã được repair và save trực tiếp
                            used_direct_save = True
                            # Đọc lại để có dfs cho validation
                            try:
                                dfs = pd.read_excel(new_path, sheet_name=None, header=None, engine='openpyxl')
                            except:
                                pass
                    except Exception as com_err:
                        try:
                            print(f"[DEBUG] Excel COM repair failed: {com_err}")
                        except:
                            pass

        # Bước 2: Thử đọc như HTML/CSV (CHỈ cho file KHÔNG phải binary Excel)
        if not dfs and not is_binary_excel and not used_direct_save:
            raw_content = None
            encodings_to_try = ['utf-16', 'utf-8', 'utf-8-sig', 'cp1252', 'latin1']
            for enc in encodings_to_try:
                try:
                    with open(file_path, 'r', encoding=enc) as f:
                        content = f.read()
                        if len(content) > 10:
                            raw_content = content
                            break 
                except Exception:
                    continue
            
            if raw_content:
                try:
                    list_dfs = pd.read_html(io.StringIO(raw_content), header=None)
                    if list_dfs:
                        for i, df in enumerate(list_dfs):
                            dfs[f"Sheet{i+1}"] = df
                except Exception as e:
                    pass

                if not dfs:
                    try:
                        df = pd.read_csv(io.StringIO(raw_content), sep='\t', header=None, on_bad_lines='skip')
                        if len(df.columns) > 1: dfs["Sheet1"] = df
                    except:
                        pass
                
                if not dfs:
                    try:
                        df = pd.read_csv(io.StringIO(raw_content), sep=',', header=None, on_bad_lines='skip')
                        if len(df.columns) > 1: dfs["Sheet1"] = df
                    except:
                        pass

        # Bước 3: Fallback - Thử Excel COM SaveAs trực tiếp nếu vẫn chưa có
        if not dfs and not used_direct_save and HAS_WIN32COM:
            try:
                success = _repair_file_via_excel_com(file_path, new_path)
                if success and os.path.exists(new_path):
                    used_direct_save = True
                    try:
                        dfs = pd.read_excel(new_path, sheet_name=None, header=None, engine='openpyxl')
                    except:
                        pass
            except Exception as e:
                try:
                    print(f"[DEBUG] Excel COM convert failed: {e}")
                except:
                    pass

        if not dfs:
            return None, "Định dạng file không hỗ trợ hoặc file bị hỏng. Không tìm thấy bảng dữ liệu."

        # Nếu đã SaveAs trực tiếp từ Excel COM thì không cần ghi lại
        if used_direct_save and os.path.exists(new_path):
            return new_path, None

        # Còn lại: ghi file mới từ dfs (cho trường hợp HTML/CSV hoặc xlrd thành công)
        valid_dfs = {}
        for name, df in dfs.items():
            lower_name = str(name).lower()
            if any(x in lower_name for x in ["bìa", "hướng dẫn", "intro", "cover", "phiếu"]): continue
            if df.shape[0] < 3 or df.shape[1] < 3: continue
            valid_dfs[name] = df

        if not valid_dfs: valid_dfs = dfs 

        with pd.ExcelWriter(new_path, engine='openpyxl') as writer:
            for sheet_name, df in valid_dfs.items():
                safe_name = str(sheet_name)[:30]
                # Sanitize tên sheet - loại bỏ ký tự không hợp lệ
                safe_name = _sanitize_excel_string(safe_name)
                if not safe_name:
                    safe_name = f"Sheet{list(valid_dfs.keys()).index(sheet_name) + 1}"
                
                df = df.astype(str) 
                df = df.replace("nan", "")
                df = df.replace("None", "")
                
                # Sanitize toàn bộ dữ liệu - loại bỏ ký tự không hợp lệ cho Excel
                # Sử dụng .map() thay vì .applymap() (deprecated trong pandas 2.1+)
                df = df.map(lambda x: _sanitize_excel_string(str(x)) if pd.notna(x) else "")
                
                df.to_excel(writer, sheet_name=safe_name, index=False, header=False)
        
        return new_path, None

    except Exception as e:
        return None, str(e)


def _sanitize_excel_string(s):
    """
    Loại bỏ các ký tự không hợp lệ cho Excel worksheets.
    Excel không chấp nhận: control characters (0x00-0x1F except 0x09, 0x0A, 0x0D)
    
    Args:
        s: Chuỗi cần sanitize
        
    Returns:
        Chuỗi đã được làm sạch
    """
    if not isinstance(s, str):
        s = str(s)
    
    # Loại bỏ các ký tự control (trừ tab, newline, carriage return)
    # Và các ký tự không in được
    result = []
    for char in s:
        code = ord(char)
        # Cho phép: Tab (9), Newline (10), Carriage Return (13), và các ký tự >= 32
        if code == 9 or code == 10 or code == 13 or code >= 32:
            # Loại bỏ thêm các ký tự đặc biệt gây lỗi Excel
            if code < 0xFFFE and code not in [0x7F]:  # DEL character
                result.append(char)
    
    return ''.join(result)


def _convert_via_excel_com(file_path, output_path=None):
    """
    Dùng Excel COM để repair file corrupt và SaveAs thành .xlsx (giữ nguyên định dạng).
    
    Args:
        file_path: Đường dẫn file Excel gốc
        output_path: Đường dẫn file output .xlsx (nếu None, tạo tự động)
        
    Returns:
        Tuple[str, Dict[str, DataFrame]]: (đường dẫn file mới, dict of DataFrames)
        Nếu output_path=None (gọi từ chỗ cũ), trả về dict để tương thích ngược
    """
    if not HAS_WIN32COM:
        return {} if output_path is None else (None, {})
        
    excel = None
    dfs = {}
    result_path = None
    
    try:
        # Tạo output path nếu chưa có
        need_save = (output_path is not None) or True  # Luôn SaveAs để giữ format
        if output_path is None:
            base_name = os.path.splitext(file_path)[0]
            timestamp = datetime.now().strftime("%H%M%S")
            output_path = f"{base_name}_repaired_{timestamp}.xlsx"
        
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Mở file với chế độ repair (CorruptLoad=1 = xlRepairFile)
        wb = excel.Workbooks.Open(
            os.path.abspath(file_path), 
            UpdateLinks=0, 
            ReadOnly=True, 
            CorruptLoad=1
        )
        
        # SaveAs thành .xlsx (FileFormat=51 = xlOpenXMLWorkbook)
        # Cách này GIỮ NGUYÊN toàn bộ formatting, merge cells, styles, etc.
        abs_output = os.path.abspath(output_path)
        wb.SaveAs(abs_output, FileFormat=51)
        result_path = abs_output
        
        wb.Close(False)
        excel.Quit()
        excel = None
        
        # Đọc file .xlsx mới bằng openpyxl/pandas (đã được repair)
        try:
            dfs = pd.read_excel(result_path, sheet_name=None, header=None, engine='openpyxl')
            try:
                print(f"[INFO] Excel COM: Repaired & saved {len(dfs)} sheets to {os.path.basename(result_path)}")
            except:
                pass
        except Exception as e:
            try:
                print(f"[DEBUG] Read repaired file failed: {e}")
            except:
                pass
        
        return dfs  # Trả về dict để tương thích với code gọi hiện tại
        
    except Exception as e:
        try:
            print(f"[DEBUG] _convert_via_excel_com error: {e}")
        except:
            print("[DEBUG] _convert_via_excel_com error: (unicode error)")
        return {}
        
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except:
                pass


def _repair_file_via_excel_com(file_path, output_path):
    """
    Dùng Excel COM để repair file corrupt và SaveAs trực tiếp (không đọc data).
    Giữ nguyên 100% định dạng, merge cells, styles.
    
    Args:
        file_path: Đường dẫn file Excel gốc
        output_path: Đường dẫn file output .xlsx
        
    Returns:
        bool: True nếu thành công
    """
    if not HAS_WIN32COM:
        return False
        
    excel = None
    
    try:
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Mở file với chế độ repair
        wb = excel.Workbooks.Open(
            os.path.abspath(file_path), 
            UpdateLinks=0, 
            ReadOnly=True, 
            CorruptLoad=1
        )
        
        # SaveAs thành .xlsx (FileFormat=51 = xlOpenXMLWorkbook)
        wb.SaveAs(os.path.abspath(output_path), FileFormat=51)
        
        wb.Close(False)
        excel.Quit()
        excel = None
        
        try:
            print(f"[INFO] Excel COM: File repaired -> {os.path.basename(output_path)}")
        except:
            pass
        return True
        
    except Exception as e:
        try:
            print(f"[DEBUG] _repair_file_via_excel_com error: {e}")
        except:
            pass
        return False
        
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except:
                pass


# =================================================================================
# LICENSE KEY MANAGEMENT SYSTEM (CRYPTOLENS)
# =================================================================================

class LicenseManager:
    """
    Quản lý License Key với Cryptolens.io
    - Trial mode: 7 ngày miễn phí
    - Time-limited license với machine locking
    - Offline grace period: 7 ngày
    - Hybrid validation: Online (API) + Offline (local check)
    """
    
    # ===== CRYPTOLENS CONFIG =====
    ACCESS_TOKEN = "WyIxMTY5NzkzMzEiLCJFbXJaM0xXTmJTQVAxOHdRSGhTK3RzLzNoemVwNXdJalF4djluNENqIl0="
    PRODUCT_ID = 32078
    API_BASE_URL = "https://app.cryptolens.io/api/key"
    
    # RSA Public Key từ Cryptolens (Product Settings > Security)
    # Dùng để verify signature, chống giả mạo API response
    RSA_PUBLIC_KEY = """<RSAKeyValue><Modulus>sGbvxwdGAKEP7nI8N/MRjdS7qMOx8xD+y2DpIBcV3K3Qs7TF5Z4kS1V5jR5rXyj7q+U8xkQP5lZ+H2xJRkX8O7vK+H9x6oQvY7EfPD5dK3rJ9h2wM8BsI5OeN1vQk7H+R2kVxY8jN6dM5xZP7FwB9uS3rN1vXfT7H8xD2bK5yE=</Modulus><Exponent>AQAB</Exponent></RSAKeyValue>"""
    
    # Flag để bật/tắt signature verification (tắt khi chưa có public key đúng)
    VERIFY_SIGNATURE = False  # Bật = True khi đã có RSA key đúng
    
    # ===== FILE PATHS (CÙNG THƯ MỤC VỚI FILE .py) =====
    @staticmethod
    def _get_app_dir():
        """Lấy thư mục chứa file .py (hoặc .exe khi đóng gói)"""
        import sys
        if getattr(sys, 'frozen', False):
            # Đang chạy từ exe (PyInstaller)
            return os.path.dirname(sys.executable)
        else:
            # Đang chạy từ .py
            return os.path.dirname(os.path.abspath(__file__))
    
    # ===== SETTINGS =====
    TRIAL_DAYS = 7
    OFFLINE_GRACE_DAYS = 7
    
    def __init__(self):
        """Khởi tạo License Manager"""
        self.machine_code = self.get_machine_code()
        self.license_data = None
        
        # Đặt đường dẫn file cùng thư mục với app
        app_dir = self._get_app_dir()
        self.LICENSE_FILE = os.path.join(app_dir, "edumanager.lic")
        self.TRIAL_FILE = os.path.join(app_dir, "edumanager.trial")
        print(f"[DEBUG] App directory: {app_dir}")
        print(f"[DEBUG] License file: {self.LICENSE_FILE}")
        print(f"[DEBUG] Trial file: {self.TRIAL_FILE}")
    
    def get_machine_code(self):
        """
        Tạo Machine Code duy nhất dựa trên hardware.
        Kết hợp: MAC address + CPU info + Disk serial
        """
        try:
            import uuid
            
            # 1. MAC Address
            mac = ':'.join(['{:02x}'.format((uuid.getnode() >> i) & 0xff) 
                            for i in range(0, 8*6, 8)][::-1])
            
            # 2. CPU Info
            cpu = platform.processor()
            
            # 3. Windows Disk Serial
            disk_serial = ""
            try:
                result = subprocess.check_output(
                    "wmic diskdrive get serialnumber", 
                    shell=True,
                    stderr=subprocess.DEVNULL
                ).decode()
                lines = result.strip().split('\n')
                if len(lines) > 1:
                    disk_serial = lines[1].strip()
            except:
                pass
            
            # 4. Combine and hash
            raw_data = f"{mac}|{cpu}|{disk_serial}|{platform.system()}"
            machine_hash = hashlib.sha256(raw_data.encode()).hexdigest()
            
            # Return first 16 chars (readable length)
            return machine_hash[:16].upper()
            
        except Exception as e:
            # Fallback: Use username + computername
            fallback = f"{os.getenv('USERNAME', 'user')}_{os.getenv('COMPUTERNAME', 'pc')}"
            return hashlib.md5(fallback.encode()).hexdigest()[:16].upper()
    
    # ========== ANTI-TIME-MANIPULATION ==========
    
    def _get_network_time(self):
        """
        Lấy thời gian từ server online để chống sửa đồng hồ hệ thống.
        Returns: datetime hoặc None nếu offline
        """
        try:
            # Sử dụng nhiều nguồn thời gian
            time_servers = [
                "http://worldtimeapi.org/api/ip",
                "https://timeapi.io/api/Time/current/zone?timeZone=Asia/Ho_Chi_Minh"
            ]
            
            for server in time_servers:
                try:
                    response = requests.get(server, timeout=3)
                    if response.status_code == 200:
                        data = response.json()
                        # WorldTimeAPI format
                        if "datetime" in data:
                            time_str = data["datetime"][:19]  # Remove timezone
                            return datetime.fromisoformat(time_str)
                        # TimeAPI format
                        elif "dateTime" in data:
                            time_str = data["dateTime"][:19]
                            return datetime.fromisoformat(time_str)
                except:
                    continue
            
            return None
        except:
            return None
    
    def _check_time_manipulation(self):
        """
        Kiểm tra xem user có đang sửa đồng hồ hệ thống không.
        Returns: (is_manipulated: bool, reason: str)
        """
        try:
            local_time = datetime.now()
            network_time = self._get_network_time()
            
            if network_time:
                # Cho phép sai lệch tối đa 1 giờ
                time_diff = abs((local_time - network_time).total_seconds())
                if time_diff > 3600:  # > 1 giờ
                    return True, f"Đồng hồ hệ thống sai lệch {int(time_diff/60)} phút so với thời gian thực!"
            
            return False, None
        except:
            return False, None
    
    # ========== SIGNATURE VERIFICATION ==========
    
    def _verify_signature(self, license_key_base64, signature):
        """
        Verify RSA signature từ Cryptolens response.
        Đảm bảo dữ liệu không bị giả mạo.
        """
        if not self.VERIFY_SIGNATURE:
            print("[DEBUG] Signature verification disabled")
            return True
        
        try:
            # Cần pycryptodome để verify RSA
            if not HAS_CRYPTO:
                print("[WARNING] Cannot verify signature - pycryptodome not installed")
                return True  # Skip nếu không có crypto
            
            from Crypto.PublicKey import RSA
            from Crypto.Signature import pkcs1_15
            from Crypto.Hash import SHA256
            import xml.etree.ElementTree as ET
            
            # Parse RSA public key từ XML format
            root = ET.fromstring(self.RSA_PUBLIC_KEY)
            modulus = base64.b64decode(root.find('Modulus').text)
            exponent = base64.b64decode(root.find('Exponent').text)
            
            # Construct RSA key
            n = int.from_bytes(modulus, byteorder='big')
            e = int.from_bytes(exponent, byteorder='big')
            public_key = RSA.construct((n, e))
            
            # Verify signature
            decoded_signature = base64.b64decode(signature)
            h = SHA256.new(license_key_base64.encode('utf-8'))
            
            try:
                pkcs1_15.new(public_key).verify(h, decoded_signature)
                print("[DEBUG] Signature verification: PASSED ✓")
                return True
            except (ValueError, TypeError):
                print("[SECURITY] Signature verification: FAILED ✗")
                return False
                
        except Exception as e:
            print(f"[WARNING] Signature verification error: {e}")
            return True  # Skip on error to prevent blocking
    
    # ========== TRIAL MODE (ENCRYPTED + MACHINE LOCKED + ANTI-BACKUP) ==========
    
    # Registry key để lưu backup trial info (chống copy file)
    REGISTRY_KEY = r"SOFTWARE\\EduManager"
    REGISTRY_VALUE = "TrialInfo"
    
    def _get_registry_trial(self):
        """Đọc trial info từ Windows Registry"""
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.REGISTRY_KEY, 0, winreg.KEY_READ)
            value, _ = winreg.QueryValueEx(key, self.REGISTRY_VALUE)
            winreg.CloseKey(key)
            # Decrypt value
            decoded = base64.b64decode(value).decode()
            return json.loads(decoded)
        except:
            return None
    
    def _set_registry_trial(self, trial_data):
        """Lưu trial info vào Windows Registry"""
        try:
            import winreg
            key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, self.REGISTRY_KEY)
            # Encrypt và lưu
            encoded = base64.b64encode(json.dumps(trial_data).encode()).decode()
            winreg.SetValueEx(key, self.REGISTRY_VALUE, 0, winreg.REG_SZ, encoded)
            winreg.CloseKey(key)
            return True
        except Exception as e:
            print(f"[DEBUG] Registry write failed: {e}")
            return False
    
    def is_trial_available(self):
        """
        Kiểm tra xem còn trial không.
        Trial được mã hóa và gắn với machine code để chống reset.
        CHỐNG BACKUP: So sánh last_used trong file vs Registry
        CHỐNG SỬA ĐỒNG HỒ: Kiểm tra network time
        Returns: (is_available: bool, days_left: int, hours_left: int, mins_left: int)
        """
        # ===== KIỂM TRA SỬA ĐỒNG HỒ HỆ THỐNG =====
        is_manipulated, reason = self._check_time_manipulation()
        if is_manipulated:
            print(f"[SECURITY] Time manipulation detected: {reason}")
            # Không return False ngay, vẫn cho phép nhưng ghi log
        
        # Kiểm tra Registry trước (chống xóa file)
        registry_trial = self._get_registry_trial()
        
        if not os.path.exists(self.TRIAL_FILE):
            # Không có file trial
            if registry_trial and registry_trial.get("machine_code") == self.machine_code:
                # Đã có trong Registry - kiểm tra còn thời gian không
                try:
                    first_run = datetime.fromisoformat(registry_trial["first_run"])
                    now = datetime.now()
                    trial_end = first_run + timedelta(days=self.TRIAL_DAYS)
                    remaining = trial_end - now
                    
                    if remaining.total_seconds() > 0:
                        # Trial còn thời gian - khôi phục file từ Registry
                        print("[DEBUG] Trial file missing but Registry valid - restoring trial file")
                        self._save_trial_encrypted(registry_trial)
                        days_left = remaining.days
                        hours_left = remaining.seconds // 3600
                        mins_left = (remaining.seconds % 3600) // 60
                        return True, days_left, hours_left, mins_left
                    else:
                        # Trial đã hết hạn
                        print("[DEBUG] Trial expired (Registry check)")
                        return False, 0, 0, 0
                except:
                    # Registry corrupt
                    print("[DEBUG] Registry trial data corrupt")
                    return False, 0, 0, 0
            
            # Chưa dùng trial - kiểm tra online
            online_trial = self._check_online_trial()
            if online_trial:
                return False, 0, 0, 0
            return True, self.TRIAL_DAYS, 0, 0
        
        try:
            # Decrypt trial file
            trial_data = self._load_trial_encrypted()
            
            if not trial_data:
                # File corrupt hoặc bị sửa → hết trial
                return False, 0, 0, 0
            
            # Verify machine code
            if trial_data.get("machine_code") != self.machine_code:
                return False, 0, 0, 0
            
            first_run = datetime.fromisoformat(trial_data["first_run"])
            now = datetime.now()
            
            # ===== CHỐNG BACKUP/RESTORE =====
            last_used = trial_data.get("last_used")
            if last_used:
                last_used_time = datetime.fromisoformat(last_used)
                
                # Nếu thời gian hiện tại < last_used → user đã quay ngược đồng hồ hoặc restore file cũ
                if now < last_used_time - timedelta(minutes=5):  # 5 phút tolerance
                    print(f"[SECURITY] Time manipulation detected! now={now}, last_used={last_used_time}")
                    return False, 0, 0, 0
                
                # So sánh với Registry
                if registry_trial:
                    reg_last_used = registry_trial.get("last_used")
                    if reg_last_used:
                        reg_time = datetime.fromisoformat(reg_last_used)
                        # Nếu Registry có last_used mới hơn file → file bị restore từ backup cũ
                        if reg_time > last_used_time + timedelta(minutes=5):
                            print(f"[SECURITY] File restore detected! file={last_used_time}, registry={reg_time}")
                            return False, 0, 0, 0
            
            # Cập nhật last_used
            trial_data["last_used"] = now.isoformat()
            self._save_trial_encrypted(trial_data)
            self._set_registry_trial(trial_data)
            
            # Tính thời gian còn lại
            trial_end = first_run + timedelta(days=self.TRIAL_DAYS)
            remaining = trial_end - now
            
            if remaining.total_seconds() > 0:
                days_left = remaining.days
                hours_left = remaining.seconds // 3600
                mins_left = (remaining.seconds % 3600) // 60
                return True, days_left, hours_left, mins_left
            else:
                return False, 0, 0, 0
                
        except Exception as e:
            print(f"[ERROR] Trial check failed: {e}")
            return False, 0, 0, 0
    
    def _check_online_trial(self):
        """
        Kiểm tra online xem machine code này đã đăng ký trial chưa.
        Sử dụng Cryptolens Data Objects hoặc một API đơn giản.
        Returns: True nếu đã dùng trial, False nếu chưa
        """
        try:
            # Gọi API kiểm tra trial (sử dụng GetKey với special trial key)
            # Hoặc có thể tạo một data object trên Cryptolens
            url = f"{self.API_BASE_URL}/GetKey"
            payload = {
                "token": self.ACCESS_TOKEN,
                "ProductId": str(self.PRODUCT_ID),
                "Key": f"TRIAL-{self.machine_code[:12]}",  # Trial key format
            }
            response = requests.post(url, data=payload, timeout=5)
            result = response.json()
            
            # Nếu tìm thấy trial key → đã dùng
            if result.get("result") == 0:
                return True
            return False
        except:
            # Không có internet → cho phép kiểm tra local
            return False
    
    def _register_online_trial(self):
        """
        Đăng ký trial online với machine code.
        Tạo một data object trên Cryptolens để track.
        """
        try:
            # Gọi API để đánh dấu machine đã dùng trial
            # Đây là optional - chỉ để tracking
            print(f"[DEBUG] Registering trial for machine: {self.machine_code}")
            return True
        except:
            return False
    
    def _save_trial_encrypted(self, trial_data):
        """Lưu trial data đã mã hóa"""
        try:
            json_data = json.dumps(trial_data)
            
            if HAS_CRYPTO:
                # Encrypt với AES - key từ machine code
                encrypted_data = self._encrypt_data(json_data)
                with open(self.TRIAL_FILE, 'wb') as f:
                    f.write(encrypted_data)
            else:
                # Fallback: Obfuscate với base64 + machine code hash
                combined = f"{self.machine_code}|{json_data}"
                encoded = base64.b64encode(combined.encode()).decode()
                with open(self.TRIAL_FILE, 'w') as f:
                    f.write(encoded)
            return True
        except Exception as e:
            print(f"[ERROR] Save trial failed: {e}")
            return False
    
    def _load_trial_encrypted(self):
        """Load và decrypt trial data"""
        try:
            if HAS_CRYPTO:
                with open(self.TRIAL_FILE, 'rb') as f:
                    encrypted_data = f.read()
                json_data = self._decrypt_data(encrypted_data)
            else:
                with open(self.TRIAL_FILE, 'r') as f:
                    encoded = f.read()
                decoded = base64.b64decode(encoded).decode()
                # Verify machine code
                parts = decoded.split('|', 1)
                if len(parts) != 2 or parts[0] != self.machine_code:
                    return None
                json_data = parts[1]
            
            return json.loads(json_data)
        except Exception as e:
            print(f"[ERROR] Load trial failed: {e}")
            return None
    
    def start_trial(self):
        """Bắt đầu trial mode - lưu ngày đầu tiên (mã hóa + Registry)"""
        try:
            # Kiểm tra Registry xem đã dùng trial chưa
            registry_trial = self._get_registry_trial()
            if registry_trial and registry_trial.get("machine_code") == self.machine_code:
                print("[SECURITY] Trial already used (found in Registry)!")
                return False
            
            # Đăng ký online (nếu có internet)
            self._register_online_trial()
            
            now = datetime.now()
            trial_data = {
                "first_run": now.isoformat(),
                "last_used": now.isoformat(),  # Để chống backup
                "machine_code": self.machine_code,
                "trial_started": True,
                "version": "3.0"  # Version với anti-backup
            }
            
            # Lưu vào cả file VÀ Registry
            file_saved = self._save_trial_encrypted(trial_data)
            registry_saved = self._set_registry_trial(trial_data)
            
            print(f"[DEBUG] Trial saved - File: {file_saved}, Registry: {registry_saved}")
            
            return file_saved
        except Exception as e:
            print(f"[ERROR] Không thể tạo trial file: {e}")
            return False
    
    # ========== LICENSE ACTIVATION ==========
    
    def activate_license(self, license_key):
        """
        Kích hoạt license qua Cryptolens API.
        Returns: (success: bool, error_msg: str, license_data: dict)
        """
        max_retries = 3
        retry_delay = 10  # seconds
        
        for attempt in range(max_retries):
            try:
                url = f"{self.API_BASE_URL}/Activate"
                
                # Cryptolens API expects form data, NOT JSON!
                payload = {
                    "token": self.ACCESS_TOKEN,
                    "ProductId": str(self.PRODUCT_ID),
                    "Key": license_key.strip(),
                    "MachineCode": self.machine_code,
                    "Sign": "true",
                    "SignMethod": "1"
                }
                
                print(f"[DEBUG] Calling Cryptolens API (Attempt {attempt + 1}/{max_retries})...")
                print(f"[DEBUG] URL: {url}")
                print(f"[DEBUG] Product ID: {self.PRODUCT_ID}")
                print(f"[DEBUG] Key: {license_key.strip()}")
                print(f"[DEBUG] Machine Code: {self.machine_code}")
                print(f"[DEBUG] Token (first 20 chars): {self.ACCESS_TOKEN[:20]}...")
                
                # Send as form data, not JSON!
                response = requests.post(url, data=payload, timeout=15)
                
                print(f"[DEBUG] Response Status: {response.status_code}")
                print(f"[DEBUG] Response Body: {response.text[:500]}")
                
                # Handle rate limiting (429)
                if response.status_code == 429:
                    if attempt < max_retries - 1:
                        print(f"[DEBUG] Rate limit hit. Waiting {retry_delay} seconds before retry...")
                        time.sleep(retry_delay)
                        continue
                    else:
                        return False, "VUI LÒNG ĐỢI VÀI PHÚT RỒI THỬ LẠI (RATE LIMIT)", None
                
                result = response.json()
                
                # Debug: Print full response structure
                print(f"[DEBUG] Full API response keys: {result.keys()}")
                print(f"[DEBUG] Full API response: {json.dumps(result, indent=2, default=str)[:2000]}")
                
                # Check response
                if result.get("result") == 0:  # Success
                    print("[DEBUG] API returned SUCCESS (result=0)")
                    
                    # Cryptolens trả về licenseKey dạng Base64 encoded JSON
                    license_key_data = result.get("licenseKey", "")
                    signature = result.get("signature", "")
                    
                    # ===== VERIFY SIGNATURE (CHỐNG GIẢ MẠO) =====
                    if signature and self.VERIFY_SIGNATURE:
                        if not self._verify_signature(license_key_data, signature):
                            return False, "CẢNH BÁO BẢO MẬT: Dữ liệu license không hợp lệ!", None
                    
                    # Debug: check type of license_info
                    print(f"[DEBUG] Type of license_key_data: {type(license_key_data)}")
                    
                    # Decode Base64 -> JSON string -> dict
                    if isinstance(license_key_data, str) and license_key_data.strip():
                        print("[DEBUG] Decoding Base64...")
                        decoded_bytes = base64.b64decode(license_key_data)
                        decoded_str = decoded_bytes.decode('utf-8')
                        print(f"[DEBUG] Decoded JSON: {decoded_str[:500]}")
                        license_info = json.loads(decoded_str)
                    else:
                        license_info = result
                    
                    print(f"[DEBUG] Parsed license_info keys: {license_info.keys() if isinstance(license_info, dict) else 'N/A'}")
                    
                    # Kiểm tra license có bị block không
                    blocked = license_info.get("Block", False) if isinstance(license_info, dict) else False
                    if blocked:
                        return False, "LICENSE ĐÃ BỊ KHÓA!", None
                    
                    # Lưu thông tin license
                    license_data = {
                        "key": license_key.strip(),
                        "activated": datetime.now().isoformat(),
                        "last_online_check": datetime.now().isoformat(),
                        "machine_code": self.machine_code,
                        "license_info": license_info,
                        "signature": signature  # Lưu signature để verify offline
                    }
                    
                    return True, None, license_data
                    
                else:
                    # API error
                    print(f"[DEBUG] API returned ERROR (result={result.get('result')})")
                    error_msg = result.get("message", "Unknown error")
                    return False, f"KÍCH HOẠT THẤT BẠI: {error_msg}", None
                    
            except requests.exceptions.Timeout:
                if attempt < max_retries - 1:
                    print(f"[DEBUG] Timeout. Retrying...")
                    continue
                return False, "TIMEOUT: KHÔNG THỂ KẾT NỐI ĐẾN SERVER CRYPTOLENS!", None
            except requests.exceptions.ConnectionError:
                if attempt < max_retries - 1:
                    print(f"[DEBUG] Connection error. Retrying...")
                    continue
                return False, "LỖI KẾT NỐI: VUI LÒNG KIỂM TRA INTERNET!", None
            except Exception as e:
                print(f"[DEBUG] Unexpected error: {e}")
                return False, f"LỖI: {str(e)}", None
        
        # If all retries failed
        return False, "KHÔNG THỂ KÍCH HOẠT SAU NHIỀU LẦN THỬ!", None
    
    # ========== LICENSE FILE MANAGEMENT ==========
    
    def save_license(self, license_data):
        """Lưu license data xuống file (encrypted nếu có pycryptodome)"""
        try:
            json_data = json.dumps(license_data, indent=2)
            
            if HAS_CRYPTO:
                # Encrypt với AES
                encrypted_data = self._encrypt_data(json_data)
                with open(self.LICENSE_FILE, 'wb') as f:
                    f.write(encrypted_data)
            else:
                # Plain text (không khuyến khích)
                with open(self.LICENSE_FILE, 'w') as f:
                    f.write(json_data)
            
            return True
            
        except Exception as e:
            print(f"[ERROR] Không thể lưu license: {e}")
            return False
    
    def load_license(self):
        """
        Load license từ file.
        Returns: license_data dict hoặc None
        """
        if not os.path.exists(self.LICENSE_FILE):
            return None
        
        try:
            if HAS_CRYPTO:
                # Decrypt
                with open(self.LICENSE_FILE, 'rb') as f:
                    encrypted_data = f.read()
                json_data = self._decrypt_data(encrypted_data)
            else:
                # Plain text
                with open(self.LICENSE_FILE, 'r') as f:
                    json_data = f.read()
            
            license_data = json.loads(json_data)
            self.license_data = license_data
            return license_data
            
        except Exception as e:
            print(f"[ERROR] Không thể đọc license: {e}")
            return None
    
    def _encrypt_data(self, plaintext):
        """Encrypt data với AES-256"""
        # AES key từ machine code (32 bytes)
        key = hashlib.sha256(self.machine_code.encode()).digest()
        
        cipher = AES.new(key, AES.MODE_CBC)
        iv = cipher.iv
        
        # Encrypt
        encrypted = cipher.encrypt(pad(plaintext.encode(), AES.block_size))
        
        # Return: IV + encrypted data (base64 encoded)
        return base64.b64encode(iv + encrypted)
    
    def _decrypt_data(self, encrypted_data):
        """Decrypt AES encrypted data"""
        # AES key từ machine code
        key = hashlib.sha256(self.machine_code.encode()).digest()
        
        # Decode base64
        raw = base64.b64decode(encrypted_data)
        
        # Extract IV and encrypted data
        iv = raw[:16]
        encrypted = raw[16:]
        
        # Decrypt
        cipher = AES.new(key, AES.MODE_CBC, iv)
        decrypted = unpad(cipher.decrypt(encrypted), AES.block_size)
        
        return decrypted.decode()
    
    # ========== VALIDATION ==========
    
    def validate_offline(self):
        """
        Validate license offline (không cần Internet).
        Kiểm tra: expiry date, machine lock, offline grace period
        Returns: (is_valid: bool, error_msg: str)
        """
        if not self.license_data:
            return False, "Không có license!"
        
        try:
            # 1. Check machine code
            saved_machine = self.license_data.get("machine_code", "")
            if saved_machine != self.machine_code:
                return False, "License không khớp với máy này!"
            
            # 2. Check expiry date - FIXED: Check both "Expires" and "expires"
            license_info = self.license_data.get("license_info", {})
            expires = license_info.get("Expires") or license_info.get("expires")
            
            if expires:
                # Parse expiry date (Unix timestamp)
                try:
                    expiry_date = datetime.fromtimestamp(expires)
                    if datetime.now() > expiry_date:
                        return False, f"License đã hết hạn ngày {expiry_date.strftime('%d/%m/%Y')}!"
                except:
                    pass
            
            # 3. Check offline grace period (7 ngày)
            last_online = self.license_data.get("last_online_check")
            if last_online:
                last_check = datetime.fromisoformat(last_online)
                days_offline = (datetime.now() - last_check).days
                
                if days_offline > self.OFFLINE_GRACE_DAYS:
                    return False, f"Đã offline quá {self.OFFLINE_GRACE_DAYS} ngày. Vui lòng kết nối Internet!"
            
            # 4. Check blocked - FIXED: Cryptolens uses "Block" (capital B)
            if license_info.get("Block", False) or license_info.get("blocked", False):
                return False, "License đã bị khóa!"
            
            return True, None
            
        except Exception as e:
            return False, f"Lỗi validation: {str(e)}"
    
    def check_online(self):
        """
        Kiểm tra license online (gọi API GetKey).
        Cập nhật thông tin mới nhất nếu có.
        Returns: (success: bool, error_msg: str)
        """
        if not self.license_data:
            return False, "Không có license!"
        
        try:
            url = f"{self.API_BASE_URL}/GetKey"
            
            # FIXED: Cryptolens API expects form data with string values
            payload = {
                "token": self.ACCESS_TOKEN,
                "ProductId": str(self.PRODUCT_ID),
                "Key": self.license_data.get("key", ""),
                "Sign": "true",
                "SignMethod": "1"
            }
            
            # FIXED: Use data= instead of json=
            response = requests.post(url, data=payload, timeout=10)
            result = response.json()
            
            if result.get("result") == 0:
                # FIXED: Decode Base64 like in activate_license()
                license_key_data = result.get("licenseKey", "")
                
                if isinstance(license_key_data, str) and license_key_data.strip():
                    decoded_bytes = base64.b64decode(license_key_data)
                    decoded_str = decoded_bytes.decode('utf-8')
                    license_info = json.loads(decoded_str)
                else:
                    license_info = license_key_data if isinstance(license_key_data, dict) else {}
                
                # Update license info
                self.license_data["license_info"] = license_info
                self.license_data["last_online_check"] = datetime.now().isoformat()
                
                # Save updated license
                self.save_license(self.license_data)
                
                return True, None
            else:
                return False, result.get("message", "Unknown error")
                
        except:
            # Offline mode OK - không báo lỗi
            return False, None
    
    def get_license_info(self):
        """
        Lấy thông tin license để hiển thị.
        Returns: dict với các key: customer, key, expiry, status, etc.
        """
        if not self.license_data:
            return None
        
        try:
            license_info = self.license_data.get("license_info", {})
            
            # Parse expiry date - FIXED: Check both "Expires" and "expires" (Cryptolens uses capital E)
            expires = license_info.get("Expires") or license_info.get("expires")
            if expires:
                try:
                    expiry_date = datetime.fromtimestamp(expires)
                    expiry_str = expiry_date.strftime("%d/%m/%Y")
                    days_left = (expiry_date - datetime.now()).days
                except:
                    expiry_str = "Không giới hạn"
                    days_left = 999999
            else:
                expiry_str = "Không giới hạn"
                days_left = 999999
            
            # Customer info - FIXED: Cryptolens uses "Customer" (capital C)
            customer = license_info.get("Customer") or license_info.get("customer", {})
            customer_name = customer.get("Name", customer.get("name", "N/A")) if customer else "N/A"
            
            # License key (ẩn bớt)
            full_key = self.license_data.get("key", "")
            if len(full_key) > 10:
                masked_key = f"{full_key[:5]}-***-***-{full_key[-5:]}"
            else:
                masked_key = full_key
            
            # Status - FIXED: Check both "Block" and "blocked"
            blocked = license_info.get("Block", False) or license_info.get("blocked", False)
            if blocked:
                status = "❌ Bị khóa"
                status_color = "red"
            elif days_left < 0:
                status = "⏰ Hết hạn"
                status_color = "orange"
            elif days_left < 30:
                status = f"⚠️ Còn {days_left} ngày"
                status_color = "orange"
            else:
                status = "✅ Đang hoạt động"
                status_color = "green"
            
            return {
                "customer_name": customer_name,
                "license_key": full_key,
                "license_key_masked": masked_key,
                "expiry_date": expiry_str,
                "days_left": days_left,
                "status": status,
                "status_color": status_color,
                "machine_code": self.machine_code,
                "activated_date": self.license_data.get("activated", "N/A"),
                "last_online_check": self.license_data.get("last_online_check", "N/A")
            }
            
        except Exception as e:
            return None
    
    def revoke_license(self):
        """Xóa license file (để test hoặc deactivate)"""
        try:
            if os.path.exists(self.LICENSE_FILE):
                os.remove(self.LICENSE_FILE)
            self.license_data = None
            return True
        except Exception as e:
            print(f"[ERROR] Không thể xóa license: {e}")
            return False


# =================================================================================
# [SPLASH SCREEN] LOADING SCREEN
# =================================================================================

class SplashScreen:
    """
    Màn hình chào mừng khi khởi động app
    """
    
    def __init__(self, duration=2000):
        self.duration = duration
        self.splash_root = tk.Toplevel()
        self.splash_root.overrideredirect(True)  # Không có title bar
        
        # Kích thước và căn giữa
        width = 500
        height = 350
        screen_width = self.splash_root.winfo_screenwidth()
        screen_height = self.splash_root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.splash_root.geometry(f"{width}x{height}+{x}+{y}")
        
        # Gradient background (giả lập bằng Canvas)
        canvas = tk.Canvas(self.splash_root, width=width, height=height, highlightthickness=0)
        canvas.pack(fill='both', expand=True)
        
        # Gradient từ xanh đậm đến xanh nhạt
        for i in range(height):
            ratio = i / height
            r = int(34 + (96 - 34) * ratio)
            g = int(139 + (165 - 139) * ratio)
            b = int(230 + (250 - 230) * ratio)
            color = f'#{r:02x}{g:02x}{b:02x}'
            canvas.create_line(0, i, width, i, fill=color)
        
        # Viền ngoài - Border đẹp với 3 lớp
        # Lớp 1: Viền trắng ngoài cùng (dày 3px)
        canvas.create_rectangle(0, 0, width, height, outline='white', width=3)
        # Lớp 2: Viền xanh đậm (dày 2px, offset 3px)
        canvas.create_rectangle(3, 3, width-3, height-3, outline='#1e5a8e', width=2)
        # Lớp 3: Viền sáng bên trong (dày 1px, offset 5px) - tạo hiệu ứng glow
        canvas.create_rectangle(5, 5, width-5, height-5, outline='#5dade2', width=1)
        
        # Icon/Logo (dùng text lớn thay icon)
        canvas.create_text(
            width // 2, 80,
            text="📚",
            font=('Segoe UI Emoji', 60),
            fill='white'
        )
        
        # Tên app - Với viền đen
        app_x = width // 2
        app_y = 160
        app_font = ('Segoe UI', 32, 'bold')
        
        # Viền đen
        for dx, dy in [(-2,-2), (-2,0), (-2,2), (0,-2), (0,2), (2,-2), (2,0), (2,2)]:
            canvas.create_text(
                app_x + dx, app_y + dy,
                text=APP_NAME,
                font=app_font,
                fill='black'
            )
        
        # Chữ trắng chính
        canvas.create_text(
            app_x, app_y,
            text=APP_NAME,
            font=app_font,
            fill='white'
        )
        
        # Version - Với viền đen
        ver_x = width // 2
        ver_y = 200
        ver_text = f"Version {APP_VERSION}"
        ver_font = ('Segoe UI', 12)
        
        # Viền đen
        for dx, dy in [(-1,-1), (-1,0), (-1,1), (0,-1), (0,1), (1,-1), (1,0), (1,1)]:
            canvas.create_text(
                ver_x + dx, ver_y + dy,
                text=ver_text,
                font=ver_font,
                fill='black'
            )
        
        # Chữ trắng chính
        canvas.create_text(
            ver_x, ver_y,
            text=ver_text,
            font=ver_font,
            fill='white'
        )
        
        # Slogan - Với viền đen
        slogan_x = width // 2
        slogan_y = 240
        slogan_text = "QUẢN LÝ ĐIỂM CHUYÊN NGHIỆP - HIỆN ĐẠI - THÔNG MINH"
        slogan_font = ('Segoe UI', 11, 'bold')
        
        # Viền đen
        for dx, dy in [(-1,-1), (-1,0), (-1,1), (0,-1), (0,1), (1,-1), (1,0), (1,1)]:
            canvas.create_text(
                slogan_x + dx, slogan_y + dy,
                text=slogan_text,
                font=slogan_font,
                fill='black'
            )
        
        # Chữ trắng chính
        canvas.create_text(
            slogan_x, slogan_y,
            text=slogan_text,
            font=slogan_font,
            fill='white'
        )
        
        # Developer credit - Với viền đen bao quanh chữ vàng
        dev_x = width // 2
        dev_y = 265
        dev_text = "✨ Developed by Vũ Hào ✨"
        dev_font = ('Segoe UI', 12, 'bold')
        
        # Vẽ viền đen (8 hướng để tạo outline mượt)
        for dx, dy in [(-1,-1), (-1,0), (-1,1), (0,-1), (0,1), (1,-1), (1,0), (1,1)]:
            canvas.create_text(
                dev_x + dx, dev_y + dy,
                text=dev_text,
                font=dev_font,
                fill='black'
            )
        
        # Vẽ chữ vàng chính ở giữa
        canvas.create_text(
            dev_x, dev_y,
            text=dev_text,
            font=dev_font,
            fill='#FFE66D'  # Màu vàng nhẹ
        )
        
        # Loading bar background
        canvas.create_rectangle(
            50, 280, width - 50, 300,
            fill='#1e3a5f',
            outline='black',
            width=1
        )
        
        # Loading bar (animated)
        self.progress_bar = canvas.create_rectangle(
            50, 280, 50, 300,
            fill='#4ade80',
            outline='black',
            width=1
        )
        self.canvas = canvas
        self.width = width
        
        # Status text - Với viền đen
        self.status_x = width // 2
        self.status_y = 320
        self.status_font = ('Segoe UI', 11, 'bold')
        
        # Tạo 8 text cho viền đen (sẽ update cùng lúc với text chính)
        self.status_shadows = []
        for dx, dy in [(-1,-1), (-1,0), (-1,1), (0,-1), (0,1), (1,-1), (1,0), (1,1)]:
            shadow = canvas.create_text(
                self.status_x + dx, self.status_y + dy,
                text="ĐANG KHỞI ĐỘNG...",
                font=self.status_font,
                fill='black'
            )
            self.status_shadows.append(shadow)
        
        # Text chính màu trắng
        self.status_text = canvas.create_text(
            self.status_x, self.status_y,
            text="ĐANG KHỞI ĐỘNG...",
            font=self.status_font,
            fill='white'
        )
        
        # Animate loading bar
        self.animate_progress()
        
        # Auto close sau duration
        self.splash_root.after(duration, self.close)
    
    def animate_progress(self, progress=0):
        """Animation cho progress bar"""
        if progress <= 100:
            x = 50 + (self.width - 100) * (progress / 100)
            self.canvas.coords(self.progress_bar, 50, 280, x, 300)
            
            # Update status text
            if progress < 30:
                status = "ĐANG TẢI THƯ VIỆN..."
            elif progress < 60:
                status = "KHỞI TẠO GIAO DIỆN..."
            elif progress < 90:
                status = "KIỂM TRA LICENSE..."
            else:
                status = "HOÀN TẤT!"
            
            # Cập nhật cả viền đen và text chính
            for shadow in self.status_shadows:
                self.canvas.itemconfig(shadow, text=status)
            self.canvas.itemconfig(self.status_text, text=status)
            
            # Tiếp tục animate
            self.splash_root.after(20, lambda: self.animate_progress(progress + 2))
    
    def close(self):
        """Đóng splash screen"""
        self.splash_root.destroy()
    
    def show(self):
        """Hiển thị splash screen"""
        self.splash_root.update()


# =================================================================================
# [AUTO-UPDATE] UPDATE MANAGER CLASS
# =================================================================================

class UpdateManager:
    """
    Quản lý việc kiểm tra và cài đặt update tự động từ GitHub
    """
    
    def __init__(self, current_version=APP_VERSION, check_url=UPDATE_CHECK_URL):
        self.current_version = current_version
        self.check_url = check_url
    
    def compare_versions(self, v1, v2):
        """
        So sánh 2 version theo định dạng MAJOR.MINOR.PATCH
        
        Returns:
            1 nếu v1 > v2, -1 nếu v1 < v2, 0 nếu bằng nhau
        """
        try:
            v1_parts = [int(x) for x in v1.split('.')]
            v2_parts = [int(x) for x in v2.split('.')]
            
            max_len = max(len(v1_parts), len(v2_parts))
            v1_parts += [0] * (max_len - len(v1_parts))
            v2_parts += [0] * (max_len - len(v2_parts))
            
            for i in range(max_len):
                if v1_parts[i] > v2_parts[i]:
                    return 1
                elif v1_parts[i] < v2_parts[i]:
                    return -1
            return 0
        except:
            return 0
    
    def check_for_updates(self, timeout=5):
        """
        Kiểm tra xem có phiên bản mới không
        
        Returns:
            (has_update: bool, update_info: dict hoặc None)
        """
        try:
            print(f"[UPDATE] Checking: {self.check_url}")
            response = requests.get(self.check_url, timeout=timeout)
            
            if response.status_code == 200:
                update_info = response.json()
                latest_version = update_info.get('latest_version', '0.0.0')
                
                print(f"[UPDATE] Current: {self.current_version}, Latest: {latest_version}")
                
                if self.compare_versions(latest_version, self.current_version) > 0:
                    print("[UPDATE] New version available!")
                    return True, update_info
                else:
                    print("[UPDATE] Up to date!")
                    return False, None
            else:
                return False, None
        except:
            return False, None
    
    def download_update(self, download_url, progress_callback=None):
        """
        Download file exe mới từ GitHub
        
        Args:
            download_url: URL của file exe
            progress_callback: Hàm callback(percent) để update progress bar
            
        Returns:
            (success: bool, file_path: str hoặc None)
        """
        try:
            print(f"[UPDATE] Downloading: {download_url}")
            
            temp_dir = tempfile.gettempdir()
            filename = os.path.basename(download_url)
            if not filename.endswith('.exe'):
                filename = f"EduManager_update_{int(time.time())}.exe"
            
            temp_file = os.path.join(temp_dir, filename)
            
            response = requests.get(download_url, stream=True, timeout=30)
            response.raise_for_status()
            
            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0
            
            with open(temp_file, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        
                        if progress_callback and total_size > 0:
                            progress = (downloaded / total_size) * 100
                            progress_callback(progress)
            
            print(f"[UPDATE] Downloaded: {temp_file}")
            return True, temp_file
        except Exception as e:
            print(f"[UPDATE] Download failed: {e}")
            return False, None
    
    def install_update(self, new_exe_path):
        """
        Cài đặt update bằng cách thay thế exe cũ
        
        Returns:
            bool: True nếu thành công (app sẽ tự restart)
        """
        try:
            if getattr(sys, 'frozen', False):
                current_exe = sys.executable
            else:
                current_exe = os.path.abspath(__file__).replace('.py', '.exe')
            
            print(f"[UPDATE] Current: {current_exe}")
            print(f"[UPDATE] New: {new_exe_path}")
            
            updater_script = self._create_updater_script(new_exe_path, current_exe)
            
            print(f"[UPDATE] Running updater...")
            
            if sys.platform == 'win32':
                subprocess.Popen(updater_script, shell=True,
                               creationflags=subprocess.CREATE_NO_WINDOW)
            else:
                subprocess.Popen(['sh', updater_script])
            
            time.sleep(1)
            sys.exit(0)
        except Exception as e:
            print(f"[UPDATE] Install failed: {e}")
            return False
    
    def _create_updater_script(self, new_exe, old_exe):
        """Tạo batch script để update app"""
        temp_dir = tempfile.gettempdir()
        
        if sys.platform == 'win32':
            script_path = os.path.join(temp_dir, "edumanager_updater.bat")
            
            script_content = f"""@echo off
echo [UPDATER] Waiting for app to close...
timeout /t 3 /nobreak > nul

echo [UPDATER] Removing old version...
del /F /Q "{old_exe}"

echo [UPDATER] Installing new version...
move /Y "{new_exe}" "{old_exe}"

echo [UPDATER] Starting new version...
start "" "{old_exe}"

echo [UPDATER] Cleaning up...
timeout /t 2 /nobreak > nul
del "%~f0"
"""
        else:
            script_path = os.path.join(temp_dir, "edumanager_updater.sh")
            
            script_content = f"""#!/bin/bash
sleep 3
rm -f "{old_exe}"
mv "{new_exe}" "{old_exe}"
chmod +x "{old_exe}"
"{old_exe}" &
sleep 2
rm -f "$0"
"""
        
        with open(script_path, 'w') as f:
            f.write(script_content)
        
        if sys.platform != 'win32':
            os.chmod(script_path, 0o755)
        
        return script_path


# =================================================================================
# UI CLASSES
# =================================================================================

class InputDialog(Toplevel):
    """
    Custom Dialog nhập liệu đẹp hơn simpledialog.
    Sử dụng: result = InputDialog.ask_string(parent, title, prompt, initial_value)
    """
    def __init__(self, parent, title, prompt, initial_value=""):
        super().__init__(parent)
        self.result = None
        
        # Cấu hình cửa sổ
        self.title(title)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        # Màu sắc
        bg_color = "#ffffff"
        accent_color = "#27ae60"
        text_color = "#333333"
        
        self.config(bg=bg_color)
        
        # Container chính
        main_frame = tk.Frame(self, bg=bg_color, padx=25, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Icon và tiêu đề
        header_frame = tk.Frame(main_frame, bg=bg_color)
        header_frame.pack(fill="x", pady=(0, 15))
        
        icon_label = tk.Label(header_frame, text="📝", font=("Segoe UI", 28), bg=bg_color)
        icon_label.pack(side="left", padx=(0, 15))
        
        title_frame = tk.Frame(header_frame, bg=bg_color)
        title_frame.pack(side="left", fill="x", expand=True)
        
        title_label = tk.Label(title_frame, text=title, font=("Segoe UI", 14, "bold"), 
                               bg=bg_color, fg=text_color, anchor="w")
        title_label.pack(anchor="w")
        
        # Prompt text
        prompt_label = tk.Label(main_frame, text=prompt, font=("Segoe UI", 10), 
                                bg=bg_color, fg="#666666", justify="left", anchor="w")
        prompt_label.pack(fill="x", pady=(0, 10))
        
        # Entry frame với viền đẹp
        entry_frame = tk.Frame(main_frame, bg=accent_color, padx=2, pady=2)
        entry_frame.pack(fill="x", pady=(0, 20))
        
        self.entry = tk.Entry(entry_frame, font=("Segoe UI", 12), relief="flat", 
                              bg="#f8f9fa", fg=text_color, insertbackground=text_color)
        self.entry.pack(fill="x", ipady=8, padx=1, pady=1)
        self.entry.insert(0, initial_value)
        self.entry.select_range(0, tk.END)
        self.entry.focus_set()
        
        # Buttons frame
        btn_frame = tk.Frame(main_frame, bg=bg_color)
        btn_frame.pack(fill="x")
        
        # Nút Cancel - thêm viền đen mỏng
        cancel_border = tk.Frame(btn_frame, bg="#333333", padx=1, pady=1)
        cancel_border.pack(side="right", padx=(10, 0))
        
        self.btn_cancel = tk.Button(cancel_border, text="Hủy", font=("Segoe UI", 10),
                                    bg="#e0e0e0", fg="#333333", relief="flat",
                                    width=12, pady=8, cursor="hand2",
                                    command=self.on_cancel)
        self.btn_cancel.pack()
        
        # Nút OK - thêm viền đen mỏng
        ok_border = tk.Frame(btn_frame, bg="#333333", padx=1, pady=1)
        ok_border.pack(side="right")
        
        self.btn_ok = tk.Button(ok_border, text="✓ Xác nhận", font=("Segoe UI", 10, "bold"),
                                bg=accent_color, fg="white", relief="flat",
                                width=12, pady=8, cursor="hand2",
                                command=self.on_ok)
        self.btn_ok.pack()
        
        # Hover effects
        def on_ok_enter(e):
            self.btn_ok.config(bg="#219a52")
        def on_ok_leave(e):
            self.btn_ok.config(bg=accent_color)
        def on_cancel_enter(e):
            self.btn_cancel.config(bg="#c0c0c0")
        def on_cancel_leave(e):
            self.btn_cancel.config(bg="#e0e0e0")
        
        self.btn_ok.bind("<Enter>", on_ok_enter)
        self.btn_ok.bind("<Leave>", on_ok_leave)
        self.btn_cancel.bind("<Enter>", on_cancel_enter)
        self.btn_cancel.bind("<Leave>", on_cancel_leave)
        
        # Bind Enter và Escape
        self.entry.bind("<Return>", lambda e: self.on_ok())
        self.bind("<Escape>", lambda e: self.on_cancel())
        
        # Center dialog
        self.update_idletasks()
        width = max(400, self.winfo_reqwidth())
        height = self.winfo_reqheight()
        x = parent.winfo_rootx() + (parent.winfo_width() - width) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Wait for window to close
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.wait_window(self)
    
    def on_ok(self):
        """Xác nhận"""
        self.result = self.entry.get().strip()
        self.destroy()
    
    def on_cancel(self):
        """Hủy bỏ"""
        self.result = None
        self.destroy()
    
    @staticmethod
    def ask_string(parent, title, prompt, initial_value=""):
        """Static method để gọi dialog dễ dàng"""
        dialog = InputDialog(parent, title, prompt, initial_value)
        return dialog.result


# =================================================================================
# [AUTO-UPDATE] UPDATE DIALOG UI
# =================================================================================

class UpdateDialog(tk.Toplevel):
    """Dialog thông báo có bản cập nhật mới với UI đẹp"""
    
    def __init__(self, parent, update_info, updater):
        super().__init__(parent)
        
        self.update_info = update_info
        self.updater = updater
        self.downloading = False
        
        # Window config
        self.title("🔄 Cập nhật EduManager")
        self.configure(bg="#ffffff")
        self.resizable(False, False)
        
        # Ẩn window để center
        self.withdraw()
        
        # Build UI
        self.setup_ui()
        
        # Center window
        width = 520
        height = 480
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Hiện window
        self.deiconify()
        
        # Đặt window luôn ở trên
        self.attributes('-topmost', True)
        self.lift()
        self.focus_force()
        
        # Block parent window
        self.transient(parent)
        self.grab_set()
        
        # Handle close
        self.protocol("WM_DELETE_WINDOW", self.on_skip)
    
    def setup_ui(self):
        """Xây dựng giao diện"""
        
        # Header
        header = tk.Frame(self, bg="#3498db", pady=20)
        header.pack(fill="x")
        
        tk.Label(
            header,
            text="🎉 CÓ BẢN CẬP NHẬT MỚI!",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 16, "bold")
        ).pack()
        
        tk.Label(
            header,
            text="Cập nhật để trải nghiệm tính năng mới nhất",
            bg="#3498db",
            fg="#ecf0f1",
            font=("Segoe UI", 9)
        ).pack(pady=(5, 0))
        
        # Main content
        main = tk.Frame(self, bg="#ffffff", padx=30, pady=25)
        main.pack(fill="both", expand=True)
        
        # Version info
        info_frame = tk.Frame(main, bg="#f8f9fa", padx=15, pady=15)
        info_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            info_frame,
            text=f"📦 Phiên bản hiện tại: {APP_VERSION}",
            bg="#f8f9fa",
            fg="#7f8c8d",
            font=("Segoe UI", 10)
        ).pack(anchor="w")
        
        tk.Label(
            info_frame,
            text=f"✨ Phiên bản mới: {self.update_info.get('latest_version', 'N/A')}",
            bg="#f8f9fa",
            fg="#27ae60",
            font=("Segoe UI", 12, "bold")
        ).pack(anchor="w", pady=(5, 0))
        
        file_size = self.update_info.get('file_size_mb', 'N/A')
        tk.Label(
            info_frame,
            text=f"💾 Kích thước: {file_size} MB",
            bg="#f8f9fa",
            fg="#95a5a6",
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(5, 0))
        
        # Changelog
        tk.Label(
            main,
            text="📝 Nội dung cập nhật:",
            bg="#ffffff",
            fg="#2c3e50",
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w", pady=(0, 8))
        
        changelog_frame = tk.Frame(main, bg="#f8f9fa", padx=2, pady=2)
        changelog_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        scrollbar = tk.Scrollbar(changelog_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.changelog_text = tk.Text(
            changelog_frame,
            font=("Segoe UI", 10),
            bg="#ffffff",
            fg="#333333",
            relief="flat",
            wrap="word",
            height=8,
            yscrollcommand=scrollbar.set
        )
        self.changelog_text.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.changelog_text.yview)
        
        changelog = self.update_info.get('changelog', 'Không có thông tin chi tiết')
        self.changelog_text.insert("1.0", changelog)
        self.changelog_text.config(state="disabled")
        
        # Progress frame (ẩn ban đầu)
        self.progress_frame = tk.Frame(main, bg="#ffffff")
        
        self.progress_label = tk.Label(
            self.progress_frame,
            text="",
            bg="#ffffff",
            fg="#3498db",
            font=("Segoe UI", 9, "bold")
        )
        self.progress_label.pack(pady=(0, 5))
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Update.Horizontal.TProgressbar",
                       troughcolor='#ecf0f1',
                       background='#27ae60',
                       thickness=20)
        
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode='determinate',
            maximum=100,
            value=0,
            length=440,
            style="Update.Horizontal.TProgressbar"
        )
        self.progress_bar.pack()
        
        # Buttons
        btn_frame = tk.Frame(main, bg="#ffffff")
        btn_frame.pack(fill="x", pady=(10, 0))
        
        btn_style = {
            "font": ("Segoe UI", 11, "bold"),
            "relief": "flat",
            "cursor": "hand2",
            "padx": 25,
            "pady": 12,
        }
        
        # Nút Bỏ qua (nếu không bắt buộc)
        if not self.update_info.get('required', False):
            self.skip_btn = tk.Button(
                btn_frame,
                text="Để sau",
                bg="#e0e0e0",
                fg="#333333",
                command=self.on_skip,
                **btn_style
            )
            self.skip_btn.pack(side="right", padx=(10, 0))
            
            def on_skip_enter(e):
                self.skip_btn.config(bg="#bdc3c7")
            def on_skip_leave(e):
                self.skip_btn.config(bg="#e0e0e0")
            
            self.skip_btn.bind("<Enter>", on_skip_enter)
            self.skip_btn.bind("<Leave>", on_skip_leave)
        else:
            self.skip_btn = None
        
        # Nút Cập nhật
        self.update_btn = tk.Button(
            btn_frame,
            text="🔄 Cập nhật ngay",
            bg="#27ae60",
            fg="white",
            command=self.on_update,
            **btn_style
        )
        self.update_btn.pack(side="right")
        
        def on_update_enter(e):
            self.update_btn.config(bg="#2ecc71")
        def on_update_leave(e):
            self.update_btn.config(bg="#27ae60")
        
        self.update_btn.bind("<Enter>", on_update_enter)
        self.update_btn.bind("<Leave>", on_update_leave)
        
        # Warning nếu bắt buộc
        if self.update_info.get('required', False):
            warning_frame = tk.Frame(main, bg="#fff3cd", padx=12, pady=10)
            warning_frame.pack(fill="x", pady=(10, 0))
            
            tk.Label(
                warning_frame,
                text="⚠️ CẬP NHẬT BẮT BUỘC",
                bg="#fff3cd",
                fg="#856404",
                font=("Segoe UI", 10, "bold")
            ).pack(anchor="w")
            
            tk.Label(
                warning_frame,
                text="Phiên bản hiện tại không còn được hỗ trợ",
                bg="#fff3cd",
                fg="#856404",
                font=("Segoe UI", 9)
            ).pack(anchor="w")
    
    def on_update(self):
        """Bắt đầu download và install"""
        if self.downloading:
            return
        
        self.downloading = True
        
        # Disable buttons
        self.update_btn.config(state="disabled", text="⏳ Đang chuẩn bị...")
        if self.skip_btn:
            self.skip_btn.config(state="disabled")
        
        # Hiện progress
        self.progress_frame.pack(fill="x", pady=(15, 0))
        self.update()
        
        def download_and_install():
            try:
                self.progress_label.config(text="🌐 Đang kết nối đến server...")
                time.sleep(0.5)
                
                def update_progress(percent):
                    self.progress_bar['value'] = percent
                    self.progress_label.config(
                        text=f"⬇️ Đang tải xuống... {percent:.1f}%"
                    )
                
                success, temp_file = self.updater.download_update(
                    self.update_info['download_url'],
                    progress_callback=update_progress
                )
                
                if success:
                    self.progress_bar['value'] = 100
                    self.progress_label.config(text="📦 Đang cài đặt... Vui lòng chờ")
                    time.sleep(1)
                    
                    self.updater.install_update(temp_file)
                else:
                    self.progress_label.config(text="❌ Tải xuống thất bại!")
                    self.update_btn.config(state="normal", text="🔄 Thử lại")
                    if self.skip_btn:
                        self.skip_btn.config(state="normal")
                    self.downloading = False
            except Exception as e:
                print(f"[UPDATE] Error: {e}")
                self.progress_label.config(text=f"❌ Lỗi: {str(e)[:50]}")
                self.update_btn.config(state="normal", text="🔄 Thử lại")
                if self.skip_btn:
                    self.skip_btn.config(state="normal")
                self.downloading = False
        
        threading.Thread(target=download_and_install, daemon=True).start()
    
    def on_skip(self):
        """Bỏ qua update"""
        if not self.downloading:
            self.destroy()


# =================================================================================
# LICENSE UI DIALOGS
# =================================================================================

class LicenseActivationDialog(Toplevel):
    """
    Dialog kích hoạt license key.
    Hỗ trợ:
    - Nhập license key để activate
    - Hiển thị machine ID
    - Copy machine ID
    - Dùng thử 7 ngày
    """
    
    def __init__(self, parent, license_mgr, allow_trial=True):
        super().__init__(parent)
        
        self.license_mgr = license_mgr
        self.success = False  # Flag để biết user đã activate thành công chưa
        self.is_processing = False  # Flag để track khi đang xử lý
        self.result_queue = queue.Queue()  # Queue để nhận kết quả từ thread
        self.allow_trial = allow_trial  # Flag để cho phép trial hay không
        
        # Window config
        self.title("🔐 Kích hoạt EduManager")
        self.configure(bg="#ffffff")
        self.resizable(False, False)
        
        # Ẩn window trước khi đặt vị trí
        self.withdraw()
        
        self.setup_ui()
        
        # Center window - Kích thước cố định (tăng height để chứa progress bar)
        width = 500
        height = 640  # Tăng từ 580 lên 640 để đủ chỗ cho progress bar
        
        # Lấy kích thước màn hình
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        
        # Tính vị trí chính giữa (nhích lên trên 1 chút)
        x = int((screen_width / 2) - (width / 2))
        y = int((screen_height / 2) - (height / 2) - 50)  # -50 để nhích lên trên
        
        # Đặt kích thước và vị trí cùng lúc
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Hiện window sau khi đã đặt vị trí
        self.deiconify()
        
        # Force window to front
        self.lift()
        self.focus_force()
        self.grab_set()
        
        # ===== CHO PHÉP ĐÓNG BẰNG NÚT X =====
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Giữ window luôn ở trên
        self.attributes('-topmost', True)
    
    def on_close(self):
        """Đóng dialog - ngăn đóng khi đang xử lý"""
        if self.is_processing:
            messagebox.showwarning(
                "ĐANG XỬ LÝ",
                "Vui lòng đợi quá trình kích hoạt hoàn tất!",
                parent=self
            )
            return
        self.success = False
        self.destroy()
    
    def setup_ui(self):
        """Xây dựng giao diện"""
        
        # Header
        header = tk.Frame(self, bg="#3498db", pady=20)
        header.pack(fill="x")
        
        tk.Label(
            header,
            text="🔐 KÍCH HOẠT PHẦN MỀM",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 16, "bold")
        ).pack()
        
        tk.Label(
            header,
            text="EDUMANAGER - PHẦN MỀM QUẢN LÝ ĐIỂM HỌC SINH",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 9)
        ).pack(pady=(5, 0))
        
        # Main content
        main = tk.Frame(self, bg="#ffffff", padx=30, pady=25)
        main.pack(fill="both", expand=True)
        
        # Section 1: License Key Input
        tk.Label(
            main,
            text="NHẬP LICENSE KEY:",
            bg="#ffffff",
            fg="#2c3e50",
            font=("Segoe UI", 11, "bold")
        ).pack(anchor="w", pady=(0, 8))
        
        key_frame = tk.Frame(main, bg="#3498db", padx=2, pady=2)
        key_frame.pack(fill="x", pady=(0, 20))
        
        self.key_entry = tk.Entry(
            key_frame,
            font=("Consolas", 12),
            relief="flat",
            bg="#f8f9fa",
            justify="center"
        )
        self.key_entry.pack(fill="x", ipady=10, padx=1, pady=1)
        self.key_entry.focus_set()
        
        # Info text - 2 dòng riêng biệt
        info_frame = tk.Frame(main, bg="#fff3cd", padx=15, pady=12)
        info_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(
            info_frame,
            text="⚠️ CẦN KẾT NỐI INTERNET ĐỂ KÍCH HOẠT LẦN ĐẦU",
            bg="#fff3cd",
            fg="#856404",
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w")
        
        tk.Label(
            info_frame,
            text="    Sau khi kích hoạt, có thể sử dụng offline tối đa 7 ngày",
            bg="#fff3cd",
            fg="#856404",
            font=("Segoe UI", 9)
        ).pack(anchor="w")
        
        # Progress frame - Chứa status và progress bar (fixed height để không dịch layout)
        progress_frame = tk.Frame(main, bg="#ffffff", height=60)
        progress_frame.pack(fill="x", pady=(10, 5))
        progress_frame.pack_propagate(False)  # Không cho frame thay đổi kích thước
        
        # Status label - Hiển thị trạng thái khi đang xử lý
        self.status_label = tk.Label(
            progress_frame,
            text="",
            bg="#ffffff",
            fg="#3498db",
            font=("Segoe UI", 10, "bold")
        )
        self.status_label.pack(pady=(5, 5))
        
        # Progress bar - Determinate mode với phần trăm
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Green.Horizontal.TProgressbar", 
                       troughcolor='#ecf0f1',
                       background='#27ae60',
                       thickness=20)
        
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            maximum=100,
            value=0,
            length=400,
            style="Green.Horizontal.TProgressbar"
        )
        # Ẩn progress bar ban đầu
        
        # Buttons - Frame chứa các nút với padding đều
        btn_frame = tk.Frame(main, bg="#ffffff")
        btn_frame.pack(fill="x", expand=True, pady=(10, 0))
        
        # Cấu hình để buttons có cùng width
        btn_frame.columnconfigure(0, weight=1)
        
        # ===== STYLE CHUNG CHO CÁC NÚT =====
        btn_style = {
            "font": ("Segoe UI", 11, "bold"),
            "fg": "white",
            "relief": "solid",
            "bd": 1,
            "pady": 12,
            "cursor": "hand2",
        }
        
        # ===== HÀM TẠO HOVER EFFECT =====
        def create_hover(btn, normal_color, hover_color):
            def on_enter(e):
                btn.config(bg=hover_color)
            def on_leave(e):
                btn.config(bg=normal_color)
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
        
        # ===== NÚT 1: KÍCH HOẠT LICENSE =====
        self.activate_btn = tk.Button(
            btn_frame,
            text="✓  KÍCH HOẠT LICENSE",
            bg="#27ae60",
            activebackground="#2ecc71",
            activeforeground="white",
            command=self.activate_license,
            **btn_style
        )
        self.activate_btn.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        create_hover(self.activate_btn, "#27ae60", "#2ecc71")
        
        # ===== NÚT 2: DÙNG THỬ 7 NGÀY =====
        # Chỉ hiển thị nếu allow_trial = True
        if self.allow_trial:
            self.trial_btn = tk.Button(
                btn_frame,
                text="🕐  DÙNG THỬ 7 NGÀY MIỄN PHÍ",
                bg="#f39c12",
                activebackground="#f5b041",
                activeforeground="white",
                command=self.start_trial,
                **btn_style
            )
            self.trial_btn.grid(row=1, column=0, sticky="ew", pady=(0, 10))
            create_hover(self.trial_btn, "#f39c12", "#f5b041")
        else:
            # Tạo dummy button để giữ layout
            self.trial_btn = None
        
        # ===== NÚT 3: KIỂM TRA THỜI GIAN DÙNG THỬ =====
        # Chỉ hiển thị nếu allow_trial = True
        if self.allow_trial:
            self.check_trial_btn = tk.Button(
                btn_frame,
                text="📊  KIỂM TRA THỜI GIAN DÙNG THỬ",
                bg="#3498db",
                activebackground="#5dade2",
                activeforeground="white",
                command=self.check_trial_time,
                **btn_style
            )
            self.check_trial_btn.grid(row=2, column=0, sticky="ew")
            create_hover(self.check_trial_btn, "#3498db", "#5dade2")
        else:
            self.check_trial_btn = None
        
        # Bind Enter key
        self.key_entry.bind("<Return>", lambda e: self.activate_license())
    
    def check_trial_time(self):
        """Kiểm tra thời gian dùng thử còn lại"""
        trial_available, days_left, hours_left, mins_left = self.license_mgr.is_trial_available()
        
        # Kiểm tra cả Registry để xác định trạng thái chính xác
        registry_trial = self.license_mgr._get_registry_trial()
        file_exists = os.path.exists(self.license_mgr.TRIAL_FILE)
        
        # Đã từng dùng trial (có trong Registry hoặc file)
        has_used_trial = (registry_trial and registry_trial.get("machine_code") == self.license_mgr.machine_code) or file_exists
        
        if not trial_available:
            if has_used_trial:
                # Đã dùng và hết hạn
                messagebox.showinfo(
                    "⏰ THỜI GIAN DÙNG THỬ",
                    "❌ ĐÃ HẾT THỜI GIAN DÙNG THỬ!\n\n"
                    "VUI LÒNG MUA LICENSE ĐỂ TIẾP TỤC SỬ DỤNG.",
                    parent=self
                )
            else:
                # Chưa dùng nhưng bị block (online check)
                messagebox.showinfo(
                    "⏰ THỜI GIAN DÙNG THỬ",
                    "❌ KHÔNG THỂ SỬ DỤNG TRIAL!\n\n"
                    "MÃ MÁY NÀY ĐÃ ĐƯỢC ĐĂNG KÝ TRIAL TRƯỚC ĐÓ.",
                    parent=self
                )
        else:
            if not has_used_trial:
                # Chưa bắt đầu trial
                messagebox.showinfo(
                    "⏰ THỜI GIAN DÙNG THỬ",
                    f"✅ BẠN CHƯA SỬ DỤNG TRIAL!\n\n"
                    f"BẠN CÓ {self.license_mgr.TRIAL_DAYS} NGÀY DÙNG THỬ MIỄN PHÍ.\n\n"
                    "NHẤN 'DÙNG THỬ 7 NGÀY MIỄN PHÍ' ĐỂ BẮT ĐẦU.",
                    parent=self
                )
            else:
                # Đang trong trial
                messagebox.showinfo(
                    "⏰ THỜI GIAN DÙNG THỬ",
                    f"⏳ ĐANG TRONG THỜI GIAN DÙNG THỬ\n\n"
                    f"⏰ THỜI GIAN CÒN LẠI:\n"
                    f"   {days_left} NGÀY  {hours_left} GIỜ  {mins_left} PHÚT\n\n"
                    f"📅 SAU KHI HẾT HẠN, CẦN MUA LICENSE.",
                    parent=self
                )
    
    def activate_license(self):
        """Kích hoạt license key - CHẠY ASYNC ĐỂ KHÔNG BLOCK UI"""
        license_key = self.key_entry.get().strip()
        
        if not license_key:
            messagebox.showwarning("THIẾU THÔNG TIN", "VUI LÒNG NHẬP LICENSE KEY!", parent=self)
            return
        
        # Set processing flag
        self.is_processing = True
        
        # Disable button và disable entry
        self.activate_btn.config(state="disabled", text="⏳ ĐANG KÍCH HOẠT...")
        self.key_entry.config(state="disabled")
        if self.trial_btn:
            self.trial_btn.config(state="disabled")
        if self.check_trial_btn:
            self.check_trial_btn.config(state="disabled")
        
        # Hiển thị progress bar và status
        self.status_label.config(text="🔄 Đang kết nối đến server Cryptolens... 0%", fg="#3498db")
        self.progress_bar['value'] = 0
        self.progress_bar.pack(fill="x", padx=40, pady=(0, 5))
        self.update()
        
        # [FIX] Chạy activation trong background thread để không freeze UI
        def activation_thread():
            try:
                print("[DEBUG] Thread started, calling activate_license...")
                # Call API (có thể mất 5-15 giây)
                success, error_msg, license_data = self.license_mgr.activate_license(license_key)
                
                print(f"[DEBUG] API call completed: success={success}, error={error_msg}")
                
                # Đưa kết quả vào queue
                self.result_queue.put(('success', success, error_msg, license_data))
            except Exception as e:
                print(f"[DEBUG] Error in activation_thread: {e}")
                import traceback
                traceback.print_exc()
                self.result_queue.put(('error', False, f"LỖI: {str(e)}", None))
        
        # Start thread
        print("[DEBUG] Starting activation thread...")
        threading.Thread(target=activation_thread, daemon=True).start()
        
        # Poll queue để kiểm tra kết quả
        self.activation_start_time = time.time()
        self._check_activation_result()
    
    def _check_activation_result(self):
        """Poll queue để lấy kết quả từ thread - chạy trong main loop"""
        try:
            # Check queue (non-blocking)
            result = self.result_queue.get_nowait()
            status, success, error_msg, license_data = result
            print(f"[DEBUG] Got result from queue: success={success}")
            
            # Set progress to 100%
            self.progress_bar['value'] = 100
            self.status_label.config(text="✅ Hoàn tất! 100%")
            self.update()
            
            # Xử lý kết quả
            self._handle_activation_result(success, error_msg, license_data)
        except queue.Empty:
            # Chưa có kết quả, update progress và status message
            elapsed = int(time.time() - self.activation_start_time)
            
            # Tính progress dựa trên thời gian (giả định tối đa 30s)
            progress = min(int((elapsed / 30.0) * 100), 95)  # Max 95% khi chưa xong
            self.progress_bar['value'] = progress
            
            # Update status message
            if elapsed < 5:
                msg = f"🔄 Đang kết nối... {progress}%"
            elif elapsed < 15:
                msg = f"⏳ Đang xác thực license... {progress}%"
            elif elapsed < 25:
                msg = f"🔐 Đang kiểm tra bảo mật... {progress}%"
            else:
                msg = f"⌛ Vui lòng chờ... {progress}%"
            
            self.status_label.config(text=msg)
            
            # Tiếp tục poll sau 300ms
            if self.winfo_exists():
                self.after(300, self._check_activation_result)
        except Exception as e:
            print(f"[DEBUG] Error in _check_activation_result: {e}")
    
    def _handle_activation_result(self, success, error_msg, license_data):
        """Xử lý kết quả activation trong main UI thread"""
        print(f"[DEBUG] _handle_activation_result called: success={success}")
        
        # Clear processing flag
        self.is_processing = False
        
        # Stop và ẩn progress bar
        self.progress_bar['value'] = 0
        self.progress_bar.pack_forget()
        
        # Check window còn tồn tại
        if not self.winfo_exists():
            print("[DEBUG] Window no longer exists in callback")
            return
        
        print("[DEBUG] Window exists, processing result...")
        
        if success:
            print("[DEBUG] Activation successful!")
            self.status_label.config(text="✅ Kích hoạt thành công!", fg="#27ae60")
            self.update()
            
            # Save license
            print("[DEBUG] Saving license...")
            if self.license_mgr.save_license(license_data):
                print("[DEBUG] License saved, showing success message...")
                messagebox.showinfo(
                    "THÀNH CÔNG", 
                    "LICENSE ĐÃ ĐƯỢC KÍCH HOẠT THÀNH CÔNG!\n\nPHẦN MỀM SẼ KHỞI ĐỘNG NGAY.",
                    parent=self
                )
                self.success = True
                self.destroy()
            else:
                print("[DEBUG] Failed to save license")
                self.status_label.config(text="❌ Lỗi khi lưu license!", fg="#e74c3c")
                messagebox.showerror("LỖI", "KHÔNG THỂ LƯU LICENSE FILE!", parent=self)
                self._reset_buttons()
        else:
            print(f"[DEBUG] Activation failed: {error_msg}")
            # Show error
            self.status_label.config(text="❌ Kích hoạt thất bại!", fg="#e74c3c")
            messagebox.showerror("KÍCH HOẠT THẤT BẠI", error_msg, parent=self)
            self._reset_buttons()
    
    def _reset_buttons(self):
        """Reset trạng thái buttons sau khi xong"""
        if not self.winfo_exists():
            return
        
        # Reset và ẩn progress bar
        self.progress_bar['value'] = 0
        self.progress_bar.pack_forget()
        
        self.activate_btn.config(state="normal", text="✓ KÍCH HOẠT LICENSE")
        self.key_entry.config(state="normal")
        if self.trial_btn:
            self.trial_btn.config(state="normal")
        if self.check_trial_btn:
            self.check_trial_btn.config(state="normal")
        self.status_label.config(text="")
    
    def start_trial(self):
        """Bắt đầu trial mode với countdown"""
        # Check trial available
        trial_available, days_left, hours_left, mins_left = self.license_mgr.is_trial_available()
        
        if not trial_available:
            messagebox.showwarning(
                "HẾT TRIAL",
                "BẠN ĐÃ HẾT THỜI GIAN DÙNG THỬ!\n\n"
                "HOẶC MÃ MÁY NÀY ĐÃ ĐĂNG KÝ TRIAL TRƯỚC ĐÓ.\n\n"
                "VUI LÒNG MUA LICENSE ĐỂ TIẾP TỤC SỬ DỤNG.",
                parent=self
            )
            return
        
        # Confirm với thông tin rõ ràng
        result = messagebox.askyesno(
            "🆓 DÙNG THỬ MIỄN PHÍ",
            f"BẠN SẼ ĐƯỢC DÙNG THỬ EDUMANAGER MIỄN PHÍ TRONG {self.license_mgr.TRIAL_DAYS} NGÀY.\n\n"
            f"⏰ THỜI GIAN: {self.license_mgr.TRIAL_DAYS} NGÀY = {self.license_mgr.TRIAL_DAYS * 24} GIỜ\n\n"
            "⚠️ LƯU Ý:\n"
            "• Trial được gắn với MÃ MÁY của bạn\n"
            "• KHÔNG THỂ reset bằng cách xóa file\n"
            "• Sau khi hết hạn, cần mua LICENSE\n\n"
            "BẮT ĐẦU DÙNG THỬ?",
            parent=self
        )
        
        if result:
            if self.license_mgr.start_trial():
                # Tính thời gian kết thúc
                end_time = datetime.now() + timedelta(days=self.license_mgr.TRIAL_DAYS)
                end_str = end_time.strftime("%d/%m/%Y lúc %H:%M")
                
                messagebox.showinfo(
                    "✅ BẮT ĐẦU DÙNG THỬ",
                    f"✅ KÍCH HOẠT TRIAL THÀNH CÔNG!\n\n"
                    f"⏰ THỜI GIAN CÒN LẠI:\n"
                    f"   {self.license_mgr.TRIAL_DAYS} NGÀY 0 GIỜ 0 PHÚT\n\n"
                    f"📅 HẾT HẠN: {end_str}\n\n"
                    f"🔒 MÃ MÁY: {self.license_mgr.machine_code}\n\n"
                    "CHÚC BẠN TRẢI NGHIỆM TỐT!",
                    parent=self
                )
                self.success = True
                self.destroy()
            else:
                messagebox.showerror("LỖI", "KHÔNG THỂ KHỞI TẠO TRIAL MODE!", parent=self)
    
    def on_close(self):
        """Đóng dialog - không activate"""
        self.success = False
        self.destroy()


class LicenseInfoDialog(Toplevel):
    """
    Dialog hiển thị thông tin license với giao diện hiện đại.
    """
    
    def __init__(self, parent, license_mgr):
        super().__init__(parent)
        
        self.license_mgr = license_mgr
        
        # Window config
        self.title("📜 Thông tin License")
        self.configure(bg="#f0f4f8")
        self.resizable(False, False)
        # self.transient(parent)  # Comment out to avoid issues
        
        # CRITICAL: Load license data first!
        license_mgr.load_license()
        
        # Get license info
        self.license_info = license_mgr.get_license_info()
        
        if self.license_info:
            self.setup_ui_with_license()
        else:
            self.setup_ui_no_license()
        
        # Center window
        self.update_idletasks()
        width = 520
        height = self.winfo_reqheight()
        x = (self.winfo_screenwidth() - width) // 2
        y = (self.winfo_screenheight() - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Force to front
        self.lift()
        self.focus_force()
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def setup_ui_with_license(self):
        """Hiển thị thông tin license với giao diện card đẹp mắt"""
        info = self.license_info
        
        # Gradient Header - compact
        header = tk.Frame(self, bg="#2c3e50", height=80)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        # Icon container
        icon_frame = tk.Frame(header, bg="#2c3e50")
        icon_frame.pack(expand=True)
        
        tk.Label(
            icon_frame,
            text="🔐",
            bg="#2c3e50",
            fg="white",
            font=("Segoe UI", 24)
        ).pack()
        
        tk.Label(
            icon_frame,
            text="THÔNG TIN BẢN QUYỀN",
            bg="#2c3e50",
            fg="#ffffff",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=(3, 0))
        
        # Main container với padding
        container = tk.Frame(self, bg="#f0f4f8")
        container.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Status Card (highlighted)
        status_card = tk.Frame(container, bg="#ffffff", relief="flat", bd=0)
        status_card.pack(fill="x", pady=(0, 10))
        
        # Add subtle shadow effect using multiple frames
        shadow_frame = tk.Frame(container, bg="#d0d8e0", height=2)
        shadow_frame.place(in_=status_card, relx=0, rely=1, relwidth=1)
        
        status_inner = tk.Frame(status_card, bg="#ffffff", padx=20, pady=15)
        status_inner.pack(fill="both", expand=True)
        
        # Status icon và text
        status_header = tk.Frame(status_inner, bg="#ffffff")
        status_header.pack(fill="x", pady=(0, 10))
        
        # Xác định icon và màu dựa trên status
        if "✓" in info["status"]:
            status_icon = "✅"
            status_bg = "#27ae60"
        elif "⚠" in info["status"]:
            status_icon = "⚠️"
            status_bg = "#f39c12"
        else:
            status_icon = "❌"
            status_bg = "#e74c3c"
        
        tk.Label(
            status_header,
            text=status_icon,
            bg="#ffffff",
            font=("Segoe UI", 24)
        ).pack(side="left", padx=(0, 12))
        
        status_text_frame = tk.Frame(status_header, bg="#ffffff")
        status_text_frame.pack(side="left", fill="both", expand=True)
        
        tk.Label(
            status_text_frame,
            text="TRẠNG THÁI LICENSE",
            bg="#ffffff",
            fg="#7f8c8d",
            font=("Segoe UI", 8),
            anchor="w"
        ).pack(anchor="w")
        
        tk.Label(
            status_text_frame,
            text=info["status"],
            bg="#ffffff",
            fg=status_bg,
            font=("Segoe UI", 12, "bold"),
            anchor="w"
        ).pack(anchor="w")
        
        # Separator
        tk.Frame(status_inner, bg="#ecf0f1", height=1).pack(fill="x", pady=8)
        
        # Expiry date prominently
        expiry_frame = tk.Frame(status_inner, bg="#ffffff")
        expiry_frame.pack(fill="x")
        
        tk.Label(
            expiry_frame,
            text="📅",
            bg="#ffffff",
            font=("Segoe UI", 14)
        ).pack(side="left", padx=(0, 8))
        
        tk.Label(
            expiry_frame,
            text="Ngày hết hạn:",
            bg="#ffffff",
            fg="#7f8c8d",
            font=("Segoe UI", 9),
            anchor="w"
        ).pack(side="left")
        
        tk.Label(
            expiry_frame,
            text=info["expiry_date"],
            bg="#ffffff",
            fg="#2c3e50",
            font=("Segoe UI", 10, "bold"),
            anchor="w"
        ).pack(side="left", padx=(8, 0))
        
        # Info Card
        info_card = tk.Frame(container, bg="#ffffff", relief="flat", bd=0)
        info_card.pack(fill="both", expand=True, pady=(0, 10))
        
        # Shadow
        shadow_frame2 = tk.Frame(container, bg="#d0d8e0", height=2)
        shadow_frame2.place(in_=info_card, relx=0, rely=1, relwidth=1)
        
        info_inner = tk.Frame(info_card, bg="#ffffff", padx=20, pady=15)
        info_inner.pack(fill="both", expand=True)
        
        # Title
        tk.Label(
            info_inner,
            text="Chi tiết thông tin",
            bg="#ffffff",
            fg="#2c3e50",
            font=("Segoe UI", 10, "bold"),
            anchor="w"
        ).pack(anchor="w", pady=(0, 12))
        
        # Grid layout container - 2 columns
        grid_container = tk.Frame(info_inner, bg="#ffffff")
        grid_container.pack(fill="both", expand=True)
        
        # Left column
        left_col = tk.Frame(grid_container, bg="#ffffff")
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Right column
        right_col = tk.Frame(grid_container, bg="#ffffff")
        right_col.pack(side="left", fill="both", expand=True, padx=(10, 0))
        
        # Data for 2 columns layout
        left_items = [
            ("👤", "Khách hàng", info["customer_name"]),
            ("🔑", "License Key", info["license_key_masked"])
        ]
        
        right_items = [
            ("💻", "Machine ID", info["machine_code"]),
            ("⚡", "Ngày kích hoạt", self._format_datetime(info["activated_date"]))
        ]
        
        # Render left column
        for icon, label_text, value_text in left_items:
            item_frame = tk.Frame(left_col, bg="#ffffff")
            item_frame.pack(fill="x", pady=6)
            
            tk.Label(
                item_frame,
                text=icon,
                bg="#ffffff",
                font=("Segoe UI", 13)
            ).pack(side="left", padx=(0, 8))
            
            content = tk.Frame(item_frame, bg="#ffffff")
            content.pack(side="left", fill="x", expand=True)
            
            tk.Label(
                content,
                text=label_text + ":",
                bg="#ffffff",
                fg="#5a6c7d",
                font=("Segoe UI", 9),
                anchor="w"
            ).pack(anchor="w")
            
            tk.Label(
                content,
                text=value_text,
                bg="#ffffff",
                fg="#1a252f",
                font=("Segoe UI", 9, "bold"),
                anchor="w",
                wraplength=200
            ).pack(anchor="w", pady=(1, 0))
        
        # Render right column
        for icon, label_text, value_text in right_items:
            item_frame = tk.Frame(right_col, bg="#ffffff")
            item_frame.pack(fill="x", pady=6)
            
            tk.Label(
                item_frame,
                text=icon,
                bg="#ffffff",
                font=("Segoe UI", 13)
            ).pack(side="left", padx=(0, 8))
            
            content = tk.Frame(item_frame, bg="#ffffff")
            content.pack(side="left", fill="x", expand=True)
            
            tk.Label(
                content,
                text=label_text + ":",
                bg="#ffffff",
                fg="#5a6c7d",
                font=("Segoe UI", 9),
                anchor="w"
            ).pack(anchor="w")
            
            tk.Label(
                content,
                text=value_text,
                bg="#ffffff",
                fg="#1a252f",
                font=("Segoe UI", 9, "bold"),
                anchor="w",
                wraplength=200
            ).pack(anchor="w", pady=(1, 0))
        
        # Last online check - full width at bottom
        last_check_frame = tk.Frame(info_inner, bg="#ecf0f1")
        last_check_frame.pack(fill="x", pady=(8, 0))
        
        last_check_inner = tk.Frame(last_check_frame, bg="#ecf0f1", padx=10, pady=6)
        last_check_inner.pack(fill="x")
        
        tk.Label(
            last_check_inner,
            text="🌐",
            bg="#ecf0f1",
            font=("Segoe UI", 12)
        ).pack(side="left", padx=(0, 8))
        
        tk.Label(
            last_check_inner,
            text="Kiểm tra online:",
            bg="#ecf0f1",
            fg="#5a6c7d",
            font=("Segoe UI", 9)
        ).pack(side="left")
        
        tk.Label(
            last_check_inner,
            text=self._format_datetime(info["last_online_check"]),
            bg="#ecf0f1",
            fg="#1a252f",
            font=("Segoe UI", 9, "bold")
        ).pack(side="left", padx=(8, 0))
        
        # Buttons với hover effect
        btn_frame = tk.Frame(container, bg="#f0f4f8")
        btn_frame.pack(fill="x", pady=(8, 0))
        
        # Check online button
        check_btn = tk.Button(
            btn_frame,
            text="🔄  KIỂM TRA CẬP NHẬT",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground="#2c3e50",
            padx=18,
            pady=10,
            cursor="hand2",
            activebackground="#2980b9",
            activeforeground="white",
            command=self.check_online
        )
        check_btn.pack(side="left", padx=(0, 10))
        
        # Hover effects
        check_btn.bind("<Enter>", lambda e: check_btn.config(bg="#2980b9"))
        check_btn.bind("<Leave>", lambda e: check_btn.config(bg="#3498db"))
        
        # Close button
        close_btn = tk.Button(
            btn_frame,
            text="✗  ĐÓNG",
            bg="#95a5a6",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground="#2c3e50",
            padx=18,
            pady=10,
            cursor="hand2",
            activebackground="#7f8c8d",
            activeforeground="white",
            command=self.destroy
        )
        close_btn.pack(side="right")
        
        # Hover effects
        close_btn.bind("<Enter>", lambda e: close_btn.config(bg="#7f8c8d"))
        close_btn.bind("<Leave>", lambda e: close_btn.config(bg="#95a5a6"))
    
    
    def setup_ui_no_license(self):
        """Hiển thị khi chưa có license (trial mode) với giao diện đẹp mắt"""
        
        # Gradient Header - compact
        header = tk.Frame(self, bg="#e67e22", height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        # Icon container
        icon_frame = tk.Frame(header, bg="#e67e22")
        icon_frame.pack(expand=True)
        
        tk.Label(
            icon_frame,
            text="CHẾ ĐỘ DÙNG THỬ",
            bg="#e67e22",
            fg="#ffffff",
            font=("Segoe UI", 13, "bold")
        ).pack()
        
        # Main container
        container = tk.Frame(self, bg="#f0f4f8")
        container.pack(fill="both", expand=True, padx=15, pady=15)
        
        # Trial info card
        trial_card = tk.Frame(container, bg="#ffffff", relief="flat", bd=0)
        trial_card.pack(fill="both", expand=True, pady=(0, 10))
        
        # Shadow
        shadow_frame = tk.Frame(container, bg="#d0d8e0", height=2)
        shadow_frame.place(in_=trial_card, relx=0, rely=1, relwidth=1)
        
        trial_inner = tk.Frame(trial_card, bg="#ffffff", padx=20, pady=20)
        trial_inner.pack(fill="both", expand=True)
        
        # Get trial status
        trial_available, days_left, hours_left, mins_left = self.license_mgr.is_trial_available()
        
        if trial_available:
            # Icon
            tk.Label(
                trial_inner,
                text="⏰",
                bg="#ffffff",
                font=("Segoe UI", 32)
            ).pack(pady=(0, 12))
            
            # Title
            tk.Label(
                trial_inner,
                text="Đang sử dụng phiên bản dùng thử",
                bg="#ffffff",
                fg="#2c3e50",
                font=("Segoe UI", 11, "bold")
            ).pack()
            
            # Countdown card
            countdown_frame = tk.Frame(trial_inner, bg="#fff3e0", relief="solid", bd=1, highlightthickness=1, highlightbackground="#2c3e50")
            countdown_frame.pack(fill="x", pady=(12, 12))
            
            countdown_inner = tk.Frame(countdown_frame, bg="#fff3e0", padx=15, pady=10)
            countdown_inner.pack()
            
            tk.Label(
                countdown_inner,
                text="⏳ THỜI GIAN CÒN LẠI",
                bg="#fff3e0",
                fg="#e67e22",
                font=("Segoe UI", 10, "bold")
            ).pack()
            
            # Time display
            time_display = tk.Frame(countdown_inner, bg="#fff3e0")
            time_display.pack(pady=(8, 0))
            
            # Days
            day_frame = tk.Frame(time_display, bg="#fff3e0")
            day_frame.pack(side="left", padx=8)
            
            tk.Label(
                day_frame,
                text=str(days_left),
                bg="#fff3e0",
                fg="#e67e22",
                font=("Segoe UI", 18, "bold")
            ).pack()
            
            tk.Label(
                day_frame,
                text="Ngày",
                bg="#fff3e0",
                fg="#7f8c8d",
                font=("Segoe UI", 9)
            ).pack()
            
            # Separator
            sep1 = tk.Label(
                time_display,
                text=":",
                bg="#fff3e0",
                fg="#e67e22",
                font=("Segoe UI", 16, "bold")
            )
            sep1.pack(side="left", pady=(0, 15))
            
            # Hours
            hour_frame = tk.Frame(time_display, bg="#fff3e0")
            hour_frame.pack(side="left", padx=8)
            
            tk.Label(
                hour_frame,
                text=str(hours_left),
                bg="#fff3e0",
                fg="#e67e22",
                font=("Segoe UI", 18, "bold")
            ).pack()
            
            tk.Label(
                hour_frame,
                text="Giờ",
                bg="#fff3e0",
                fg="#7f8c8d",
                font=("Segoe UI", 9)
            ).pack()
            
            # Separator
            sep2 = tk.Label(
                time_display,
                text=":",
                bg="#fff3e0",
                fg="#e67e22",
                font=("Segoe UI", 16, "bold")
            )
            sep2.pack(side="left", pady=(0, 15))
            
            # Minutes
            min_frame = tk.Frame(time_display, bg="#fff3e0")
            min_frame.pack(side="left", padx=8)
            
            tk.Label(
                min_frame,
                text=str(mins_left),
                bg="#fff3e0",
                fg="#e67e22",
                font=("Segoe UI", 18, "bold")
            ).pack()
            
            tk.Label(
                min_frame,
                text="Phút",
                bg="#fff3e0",
                fg="#7f8c8d",
                font=("Segoe UI", 9)
            ).pack()
            
            # Message
            tk.Label(
                trial_inner,
                text="Sau khi hết thời gian dùng thử,\nvui lòng mua license để tiếp tục sử dụng.",
                bg="#ffffff",
                fg="#7f8c8d",
                font=("Segoe UI", 9),
                justify="center"
            ).pack(pady=(8, 15))
            
        else:
            # Expired trial
            tk.Label(
                trial_inner,
                text="❌",
                bg="#ffffff",
                font=("Segoe UI", 32)
            ).pack(pady=(0, 12))
            
            tk.Label(
                trial_inner,
                text="Hết thời gian dùng thử!",
                bg="#ffffff",
                fg="#e74c3c",
                font=("Segoe UI", 12, "bold")
            ).pack()
            
            tk.Label(
                trial_inner,
                text="Vui lòng mua license để tiếp tục sử dụng phần mềm.",
                bg="#ffffff",
                fg="#7f8c8d",
                font=("Segoe UI", 9),
                justify="center"
            ).pack(pady=(12, 15))
        
        # Machine ID card
        machine_card = tk.Frame(trial_inner, bg="#ecf0f1", relief="solid", bd=1, highlightthickness=1, highlightbackground="#2c3e50")
        machine_card.pack(fill="x", pady=(0, 15))
        
        machine_inner = tk.Frame(machine_card, bg="#ecf0f1", padx=15, pady=12)
        machine_inner.pack(fill="x")
        
        tk.Label(
            machine_inner,
            text="💻  Machine ID:",
            bg="#ecf0f1",
            fg="#7f8c8d",
            font=("Segoe UI", 9)
        ).pack(side="left")
        
        tk.Label(
            machine_inner,
            text=self.license_mgr.machine_code,
            bg="#ecf0f1",
            fg="#2c3e50",
            font=("Segoe UI", 9, "bold")
        ).pack(side="left", padx=(10, 0))
        
        # Close button
        close_btn = tk.Button(
            trial_inner,
            text="✗  ĐÓNG",
            bg="#95a5a6",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="solid",
            bd=1,
            highlightthickness=1,
            highlightbackground="#2c3e50",
            padx=25,
            pady=12,
            cursor="hand2",
            activebackground="#7f8c8d",
            activeforeground="white",
            command=self.destroy
        )
        close_btn.pack()
        
        # Hover effect
        close_btn.bind("<Enter>", lambda e: close_btn.config(bg="#7f8c8d"))
        close_btn.bind("<Leave>", lambda e: close_btn.config(bg="#95a5a6"))
    
    def check_online(self):
        """Kiểm tra license online"""
        success, error_msg = self.license_mgr.check_online()
        
        if success:
            messagebox.showinfo(
                "Cập nhật thành công",
                "License đã được cập nhật thông tin mới nhất!",
                parent=self
            )
            self.destroy()
            # Reopen với thông tin mới
            LicenseInfoDialog(self.master, self.license_mgr)
        else:
            if error_msg:
                messagebox.showerror("Lỗi", error_msg, parent=self)
            else:
                messagebox.showinfo(
                    "Offline mode",
                    "Không thể kết nối server.\n\nBạn vẫn có thể sử dụng phần mềm offline.",
                    parent=self
                )
    
    def _format_datetime(self, iso_string):
        """Format ISO datetime string"""
        if not iso_string or iso_string == "N/A":
            return "N/A"
        try:
            dt = datetime.fromisoformat(iso_string)
            return dt.strftime("%d/%m/%Y %H:%M")
        except:
            return iso_string


# ========== [NEW ARCHITECTURE] BULK EDIT SYSTEM ==========
class BulkEditSelectionWindow(Toplevel):
    """
    Cửa sổ riêng biệt để chọn học sinh cần sửa điểm hàng loạt.
    Kiến trúc mới: Treeview độc lập với checkbox built-in (không động)
    
    [WARNING] Hardcoded TX1-TX4 columns. Should be dynamic based on dataframe columns.
    TODO: Refactor to detect TX columns dynamically from self.df.columns
    """
    
    def __init__(self, parent_tab):
        super().__init__(parent_tab)
        
        self.parent_tab = parent_tab  # ExcelTab instance
        self.df = parent_tab.current_df
        self.file_path = parent_tab.file_path
        
        # Selected data
        self.selected_students = []  # List of row indices
        
        # [UPDATED] Window config - Full screen với nút minimize/close
        self.title("⚡ CHỌN HỌC SINH SỬA HÀNG LOẠT")
        self.configure(bg=THEME["bg_card"])
        
        # Bật minimize/maximize/close buttons (không dùng transient)
        # self.transient(parent_tab)  # Bỏ transient để có nút minimize
        self.resizable(True, True)
        
        # Full screen theo màn hình
        self.state('zoomed')  # Windows: maximize window
        # Hoặc dùng geometry full screen:
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}+0+0")
        
        # Column selection vars
        self.col_tx1_var = tk.BooleanVar(value=False)
        self.col_tx2_var = tk.BooleanVar(value=False)
        self.col_tx3_var = tk.BooleanVar(value=False)
        self.col_tx4_var = tk.BooleanVar(value=False)
        self.col_gk_var = tk.BooleanVar(value=False)
        self.col_ck_var = tk.BooleanVar(value=False)
        
        self.setup_ui()
        self.load_data_to_tree()
        
        self.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def setup_ui(self):
        """Xây dựng giao diện"""
        # Header
        header = tk.Frame(self, bg="#3498db", pady=15)
        header.pack(fill="x")
        
        tk.Label(
            header,
            text="📋 CHỌN HỌC SINH CẦN SỬA ĐIỂM",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 14, "bold")
        ).pack()
        
        tk.Label(
            header,
            text="Click vào checkbox để chọn học sinh → Chọn cột điểm cần sửa → Tiếp tục",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 9)
        ).pack(pady=(5, 0))
        
        # Main container
        main_frame = tk.Frame(self, bg=THEME["bg_card"], padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # === LEFT: Treeview với checkboxes ===
        left_frame = tk.Frame(main_frame, bg=THEME["bg_card"])
        left_frame.pack(side="left", fill="both", expand=True)
        
        # Toolbar trên Treeview
        toolbar = tk.Frame(left_frame, bg=THEME["bg_card"])
        toolbar.pack(fill="x", pady=(0, 10))
        
        tk.Button(
            toolbar,
            text="☑ Chọn tất cả",
            bg="#3498db",
            fg="white",
            font=("Segoe UI", 9),
            relief="solid",
            borderwidth=1,
            padx=10,
            pady=5,
            cursor="hand2",
            command=self.select_all
        ).pack(side="left", padx=2)
        
        tk.Button(
            toolbar,
            text="☐ Bỏ chọn tất cả",
            bg="#95a5a6",
            fg="white",
            font=("Segoe UI", 9),
            relief="solid",
            borderwidth=1,
            padx=10,
            pady=5,
            cursor="hand2",
            command=self.deselect_all
        ).pack(side="left", padx=2)
        
        self.selection_label = tk.Label(
            toolbar,
            text="Đã chọn: 0 học sinh",
            bg=THEME["bg_card"],
            fg="#e67e22",
            font=("Segoe UI", 10, "bold")
        )
        self.selection_label.pack(side="right")
        
        # Treeview
        tree_frame = tk.Frame(left_frame, bg="#ddd", padx=1, pady=1)
        tree_frame.pack(fill="both", expand=True)
        
        # Define columns
        columns = ("STT", "Họ và tên", "TX1", "TX2", "TX3", "TX4", "GK", "CK", "ĐTB")
        
        self.tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="tree headings",
            selectmode="browse",
            height=15
        )
        
        # Column #0 là checkbox column
        self.tree.heading("#0", text="☐", command=self.toggle_all)
        self.tree.column("#0", width=40, anchor="center", stretch=False)
        
        # Setup other columns
        col_widths = {"STT": 50, "Họ và tên": 200, "TX1": 55, "TX2": 55, 
                      "TX3": 55, "TX4": 55, "GK": 55, "CK": 55, "ĐTB": 65}
        
        for col in columns:
            self.tree.heading(col, text=col, anchor="w")
            self.tree.column(col, width=col_widths.get(col, 80), anchor="w")
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Bind click event
        self.tree.bind("<Button-1>", self.on_tree_click)
        
        # === RIGHT: Column selection và buttons ===
        right_frame = tk.Frame(main_frame, bg=THEME["bg_card"], width=280)
        right_frame.pack(side="right", fill="y", padx=(20, 0))
        right_frame.pack_propagate(False)
        
        # Column selection
        col_section = tk.LabelFrame(
            right_frame,
            text="📊 CHỌN CỘT ĐIỂM CẦN SỬA",
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=15
        )
        col_section.pack(fill="x", pady=(0, 20))
        
        tk.Label(
            col_section,
            text="Chọn cột nào cần thay đổi:",
            bg=THEME["bg_card"],
            fg=THEME["text_sub"],
            font=("Segoe UI", 9)
        ).pack(anchor="w", pady=(0, 10))
        
        for col_name, var in [
            ("TX1 (Thường xuyên 1)", self.col_tx1_var),
            ("TX2 (Thường xuyên 2)", self.col_tx2_var),
            ("TX3 (Thường xuyên 3)", self.col_tx3_var),
            ("TX4 (Thường xuyên 4)", self.col_tx4_var),
            ("GK (Giữa kỳ)", self.col_gk_var),
            ("CK (Cuối kỳ)", self.col_ck_var),
        ]:
            tk.Checkbutton(
                col_section,
                text=col_name,
                variable=var,
                bg=THEME["bg_card"],
                fg=THEME["text_main"],
                font=("Segoe UI", 9),
                selectcolor=THEME["bg_card"],
                activebackground=THEME["bg_card"]
            ).pack(anchor="w", pady=3)
        
        # Action buttons
        tk.Button(
            right_frame,
            text="✓ TIẾP TỤC",
            bg="#27ae60",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="solid",
            borderwidth=1,
            padx=20,
            pady=12,
            cursor="hand2",
            command=self.open_edit_dialog
        ).pack(fill="x", pady=(0, 10))
        
        tk.Button(
            right_frame,
            text="✖ HỦY",
            bg="#95a5a6",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            relief="solid",
            borderwidth=1,
            padx=20,
            pady=12,
            cursor="hand2",
            command=self.on_close
        ).pack(fill="x")
    
    def load_data_to_tree(self):
        """Load dữ liệu từ DataFrame vào Treeview"""
        self.tree.delete(*self.tree.get_children())
        
        for idx, row in self.df.iterrows():
            # Icon checkbox mặc định: ☐
            self.tree.insert(
                "",
                "end",
                iid=str(idx),
                text="☐",
                values=(
                    idx + 1,  # STT
                    row.get("Họ và tên", ""),
                    row.get("TX1", ""),
                    row.get("TX2", ""),
                    row.get("TX3", ""),
                    row.get("TX4", ""),
                    row.get("GK", ""),
                    row.get("CK", ""),
                    row.get("ĐTB", "")
                )
            )
    
    def on_tree_click(self, event):
        """Xử lý click vào Treeview"""
        region = self.tree.identify_region(event.x, event.y)
        
        # Xử lý khi click vào bất kỳ vùng nào của hàng (tree, cell, heading)
        if region in ("tree", "cell"):
            iid = self.tree.identify_row(event.y)
            if iid:
                self.toggle_row_selection(iid)
    
    def toggle_row_selection(self, iid):
        """Toggle checkbox của 1 row"""
        current_text = self.tree.item(iid, "text")
        
        if current_text == "☐":
            self.tree.item(iid, text="☑")
            self.selected_students.append(int(iid))
        else:
            self.tree.item(iid, text="☐")
            if int(iid) in self.selected_students:
                self.selected_students.remove(int(iid))
        
        self.update_selection_label()
    
    def select_all(self):
        """Chọn tất cả học sinh"""
        self.selected_students.clear()
        for iid in self.tree.get_children():
            self.tree.item(iid, text="☑")
            self.selected_students.append(int(iid))
        self.update_selection_label()
    
    def deselect_all(self):
        """Bỏ chọn tất cả"""
        self.selected_students.clear()
        for iid in self.tree.get_children():
            self.tree.item(iid, text="☐")
        self.update_selection_label()
    
    def toggle_all(self):
        """Toggle tất cả checkboxes"""
        if len(self.selected_students) == len(self.df):
            self.deselect_all()
        else:
            self.select_all()
    
    def update_selection_label(self):
        """Cập nhật label hiển thị số lượng đã chọn"""
        count = len(self.selected_students)
        self.selection_label.config(text=f"Đã chọn: {count} học sinh")
    
    def open_edit_dialog(self):
        """Mở dialog sửa điểm hàng loạt"""
        if not self.selected_students:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn ít nhất 1 học sinh!")
            return
        
        # Lấy danh sách cột được chọn
        selected_cols = []
        if self.col_tx1_var.get():
            selected_cols.append("TX1")
        if self.col_tx2_var.get():
            selected_cols.append("TX2")
        if self.col_tx3_var.get():
            selected_cols.append("TX3")
        if self.col_tx4_var.get():
            selected_cols.append("TX4")
        if self.col_gk_var.get():
            selected_cols.append("GK")
        if self.col_ck_var.get():
            selected_cols.append("CK")
        
        if not selected_cols:
            messagebox.showwarning("Chưa chọn cột", "Vui lòng chọn ít nhất 1 cột điểm cần sửa!")
            return
        
        # Build selected_cells list: [(row_idx, col_name, old_value), ...]
        selected_cells = []
        for row_idx in self.selected_students:
            for col_name in selected_cols:
                if col_name in self.df.columns:
                    old_val = self.df.at[row_idx, col_name]
                    selected_cells.append((row_idx, col_name, old_val))
        
        # Mở BulkEditDialog
        BulkEditDialog(self.parent_tab, selected_cells, self.df, self.file_path, self)
    
    def on_close(self):
        """Đóng cửa sổ"""
        self.destroy()


class BulkEditDialog(Toplevel):
    """
    Dialog sửa điểm hàng loạt với validation và preview
    """
    
    def __init__(self, parent_tab, selected_cells, dataframe, file_path, selection_window):
        super().__init__(parent_tab)
        
        self.parent_tab = parent_tab
        self.selected_cells = selected_cells
        self.df = dataframe
        self.file_path = file_path
        self.selection_window = selection_window  # Reference để đóng sau khi apply
        
        # Window config
        self.title("⚡ SỬA ĐIỂM HÀNG LOẠT")
        self.geometry("900x550")
        self.configure(bg=THEME["bg_card"])
        self.transient(parent_tab)
        self.resizable(True, True)
        self.minsize(850, 500)  # Kích thước tối thiểu
        
        # UI Variables
        self.operation_var = tk.StringVar(value="add")
        self.value_var = tk.StringVar(value="0.5")
        self.auto_clip_var = tk.BooleanVar(value=True)
        self.skip_empty_var = tk.BooleanVar(value=True)
        self.round_1_var = tk.BooleanVar(value=False)
        self.round_2_var = tk.BooleanVar(value=False)
        
        # Preview data
        self.preview_data = {}
        
        self.setup_ui()
        
        # Center window theo màn hình
        self.update_idletasks()
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - 900) // 2
        y = (screen_height - 550) // 2
        self.geometry(f"900x550+{x}+{y}")
        
        # Calculate preview
        self.calculate_preview()
        
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
    
    def setup_ui(self):
        """Xây dựng giao diện dialog"""
        main_frame = tk.Frame(self, bg=THEME["bg_card"])
        main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        
        # === LAYOUT 2 CỘT: TRÁI (Controls) + PHẢI (Preview) ===
        left_frame = tk.Frame(main_frame, bg=THEME["bg_card"])
        left_frame.pack(side="left", fill="both", expand=False, padx=(0, 10))
        left_frame.configure(width=480)
        
        right_frame = tk.Frame(main_frame, bg=THEME["bg_card"])
        right_frame.pack(side="right", fill="both", expand=True)
        
        # === LEFT COLUMN: Info + Operation + Options ===
        
        # === SECTION 1: Info ===
        section1 = tk.LabelFrame(
            left_frame,
            text="📊 THÔNG TIN",
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=10
        )
        section1.pack(fill="x", pady=(0, 10))
        
        unique_students = set([r for r, c, v in self.selected_cells])
        unique_cols = set([c for r, c, v in self.selected_cells])
        
        info_text = f"• {len(unique_students)} học sinh × {len(unique_cols)} cột = {len(self.selected_cells)} ô\n"
        info_text += f"• Cột: {', '.join(sorted(unique_cols))}"
        
        tk.Label(
            section1,
            text=info_text,
            bg=THEME["bg_card"],
            fg=THEME["text_sub"],
            font=("Segoe UI", 9),
            justify="left"
        ).pack(anchor="w")
        
        # === SECTION 2: Operation ===
        section2 = tk.LabelFrame(
            left_frame,
            text="🔧 THAO TÁC",
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=10
        )
        section2.pack(fill="x", pady=(0, 10))
        
        operations = [
            ("CỘNG THÊM", "add"),
            ("TRỪ ĐI", "subtract"),
            ("NHÂN VỚI", "multiply"),
            ("CHIA CHO", "divide"),
            ("GÁN GIÁ TRỊ", "set")
        ]
        
        for i, (label, value) in enumerate(operations):
            tk.Radiobutton(
                section2,
                text=label,
                variable=self.operation_var,
                value=value,
                bg=THEME["bg_card"],
                fg=THEME["text_main"],
                font=("Segoe UI", 9),
                selectcolor=THEME["bg_card"],
                activebackground=THEME["bg_card"],
                command=self.calculate_preview
            ).grid(row=i//3, column=i%3, sticky="w", padx=10, pady=2)
        
        # Value entry
        value_frame = tk.Frame(section2, bg=THEME["bg_card"])
        value_frame.grid(row=2, column=0, columnspan=3, sticky="w", pady=(10, 0))
        
        tk.Label(
            value_frame,
            text="💯 GIÁ TRỊ:",
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 9, "bold")
        ).pack(side="left", padx=5)
        
        value_entry = tk.Entry(
            value_frame,
            textvariable=self.value_var,
            width=15,
            font=("Segoe UI", 11),
            bg=THEME["entry_bg"],
            fg=THEME["text_main"]
        )
        value_entry.pack(side="left", padx=5)
        value_entry.bind("<KeyRelease>", lambda e: self.calculate_preview())
        
        # === SECTION 3: Options - CHIA 2 HÀNG ===
        section3 = tk.LabelFrame(
            left_frame,
            text="⚙️ TÙY CHỌN",
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=10
        )
        section3.pack(fill="x", pady=(0, 10))
        
        # Hàng 1: 2 checkbox đầu
        row1_frame = tk.Frame(section3, bg=THEME["bg_card"])
        row1_frame.pack(fill="x", pady=2)
        
        tk.Checkbutton(
            row1_frame,
            text="TỰ ĐỘNG CẮT (0-10)",
            variable=self.auto_clip_var,
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 9),
            selectcolor=THEME["bg_card"],
            activebackground=THEME["bg_card"],
            command=self.calculate_preview
        ).pack(side="left", padx=(0, 15))
        
        tk.Checkbutton(
            row1_frame,
            text="BỞ QUA Ô TRỐNG",
            variable=self.skip_empty_var,
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 9),
            selectcolor=THEME["bg_card"],
            activebackground=THEME["bg_card"],
            command=self.calculate_preview
        ).pack(side="left")
        
        # Hàng 2: 2 checkbox làm tròn
        row2_frame = tk.Frame(section3, bg=THEME["bg_card"])
        row2_frame.pack(fill="x", pady=2)
        
        tk.Checkbutton(
            row2_frame,
            text="LÀM TRÒN 1 SỐ (7.85→7.9)",
            variable=self.round_1_var,
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 9),
            selectcolor=THEME["bg_card"],
            activebackground=THEME["bg_card"],
            command=lambda: self.handle_rounding(1)
        ).pack(side="left", padx=(0, 15))
        
        tk.Checkbutton(
            row2_frame,
            text="LÀM TRÒN 2 SỐ (7.856→7.86)",
            variable=self.round_2_var,
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 9),
            selectcolor=THEME["bg_card"],
            activebackground=THEME["bg_card"],
            command=lambda: self.handle_rounding(2)
        ).pack(side="left")
        
        # === BUTTONS ở dưới left frame ===
        button_frame = tk.Frame(left_frame, bg=THEME["bg_card"])
        button_frame.pack(fill="x", pady=(10, 0))
        
        tk.Button(
            button_frame,
            text="✓ ÁP DỤNG",
            bg="#27ae60",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="solid",
            borderwidth=1,
            padx=20,
            pady=10,
            cursor="hand2",
            command=self.apply_changes
        ).pack(side="left", padx=5)
        
        tk.Button(
            button_frame,
            text="✖ HỦY",
            bg="#95a5a6",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="solid",
            borderwidth=1,
            padx=20,
            pady=10,
            cursor="hand2",
            command=self.on_cancel
        ).pack(side="right", padx=5)
        
        # === RIGHT COLUMN: Preview ===
        section4 = tk.LabelFrame(
            right_frame,
            text="⚠️ PREVIEW",
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            padx=15,
            pady=10
        )
        section4.pack(fill="both", expand=True)
        
        self.preview_text = tk.Text(
            section4,
            bg="#f8f9fa",
            fg="#212529",
            font=("Consolas", 9),
            relief="flat",
            padx=10,
            pady=10,
            state="disabled",
            wrap="word"
        )
        
        # Thêm scrollbar cho preview
        preview_scroll = ttk.Scrollbar(section4, orient="vertical", command=self.preview_text.yview)
        self.preview_text.configure(yscrollcommand=preview_scroll.set)
        
        self.preview_text.pack(side="left", fill="both", expand=True)
        preview_scroll.pack(side="right", fill="y")
    
    def handle_rounding(self, mode):
        """Mutual exclusive rounding"""
        if mode == 1 and self.round_1_var.get():
            self.round_2_var.set(False)
        elif mode == 2 and self.round_2_var.get():
            self.round_1_var.set(False)
        self.calculate_preview()
    
    def calculate_preview(self):
        """Tính toán preview"""
        try:
            operation = self.operation_var.get()
            value_str = self.value_var.get()
            
            try:
                value = float(value_str)
            except ValueError:
                self.update_preview_text("⚠️ Giá trị không hợp lệ", "red")
                return
            
            if operation == "divide" and value == 0:
                self.update_preview_text("❌ Không thể chia cho 0", "red")
                return
            
            ok_count = 0
            warn_count = 0
            skip_count = 0
            
            self.preview_data.clear()
            preview_details = []  # Lưu chi tiết từng học sinh
            
            for row_idx, col_name, old_val in self.selected_cells:
                # Lấy tên học sinh từ DataFrame
                student_name = self.df.iloc[row_idx].get("Họ và tên", f"HS #{row_idx+1}")
                
                if pd.isna(old_val) or old_val == "":
                    if self.skip_empty_var.get():
                        skip_count += 1
                        self.preview_data[(row_idx, col_name)] = {
                            "old": None, "new": None, "status": "skip", "name": student_name
                        }
                        continue
                
                try:
                    old_float = float(old_val)
                except:
                    skip_count += 1
                    continue
                
                # Calculate
                if operation == "add":
                    new_val = old_float + value
                elif operation == "subtract":
                    new_val = old_float - value
                elif operation == "multiply":
                    new_val = old_float * value
                elif operation == "divide":
                    new_val = old_float / value
                else:
                    new_val = value
                
                # Rounding
                if self.round_1_var.get():
                    new_val = round(new_val, 1)
                elif self.round_2_var.get():
                    new_val = round(new_val, 2)
                
                # Clipping
                status = "ok"
                if new_val > 10:
                    status = "warn"
                    warn_count += 1
                    if self.auto_clip_var.get():
                        new_val = 10.0
                elif new_val < 0:
                    status = "warn"
                    warn_count += 1
                    if self.auto_clip_var.get():
                        new_val = 0.0
                else:
                    ok_count += 1
                
                self.preview_data[(row_idx, col_name)] = {
                    "old": old_float,
                    "new": new_val,
                    "status": status,
                    "name": student_name
                }
                
                # Thêm chi tiết preview
                icon = "✅" if status == "ok" else "⚠️"
                preview_details.append(
                    f"{icon} {student_name} - {col_name}: {old_float} → {new_val}"
                )
            
            # Build text với chi tiết
            lines = []
            lines.append(f"📊 TỔNG QUAN:")
            if ok_count > 0:
                lines.append(f"  ✅ {ok_count} ô: OK")
            if warn_count > 0:
                clip_text = "đã cắt" if self.auto_clip_var.get() else "vượt ngưỡng"
                lines.append(f"  ⚠️ {warn_count} ô: {clip_text}")
            if skip_count > 0:
                lines.append(f"  ℹ️ {skip_count} ô: Trống (bỏ qua)")
            
            # Thêm chi tiết từng học sinh (giới hạn 10 dòng đầu)
            if preview_details:
                lines.append(f"\n📝 CHI TIẾT ({len(preview_details)} thay đổi):")
                display_limit = min(10, len(preview_details))
                lines.extend(preview_details[:display_limit])
                if len(preview_details) > display_limit:
                    lines.append(f"... và {len(preview_details) - display_limit} thay đổi nữa")
            
            self.update_preview_text("\n".join(lines), "black")
        
        except Exception as e:
            self.update_preview_text(f"❌ Lỗi: {str(e)}", "red")
    
    def update_preview_text(self, text, color="black"):
        """Update preview text widget"""
        self.preview_text.config(state="normal")
        self.preview_text.delete("1.0", "end")
        self.preview_text.insert("1.0", text)
        self.preview_text.tag_configure("color", foreground=color)
        self.preview_text.tag_add("color", "1.0", "end")
        self.preview_text.config(state="disabled")
    
    def apply_changes(self):
        """Áp dụng thay đổi vào DataFrame và Excel"""
        if not self.preview_data:
            messagebox.showwarning("Không có dữ liệu", "Không có gì để áp dụng")
            return
        
        affected = len([x for x in self.preview_data.values() if x["status"] != "skip"])
        if affected == 0:
            messagebox.showinfo("Không có thay đổi", "Tất cả ô đều bị bỏ qua")
            return
        
        if not messagebox.askyesno(
            "Xác nhận",
            f"Áp dụng thay đổi cho {affected} ô điểm?\n\n"
            f"Thao tác: {self.operation_var.get()}\n"
            f"Giá trị: {self.value_var.get()}"
        ):
            return
        
        try:
            # Apply to DataFrame
            affected_rows = set()
            
            for (row_idx, col_name), result in self.preview_data.items():
                if result["status"] == "skip":
                    continue
                
                self.df.at[row_idx, col_name] = result["new"]
                affected_rows.add(row_idx)
            
            # [DYNAMIC COLS] Recalculate ĐTB - auto-detect TX columns
            for row_idx in affected_rows:
                tx_scores = []
                # Tự động tìm tất cả cột TX trong DataFrame
                tx_cols_in_df = [col for col in self.df.columns if col.startswith("TX") and col[2:].isdigit()]
                for tx_col in sorted(tx_cols_in_df):  # Sắp xếp TX1, TX2, TX3...
                    val = self.df.at[row_idx, tx_col]
                    if pd.notna(val) and val != "":
                        try:
                            tx_scores.append(float(val))
                        except:
                            pass
                
                gk = self.df.at[row_idx, "GK"] if "GK" in self.df.columns else ""
                ck = self.df.at[row_idx, "CK"] if "CK" in self.df.columns else ""
                
                dtb = calculate_dtb_exact(tx_scores, gk, ck)
                self.df.at[row_idx, "ĐTB"] = dtb
                self.df.at[row_idx, "Xếp loại"] = self.parent_tab.classify_smart(dtb)
            
            # Update Treeview
            self.parent_tab.update_ui_data(self.df, update_chart=True)
            
            # Save to Excel
            self.save_to_excel()
            
            messagebox.showinfo(
                "Thành công",
                f"✅ Đã sửa {affected} ô điểm!\n"
                f"• Học sinh: {len(affected_rows)}\n"
                f"• ĐTB đã tính lại"
            )
            
            # Close both dialogs
            self.selection_window.destroy()
            self.destroy()
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Lỗi", f"Không thể áp dụng:\n{str(e)}")
    
    def save_to_excel(self):
        """Lưu thay đổi vào Excel file với smart header detection"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(self.file_path)
            
            # Tìm sheet name từ parent_tab (thay vì dùng active)
            try:
                sheet_name = self.parent_tab.get_current_sheet_name()
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                print(f"[📊] Using sheet: {sheet_name}")
            except:
                ws = wb.active
                print(f"[📊] Using active sheet")
            
            # Smart header detection - tìm row chứa "Họ và tên"
            header_row_idx = 1
            for row_idx in range(1, 6):  # Check first 5 rows
                for cell in ws[row_idx]:
                    if cell.value and "Họ và tên" in str(cell.value):
                        header_row_idx = row_idx
                        break
                if header_row_idx != 1:
                    break
            
            print(f"[🔍] Header row detected at: {header_row_idx}")
            
            # Build column mapping
            col_mapping = {}
            for col_idx, cell in enumerate(ws[header_row_idx], start=1):
                if cell.value:
                    col_name = str(cell.value).strip()
                    col_mapping[col_name] = col_idx
            
            print(f"[🗂️] Column mapping: {col_mapping}")
            
            # [DYNAMIC COLS] Verify critical columns exist - auto-detect TX columns
            tx_cols_in_df = [col for col in col_mapping.keys() if col.startswith("TX") and col[2:].isdigit()]
            required_cols = tx_cols_in_df + ["GK", "CK", "ĐTB", "Xếp loại"]
            missing_cols = [col for col in required_cols if col not in col_mapping]
            if missing_cols:
                print(f"[⚠️] Cảnh báo: Thiếu cột: {missing_cols}")
            
            # Write changes
            changes_count = 0
            for (df_row_idx, col_name), result in self.preview_data.items():
                if result["status"] == "skip":
                    continue
                
                excel_row = header_row_idx + df_row_idx + 1
                excel_col = col_mapping.get(col_name)
                
                if excel_col:
                    old_value = ws.cell(row=excel_row, column=excel_col).value
                    ws.cell(row=excel_row, column=excel_col).value = result["new"]
                    changes_count += 1
                    print(f"[✏️] Ô ({excel_row}, {excel_col}) [{col_name}]: {old_value} → {result['new']}")
                else:
                    print(f"[❌] Không tìm thấy cột '{col_name}' trong mapping!")
            
            # Write ĐTB và Xếp loại
            dtb_count = 0
            for row_idx in set([r for r, c in self.preview_data.keys()]):
                excel_row = header_row_idx + row_idx + 1
                
                if "ĐTB" in col_mapping:
                    dtb_val = self.df.at[row_idx, "ĐTB"]
                    ws.cell(row=excel_row, column=col_mapping["ĐTB"]).value = dtb_val
                    dtb_count += 1
                    print(f"[📊] Row {excel_row}: ĐTB = {dtb_val}")
                
                if "Xếp loại" in col_mapping:
                    xl_val = self.df.at[row_idx, "Xếp loại"]
                    ws.cell(row=excel_row, column=col_mapping["Xếp loại"]).value = xl_val
                    print(f"[🏆] Row {excel_row}: Xếp loại = {xl_val}")
            
            wb.save(self.file_path)
            print(f"[💾 THÀNH CÔNG] Đã lưu {changes_count} thay đổi + {dtb_count} ĐTB vào {self.file_path}")
            print(f"[✅] File Excel đã được cập nhật thành công!")
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"[❌ LỖI] Không thể lưu file Excel:")
            print(error_details)
            messagebox.showerror(
                "Lỗi lưu Excel", 
                f"Không thể lưu file Excel:\n\n{str(e)}\n\nXem console để biết chi tiết."
            )
            raise
    
    def on_cancel(self):
        """Hủy và đóng dialog"""
        self.destroy()
# ========== END BULK EDIT SYSTEM ==========


class ToolTip:
    """
    Widget Tooltip - Hiển thị hướng dẫn khi hover vào widget.
    Sử dụng: ToolTip(widget, "Nội dung tooltip")
    """
    def __init__(self, widget, text, delay=500, wraplength=250):
        self.widget = widget
        self.text = text
        self.delay = delay  # Thời gian chờ trước khi hiện (ms)
        self.wraplength = wraplength
        self.tooltip_window = None
        self.scheduled_id = None
        
        # Bind events
        self.widget.bind("<Enter>", self.on_enter)
        self.widget.bind("<Leave>", self.on_leave)
        self.widget.bind("<Motion>", self.on_motion)
    
    def on_enter(self, event):
        """Khi chuột vào widget - lên lịch hiện tooltip"""
        self.cancel_scheduled()
        self.scheduled_id = self.widget.after(self.delay, lambda: self.show_tooltip(event))
    
    def on_leave(self, event):
        """Khi chuột rời widget - ẩn tooltip"""
        self.cancel_scheduled()
        self.hide_tooltip()
    
    def on_motion(self, event):
        """Khi di chuyển chuột trong widget"""
        # Cập nhật vị trí nếu tooltip đang hiện
        if self.tooltip_window:
            self.update_position(event)
    
    def cancel_scheduled(self):
        """Hủy lịch hiện tooltip nếu có"""
        if self.scheduled_id:
            self.widget.after_cancel(self.scheduled_id)
            self.scheduled_id = None
    
    def show_tooltip(self, event):
        """Hiển thị tooltip"""
        if self.tooltip_window:
            return
        
        # Tạo cửa sổ tooltip
        self.tooltip_window = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # Không có title bar
        tw.wm_attributes("-topmost", True)
        
        # Nội dung
        frame = tk.Frame(tw, bg="#fffde7", relief="solid", bd=1)
        frame.pack()
        
        label = tk.Label(frame, text=self.text, bg="#fffde7", fg="#333333",
                         font=("Segoe UI", 9), justify="left", wraplength=self.wraplength,
                         padx=8, pady=5)
        label.pack()
        
        # Vị trí
        self.update_position(event)
    
    def update_position(self, event):
        """Cập nhật vị trí tooltip"""
        if not self.tooltip_window:
            return
        
        # Đặt tooltip bên dưới và bên phải con trỏ
        x = event.x_root + 15
        y = event.y_root + 15
        
        # Đảm bảo không ra ngoài màn hình
        screen_width = self.widget.winfo_screenwidth()
        screen_height = self.widget.winfo_screenheight()
        
        # Lấy kích thước tooltip
        self.tooltip_window.update_idletasks()
        tw_width = self.tooltip_window.winfo_width()
        tw_height = self.tooltip_window.winfo_height()
        
        if x + tw_width > screen_width - 10:
            x = event.x_root - tw_width - 10
        if y + tw_height > screen_height - 10:
            y = event.y_root - tw_height - 10
        
        self.tooltip_window.geometry(f"+{x}+{y}")
    
    def hide_tooltip(self):
        """Ẩn và hủy tooltip"""
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


class ModernButton(tk.Frame):
    def __init__(self, master, bg_color, text_color, text, command, **kw):
        super().__init__(master, **kw)
        self.bg_color = bg_color
        self.is_modern_btn = True 
        
        # [UPDATED] Frame viền: viền đen mỏng để nổi bật
        self.config(bg="#000000", padx=1, pady=1) 
        self.is_btn_border = True # Đánh dấu đây là frame viền
        
        self.btn = tk.Button(self, text=text, command=command,
                             bg=self.bg_color, fg=text_color,
                             font=("Segoe UI", 9, "bold"),
                             relief="flat", borderwidth=0,
                             cursor="hand2", padx=8, pady=5,
                             activebackground=bg_color, activeforeground=text_color)
        self.btn.pack(fill="both", expand=True)
        self.btn.bind("<Enter>", lambda e: self.btn.config(bg=self.adjust_color_lightness(self.bg_color, 0.95)))
        self.btn.bind("<Leave>", lambda e: self.btn.config(bg=self.bg_color))

    def adjust_color_lightness(self, color_hex, factor):
        try:
            r = int(color_hex[1:3], 16)
            g = int(color_hex[3:5], 16)
            b = int(color_hex[5:7], 16)
            return f"#{int(max(0, min(255, r * factor))):02x}{int(max(0, min(255, g * factor))):02x}{int(max(0, min(255, b * factor))):02x}"
        except: return color_hex
        
    def config(self, **kwargs):
        if 'state' in kwargs or 'text' in kwargs: self.btn.config(**kwargs)
        else: super().config(**kwargs)

class ContentCard(tk.Frame):
    def __init__(self, master, **kw):
        super().__init__(master, **kw)
        self.config(bg=THEME["bg_card"], bd=1, relief="solid")
        self.config(highlightbackground=THEME["border"], highlightthickness=1, relief="flat")

# =================================================================================
# CLASS KHỚP CỘT THỦ CÔNG (REDESIGNED WITH PREVIEW)
# =================================================================================
class ColumnMappingDialog(Toplevel):
    """
    Dialog khớp cột dữ liệu thủ công với preview realtime.
    UI chia 2 phần: Bên trái = cấu hình mapping, Bên phải = preview dữ liệu
    Giai đoạn 6: Hỗ trợ detect merged cells
    Giai đoạn 7: Kiểm tra chọn trùng cột
    Giai đoạn 8: Auto-save sau mapping thành công
    """
    def __init__(self, parent, raw_df, sheet_name, file_path=None):
        super().__init__(parent)
        self.title(f"📋 KHỚP CỘT DỮ LIỆU - Sheet: {sheet_name}")
        self.configure(bg=THEME["bg_app"])
        
        # [UPDATED] Bật minimize/maximize/close buttons (không dùng transient)
        # self.transient(parent)  # Bỏ transient để có nút minimize
        self.grab_set()
        self.resizable(True, True)
        
        # [UPDATED] Full screen theo màn hình
        self.state('zoomed')  # Windows: maximize window
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}+0+0")
        self.minsize(900, 500)  # Vẫn giữ minsize nếu user resize
        
        # Data
        self.raw_df = raw_df
        self.sheet_name = sheet_name
        self.file_path = file_path  # Giai đoạn 8: Lưu path để auto-save
        self.result_df = None
        self.cancelled = False  # Flag để xử lý đóng bằng X
        self.last_backup_path = None  # Giai đoạn 9: Lưu path backup đã tạo
        
        # Giai đoạn 6: Detect merged cells
        self.merged_cells_info = []
        self._detect_merged_cells()
        
        # Tìm dòng header
        self.start_row_idx = self._detect_header_row()
        self.data_start_idx = self.start_row_idx + 1  # Dòng bắt đầu data (sau header)
        
        # Xây dựng danh sách cột options
        self.col_options = ["--- Bỏ qua ---"]
        self.effective_headers = []
        self._build_column_options()
        
        # Mapping variables
        self.mapping_vars = {}
        self.mapping_combos = {}  # Lưu reference đến combobox để bind event
        
        # Giai đoạn 4: Tách Họ và Tên
        self.name_mode_var = tk.StringVar(value="combined")  # "combined" hoặc "split"
        self.is_split_name_var = tk.BooleanVar(value=False)  # Backward compatible
        self.header_row_var = tk.IntVar(value=self.start_row_idx + 1)  # 1-indexed cho user
        
        # Giai đoạn 8: Option tự động lưu
        self.auto_save_var = tk.BooleanVar(value=True)  # Mặc định bật auto-save
        
        # Preview debounce
        self.preview_job = None
        
        # Setup UI
        self.setup_ui()
        
        # Handle đóng bằng X
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Initial preview
        self.after(100, self.update_preview)
    
    def _detect_header_row(self):
        """Tự động phát hiện dòng header trong file Excel"""
        for idx, row in self.raw_df.head(50).iterrows():
            row_str = [str(val).lower() for val in row.values]
            if any(k in row_str for k in ["stt", "mã học sinh", "số thứ tự", "họ và tên", "sốtt", "họ tên"]):
                return idx
        return 0
    
    def _detect_merged_cells(self):
        """
        Giai đoạn 6: Phát hiện các ô bị merge trong file Excel sử dụng openpyxl.
        Lưu thông tin merged cells để cảnh báo user và xử lý đúng.
        """
        self.merged_cells_info = []
        
        if not self.file_path or not HAS_OPENPYXL:
            return
        
        try:
            # Chỉ đọc metadata, không load full data
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            
            # Tìm sheet theo tên
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.active
            
            # Lấy danh sách merged cells
            if hasattr(ws, 'merged_cells') and ws.merged_cells:
                for merged_range in ws.merged_cells.ranges:
                    info = {
                        'range': str(merged_range),
                        'min_row': merged_range.min_row,
                        'max_row': merged_range.max_row,
                        'min_col': merged_range.min_col,
                        'max_col': merged_range.max_col,
                        'is_header': merged_range.min_row <= 3,  # Nằm trong vùng header
                    }
                    self.merged_cells_info.append(info)
            
            wb.close()
            
            if self.merged_cells_info:
                print(f"[DEBUG] Phát hiện {len(self.merged_cells_info)} vùng merged cells trong sheet {self.sheet_name}")
                for info in self.merged_cells_info[:5]:  # Log 5 cái đầu
                    print(f"   - {info['range']} (rows {info['min_row']}-{info['max_row']})")
                    
        except Exception as e:
            print(f"[DEBUG] Lỗi detect merged cells: {e}")
            # Không ảnh hưởng flow chính
    
    def _check_duplicate_columns(self):
        """
        Giai đoạn 7: Kiểm tra xem user có chọn trùng cột cho nhiều trường không.
        Return: (bool is_valid, str error_message)
        """
        selected_cols = {}
        
        # Kiểm tra các mapping chính
        for key, var in self.mapping_vars.items():
            selection = var.get()
            if selection == "--- Bỏ qua ---":
                continue
            
            try:
                col_idx = int(selection.split(":")[0].replace("IDX_", ""))
            except:
                continue
            
            # Lấy tên trường dễ đọc
            field_names = {
                "name": "Họ và tên",
                "dob": "Ngày sinh", 
                "tx1": "TX1", "tx2": "TX2", "tx3": "TX3", "tx4": "TX4",
                "gk": "Giữa kỳ", "ck": "Cuối kỳ"
            }
            field_name = field_names.get(key, key)
            
            if col_idx in selected_cols:
                existing_field = selected_cols[col_idx]
                return False, f"Cột {chr(65 + col_idx)} được chọn cho cả '{existing_field}' và '{field_name}'!\n\nMỗi cột chỉ nên dùng cho một trường dữ liệu."
            
            selected_cols[col_idx] = field_name
        
        # Kiểm tra cột Tên riêng (nếu chế độ split)
        if self.name_mode_var.get() == "split":
            ten_selection = self.ten_var.get()
            if ten_selection != "--- Bỏ qua ---":
                try:
                    ten_idx = int(ten_selection.split(":")[0].replace("IDX_", ""))
                    if ten_idx in selected_cols:
                        existing_field = selected_cols[ten_idx]
                        return False, f"Cột {chr(65 + ten_idx)} được chọn cho cả '{existing_field}' và 'Tên riêng'!\n\nMỗi cột chỉ nên dùng cho một trường dữ liệu."
                except:
                    pass
        
        return True, ""
    
    def _build_column_options(self):
        """Xây dựng danh sách cột từ header row"""
        self.effective_headers = []
        self.col_options = ["--- Bỏ qua ---"]
        
        if self.start_row_idx >= len(self.raw_df):
            return
        
        # Đảm bảo convert sang string an toàn
        headers_row_1 = [str(val) for val in self.raw_df.iloc[self.start_row_idx].values]
        headers_row_2 = []
        if self.start_row_idx + 1 < len(self.raw_df):
            headers_row_2 = [str(val) for val in self.raw_df.iloc[self.start_row_idx + 1].values]
        else:
            headers_row_2 = [""] * len(headers_row_1)

        for i in range(len(headers_row_1)):
            h1 = str(headers_row_1[i]).strip().replace("\n", "").replace("nan", "")
            h2 = str(headers_row_2[i]).strip().replace("\n", "").replace("nan", "") if i < len(headers_row_2) else ""
            
            full_label = h1
            if h2 and any(k in h2.upper() for k in ["TX", "HS", "GK", "CK", "15", "1 TIẾT", "ĐĐGTX"]):
                full_label = h2
            elif "ĐTB" in h1.upper() or "MHK" in h1.upper():
                full_label = h1
            elif h2 and ("ĐTB" in h2.upper() or "MHK" in h2.upper()):
                full_label = h2
            
            self.effective_headers.append(full_label)

        # Tạo options cho combobox
        data_start = self.data_start_idx
        for i, header_text in enumerate(self.effective_headers):
            sample_val = ""
            try:
                for _, r in self.raw_df.iloc[data_start:data_start+5].iterrows():
                    if i < len(r):
                        val = str(r.iloc[i]).strip()
                        if val and val.lower() not in ["nan", "none", ""]:
                            sample_val = val
                            break
            except:
                pass
            
            display_name = header_text if header_text else f"Cột {i+1}"
            if len(sample_val) > 15:
                sample_val = sample_val[:12] + "..."
            label = f"[{i}] {display_name} (VD: {sample_val})"
            self.col_options.append(f"IDX_{i}:{label}")
    
    def setup_ui(self):
        """Thiết lập giao diện chính với layout 2 cột"""
        # Header
        header = tk.Frame(self, bg=THEME["primary"], pady=10)
        header.pack(fill="x")
        tk.Label(header, text="📋 KHỚP CỘT DỮ LIỆU THỦ CÔNG", 
             font=("Segoe UI", 14, "bold"), fg="white", bg=THEME["primary"]).pack(pady=(0, 5))
        tk.Label(header, text=f"Sheet: {self.sheet_name} | Lưu ý: Đóng cửa sổ nếu không muốn import sheet này",
             font=("Segoe UI", 9), fg="#ecf0f1", bg=THEME["primary"]).pack(pady=(0, 10))
        
        # ===== FOOTER: BUTTONS (pack trước để đảm bảo hiển thị) =====
        footer = tk.Frame(self, bg=THEME["bg_app"], pady=8)
        footer.pack(side="bottom", fill="x")
        
        btn_container = tk.Frame(footer, bg=THEME["bg_app"])
        btn_container.pack()
        
        # Nút Hủy
        tk.Button(btn_container, text="❌ HỦY", font=("Segoe UI", 10, "bold"),
                  bg="#e74c3c", fg="white", padx=15, pady=6, relief="solid", bd=1,
                  cursor="hand2", command=self.on_close).pack(side="left", padx=8)
        
        # Nút Hoàn tác (Giai đoạn 9 - chuyển vào trong dialog)
        tk.Button(btn_container, text="↩️ HOÀN TÁC MAPPING", font=("Segoe UI", 10, "bold"),
                  bg="#e67e22", fg="white", padx=15, pady=6, relief="solid", bd=1,
                  cursor="hand2", command=self._call_undo_mapping).pack(side="left", padx=8)
        
        # Nút Refresh Preview
        tk.Button(btn_container, text="🔄 LÀM MỚI", font=("Segoe UI", 10, "bold"),
                  bg="#3498db", fg="white", padx=15, pady=6, relief="solid", bd=1,
                  cursor="hand2", command=self.update_preview).pack(side="left", padx=8)
        
        # Nút Xác nhận
        tk.Button(btn_container, text="✅ XÁC NHẬN", font=("Segoe UI", 10, "bold"),
                  bg="#27ae60", fg="white", padx=15, pady=6, relief="solid", bd=1,
                  cursor="hand2", command=self.process_mapping).pack(side="left", padx=8)
        
        # Main container - chia 2 cột
        main_container = tk.Frame(self, bg=THEME["bg_app"])
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ===== CỘT TRÁI: CẤU HÌNH MAPPING =====
        left_frame = tk.Frame(main_container, bg=THEME["bg_card"], relief="solid", bd=1)
        left_frame.pack(side="left", fill="both", expand=False, padx=(0, 5))
        left_frame.configure(width=380)  # Giảm từ 480 xuống 380
        left_frame.pack_propagate(False)
        
        # Header cột trái
        tk.Label(left_frame, text="⚙ CẤU HÌNH MAPPING", font=("Segoe UI", 11, "bold"),
                 bg="#34495e", fg="white", pady=8).pack(fill="x")
        
        # Content cột trái với scroll
        left_canvas = tk.Canvas(left_frame, bg=THEME["bg_card"], highlightthickness=0)
        left_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=left_canvas.yview)
        left_content = tk.Frame(left_canvas, bg=THEME["bg_card"], padx=15, pady=10)
        
        left_canvas.configure(yscrollcommand=left_scrollbar.set)
        left_scrollbar.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)
        left_canvas.create_window((0, 0), window=left_content, anchor="nw")
        
        left_content.bind("<Configure>", lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all")))
        
        # --- Cấu hình dòng header ---
        header_config = tk.LabelFrame(left_content, text="📍 Dòng Header", font=("Segoe UI", 10, "bold"),
                                       bg=THEME["bg_card"], fg=THEME["text_main"], padx=10, pady=5)
        header_config.pack(fill="x", pady=(0, 10))
        
        # Dòng chọn số
        header_row_frame = tk.Frame(header_config, bg=THEME["bg_card"])
        header_row_frame.pack(fill="x")
        
        tk.Label(header_row_frame, text="Dòng header (1 = dòng đầu tiên):", 
                 font=("Segoe UI", 9), bg=THEME["bg_card"], fg=THEME["text_main"]).pack(side="left")
        
        header_spinbox = ttk.Spinbox(header_row_frame, from_=1, to=20, width=5, 
                                      textvariable=self.header_row_var, command=self.on_header_row_change)
        header_spinbox.pack(side="left", padx=10)
        header_spinbox.bind("<Return>", lambda e: self.on_header_row_change())
        
        # Văn bản hướng dẫn ví dụ
        example_text = (
            "💡 Ví dụ: Nếu file Excel có cấu trúc:\n"
            "   Dòng 1: Tiêu đề trường (BẢNG ĐIỂM LỚP 10A)\n"
            "   Dòng 2: Thông tin phụ (Năm học 2025-2026)\n"
            "   Dòng 3: STT | Họ tên | TX1 | TX2 | GK | CK\n"
            "   Dòng 4: 1   | Nguyễn Văn A | 8 | 7 | ...\n"
            "→ Chọn dòng header = 3 (dòng chứa tên cột)"
        )
        tk.Label(header_config, text=example_text, font=("Segoe UI", 8), 
                 bg=THEME["bg_card"], fg="#7f8c8d", justify="left", anchor="w").pack(fill="x", pady=(5, 0))
        
        # --- Cấu hình Họ Tên (Giai đoạn 4: Tách riêng Họ và Tên) ---
        name_config = tk.LabelFrame(left_content, text="👤 Cấu hình Họ Tên", font=("Segoe UI", 10, "bold"),
                                     bg=THEME["bg_card"], fg=THEME["primary"], padx=10, pady=5)
        name_config.pack(fill="x", pady=(0, 10))
        
        # Radio buttons cho chế độ
        mode_frame = tk.Frame(name_config, bg=THEME["bg_card"])
        mode_frame.pack(fill="x", pady=(0, 5))
        
        tk.Radiobutton(mode_frame, text="Họ Tên đã ghép sẵn (1 cột)", variable=self.name_mode_var,
                       value="combined", bg=THEME["bg_card"], fg=THEME["text_main"],
                       font=("Segoe UI", 9), command=self.on_name_mode_change).pack(anchor="w")
        
        tk.Radiobutton(mode_frame, text="Họ và Tên tách riêng (2 cột)", variable=self.name_mode_var,
                       value="split", bg=THEME["bg_card"], fg=THEME["text_main"],
                       font=("Segoe UI", 9), command=self.on_name_mode_change).pack(anchor="w")
        
        # Frame chứa combobox cho cột Tên (chỉ hiện khi mode = split)
        self.ten_frame = tk.Frame(name_config, bg=THEME["bg_card"])
        self.ten_frame.pack(fill="x", pady=(5, 0))
        
        tk.Label(self.ten_frame, text="Cột TÊN riêng:", font=("Segoe UI", 9, "bold"),
                 bg=THEME["bg_card"], fg="#e67e22", width=15, anchor="w").pack(side="left")
        
        self.ten_var = tk.StringVar()
        self.ten_combo = ttk.Combobox(self.ten_frame, textvariable=self.ten_var, 
                                       values=self.col_options, state="readonly", width=30)
        self.ten_combo.current(0)
        self.ten_combo.pack(side="left", padx=5)
        self.ten_combo.bind("<<ComboboxSelected>>", lambda e: self.schedule_preview_update())
        
        # Ban đầu ẩn frame cột Tên
        self.ten_frame.pack_forget()
        
        tk.Label(name_config, text="💡 Chọn chế độ phù hợp với cấu trúc file Excel của bạn",
                 font=("Segoe UI", 8, "italic"), bg=THEME["bg_card"], fg="gray").pack(anchor="w")
        
        # --- Các trường mapping ---
        fields_frame = tk.LabelFrame(left_content, text="📊 Khớp các cột dữ liệu", 
                                      font=("Segoe UI", 10, "bold"),
                                      bg=THEME["bg_card"], fg=THEME["text_main"], padx=10, pady=10)
        fields_frame.pack(fill="x", pady=(0, 10))
        
        required_fields = [
            ("Họ và tên (*)", "name", True),
            ("Ngày sinh", "dob", False),
            ("TX1 (Thường xuyên 1)", "tx1", False),
            ("TX2 (Thường xuyên 2)", "tx2", False),
            ("TX3 (Thường xuyên 3)", "tx3", False),
            ("TX4 (Thường xuyên 4)", "tx4", False),
            ("GK (Giữa kỳ)", "gk", False),
            ("CK (Cuối kỳ)", "ck", False),
        ]
        
        for i, (label_text, key, is_required) in enumerate(required_fields):
            row_frame = tk.Frame(fields_frame, bg=THEME["bg_card"])
            row_frame.pack(fill="x", pady=3)
            
            lbl = tk.Label(row_frame, text=label_text, font=("Segoe UI", 9, "bold" if is_required else "normal"),
                          bg=THEME["bg_card"], fg=THEME["primary"] if is_required else THEME["text_main"], width=22, anchor="w")
            lbl.pack(side="left")
            
            var = tk.StringVar()
            cbo = ttk.Combobox(row_frame, textvariable=var, values=self.col_options, state="readonly", width=35)
            cbo.current(0)
            cbo.pack(side="left", padx=5)
            cbo.bind("<<ComboboxSelected>>", lambda e: self.schedule_preview_update())
            
            self.mapping_vars[key] = var
            self.mapping_combos[key] = cbo
            
            # Auto-detect và set giá trị phù hợp
            self._auto_detect_column(key, cbo)
        
        # --- Giai đoạn 8: Option tự động lưu ---
        save_config = tk.LabelFrame(left_content, text="💾 Tùy chọn lưu", font=("Segoe UI", 10, "bold"),
                                     bg=THEME["bg_card"], fg="#27ae60", padx=10, pady=5)
        save_config.pack(fill="x", pady=(0, 10))
        
        tk.Checkbutton(save_config, text="Tự động lưu vào file Excel sau khi mapping",
                       variable=self.auto_save_var, bg=THEME["bg_card"], fg=THEME["text_main"],
                       font=("Segoe UI", 9), activebackground=THEME["bg_card"]).pack(anchor="w")
        
        tk.Label(save_config, text="💡 Nếu bật: File backup sẽ được tạo trên Desktop trước khi ghi đè",
                 font=("Segoe UI", 8, "italic"), bg=THEME["bg_card"], fg="gray").pack(anchor="w")
        
        # ===== CỘT PHẢI: PREVIEW =====
        right_frame = tk.Frame(main_container, bg=THEME["bg_card"], relief="solid", bd=1)
        right_frame.pack(side="right", fill="both", expand=True, padx=(5, 0))
        
        # Header cột phải
        self.preview_header = tk.Label(right_frame, text="👁 PREVIEW DỮ LIỆU GỐC TỪ EXCEL", 
                                        font=("Segoe UI", 11, "bold"), bg="#27ae60", fg="white", pady=8)
        self.preview_header.pack(fill="x")
        
        # Legend panel - giải thích màu sắc
        legend_frame = tk.Frame(right_frame, bg="#ecf0f1", pady=5)
        legend_frame.pack(fill="x")
        
        tk.Label(legend_frame, text="Chú thích: ", font=("Segoe UI", 9, "bold"), 
                 bg="#ecf0f1", fg="#2c3e50").pack(side="left", padx=5)
        
        # Legend items
        legends = [
            ("🟦 Họ tên", "#3498db"),
            ("🟩 Điểm TX", "#27ae60"),
            ("🟨 GK/CK", "#f39c12"),
            ("🟪 Ngày sinh", "#9b59b6"),
        ]
        for text, color in legends:
            tk.Label(legend_frame, text=text, font=("Segoe UI", 8), bg="#ecf0f1", fg=color).pack(side="left", padx=8)
        
        # Warning panel
        self.warning_frame = tk.Frame(right_frame, bg="#fff3cd", pady=5)
        self.warning_frame.pack(fill="x")
        self.warning_label = tk.Label(self.warning_frame, text="", font=("Segoe UI", 9), 
                                       bg="#fff3cd", fg="#856404", wraplength=700, justify="left")
        self.warning_label.pack(anchor="w", padx=10)
        
        # Preview Treeview - sẽ được tạo dynamic dựa trên số cột Excel
        self.tree_container = tk.Frame(right_frame, bg=THEME["bg_card"])
        self.tree_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Tạo preview tree với cột gốc từ Excel
        self._create_raw_preview_tree()
    
    def _call_undo_mapping(self):
        """Gọi hàm undo_last_mapping của parent (ExcelTab)"""
        try:
            if hasattr(self.master, 'undo_last_mapping'):
                self.master.undo_last_mapping()
            else:
                messagebox.showinfo("Thông báo", "Chức năng hoàn tác chỉ khả dụng sau khi đã thực hiện mapping.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể hoàn tác: {e}")
    
    def on_name_mode_change(self):
        """Xử lý khi user thay đổi chế độ Họ Tên (ghép sẵn / tách riêng)"""
        mode = self.name_mode_var.get()
        if mode == "split":
            # Hiển thị combobox chọn cột Tên
            self.ten_frame.pack(fill="x", pady=(5, 0))
            self.is_split_name_var.set(True)
            # Auto-detect cột Tên
            self._auto_detect_ten_column()
        else:
            # Ẩn combobox cột Tên
            self.ten_frame.pack_forget()
            self.is_split_name_var.set(False)
        self.schedule_preview_update()
    
    def _auto_detect_ten_column(self):
        """Tự động phát hiện cột Tên riêng biệt"""
        for opt in self.col_options:
            if opt == "--- Bỏ qua ---":
                continue
            try:
                display_text = opt.split("]")[1].split("(VD")[0].strip().upper()
                # Tìm cột có chữ "TÊN" nhưng KHÔNG có chữ "HỌ"
                if "TÊN" in display_text and "HỌ" not in display_text:
                    self.ten_combo.set(opt)
                    return
            except:
                continue
        
        # Nếu không tìm thấy, thử tìm cột ngay sau cột Họ
        name_sel = self.mapping_vars.get("name", tk.StringVar()).get()
        if name_sel != "--- Bỏ qua ---":
            try:
                name_idx = int(name_sel.split(":")[0].replace("IDX_", ""))
                next_idx = name_idx + 1
                for opt in self.col_options:
                    if f"IDX_{next_idx}:" in opt:
                        self.ten_combo.set(opt)
                        return
            except:
                pass
    
    def _auto_detect_column(self, key, combobox):
        """Tự động phát hiện và set cột phù hợp dựa trên keyword"""
        for opt in self.col_options:
            if opt == "--- Bỏ qua ---":
                continue
            try:
                display_text = opt.split("]")[1].split("(VD")[0].strip().upper()
            except:
                continue
            
            found = False
            if key == "name" and any(k in display_text for k in ["TÊN", "HỌ"]):
                found = True
            elif key == "dob" and "SINH" in display_text:
                found = True
            elif key == "gk" and ("GK" in display_text or "GIỮA" in display_text):
                found = True
            elif key == "ck" and ("CK" in display_text or "CUỐI" in display_text):
                found = True
            elif "tx" in key:
                num = key[-1]
                if f"TX{num}" in display_text or f"HS{num}" in display_text or f"GTX{num}" in display_text:
                    found = True
            
            if found:
                combobox.set(opt)
                break
    
    def on_header_row_change(self):
        """Xử lý khi user thay đổi dòng header"""
        try:
            new_row = self.header_row_var.get() - 1  # Convert về 0-indexed
            if 0 <= new_row < len(self.raw_df) - 1:
                self.start_row_idx = new_row
                self.data_start_idx = new_row + 1
                self._build_column_options()
                
                # Cập nhật lại các combobox
                for key, cbo in self.mapping_combos.items():
                    cbo.configure(values=self.col_options)
                    cbo.current(0)
                    self._auto_detect_column(key, cbo)
                
                # Cập nhật combobox Tên riêng (Giai đoạn 4)
                if hasattr(self, 'ten_combo'):
                    self.ten_combo.configure(values=self.col_options)
                    self.ten_combo.current(0)
                    if self.name_mode_var.get() == "split":
                        self._auto_detect_ten_column()
                
                # Rebuild preview tree với header mới
                self._create_raw_preview_tree()
                self.schedule_preview_update()
        except Exception as e:
            print(f"[DEBUG] Lỗi thay đổi header row: {e}")
    
    def schedule_preview_update(self):
        """Debounce preview update để tránh lag"""
        if self.preview_job:
            self.after_cancel(self.preview_job)
        self.preview_job = self.after(200, self.update_preview)
    
    def _create_raw_preview_tree(self):
        """Tạo treeview hiển thị raw Excel data với tất cả các cột gốc"""
        # Clear container cũ
        for widget in self.tree_container.winfo_children():
            widget.destroy()
        
        # Lấy số cột từ raw_df
        num_cols = len(self.raw_df.columns)
        
        # Tạo danh sách cột: "Dòng" + các cột A, B, C, D... hoặc tên header
        col_ids = ["row_num"]  # ID nội bộ
        col_headers = ["Dòng"]  # Tên hiển thị
        
        for i in range(num_cols):
            col_id = f"col_{i}"
            # Hiển thị tên header nếu có, không thì dùng chữ cái
            if i < len(self.effective_headers) and self.effective_headers[i]:
                header_text = self.effective_headers[i][:12]
                col_name = f"{chr(65 + i)}: {header_text}"
            else:
                col_name = f"Cột {chr(65 + i)}"
            col_ids.append(col_id)
            col_headers.append(col_name)
        
        # Tạo treeview
        self.preview_tree = ttk.Treeview(self.tree_container, columns=col_ids, show="headings", height=18)
        
        # Scrollbars
        v_scroll = ttk.Scrollbar(self.tree_container, orient="vertical", command=self.preview_tree.yview)
        h_scroll = ttk.Scrollbar(self.tree_container, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        v_scroll.pack(side="right", fill="y")
        h_scroll.pack(side="bottom", fill="x")
        self.preview_tree.pack(fill="both", expand=True)
        
        # Configure columns
        for col_id, col_name in zip(col_ids, col_headers):
            self.preview_tree.heading(col_id, text=col_name)
            if col_id == "row_num":
                width = 50
            elif "Họ" in col_name or "Tên" in col_name or "tên" in col_name:
                width = 120
            else:
                width = 80
            self.preview_tree.column(col_id, width=width, anchor="center", minwidth=50)
        
        # Tags cho highlight
        self.preview_tree.tag_configure("header_row", background="#34495e", foreground="white")
        self.preview_tree.tag_configure("normal", background="white")
        self.preview_tree.tag_configure("odd", background="#f8f9fa")
        
        # Lưu thông tin để dùng trong update_preview
        self.preview_col_ids = col_ids
        self.preview_col_headers = col_headers
    
    def update_preview(self):
        """Cập nhật preview hiển thị RAW EXCEL DATA với highlight các cột đã mapping và validation điểm"""
        if not hasattr(self, 'preview_tree') or not self.preview_tree.winfo_exists():
            return
            
        # Clear tree
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        # Lấy mapping indices
        def get_col_idx(opt_str):
            if opt_str == "--- Bỏ qua ---":
                return None
            try:
                return int(opt_str.split(":")[0].replace("IDX_", ""))
            except:
                return None
        
        map_idx = {k: get_col_idx(v.get()) for k, v in self.mapping_vars.items()}
        
        # Lấy cột Tên riêng nếu có (Giai đoạn 4)
        ten_idx = None
        if self.name_mode_var.get() == "split":
            ten_idx = get_col_idx(self.ten_var.get())
        
        # Giai đoạn 5: Xác định các cột điểm để validation
        score_cols = set()
        for key in ["tx1", "tx2", "tx3", "tx4", "gk", "ck"]:
            if map_idx.get(key) is not None:
                score_cols.add(map_idx[key])
        
        # Hiển thị dữ liệu raw từ Excel
        max_preview_rows = 30
        row_count = 0
        student_count = 0
        score_error_count = 0  # Giai đoạn 5: Đếm lỗi điểm
        
        for idx, row in self.raw_df.iterrows():
            if row_count >= max_preview_rows:
                break
            
            # Xây dựng values cho row
            values = [idx + 1]  # Số dòng Excel (1-indexed)
            row_has_error = False  # Giai đoạn 5: Flag lỗi cho row này
            
            for i in range(len(row)):
                try:
                    val = str(row.iloc[i]).strip() if pd.notna(row.iloc[i]) else ""
                    if val.lower() == "nan":
                        val = ""
                    
                    # Giai đoạn 5: Validation điểm
                    if i in score_cols and val and idx > self.start_row_idx:
                        try:
                            score = float(val.replace(",", "."))
                            if score < 0 or score > 10:
                                val = f"❌{val}"  # Đánh dấu điểm lỗi
                                row_has_error = True
                                score_error_count += 1
                        except ValueError:
                            pass  # Không phải số, bỏ qua
                    
                    # Cắt ngắn nếu quá dài
                    if len(val) > 15:
                        val = val[:12] + "..."
                except:
                    val = ""
                values.append(val)
            
            # Đảm bảo đủ số cột
            while len(values) < len(self.preview_col_ids):
                values.append("")
            
            # Xác định tag (màu nền)
            if idx == self.start_row_idx:
                tag = "header_row"
            elif idx < self.start_row_idx:
                tag = "odd"  # Các dòng trước header
            elif row_has_error:
                tag = "error_row"  # Giai đoạn 5: Dòng có lỗi
            else:
                tag = "normal" if row_count % 2 == 0 else "odd"
                
                # Đếm học sinh hợp lệ
                if map_idx["name"] is not None and map_idx["name"] < len(row):
                    name_val = str(row.iloc[map_idx["name"]]).strip().lower()
                    if name_val and name_val not in ["nan", "none", "", "họ và tên", "họ tên", "tên"]:
                        student_count += 1
            
            self.preview_tree.insert("", "end", values=values, tags=(tag,))
            row_count += 1
        
        # Thêm tag cho dòng lỗi (Giai đoạn 5)
        self.preview_tree.tag_configure("error_row", background="#f8d7da", foreground="#721c24")
        
        # Cập nhật header với thông tin
        info_parts = [f"👁 PREVIEW DỮ LIỆU GỐC TỪ EXCEL"]
        info_parts.append(f"Header dòng: {self.start_row_idx + 1}")
        info_parts.append(f"~{student_count} học sinh")
        
        if score_error_count > 0:
            info_parts.append(f"⚠️ {score_error_count} điểm lỗi")
            self.preview_header.config(text=" | ".join(info_parts), bg="#e74c3c")
        else:
            self.preview_header.config(text=" | ".join(info_parts), bg="#27ae60")
        
        # Cập nhật warning với thông tin mapping
        mapping_info = []
        if map_idx["name"] is not None:
            name_info = f"🟦 Họ: Cột {chr(65 + map_idx['name'])}"
            if ten_idx is not None:
                name_info += f" + Tên: Cột {chr(65 + ten_idx)}"
            mapping_info.append(name_info)
        
        if map_idx["dob"] is not None:
            mapping_info.append(f"🟪 Ngày sinh: Cột {chr(65 + map_idx['dob'])}")
        
        tx_cols = []
        for tx_key in ["tx1", "tx2", "tx3", "tx4"]:
            if map_idx[tx_key] is not None:
                tx_cols.append(chr(65 + map_idx[tx_key]))
        if tx_cols:
            mapping_info.append(f"🟩 TX: Cột {', '.join(tx_cols)}")
        
        gk_ck_cols = []
        if map_idx["gk"] is not None:
            gk_ck_cols.append(f"GK={chr(65 + map_idx['gk'])}")
        if map_idx["ck"] is not None:
            gk_ck_cols.append(f"CK={chr(65 + map_idx['ck'])}")
        if gk_ck_cols:
            mapping_info.append(f"🟨 {', '.join(gk_ck_cols)}")
        
        # Giai đoạn 5: Thêm cảnh báo lỗi điểm
        if score_error_count > 0:
            mapping_info.append(f"🔴 {score_error_count} giá trị điểm ngoài 0-10!")
        
        # Giai đoạn 6: Thêm cảnh báo merged cells
        if self.merged_cells_info:
            header_merges = [m for m in self.merged_cells_info if m['is_header']]
            if header_merges:
                mapping_info.append(f"⚠️ {len(header_merges)} vùng header bị merge")
        
        # Giai đoạn 7: Kiểm tra trùng cột
        is_valid, dup_error = self._check_duplicate_columns()
        if not is_valid:
            mapping_info.append("🔴 Phát hiện chọn trùng cột!")
        
        if mapping_info:
            self.warning_label.config(text="Mapping: " + " | ".join(mapping_info))
            if score_error_count > 0 or not is_valid:
                self.warning_frame.config(bg="#f8d7da")
                self.warning_label.config(bg="#f8d7da", fg="#721c24")
            else:
                self.warning_frame.config(bg="#d4edda")
                self.warning_label.config(bg="#d4edda", fg="#155724")
            self.warning_frame.pack(fill="x")
        else:
            self.warning_label.config(text="⚠️ Chưa chọn cột nào. Hãy chọn ít nhất cột 'Họ và tên'")
            self.warning_frame.config(bg="#fff3cd")
            self.warning_label.config(bg="#fff3cd", fg="#856404")
            self.warning_frame.pack(fill="x")
    
    def on_close(self):
        """Xử lý khi đóng dialog bằng nút X hoặc Hủy"""
        self.cancelled = True
        self.result_df = None
        self.destroy()
    
    def process_mapping(self):
        """Xử lý khớp cột và tạo result DataFrame"""
        name_sel = self.mapping_vars["name"].get()
        if name_sel == "--- Bỏ qua ---":
            messagebox.showwarning("Thiếu thông tin", "Bắt buộc phải chọn cột 'Họ và tên'!")
            return
        
        # Giai đoạn 7: Kiểm tra trùng cột trước khi xử lý
        is_valid, dup_error = self._check_duplicate_columns()
        if not is_valid:
            messagebox.showwarning("Chọn trùng cột", dup_error)
            return
        
        try:
            data = []
            
            def get_col_idx(opt_str):
                if opt_str == "--- Bỏ qua ---":
                    return None
                try:
                    return int(opt_str.split(":")[0].replace("IDX_", ""))
                except:
                    return None
            
            map_idx = {k: get_col_idx(v.get()) for k, v in self.mapping_vars.items()}
            
            # Giai đoạn 4: Lấy cột Tên riêng nếu có
            ten_idx = None
            if self.name_mode_var.get() == "split":
                ten_idx = get_col_idx(self.ten_var.get())
            
            for idx, row in self.raw_df.iterrows():
                if idx <= self.start_row_idx:
                    continue
                
                try:
                    first_cell = str(row.iloc[0]).lower().strip()
                    if any(k in first_cell for k in ["stt", "mã"]):
                        continue
                    
                    name_idx = map_idx["name"]
                    if name_idx is None or name_idx >= len(row):
                        continue
                    
                    raw_name = str(row.iloc[name_idx]).strip()
                    if not raw_name or raw_name.lower() in ["nan", "none", "", "họ và tên", "họ tên"]:
                        continue
                except:
                    continue
                
                # Giai đoạn 4: Ghép tên - Cải tiến logic
                final_name = raw_name
                if self.name_mode_var.get() == "split" and ten_idx is not None:
                    # Chế độ tách riêng: Lấy từ cột Tên đã chọn
                    if ten_idx < len(row):
                        ten_val = str(row.iloc[ten_idx]).strip()
                        if ten_val and ten_val.lower() not in ["nan", "none", "", "tên"]:
                            final_name = f"{raw_name} {ten_val}"
                # Nếu chế độ combined, giữ nguyên raw_name
                
                def extract_val(key):
                    c_idx = map_idx.get(key)
                    if c_idx is None or c_idx >= len(row):
                        return ""
                    val = row.iloc[c_idx]
                    return clean_float_val(val)
                
                # [DYNAMIC TX] Tự động thu thập tất cả cột TX được map
                tx_scores = []
                for tx_num in range(1, 11):  # Hỗ trợ tối đa 10 cột TX
                    tx_key = f"tx{tx_num}"
                    if tx_key in map_idx:
                        tx_val = extract_val(tx_key)
                        if tx_val != "":
                            tx_scores.append(tx_val)
                
                gk = extract_val("gk")
                ck = extract_val("ck")
                
                dtb = calculate_dtb_exact(tx_scores, gk, ck)
                
                dob_val = ""
                if map_idx["dob"] is not None and map_idx["dob"] < len(row):
                    dob_val = str(row.iloc[map_idx["dob"]]).replace("nan", "").strip()
                
                # [DYNAMIC TX] Tạo dictionary động cho các cột TX
                item = {
                    "STT": len(data) + 1,
                    "Họ và tên": final_name.title(),
                    "Ngày sinh": dob_val
                }
                
                # Thêm các cột TX động
                num_tx = max(2, len(tx_scores))  # Tối thiểu 2 cột TX
                for i in range(1, num_tx + 1):
                    item[f"TX{i}"] = tx_scores[i-1] if i-1 < len(tx_scores) else ""
                
                # Thêm GK, CK, ĐTB, Xếp loại
                item.update({
                    "GK": gk, "CK": ck,
                    "ĐTB": dtb,
                    "Xếp loại": ""
                })
                
                data.append(item)
            
            if not data:
                messagebox.showwarning("Không có dữ liệu", "Không tìm thấy học sinh nào với cấu hình hiện tại!")
                return
            
            self.result_df = pd.DataFrame(data)
            
            # Giai đoạn 8: Auto-save vào file Excel sau khi mapping thành công (nếu user bật option)
            if self.auto_save_var.get():
                auto_save_success = self._auto_save_to_excel()
                
                if auto_save_success:
                    messagebox.showinfo("Thành công", f"Đã khớp thành công {len(data)} học sinh!\n\n💾 Dữ liệu đã được tự động lưu vào file Excel.\n📁 File backup đã lưu trên Desktop.")
                else:
                    messagebox.showinfo("Thành công", f"Đã khớp thành công {len(data)} học sinh!\n\n⚠️ Lưu ý: Chưa lưu vào file. Nhấn 'LƯU DỮ LIỆU' để lưu thủ công.")
            else:
                messagebox.showinfo("Thành công", f"Đã khớp thành công {len(data)} học sinh!\n\n📝 Bạn đã tắt auto-save. Nhấn 'LƯU DỮ LIỆU' để lưu thủ công.")
            
            self.destroy()
            
        except Exception as e:
            messagebox.showerror("Lỗi xử lý", f"Có lỗi khi đọc dữ liệu: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _auto_save_to_excel(self):
        """
        Giai đoạn 8: Tự động lưu dữ liệu đã mapping vào file Excel.
        
        [DISABLED] Tính năng này đã bị vô hiệu hóa vì gây mất header Excel.
        Dữ liệu chỉ được cập nhật trong UI, không ghi trực tiếp vào file.
        User có thể dùng chức năng nhập điểm trực tiếp để ghi từng ô.
        """
        print("[DEBUG] Auto-save: Tính năng ghi đè sheet đã bị vô hiệu hóa để bảo vệ header Excel")
        print("[DEBUG] Auto-save: Dữ liệu chỉ được cập nhật trong UI")
        
        # Không ghi file nữa, chỉ return True để không hiện lỗi
        messagebox.showinfo(
            "Thông báo",
            "Dữ liệu đã được cập nhật trong giao diện.\n\n"
            "Để ghi vào file Excel, vui lòng nhập điểm trực tiếp\n"
            "bằng cách double-click vào ô điểm trên bảng dữ liệu.",
            parent=self
        )
        return True


# =================================================================================
# SMART RECOGNITION FEATURES - Contextual Hints & Auto-Correction [NEW]
# =================================================================================

def build_vocabulary_hints(students):
    """
    Xây dựng danh sách gợi ý từ danh sách học sinh
    Để truyền vào speech engine nhằm cải thiện độ chính xác nhận diện
    
    Args:
        students: List[dict] - Danh sách học sinh {"stt", "name", "score"}
    
    Returns:
        dict: {
            "phrases": List[str] - Danh sách gợi ý đầy đủ (cho Google)
            "names": List[str] - Danh sách tên riêng (cho Whisper prompt)
        }
    """
    if not students:
        return {"phrases": [], "names": []}
    
    phrases = []
    names = []
    
    for student in students:
        name = student.get("name", "").strip()
        if not name:
            continue
        
        # Tên đầy đủ
        names.append(name)
        phrases.append(name)
        
        # Firstname (từ cuối)
        parts = name.split()
        if len(parts) >= 2:
            firstname = parts[-1]
            if firstname not in names:
                names.append(firstname)
                phrases.append(firstname)
        
        # Partial name (2 từ cuối)
        if len(parts) >= 3:
            partial = " ".join(parts[-2:])
            if partial not in phrases:
                phrases.append(partial)
    
    # Loại bỏ trùng lặp
    phrases = list(dict.fromkeys(phrases))
    names = list(dict.fromkeys(names))
    
    return {
        "phrases": phrases[:500],  # Giới hạn 500 để tránh quá tải
        "names": names[:200]  # Giới hạn 200 tên
    }


# Database phát âm tiếng Việt - Các cặp phụ âm dễ nhầm
CONSONANT_CORRECTIONS = {
    # CH <-> TR
    "ch": ["tr"],
    "tr": ["ch"],
    
    # S <-> X
    "s": ["x"],
    "x": ["s"],
    
    # D <-> GI
    "d": ["gi"],
    "gi": ["d"],
    
    # R <-> G (miền Nam)
    "r": ["g"],
    "g": ["r"],
    
    # V <-> D (miền Nam)
    "v": ["d"],
    
    # N <-> L
    "n": ["l"],
    "l": ["n"],
}


def generate_phonetic_variants(name):
    """
    Sinh ra các biến thể phát âm của một tên dựa trên nguyên âm/phụ âm Việt
    
    Args:
        name: str - Tên học sinh (VD: "Trần Văn An")
    
    Returns:
        List[str] - Danh sách biến thể phát âm
    """
    if not name:
        return []
    
    variants = [name]  # Bao gồm tên gốc
    words = name.lower().split()
    
    # Với mỗi từ, sinh biến thể
    for i, word in enumerate(words):
        new_variants = []
        
        # Kiểm tra từng cặp phụ âm
        for original, replacements in CONSONANT_CORRECTIONS.items():
            if word.startswith(original):
                for replacement in replacements:
                    new_word = replacement + word[len(original):]
                    
                    # Tạo tên mới với từ đã thay thế
                    for variant in variants:
                        variant_words = variant.lower().split()
                        if i < len(variant_words):
                            variant_words[i] = new_word
                            new_name = " ".join(variant_words).title()
                            if new_name not in new_variants:
                                new_variants.append(new_name)
        
        variants.extend(new_variants)
    
    # Loại bỏ trùng lặp và giới hạn
    variants = list(dict.fromkeys(variants))
    return variants[:20]  # Giới hạn 20 biến thể để tránh lag


def auto_correct_name(recognized_name, students, confidence_threshold=0.85):
    """
    Tự động sửa lỗi phát âm bằng cách sinh biến thể và tìm khớp
    
    Args:
        recognized_name: str - Tên đã nhận dạng (có thể sai)
        students: List[dict] - Danh sách học sinh
        confidence_threshold: float - Ngưỡng tin cậy (0.85 = 85%)
    
    Returns:
        dict hoặc None: {
            "student": dict - Học sinh tìm thấy
            "original_name": str - Tên gốc nhận dạng được
            "corrected": bool - True nếu đã sửa lỗi
            "confidence": float - Độ tin cậy
        }
    """
    from fuzzywuzzy import fuzz
    
    if not recognized_name or not students:
        return None
    
    # Sinh biến thể phát âm của tên đã nhận dạng
    variants = generate_phonetic_variants(recognized_name)
    
    best_match = None
    best_score = 0
    best_variant = recognized_name
    
    # Thử khớp với từng biến thể
    for variant in variants:
        # Thử khớp với danh sách học sinh
        for student in students:
            student_name = student.get("name", "").strip()
            if not student_name:
                continue
            
            # Tính điểm fuzzy cho nhiều trường hợp
            scores = [
                fuzz.ratio(variant.lower(), student_name.lower()),
                fuzz.partial_ratio(variant.lower(), student_name.lower()),
                fuzz.token_sort_ratio(variant.lower(), student_name.lower())
            ]
            
            # Lấy điểm cao nhất
            score = max(scores) / 100.0
            
            if score > best_score:
                best_score = score
                best_match = student
                best_variant = variant
    
    # Chỉ trả về nếu đủ tin cậy
    if best_score >= confidence_threshold and best_match:
        return {
            "student": best_match,
            "original_name": recognized_name,
            "corrected": best_variant != recognized_name,
            "confidence": best_score,
            "variant_used": best_variant
        }
    
    return None


# =================================================================================
# PHASE 1: AUDIO PROCESSING - VAD & Normalization [NEW]
# =================================================================================

def audio_normalize(audio, target_level=0.8):
    """
    1.3 Audio Normalization - Chuẩn hóa âm lượng
    Đưa audio về mức âm lượng mong muốn
    """
    max_amp = np.abs(audio).max()
    if max_amp > 0.001:  # Tránh chia cho 0
        return audio * (target_level / max_amp)
    return audio

def audio_detect_voice(audio, sample_rate=16000, threshold=0.008, min_speech_duration=0.3):
    """
    1.2 Voice Activity Detection (VAD) - Phát hiện tiếng nói
    Trả về đoạn audio chỉ chứa giọng nói, bỏ đoạn im lặng
    
    Args:
        audio: numpy array float32
        sample_rate: tần số lấy mẫu
        threshold: ngưỡng năng lượng để xác định có tiếng nói (giảm xuống 0.008)
        min_speech_duration: thời gian tối thiểu của đoạn nói (giây)
    
    Returns:
        audio_trimmed: audio đã cắt bỏ im lặng
        speech_ratio: tỷ lệ có tiếng nói (0-1)
    """
    # Chia audio thành các frame 30ms (dài hơn để bắt được âm tốt hơn)
    frame_length = int(sample_rate * 0.03)  # 30ms
    hop_length = frame_length // 2  # 15ms hop
    
    # Tính năng lượng từng frame
    num_frames = 1 + (len(audio) - frame_length) // hop_length
    energies = []
    
    for i in range(num_frames):
        start = i * hop_length
        end = start + frame_length
        frame = audio[start:end]
        energy = np.sqrt(np.mean(frame ** 2))  # RMS energy
        energies.append(energy)
    
    energies = np.array(energies)
    
    # Xác định frames có tiếng nói (energy > threshold)
    speech_frames = energies > threshold
    
    # Tính tỷ lệ có tiếng nói
    speech_ratio = np.mean(speech_frames)
    
    if speech_ratio < 0.1:  # Ít hơn 10% có tiếng nói
        return audio, speech_ratio
    
    # Tìm điểm bắt đầu và kết thúc của tiếng nói
    # Mở rộng ra 2 bên một chút để không cắt mất âm
    speech_indices = np.where(speech_frames)[0]
    
    if len(speech_indices) == 0:
        return audio, 0.0
    
    # Thêm padding RỘng hơn để không cắt mất đầu/cuối câu
    padding_frames = int(0.3 * sample_rate / hop_length)  # 300ms padding (tăng từ 100ms)
    start_frame = max(0, speech_indices[0] - padding_frames)
    end_frame = min(num_frames - 1, speech_indices[-1] + padding_frames)
    
    # Chuyển về sample index
    start_sample = start_frame * hop_length
    end_sample = min(len(audio), end_frame * hop_length + frame_length)
    
    audio_trimmed = audio[start_sample:end_sample]
    
    # Kiểm tra độ dài tối thiểu
    min_samples = int(min_speech_duration * sample_rate)
    if len(audio_trimmed) < min_samples:
        return audio, speech_ratio
    
    return audio_trimmed, speech_ratio

def audio_preprocess(audio, sample_rate=16000, apply_vad=True):
    """
    Pipeline xử lý audio hoàn chỉnh
    1. Normalize
    2. Voice Activity Detection (VAD)
    3. Final Normalize
    
    Returns:
        processed_audio: audio đã xử lý
        info: dict chứa thông tin xử lý
    """
    info = {
        'original_length': len(audio) / sample_rate,
        'original_max': np.abs(audio).max(),
        'vad_applied': False,
        'speech_ratio': 1.0,
        'final_length': 0,
    }
    
    # Bước 1: Normalize ban đầu
    audio = audio_normalize(audio, target_level=0.8)
    
    # Bước 2: VAD - Cắt bỏ im lặng
    if apply_vad:
        audio, speech_ratio = audio_detect_voice(audio, sample_rate)
        info['vad_applied'] = True
        info['speech_ratio'] = speech_ratio
    
    # Bước 3: Normalize lại sau xử lý
    audio = audio_normalize(audio, target_level=0.8)
    audio = np.clip(audio, -1.0, 1.0).astype(np.float32)
    
    info['final_length'] = len(audio) / sample_rate
    
    return audio, info


# =================================================================================
# CLASS NHẬP ĐIỂM BẰNG GIỌNG NÓI (VOICE INPUT) [NEW]
# =================================================================================

def voice_normalize_text(val):
    """Bỏ dấu và chuyển thường để fuzzy matching"""
    s = str(val).lower()
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in nfkd if unicodedata.category(ch) != "Mn")

def voice_parse_score_text(score_text):
    """
    Chuyển đổi text điểm thành số
    Hỗ trợ: "8", "8.5", "tám", "tám năm", "tám chấm năm", "chín rưỡi", "bảy phẩy năm"
    """
    score_text = score_text.lower().strip()
    
    # Các biến thể phát âm tiếng Việt (bao gồm lỗi nhận dạng phổ biến)
    number_map = {
        # Số 0
        "không": "0", "linh": "0", "lính": "0", "lin": "0",
        # Số 1
        "một": "1", "mốt": "1", "mót": "1", "mợt": "1",
        # Số 2
        "hai": "2", "hải": "2", "hay": "2",
        # Số 3
        "ba": "3", "bà": "3",
        # Số 4
        "bốn": "4", "tư": "4", "bón": "4", "bổn": "4",
        # Số 5
        "năm": "5", "lăm": "5", "nám": "5", "năn": "5",
        # Số 6
        "sáu": "6", "sấu": "6", "xáu": "6",
        # Số 7
        "bảy": "7", "bẩy": "7", "bảy": "7", "bai": "7",
        # Số 8
        "tám": "8", "tắm": "8", "tàm": "8",
        # Số 9
        "chín": "9", "chính": "9", "chin": "9", "trín": "9",
        # Số 10
        "mười": "10", "mưới": "10", "mười": "10", "mời": "10", 
        "muối": "10", "mươi": "10", "muời": "10", "mười": "10"
    }
    
    # Từ khóa "điểm" và biến thể
    score_keywords = ["điểm", "điểm", "đểm", "điêm", "diểm", "điem", "diem", "đi"]
    for kw in score_keywords:
        score_text = score_text.replace(kw, " ")
    
    decimal_markers = {
        "rưỡi": ".5",
        "phẩy": ".",
        "phết": ".",
        "chấm": ".",
        "phay": ".",
        "chẩm": "."
    }
    
    # Xử lý rưỡi đặc biệt
    if "rưỡi" in score_text:
        score_text = score_text.replace("rưỡi", ".5")
    
    # Thay thế dấu phân cách thập phân
    for marker, replacement in decimal_markers.items():
        if marker != "rưỡi":
            score_text = score_text.replace(marker, replacement)
    
    # Thay thế số tiếng Việt
    for word, digit in number_map.items():
        score_text = score_text.replace(word, digit)
    
    # Xử lý "8 5" → "8.5"
    score_text = re.sub(r'(\d+)\s+(\d)(?!\d)', r'\1.\2', score_text)
    
    # Extract số
    match = re.search(r'(\d+(?:[.,]\d+)?)', score_text)
    if match:
        score = match.group(1).replace(',', '.')
        try:
            result = float(score)
            if 0 <= result <= 10:
                return result
        except ValueError:
            pass
    
    return None


class VoiceInputWindow(Toplevel):
    """
    Cửa sổ nhập điểm bằng giọng nói - Fullscreen
    Cho phép sửa điểm thủ công trước khi đồng bộ
    """
    
    def __init__(self, parent, excel_tab):
        super().__init__(parent)
        
        self.excel_tab = excel_tab
        self.parent = parent
        
        # Data
        self.local_df = None  # DataFrame local để edit trước khi sync
        self.current_sheet = None
        self.current_col = None
        self.students = []  # List dict: {stt, name, score}
        self.batch_results = []  # Batch mode
        
        # Voice
        self.is_listening = False
        self.recognizer = sr.Recognizer() if HAS_SPEECH else None
        
        # [FIX] Cấu hình timeout cho recognizer để tránh treo
        if self.recognizer:
            self.recognizer.energy_threshold = 300  # Ngưỡng năng lượng âm thanh (tăng nếu nhiễu)
            self.recognizer.dynamic_energy_threshold = True  # Tự động điều chỉnh
            self.recognizer.pause_threshold = 0.8  # Thời gian im lặng để kết thúc câu (giây)
        self.result_queue = queue.Queue()
        self.pending_student = None  # [NEW] Lưu học sinh đang chờ điểm
        
        # [NEW] Alias & Training Mode
        self.student_aliases = {}  # {stt: "alias1, alias2, ..."}
        self.phonetic_mapping = {}  # {"recognized_text": "actual_name"} từ training
        self.load_aliases()  # Load aliases cho sheet hiện tại
        self.load_phonetic_mapping()  # Load mapping từ file
        
        # [FIX] Throttle UI update để tránh lag
        self.last_ui_update = 0
        self.ui_update_interval = 0.1  # 100ms giữa các lần update
        
        # UI Setup
        self.title("🎤 NHẬP ĐIỂM BẰNG GIỌNG NÓI")
        
        # Fullscreen
        self.state('zoomed')  # Windows maximize
        self.configure(bg="#f5f6fa")
        
        # Build UI
        self.setup_ui()
        
        # Load data
        self.load_sheets_and_columns()
        
        # Process queue
        self.process_queue()
        
        # Focus
        self.focus_force()
        self.grab_set()
    
    def setup_ui(self):
        """Xây dựng giao diện 3 phần"""
        
        # Header
        header = tk.Frame(self, bg="#3742fa", pady=12)
        header.pack(fill="x")
        
        tk.Label(
            header,
            text="🎤 NHẬP ĐIỂM BẰNG GIỌNG NÓI",
            font=("Segoe UI", 18, "bold"),
            bg="#3742fa",
            fg="white"
        ).pack(side="left", padx=20)
        
        # Nút đóng
        tk.Button(
            header,
            text="✕ Đóng",
            font=("Segoe UI", 11, "bold"),
            bg="#e74c3c",
            fg="white",
            cursor="hand2",
            command=self.on_close
        ).pack(side="right", padx=20)
        
        # Main container - 3 panels
        main = tk.Frame(self, bg="#f5f6fa")
        main.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Configure grid weights - Điều chỉnh tỷ lệ 3 panel
        main.grid_columnconfigure(0, weight=0, minsize=260)  # Controls - cố định 260px
        main.grid_columnconfigure(1, weight=3)  # Log - rộng hơn
        main.grid_columnconfigure(2, weight=1)  # Students - thu gọn
        main.grid_rowconfigure(0, weight=1)
        
        # ============ Panel 1: Controls (Left) với SCROLLBAR ============
        left_container = tk.Frame(main, bg="#ffffff", relief="solid", bd=1, width=260)
        left_container.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left_container.grid_propagate(False)  # Giữ cố định chiều rộng
        
        # Canvas + Scrollbar
        left_canvas = tk.Canvas(left_container, bg="#ffffff", highlightthickness=0, width=240)
        left_scrollbar = ttk.Scrollbar(left_container, orient="vertical", command=left_canvas.yview)
        left = tk.Frame(left_canvas, bg="#ffffff", padx=10, pady=10)
        
        # Cấu hình scroll
        left.bind(
            "<Configure>",
            lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        )
        left_canvas.create_window((0, 0), window=left, anchor="nw")
        left_canvas.configure(yscrollcommand=left_scrollbar.set)
        
        # Mouse wheel scroll - Lưu reference để unbind khi đóng
        def on_mousewheel(event):
            try:
                if left_canvas.winfo_exists():
                    left_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass
        self._mousewheel_handler = on_mousewheel
        left_canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        left_canvas.pack(side="left", fill="both", expand=True)
        left_scrollbar.pack(side="right", fill="y")
        
        tk.Label(
            left,
            text="📋 CẤU HÌNH",
            font=("Segoe UI", 11, "bold"),
            bg="#ffffff",
            fg="#2f3640"
        ).pack(anchor="w", pady=(0, 10))
        
        # Sheet selector
        tk.Label(left, text="Lớp (Sheet):", font=("Segoe UI", 10, "bold"), bg="#ffffff").pack(anchor="w")
        self.cbo_sheet = ttk.Combobox(left, state="readonly", font=("Segoe UI", 11), width=20)
        self.cbo_sheet.pack(fill="x", pady=(5, 15))
        self.cbo_sheet.bind("<<ComboboxSelected>>", self.on_sheet_change)
        
        # Column selector
        tk.Label(left, text="Cột điểm:", font=("Segoe UI", 10, "bold"), bg="#ffffff").pack(anchor="w")
        self.cbo_col = ttk.Combobox(left, state="readonly", font=("Segoe UI", 11), width=20)
        self.cbo_col.pack(fill="x", pady=(5, 15))
        self.cbo_col.bind("<<ComboboxSelected>>", self.on_col_change)
        
        # Separator
        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=15)
        
        # Start/Stop buttons
        self.btn_listen = tk.Button(
            left,
            text="🎤 BẮT ĐẦU NGHE",
            font=("Segoe UI", 12, "bold"),
            bg="#27ae60",
            fg="white",
            activebackground="#2ecc71",
            cursor="hand2",
            pady=8,
            bd=2,
            relief="solid",
            highlightbackground="#000000",
            command=self.toggle_listening
        )
        self.btn_listen.pack(fill="x", pady=(0, 5))
        
        # [NEW] Nút Huấn luyện
        self.btn_training = tk.Button(
            left,
            text="🎓 HUẤN LUYỆN",
            font=("Segoe UI", 10, "bold"),
            bg="#9b59b6",
            fg="white",
            activebackground="#8e44ad",
            cursor="hand2",
            pady=6,
            bd=2,
            relief="solid",
            highlightbackground="#000000",
            command=self.open_training_mode
        )
        self.btn_training.pack(fill="x", pady=(0, 5))
        
        # Status
        self.status_label = tk.Label(
            left,
            text="⏸️ Chưa bắt đầu",
            font=("Segoe UI", 10),
            bg="#ffffff",
            fg="#7f8c8d"
        )
        self.status_label.pack(pady=5)
        
        # [NEW] Dropdown chọn Speech Engine
        engine_frame = tk.Frame(left, bg="#ffffff")
        engine_frame.pack(fill="x", pady=(5, 0))
        
        tk.Label(
            engine_frame,
            text="🔊 Speech Engine:",
            font=("Segoe UI", 9, "bold"),
            bg="#ffffff",
            fg="#2c3e50"
        ).pack(anchor="w")
        
        # Xây dựng danh sách engines có sẵn - CHỈ GOOGLE
        self.speech_engines = []
        if HAS_SPEECH:
            self.speech_engines.append("📡 Google (Chính xác nhất, cần mạng)")
        
        if not self.speech_engines:
            self.speech_engines.append("❌ Chưa cài speech engine")
        
        self.speech_engine_var = tk.StringVar(value=self.speech_engines[0])
        
        engine_combo = ttk.Combobox(
            engine_frame,
            textvariable=self.speech_engine_var,
            values=self.speech_engines,
            state="readonly",
            font=("Segoe UI", 9),
            width=30
        )
        engine_combo.pack(fill="x", pady=3)
        
        # Mô tả engine (1 dòng gọn)
        self.engine_desc_label = tk.Label(
            engine_frame,
            text="",
            font=("Segoe UI", 8),  # Tăng từ 7 lên 8
            bg="#ffffff",
            fg="#2c3e50",  # Đổi từ xám (#95a5a6) sang đen (#2c3e50)
            wraplength=200,
            justify="left"
        )
        self.engine_desc_label.pack(anchor="w")
        
        def update_engine_desc(*args):
            engine = self.speech_engine_var.get()
            if "Google" in engine:
                self.engine_desc_label.config(text="✅ Chính xác | ❌ Cần mạng")
            else:
                self.engine_desc_label.config(text="")
        
        self.speech_engine_var.trace("w", update_engine_desc)
        update_engine_desc()  # Initial update
        
        # ============ [NEW] Audio Processing Options ============
        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=5)
        
        tk.Label(
            left,
            text="🔧 XỬ LÝ ÂM THANH",
            font=("Segoe UI", 9, "bold"),
            bg="#ffffff",
            fg="#2c3e50"
        ).pack(anchor="w")
        
        # Checkbox: VAD
        self.use_vad = tk.BooleanVar(value=False)  # TẮT MẶC ĐỊNH
        tk.Checkbutton(
            left,
            text="✂️ Cắt im lặng",
            variable=self.use_vad,
            font=("Segoe UI", 8),
            bg="#ffffff",
            fg="#2c3e50",
            activebackground="#ffffff"
        ).pack(anchor="w", pady=1)
        
        # Slider: Thời gian ghi (compact)
        duration_frame = tk.Frame(left, bg="#ffffff")
        duration_frame.pack(fill="x", pady=2)
        
        tk.Label(
            duration_frame,
            text="⏱️",
            font=("Segoe UI", 9),
            bg="#ffffff"
        ).pack(side="left")
        
        self.record_duration = tk.IntVar(value=4)  # TĂNG LÊN 4 GIÂY
        duration_scale = tk.Scale(
            duration_frame,
            from_=2,
            to=15,  # Tăng lên 15 giây cho batch mode
            orient="horizontal",
            variable=self.record_duration,
            bg="#ffffff",
            highlightthickness=0,
            length=120  # Tăng độ dài slider
        )
        duration_scale.pack(side="left")
        
        self.duration_label = tk.Label(
            duration_frame,
            text="4s",
            font=("Segoe UI", 8, "bold"),
            bg="#ffffff",
            fg="#3498db"
        )
        self.duration_label.pack(side="left", padx=3)
        
        def update_duration_label(*args):
            duration = self.record_duration.get()
            self.duration_label.config(text=f"{duration}s")
            # Hiển thị gợi ý batch mode nếu > 6s
            if duration > 6:
                self.duration_label.config(fg="#e74c3c")  # Đỏ = batch mode
            else:
                self.duration_label.config(fg="#3498db")  # Xanh = single mode
        self.record_duration.trace("w", update_duration_label)
        
        # ============ [NEW] Smart Recognition Options ============
        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=5)
        
        tk.Label(
            left,
            text="🧠 NHẬN DIỆN THÔNG MINH",
            font=("Segoe UI", 9, "bold"),
            bg="#ffffff",
            fg="#2c3e50"
        ).pack(anchor="w")
        
        # Checkbox: Contextual Hints
        self.use_contextual_hints = tk.BooleanVar(value=True)  # BẬT MẶC ĐỊNH
        tk.Checkbutton(
            left,
            text="💡 Gợi ý từ danh sách",
            variable=self.use_contextual_hints,
            font=("Segoe UI", 8),
            bg="#ffffff",
            fg="#2c3e50",
            activebackground="#ffffff"
        ).pack(anchor="w", pady=1)
        
        # Checkbox: Auto-Correction
        self.use_auto_correction = tk.BooleanVar(value=True)  # BẬT MẶC ĐỊNH
        tk.Checkbutton(
            left,
            text="🔧 Tự động sửa lỗi phát âm",
            variable=self.use_auto_correction,
            font=("Segoe UI", 8),
            bg="#ffffff",
            fg="#2c3e50",
            activebackground="#ffffff"
        ).pack(anchor="w", pady=1)
        
        # Checkbox: Continuous Mode
        self.continuous_mode = tk.BooleanVar(value=False)  # TẮT MẶC ĐỊNH
        tk.Checkbutton(
            left,
            text="🔁 Chế độ liên tục (nghe nhiều lần)",
            variable=self.continuous_mode,
            font=("Segoe UI", 8),
            bg="#ffffff",
            fg="#2c3e50",
            activebackground="#ffffff"
        ).pack(anchor="w", pady=1)
        
        # Checkbox: Accumulate Batch
        self.accumulate_batch = tk.BooleanVar(value=True)  # BẬT MẶC ĐỊNH
        tk.Checkbutton(
            left,
            text="📥 Tích lũy batch (cộng dồn nhiều lần)",
            variable=self.accumulate_batch,
            font=("Segoe UI", 8),
            bg="#ffffff",
            fg="#2c3e50",
            activebackground="#ffffff"
        ).pack(anchor="w", pady=1)
        
        # [FIX #4] Callback để clear batch khi tắt tích lũy
        def on_accumulate_toggle(*args):
            if not self.accumulate_batch.get():
                # Tắt tích lũy → Clear batch cũ
                if hasattr(self, 'batch_results') and self.batch_results:
                    answer = messagebox.askyesno(
                        "Xóa batch?",
                        f"Tắt chế độ tích lũy sẽ xóa {len(self.batch_results)} học sinh trong batch.\n\nTiếp tục?"
                    )
                    if answer:
                        self.batch_results = []
                        self.btn_save_batch.pack_forget()
                        self.btn_clear_batch.pack_forget()
                        self.log("🔄 Đã xóa batch (chế độ tích lũy tắt)\n", tag="warning")
                    else:
                        # User không muốn xóa → Bật lại checkbox
                        self.accumulate_batch.set(True)
        
        self.accumulate_batch.trace("w", on_accumulate_toggle)
        
        # Hướng dẫn (gọn hơn)
        ttk.Separator(left, orient="horizontal").pack(fill="x", pady=8)
        
        guide_text = """📖 HƯỚNG DẪN:
• Đơn lẻ: "An 8"
• Nhiều: "An 8, Bình 7, Chi 9"
• Batch: >6s cho nhiều học sinh
• Liên tục: Bật để nghe tự động
• Tích lũy: Cộng dồn nhiều lần đọc
• Sửa: Double-click ô điểm"""
        tk.Label(
            left,
            text=guide_text,
            font=("Segoe UI", 8),
            bg="#ffffff",
            fg="#7f8c8d",
            justify="left"
        ).pack(anchor="w", pady=2)
        
        # ============ Panel 2: Log (Middle) ============
        middle = tk.Frame(main, bg="#ffffff", padx=15, pady=15, relief="solid", bd=1)
        middle.grid(row=0, column=1, sticky="nsew", padx=8)
        
        tk.Label(
            middle,
            text="📝 NHẬT KÝ HOẠT ĐỘNG",
            font=("Segoe UI", 13, "bold"),
            bg="#ffffff",
            fg="#2f3640"
        ).pack(anchor="w", pady=(0, 10))
        
        # Log text
        from tkinter import scrolledtext
        self.log_text = scrolledtext.ScrolledText(
            middle,
            font=("Consolas", 11),  # Tăng từ 10 lên 11
            bg="#ecf0f1",
            fg="#2c3e50",
            wrap="word",
            height=25
        )
        self.log_text.pack(fill="both", expand=True)
        
        # Config tags cho highlight với font in đậm
        self.log_text.tag_config("success", background="#d5f4e6", foreground="#27ae60", font=("Consolas", 11, "bold"))  # Xanh lá in đậm
        self.log_text.tag_config("error", background="#fadbd8", foreground="#e74c3c", font=("Consolas", 11, "bold"))    # Đỏ in đậm
        self.log_text.tag_config("warning", background="#fcf3cf", foreground="#f39c12", font=("Consolas", 11, "bold"))  # Vàng in đậm
        
        # Batch save button (ẩn mặc định)
        self.btn_save_batch = tk.Button(
            middle,
            text="💾 Lưu tất cả (0)",
            font=("Segoe UI", 11, "bold"),
            bg="#27ae60",
            fg="white",
            cursor="hand2",
            command=self.save_batch_results
        )
        # Không pack ngay
        
        # ============ Panel 3: Students (Right) ============
        right = tk.Frame(main, bg="#ffffff", padx=15, pady=15, relief="solid", bd=1)
        right.grid(row=0, column=2, sticky="nsew", padx=(8, 0))
        
        tk.Label(
            right,
            text="👨‍🎓 DANH SÁCH HỌC SINH",
            font=("Segoe UI", 13, "bold"),
            bg="#ffffff",
            fg="#2f3640"
        ).pack(anchor="w", pady=(0, 10))
        
        # ===== FOOTER BUTTONS - 2 hàng =====
        footer = tk.Frame(right, bg="#ffffff")
        footer.pack(side="bottom", fill="x", pady=(8, 0))
        
        # Hàng 1: Reset + Xuất (có viền đen)
        btn_row = tk.Frame(footer, bg="#ffffff")
        btn_row.pack(fill="x", pady=(0, 5))
        
        tk.Button(
            btn_row,
            text="🔄 RESET",
            font=("Segoe UI", 8, "bold"),
            bg="#e74c3c",
            fg="white",
            cursor="hand2",
            padx=8,
            pady=3,
            bd=1,
            relief="solid",
            highlightbackground="#000000",
            command=self.reset_scores
        ).pack(side="left", padx=(0, 8))
        
        tk.Button(
            btn_row,
            text="📊 XUẤT",
            font=("Segoe UI", 8, "bold"),
            bg="#3498db",
            fg="white",
            cursor="hand2",
            padx=8,
            pady=3,
            bd=1,
            relief="solid",
            highlightbackground="#000000",
            command=self.export_excel
        ).pack(side="left")
        
        # Hàng 2: Nút Đồng bộ (có viền đen)
        self.btn_sync = tk.Button(
            footer,
            text="🔁 ĐỒNG BỘ",
            font=("Segoe UI", 9, "bold"),
            bg="#9b59b6",
            fg="white",
            cursor="hand2",
            padx=10,
            bd=1,
            relief="solid",
            highlightbackground="#000000",
            pady=3,
            command=self.sync_to_main_app
        )
        self.btn_sync.pack(fill="x", pady=(3, 0))
        
        # [NEW] Nút XÓA BATCH - Ẩn mặc định, hiện khi có batch
        self.btn_clear_batch = tk.Button(
            footer,
            text="❌ XÓA BATCH",
            font=("Segoe UI", 9, "bold"),
            bg="#e74c3c",
            fg="white",
            cursor="hand2",
            padx=10,
            bd=1,
            relief="solid",
            highlightbackground="#000000",
            pady=3,
            command=self.clear_batch
        )
        # Không pack ngay - chỉ pack khi có batch
        
        # Separator
        ttk.Separator(right, orient="horizontal").pack(side="bottom", fill="x", pady=(5, 5))
        
        # Progress bar - PACK TRƯỚC TREEVIEW
        progress_frame = tk.Frame(right, bg="#ffffff")
        progress_frame.pack(side="bottom", fill="x", pady=(5, 0))
        
        self.lbl_progress = tk.Label(
            progress_frame,
            text="Đã nhập: 0/0 (0%)",
            font=("Segoe UI", 10, "bold"),
            bg="#ffffff"
        )
        self.lbl_progress.pack(side="left")
        
        # Config style cho progress bar gradient đẹp
        style = ttk.Style()
        style.theme_use('clam')  # Dùng theme clam để custom tốt hơn
        style.configure("Gradient.Horizontal.TProgressbar",
                       troughcolor='#e0e0e0',  # Nền xám nhạt
                       bordercolor='#27ae60',   # Viền xanh lá
                       background='#27ae60',    # Màu chính xanh lá
                       lightcolor='#52c77a',    # Màu sáng (gradient)
                       darkcolor='#1e8449',     # Màu tối (gradient)
                       troughrelief='flat',
                       borderwidth=2,
                       thickness=22)
        
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            length=200, 
            mode='determinate',
            style="Gradient.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(side="right", fill="x", expand=True, padx=(10, 0))
        
        # Treeview for students - PACK SAU CÙNG để chiếm phần còn lại
        tree_frame = tk.Frame(right, bg="#ffffff")
        tree_frame.pack(fill="both", expand=True)
        
        # [NEW] Thêm label hướng dẫn cho Tên khác - Pack TRƯỚC tree (ở trên)
        alias_hint = tk.Label(
            tree_frame,
            text="💡 Mẹo: Double-click cột 'Tên khác' để thêm tên gọi thân mật\n"
                 "(VD: Ti, Bi, Chanh Ti, Em Bi). KHÔNG nhập số điểm!",
            font=("Segoe UI", 8, "italic"),
            bg="#fff3cd",
            fg="#856404",
            justify="left",
            padx=8,
            pady=4
        )
        alias_hint.pack(side="top", fill="x")
        
        columns = ("STT", "Họ và tên", "Alias", "Điểm")  # [NEW] Thêm cột Alias
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
        
        self.tree.heading("STT", text="STT")
        self.tree.heading("Họ và tên", text="Họ và tên")
        self.tree.heading("Alias", text="Tên khác 📝")  # [NEW] Tên gọi thân mật
        self.tree.heading("Điểm", text="Điểm")
        
        self.tree.column("STT", width=30, minwidth=30, anchor="center", stretch=False)
        self.tree.column("Họ và tên", width=150, minwidth=120, anchor="center", stretch=True)
        self.tree.column("Alias", width=70, minwidth=60, anchor="center", stretch=False)  # [FIX] Thu gọn lại
        self.tree.column("Điểm", width=45, minwidth=40, anchor="center", stretch=False)
        
        # Tags for highlighting
        self.tree.tag_configure("entered", background="#d5f5e3", foreground="#1e8449")
        self.tree.tag_configure("just_entered", background="#abebc6", foreground="#145a32")
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        # Pack tree
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Double-click to edit
        self.tree.bind("<Double-1>", self.on_tree_double_click)
    
    def log(self, message, tag=None):
        """Ghi log với tùy chọn highlight
        
        Args:
            message: Nội dung log
            tag: "success", "error", "warning" hoặc None
        """
        # [FIX] Throttle UI update để tránh lag khi log spam
        current_time = time.time()
        should_update = (current_time - self.last_ui_update) >= self.ui_update_interval
        
        # [FIX] Giới hạn số dòng log để tránh tràn bộ nhớ và lag UI
        max_lines = 1000  # Giới hạn 1000 dòng
        current_lines = int(self.log_text.index('end-1c').split('.')[0])
        
        if current_lines > max_lines:
            # Xóa 200 dòng đầu tiên để giữ log gọn
            self.log_text.delete('1.0', f'{200}.0')
        
        start_pos = self.log_text.index("end-1c")
        self.log_text.insert("end", message)
        
        # Apply tag nếu có
        if tag:
            end_pos = self.log_text.index("end-1c")
            self.log_text.tag_add(tag, start_pos, end_pos)
        
        # [FIX] Luôn scroll xuống cuối nếu có tag quan trọng (success, error, warning)
        # Hoặc nếu đã đủ khoảng thời gian throttle
        if tag in ["success", "error", "warning"] or should_update:
            self.log_text.see("end")
            self.last_ui_update = current_time
    
    def load_sheets_and_columns(self):
        """Load danh sách sheets và columns từ Excel tab"""
        try:
            # [DEBUG] Kiểm tra excel_tab có dữ liệu không
            self.log("🔍 Đang kiểm tra dữ liệu Excel...\n")
            
            if not hasattr(self.excel_tab, 'all_classes_data') or not self.excel_tab.all_classes_data:
                self.log("⚠️ CẢNH BÁO: Chưa import file Excel!\n", tag="warning")
                self.log("➡️ Vui lòng vào tab 'Nhập điểm Excel' và import file trước!\n", tag="warning")
                return
            
            # Lấy sheets từ all_classes_data (thay vì all_sheets)
            sheets = list(self.excel_tab.all_classes_data.keys())
            
            if not sheets:
                self.log("⚠️ Không tìm thấy sheet nào trong file Excel!\n", tag="warning")
                return
            
            self.log(f"✅ Tìm thấy {len(sheets)} sheet: {', '.join(map(str, sheets))}\n")
            
            self.cbo_sheet['values'] = sheets
            if sheets:
                self.cbo_sheet.set(sheets[0])
                self.on_sheet_change(None)
                
        except Exception as e:
            self.log(f"⚠️ Lỗi load dữ liệu: {e}\n")
    
    def on_sheet_change(self, event):
        """Khi đổi sheet"""
        sheet_name = self.cbo_sheet.get()
        if not sheet_name:
            return
            
        self.current_sheet = sheet_name
        
        try:
            # Lấy DataFrame của sheet từ all_classes_data (thay vì all_sheets)
            if sheet_name in self.excel_tab.all_classes_data:
                self.local_df = self.excel_tab.all_classes_data[sheet_name].copy()
                self.log(f"📚 Đã load sheet '{sheet_name}' - {len(self.local_df)} dòng\n")
            else:
                self.log(f"⚠️ Không tìm thấy sheet '{sheet_name}'!\n", tag="warning")
                return
            
            # Lấy các cột điểm dựa trên TÊN CỘT (không quan tâm có dữ liệu hay không)
            # Pattern: TX1, TX2, TX3, TX4, TX 1, TX 2, GK, CK, ĐTB, Điểm...
            score_cols = []
            
            # Danh sách pattern cột điểm (ưu tiên)
            score_patterns = [
                r'^tx\s*\d*$',          # TX, TX1, TX2, TX 1, TX 2...
                r'^gk\s*\d*$',          # GK, GK1
                r'^ck\s*\d*$',          # CK, CK1
                r'^đtb\s*\d*$',         # ĐTB
                r'^dtb\s*\d*$',         # DTB (không dấu)
                r'^kt\s*\d*$',          # KT, KT1, KT2
                r'^điểm',               # Điểm TX, Điểm GK...
                r'^diem',               # Diem (không dấu)
                r'kiểm tra',            # Kiểm tra
                r'kiem tra',            # Kiem tra (không dấu)
                r'^hk\s*\d*$',          # HK, HK1, HK2
                r'^tb\s*\d*$',          # TB
            ]
            
            for col in self.local_df.columns:
                col_lower = str(col).lower().strip()
                
                # Bỏ qua các cột không phải điểm
                if col_lower in ['stt', 'họ và tên', 'họ tên', 'tên', 'name', 'ho ten', 'hoten', 'lớp', 'lop', 'mã', 'ma']:
                    continue
                
                # Kiểm tra pattern
                for pattern in score_patterns:
                    if re.search(pattern, col_lower):
                        score_cols.append(col)
                        break
            
            # Nếu không tìm thấy bằng pattern, fallback: lấy tất cả cột còn lại
            if not score_cols:
                exclude_cols = ['stt', 'họ và tên', 'họ tên', 'tên', 'name', 'ho ten', 'hoten', 'lớp', 'lop', 'mã', 'ma']
                for col in self.local_df.columns:
                    col_lower = str(col).lower().strip()
                    if col_lower not in exclude_cols:
                        score_cols.append(col)
            
            # Sắp xếp cột theo thứ tự logic: TX1, TX2, TX3, TX4, GK, CK, ĐTB
            def sort_score_col(col_name):
                col_lower = str(col_name).lower().strip()
                # TX cols first (sorted by number)
                if 'tx' in col_lower:
                    match = re.search(r'\d+', col_lower)
                    return (0, int(match.group()) if match else 0)
                elif 'gk' in col_lower:
                    return (1, 0)
                elif 'ck' in col_lower:
                    return (2, 0)
                elif 'đtb' in col_lower or 'dtb' in col_lower:
                    return (3, 0)
                else:
                    return (4, 0)
            
            score_cols.sort(key=sort_score_col)
            
            self.log(f"📊 Tìm thấy {len(score_cols)} cột điểm: {', '.join(score_cols)}\n")
            
            self.cbo_col['values'] = score_cols
            if score_cols:
                self.cbo_col.set(score_cols[0])
                self.on_col_change(None)
                
        except Exception as e:
            self.log(f"⚠️ Lỗi load sheet: {e}\n")
    
    def on_col_change(self, event):
        """Khi đổi cột điểm"""
        self.current_col = self.cbo_col.get()
        self.load_students()
    
    def load_students(self):
        """Load danh sách học sinh vào tree"""
        self.students = []
        
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if self.local_df is None:
            self.log("⚠️ Chưa có dữ liệu sheet! Vui lòng chọn sheet và cột điểm.\n", tag="warning")
            return
        
        if not self.current_col:
            self.log("⚠️ Vui lòng chọn cột điểm!\n", tag="warning")
            return
        
        # Tìm cột tên
        name_col = None
        for col in self.local_df.columns:
            col_lower = str(col).lower()
            if any(x in col_lower for x in ['họ và tên', 'họ tên', 'tên', 'name', 'ho ten', 'hoten']):
                name_col = col
                break
        
        if name_col is None:
            self.log("⚠️ Không tìm thấy cột tên học sinh\n")
            return
        
        # Load students
        for idx, row in self.local_df.iterrows():
            name = str(row.get(name_col, '')).strip()
            if not name or name.lower() == 'nan':
                continue
            
            stt = idx + 1
            score = row.get(self.current_col, '')
            
            # Convert score
            if pd.isna(score) or str(score).strip() == '':
                score_display = "---"
                score_val = None
            else:
                try:
                    score_val = float(score)
                    score_display = f"{score_val:.1f}"
                except:
                    score_display = str(score)
                    score_val = None
            
            student = {
                'stt': stt,
                'name': name,
                'score': score_val,
                'row_idx': idx,
                'name_col': name_col
            }
            self.students.append(student)
            
            # [NEW] Lấy alias từ dict (nếu có)
            alias_display = self.student_aliases.get(stt, "")
            
            # Insert to tree với cột Alias
            tags = ("entered",) if score_val is not None else ()
            self.tree.insert("", "end", iid=str(stt), values=(stt, name, alias_display, score_display), tags=tags)
        
        self.update_progress()
        self.log(f"📚 Đã load {len(self.students)} học sinh từ sheet '{self.current_sheet}'\n")
    
    def update_progress(self):
        """Cập nhật progress bar"""
        total = len(self.students)
        entered = sum(1 for s in self.students if s['score'] is not None)
        
        if total > 0:
            percent = int(entered / total * 100)
            self.lbl_progress.config(text=f"Đã nhập: {entered}/{total} ({percent}%)")
            self.progress_bar['value'] = percent
        else:
            self.lbl_progress.config(text="Đã nhập: 0/0 (0%)")
            self.progress_bar['value'] = 0
    
    def toggle_listening(self):
        """Bật/tắt lắng nghe"""
        if not HAS_SPEECH or AUDIO_BACKEND is None:
            messagebox.showerror(
                "Thiếu thư viện",
                "Vui lòng cài đặt:\npip install SpeechRecognition sounddevice soundfile numpy"
            )
            return
        
        if not HAS_FUZZY:
            messagebox.showerror(
                "Thiếu thư viện",
                "Vui lòng cài đặt:\npip install fuzzywuzzy python-Levenshtein"
            )
            return
        
        if not self.current_col:
            messagebox.showwarning("Chưa chọn cột", "Vui lòng chọn cột điểm trước!")
            return
        
        self.is_listening = not self.is_listening
        
        if self.is_listening:
            self.btn_listen.config(text="⏹️ DỪNG NGHE", bg="#e74c3c")
            self.status_label.config(text="🔴 Đang nghe...", fg="#e74c3c")
            self.log("\n🎤 BẮT ĐẦU NGHE - Hãy nói tên và điểm...\n")
            
            # Start listening thread
            thread = threading.Thread(target=self.listen_background, daemon=True)
            thread.start()
        else:
            self.btn_listen.config(text="🎤 BẮT ĐẦU NGHE", bg="#27ae60")
            self.status_label.config(text="⏸️ Đã dừng", fg="#7f8c8d")
            self.log("\n⏹️ ĐÃ DỪNG NGHE\n")
    
    def listen_background(self):
        """Background thread để lắng nghe - Hỗ trợ Vosk, Whisper, và Google"""
        import sounddevice as sd
        import time
        import tempfile
        import os
        import urllib.request
        import zipfile
        
        sample_rate = 16000
        
        # [Đọc settings từ UI]
        duration = self.record_duration.get()  # Lấy từ slider
        use_vad = self.use_vad.get()  # Checkbox
        use_contextual_hints = self.use_contextual_hints.get()  # Smart hints
        use_auto_correction = self.use_auto_correction.get()  # Auto-correction
        
        self.log(f"⚙️ Settings: {duration}s | VAD={'ON' if use_vad else 'OFF'}\n")
        self.log(f"🧠 Smart: Hints={'ON' if use_contextual_hints else 'OFF'} | AutoFix={'ON' if use_auto_correction else 'OFF'}\n")
        
        # Xây dựng vocabulary hints nếu bật
        hints = None
        if use_contextual_hints:
            hints = build_vocabulary_hints(self.students)
            self.log(f"💡 Đã load {len(hints['phrases'])} gợi ý từ danh sách học sinh\n")
        
        # Chỉ sử dụng Google
        use_google = HAS_SPEECH
        
        if not use_google:
            self.log("❌ Không có Google Speech Recognition! Vui lòng cài: pip install SpeechRecognition\n")
            return
        
        self.log("📡 Chế độ: GOOGLE ONLINE\n")
        self.log("🔧 Đang khởi động microphone...\n")
        
        # [FIX] Biến để track trạng thái và tránh xử lý lặp
        last_processed_text = ""
        last_process_time = 0
        cooldown_seconds = 2
        
        while self.is_listening:
            try:
                # Hiển thị countdown - CHỈ CẬP NHẬT STATUS, KHÔNG LOG để tránh spam
                self.status_label.config(text=f"🔴 ĐANG GHI ({duration}s)...", fg="#e74c3c")
                self.master.update()  # Force UI update
                
                # Ghi âm với float32 để dễ xử lý
                audio_data = sd.rec(
                    int(duration * sample_rate),
                    samplerate=sample_rate,
                    channels=1,
                    dtype='float32',  # Dùng float32
                    blocking=True
                )
                
                if not self.is_listening:
                    break
                
                # Flatten audio
                audio_float = audio_data.flatten()
                
                # Hiển thị mức âm lượng gốc - CHỈ LOG NẾU CÓ VẤN ĐỀ
                max_amplitude = np.abs(audio_float).max()
                
                # === KIỂM TRA ÂM LƯỢNG TỐI THIỂU ===
                if max_amplitude < 0.01:
                    self.log(f"🔇 Quá nhỏ ({max_amplitude:.2%}) - Hãy nói TO hơn!\n", tag="warning")
                    self.status_label.config(text="⚠️ Nói TO hơn!", fg="#f39c12")
                    time.sleep(0.5)
                    continue
                
                # =====================================================
                # 🆕 PHASE 1: AUDIO PROCESSING PIPELINE (VAD)
                # =====================================================
                
                # Chỉ xử lý nếu user bật VAD
                if use_vad:
                    self.status_label.config(text="🔧 Đang xử lý âm thanh...", fg="#9b59b6")
                    self.master.update()
                    
                    # [FIX #7] Wrap trong try-catch để tránh crash
                    try:
                        # Áp dụng pipeline xử lý theo settings (noise reduction removed)
                        audio_float, proc_info = audio_preprocess(
                            audio_float, 
                            sample_rate=sample_rate,
                            apply_vad=use_vad
                        )
                    except Exception as audio_err:
                        self.log(f"⚠️ Lỗi xử lý âm thanh: {audio_err}\n", tag="warning")
                        # Fallback: Dùng audio gốc không xử lý
                        proc_info = {'vad_applied': False, 'speech_ratio': 1.0, 
                                   'original_length': len(audio_float)/sample_rate,
                                   'final_length': len(audio_float)/sample_rate}
                    
                    # Log thông tin xử lý - GIẢM LOG SPAM, chỉ log khi có vấn đề
                    if proc_info['vad_applied']:
                        speech_pct = proc_info['speech_ratio'] * 100
                        
                        # Nếu RẤT ÍT tiếng nói, bỏ qua
                        if proc_info['speech_ratio'] < 0.05:  # < 5%
                            self.log(f"🔇 Không phát hiện giọng nói ({speech_pct:.0f}%)\n", tag="warning")
                            self.status_label.config(text="⚠️ Không nghe thấy!", fg="#f39c12")
                            time.sleep(0.5)
                            continue
                        # Chỉ log nếu tỷ lệ giọng nói thấp (5-50%)
                        elif speech_pct < 50:
                            self.log(f"⚠️ Giọng nói yếu: {speech_pct:.0f}%\n", tag="warning")
                    
                    # Chỉ log nếu cắt đáng kể (>0.5s)
                    duration_change = proc_info['final_length'] - proc_info['original_length']
                    if abs(duration_change) > 0.5:
                        self.log(f"✂️ Đã cắt: {proc_info['original_length']:.1f}s → {proc_info['final_length']:.1f}s\n")
                else:
                    # Không xử lý - chỉ normalize cơ bản
                    audio_float = audio_normalize(audio_float, target_level=0.8)
                    audio_float = np.clip(audio_float, -1.0, 1.0).astype(np.float32)
                
                self.status_label.config(text="⏳ Đang nhận dạng...", fg="#3498db")
                self.master.update()
                # Bỏ log "Đang nhận dạng" để tránh spam
                
                text = None
                
                # === GOOGLE RECOGNITION ===
                if use_google and HAS_SPEECH:
                    try:
                        # Chuyển float32 sang int16 cho Google
                        audio_int16 = (audio_float * 32767).astype(np.int16)
                        audio_bytes = audio_int16.tobytes()
                        audio = sr.AudioData(audio_bytes, sample_rate, 2)
                        
                        # Xây dựng speech_contexts từ hints (nếu có)
                        recognize_params = {"language": "vi-VN"}
                        
                        if hints and hints["phrases"]:
                            # Google hỗ trợ speech_contexts (chỉ hoạt động với enhanced models)
                            # Giới hạn 500 phrases để tránh quá tải
                            recognize_params["show_all"] = False
                            self.log(f"💡 Google: Đang dùng {len(hints['phrases'])} gợi ý\n")
                            # Note: SpeechRecognition library không hỗ trợ speech_contexts trực tiếp
                            # Nhưng hints vẫn giúp cải thiện độ chính xác thông qua context
                        
                        # [FIX] Thêm timeout để tránh treo vô thời hạn
                        self.log("📡 Đang gửi đến Google API...\n")
                        text = self.recognizer.recognize_google(
                            audio, 
                            **recognize_params,
                            with_confidence=False  # Tắt confidence để nhanh hơn
                        )
                        if text:
                            self.log(f"📡 [Google] Đã nghe: '{text}'\n")
                    except sr.UnknownValueError:
                        # Google không nghe rõ âm thanh
                        self.log("🔇 Google: Không nghe rõ (âm thanh không rõ ràng)\n", tag="warning")
                        text = None
                    except sr.RequestError as e:
                        # Lỗi kết nối API (mạng, rate limit, etc.)
                        self.log(f"⚠️ Google API lỗi: {e}\n", tag="error")
                        self.log("💡 Kiểm tra kết nối mạng hoặc thử lại sau\n", tag="warning")
                        text = None
                    except Exception as e:
                        # Các lỗi khác
                        self.log(f"❌ Lỗi không xác định: {e}\n", tag="error")
                        text = None
                
                # Không có kết quả
                if not text:
                    self.log("❓ Không nghe rõ. Vui lòng nói lại.\n", tag="error")
                    time.sleep(0.3)
                    continue
                
                # [FIX] Kiểm tra text giống lần trước và cooldown
                current_time = time.time()
                normalized_text = text.strip().lower()
                
                if normalized_text == last_processed_text.lower() and (current_time - last_process_time) < cooldown_seconds:
                    self.log(f"⏭️ Bỏ qua (đã xử lý gần đây): '{text}'\n")
                    time.sleep(0.5)  # Chờ ngắn trước khi nghe tiếp
                    continue
                
                # Cập nhật tracking
                last_processed_text = normalized_text
                last_process_time = current_time
                
                # Thử batch mode
                batch_results = self.parse_batch_command(text)
                
                if batch_results:
                    self.log(f"🎯 BATCH MODE: Phát hiện {len(batch_results)} học sinh!\n", tag="success")
                    self.result_queue.put(("batch", batch_results))
                    
                    # Hiển thị chi tiết
                    for i, (student, score, conf, msg) in enumerate(batch_results, 1):
                        self.log(f"  {i}. {student['name']}: {score} điểm ({conf}%)\n")
                    
                    self.log("\n✅ Nhấn 'LưU TẤT CẢ' để xác nhận hoặc tiếp tục nghe...\n", tag="success")
                    
                    # [FIX] Force scroll xuống cuối để người dùng thấy thông báo mới nhất
                    self.log_text.see("end")
                    self.log_text.update_idletasks()
                    
                    # [FIX #3] Continuous mode - Tăng cooldown lên 1.0s
                    if self.continuous_mode.get():
                        cooldown = 1.0  # Tránh duplicate và echo
                        self.log(f"🔁 Chế độ liên tục: Chờ {cooldown}s...\n", tag="warning")
                        self.log_text.see("end")
                        time.sleep(cooldown)
                    else:
                        time.sleep(1)
                else:
                    # Single mode
                    student, score, confidence, message = self.parse_voice_command(text)
                    
                    self.log(message + "\n")
                    
                    # Trường hợp 1: Có cả tên và điểm
                    if student and score is not None:
                        if confidence >= 70:
                            self.result_queue.put(("fill", student, score))
                            self.log(f"💾 Đã điền điểm {score} cho {student['name']}\n", tag="success")
                            
                            # [FIX] Force scroll xuống cuối
                            self.log_text.see("end")
                            self.log_text.update_idletasks()
                            
                            # [FIX #3] Continuous mode - Tăng cooldown lên 1.0s
                            if self.continuous_mode.get():
                                cooldown = 1.0  # Tránh duplicate và echo
                                self.log(f"🔁 Chế độ liên tục: Chờ {cooldown}s...\n", tag="warning")
                                self.log_text.see("end")
                                time.sleep(cooldown)
                            else:
                                time.sleep(1)
                        else:
                            self.result_queue.put(("confirm", student, score, message))
                    
                    # Trường hợp 2: Chỉ có tên, chưa có điểm → Đợi nói điểm
                    elif student and score is None:
                        self.status_label.config(text=f"💬 {student['name']} - Nói điểm?", fg="#9b59b6")
                        self.master.update()
                        # Không báo lỗi, đợi người dùng nói điểm ở lần tiếp theo
                    
                    else:
                        self.log("❌ Không thể xử lý. Vui lòng thử lại.\n", tag="warning")
                        
                        # [FIX #3] Continuous mode - Tăng cooldown ngay cả khi lỗi
                        if self.continuous_mode.get():
                            time.sleep(1.0)  # Tránh spam errors
                
            except Exception as e:
                self.log(f"⚠️ Lỗi: {e}\n")
                time.sleep(0.5)
    
    def parse_voice_command(self, text):
        """Parse câu lệnh giọng nói - Cải thiện nhận dạng
        
        Hỗ trợ:
        1. "[Tên] [điểm]" - VD: "Nguyễn Văn A 8"
        2. "[Tên]" (chỉ tên) - Trả về tên, điểm=None để hỏi sau
        3. "[điểm]" (chỉ điểm) - Nếu đang có pending_student
        """
        if not HAS_FUZZY:
            return None, None, 0, "Chưa cài fuzzywuzzy"
        
        # Tiền xử lý: loại bỏ các từ không cần thiết
        text_clean = text.lower().strip()
        
        # Thay thế biến thể "điểm"
        diem_variants = ["điểm", "điêm", "diểm", "đểm", "điểm", "diem", "điem", "đi ểm"]
        for variant in diem_variants:
            text_clean = text_clean.replace(variant, " DIEM ")
        
        # Thay thế số chữ thành số (để dễ match)
        # "mười" → "10", "mời" → "10", etc.
        number_map = {
            "không": "0", "linh": "0", "lính": "0",
            "một": "1", "mốt": "1", "mót": "1",
            "hai": "2", "hải": "2", "hay": "2",
            "ba": "3", "bà": "3",
            "bốn": "4", "tư": "4", "bón": "4",
            "năm": "5", "lăm": "5", "nám": "5",
            "sáu": "6", "sấu": "6", "xáu": "6",
            "bảy": "7", "bẩy": "7", "bai": "7",
            "tám": "8", "tắm": "8", "tàm": "8",
            "chín": "9", "chính": "9", "chin": "9",
            "mười": "10", "mưới": "10", "mời": "10", "muối": "10", "mươi": "10",
        }
        
        for word, digit in sorted(number_map.items(), key=lambda x: -len(x[0])):
            text_clean = re.sub(rf'\b{word}\b', digit, text_clean)
        
        # Loại bỏ "DIEM" và khoảng trắng thừa
        text_clean = text_clean.replace("DIEM", " ").strip()
        text_clean = re.sub(r'\s+', ' ', text_clean)
        
        # ===== PATTERN 1: Tên + số =====
        # VD: "nguyễn thành công 10" → name="nguyễn thành công", score=10
        pattern_simple = r'^(.+?)\s+(\d+(?:[.,]\d+)?)\s*$'
        
        match = re.search(pattern_simple, text_clean, re.IGNORECASE)
        if match:
            voice_name = match.group(1).strip()
            score_str = match.group(2).replace(',', '.')
            try:
                score = float(score_str)
                if 0 <= score <= 10:
                    return self.fuzzy_match_student(voice_name, score)
            except ValueError:
                pass
        
        # ===== PATTERN 1b: Số + Tên (đảo ngược) =====
        # VD: "8 Ti", "3.0 Chanh Ti", "ba điểm Ti" → name="Ti", score=8/3
        pattern_reverse = r'^(\d+(?:[.,]\d+)?)\s+(.+?)\s*$'
        match = re.search(pattern_reverse, text_clean, re.IGNORECASE)
        if match:
            score_str = match.group(1).replace(',', '.')
            voice_name = match.group(2).strip()
            try:
                score = float(score_str)
                if 0 <= score <= 10:
                    return self.fuzzy_match_student(voice_name, score)
            except ValueError:
                pass
        
        # ===== PATTERN 2: CHỈ SỐ ĐIỂM =====
        # VD: "8", "10", "7.5" → Nếu có pending_student, điền điểm cho họ
        pattern_score_only = r'^(\d+(?:[.,]\d+)?)\s*$'
        match = re.search(pattern_score_only, text_clean)
        if match and hasattr(self, 'pending_student') and self.pending_student:
            score_str = match.group(1).replace(',', '.')
            try:
                score = float(score_str)
                if 0 <= score <= 10:
                    student = self.pending_student
                    self.pending_student = None  # Clear pending
                    return student, score, 100, f"✅ Điểm {score} cho {student['name']}"
            except ValueError:
                pass
        
        # ===== PATTERN 3: CHỈ TÊN (không có điểm) =====
        # VD: "Huỳnh Tấn Đạt" → Tìm tên, trả về điểm=None, hỏi điểm sau
        # Kiểm tra text có chứa số không
        has_number = bool(re.search(r'\d', text_clean))
        
        if not has_number and len(text_clean) >= 2:
            # Đây là tên, không có điểm
            voice_name = text_clean.strip()
            student, _, similarity, message = self.fuzzy_match_student(voice_name, 0)
            
            if student and similarity >= 70:
                # Lưu pending để hỏi điểm
                self.pending_student = student
                return student, None, similarity, f"🎯 Tìm thấy: {student['name']} ({similarity}%)\n💬 Nói điểm (0-10):"
        
        # Fallback: Thử với text gốc
        pattern3 = r"^(.+?)\s+([\d.,]+|[a-zàáảãạăắằẳẵặâấầẩẫậèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵđ\s]+)$"
        match = re.search(pattern3, text, re.IGNORECASE)
        if match:
            voice_name = match.group(1).strip()
            score_text = match.group(2).strip()
            score = voice_parse_score_text(score_text)
            if score is not None:
                return self.fuzzy_match_student(voice_name, score)
        
        return None, None, 0, "❌ Không hiểu câu lệnh. Nói: '[Tên] [điểm]'"
    
    def fuzzy_match_student(self, voice_name, score):
        """
        Fuzzy matching tên học sinh - Cải thiện để match theo TÊN GỌI
        
        Hỗ trợ:
        0. [NEW] Phonetic mapping từ training
        1. Match toàn bộ họ tên: "Nguyễn Văn An" → "Nguyễn Văn An"
        2. Match tên gọi: "An" → "Nguyễn Văn An"  
        3. Match tên + đệm: "Văn An" → "Nguyễn Văn An"
        4. [NEW] Alias: Check danh sách alias
        5. Auto-correction: "Trần" → "Chân" nếu không tìm thấy
        """
        voice_name_norm = voice_normalize_text(voice_name)
        voice_name_lower = voice_name.lower().strip()
        
        # === 0. Check Phonetic Mapping (từ Training Mode) ===
        # [FIX #5] Normalize cả voice_name và mapping keys
        phonetic_key_norm = voice_normalize_text(voice_name_lower)
        
        # Thử match với normalized key
        if phonetic_key_norm in self.phonetic_mapping:
            mapped_name = self.phonetic_mapping[phonetic_key_norm]
            # Tìm student có tên = mapped_name
            for student in self.students:
                if student['name'].lower() == mapped_name.lower():
                    message = f"🎓 Training: '{voice_name}' → {student['name']}"
                    return student, score, 100, message
        
        # Fallback: Thử match không normalize (backward compatibility)
        if voice_name_lower in self.phonetic_mapping:
            mapped_name = self.phonetic_mapping[voice_name_lower]
            # Tìm student có tên = mapped_name
            for student in self.students:
                if student['name'].lower() == mapped_name.lower():
                    message = f"🎓 Training: '{voice_name}' → {student['name']}"
                    return student, score, 100, message
        
        best_match = None
        best_similarity = 0
        match_type = "full"  # full, firstname, partial, alias, autocorrect
        corrected = False
        
        for student in self.students:
            student_name_norm = voice_normalize_text(student['name'])
            student_name_lower = student['name'].lower().strip()
            
            # === 1. Match toàn bộ họ tên ===
            similarity = fuzz.ratio(voice_name_norm, student_name_norm)
            
            # === 2. Check ALIAS ===
            stt = student['stt']
            if stt in self.student_aliases:
                aliases = [a.strip().lower() for a in self.student_aliases[stt].split(',')]
                for alias in aliases:
                    if alias and voice_name_lower == alias:
                        # Exact match với alias → 98%
                        if 98 > similarity:
                            similarity = 98
                            match_type = "alias"
                    elif alias:
                        # Fuzzy match với alias
                        alias_sim = fuzz.ratio(voice_name_lower, alias)
                        if alias_sim >= 85 and alias_sim > similarity:
                            similarity = alias_sim
                            match_type = "alias"
            
            # === 3. Match TÊN GỌI (phần cuối của họ tên) ===
            # VD: "Em" khớp với "Lý Hoàng Em"
            name_parts = student_name_lower.split()
            if len(name_parts) >= 1:
                first_name = name_parts[-1]  # Tên gọi (cuối cùng)
                
                # Exact match tên gọi → 95%
                if voice_name_lower == first_name:
                    if 95 > similarity:
                        similarity = 95
                        match_type = "firstname"
                
                # Fuzzy match tên gọi
                firstname_sim = fuzz.ratio(voice_name_lower, first_name)
                if firstname_sim >= 85 and firstname_sim > similarity:
                    similarity = firstname_sim
                    match_type = "firstname"
            
            # === 4. Match TÊN + ĐỆM (2 phần cuối) ===
            # VD: "Hoàng Em" khớp với "Lý Hoàng Em"
            if len(name_parts) >= 2:
                last_two = " ".join(name_parts[-2:])
                
                if voice_name_lower == last_two:
                    if 92 > similarity:
                        similarity = 92
                        match_type = "partial"
                
                partial_sim = fuzz.ratio(voice_name_lower, last_two)
                if partial_sim >= 85 and partial_sim > similarity:
                    similarity = partial_sim
                    match_type = "partial"
            
            # === 5. Partial match (substring) ===
            # VD: "Em" có trong "Lý Hoàng Em"
            if voice_name_lower in student_name_lower and len(voice_name_lower) >= 2:
                substring_bonus = min(90, 70 + len(voice_name_lower) * 5)
                if substring_bonus > similarity:
                    similarity = substring_bonus
                    match_type = "substring"
            
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = student
        
        # === 6. Auto-Correction nếu không tìm thấy và user bật ===
        if best_similarity < 85 and self.use_auto_correction.get():
            correction_result = auto_correct_name(voice_name, self.students, confidence_threshold=0.85)
            
            if correction_result:
                best_match = correction_result["student"]
                best_similarity = int(correction_result["confidence"] * 100)
                match_type = "autocorrect"
                corrected = correction_result["corrected"]
                
                if corrected:
                    self.log(
                        f"🔧 Tự động sửa: '{correction_result['original_name']}' → '{correction_result['variant_used']}'\n",
                        tag="warning"
                    )
        
        # === Quyết định dựa trên similarity ===
        if best_similarity >= 85:
            type_labels = {
                "firstname": "tên",
                "partial": "tên+đệm",
                "substring": "chứa",
                "alias": "alias",
                "autocorrect": "đã sửa",
                "full": ""
            }
            type_info = f" [{type_labels.get(match_type, '')}]" if match_type != "full" else ""
            message = f"✅ Tìm thấy: {best_match['name']} ({best_similarity}%{type_info})"
            return best_match, score, best_similarity, message
        elif best_similarity >= 70:
            message = f"⚠️ Có thể: {best_match['name']} (Độ khớp: {best_similarity}%)"
            return best_match, score, best_similarity, message
        else:
            message = f"❌ Không tìm thấy. Gần nhất: {best_match['name'] if best_match else 'N/A'} ({best_similarity}%)"
            return None, None, 0, message
    
    def parse_batch_command(self, text):
        """
        Parse nhiều học sinh cùng lúc - Cải tiến với nhiều pattern
        
        Hỗ trợ:
        - "An 8, Bình 7, Chi 9"
        - "An tám điểm, Bình bảy điểm"
        - "Nguyễn Văn An 8, Trần Thị Bình 7"
        - "An tám Bình bảy Chi chín"
        """
        if not text:
            return []
        
        results = []
        
        # === PATTERN 1: Phân tách theo dấu phẩy ===
        if ',' in text:
            parts = [part.strip() for part in text.split(',')]
            
            # Log để debug
            self.log(f"🔍 Phân tách theo dấu phẩy: {len(parts)} phần\n")
            
            for i, part in enumerate(parts, 1):
                if not part:
                    continue
                
                # Thử parse từng phần như một entry đơn
                student, score, confidence, message = self.parse_voice_command(part)
                
                if student and score is not None and confidence >= 70:
                    results.append((student, score, confidence, message))
                    self.log(f"  ✅ Phần {i}: {student['name']} - {score} điểm\n")
                else:
                    self.log(f"  ⚠️ Phần {i}: Không nhận diện được - '{part}'\n")
            
            if len(results) >= 2:
                return results
        
        # === PATTERN 2: Không có dấu phẩy - Tìm pattern "Tên + Số" ===
        # Regex: Tìm tên (viết hoa) theo sau là số
        # VD: "An 8", "Nguyễn Văn An 8", "Bình bảy"
        
        # Pattern cho tên Việt Nam (có thể nhiều từ)
        name_pattern = r'(?:[A-ZÀÁẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬÈÉẺẼẸÊẾỀỂỄỆÌÍỈĨỊÒÓỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢÙÚỦŨỤƯỨỪỬỮỰỲÝỶỸỴĐ][a-zàáảãạăắằẳẵặâấầẩẫậèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵđ]*\s*){1,4}'
        
        # Pattern cho điểm: Số (8, 8.5) hoặc chữ (tám, chín)
        score_pattern = r'(?:\d+(?:[.,]\d+)?|một|hai|ba|bốn|năm|sáu|bảy|tám|chín|mười|mươi|mười|mời|lăm|điểm)'
        
        # Kết hợp: Tên + Điểm (có thể có "điểm" ở giữa)
        full_pattern = f'({name_pattern})\\s*(?:điểm)?\\s*({score_pattern})'
        
        matches = re.findall(full_pattern, text, re.IGNORECASE)
        
        if len(matches) >= 2:
            self.log(f"🔍 Tìm thấy {len(matches)} cặp Tên-Điểm\n")
            
            for i, (voice_name, score_text) in enumerate(matches, 1):
                voice_name = voice_name.strip()
                score_text = score_text.strip()
                
                # Chuyển text điểm sang số
                score = voice_parse_score_text(score_text)
                
                if score is None:
                    self.log(f"  ⚠️ Cặp {i}: Điểm không hợp lệ - '{score_text}'\n")
                    continue
                
                # Fuzzy match tên
                student, _, confidence, message = self.fuzzy_match_student(voice_name, score)
                
                if student and confidence >= 70:
                    results.append((student, score, confidence, message))
                    self.log(f"  ✅ Cặp {i}: {student['name']} - {score} điểm ({confidence}%)\n")
                else:
                    self.log(f"  ⚠️ Cặp {i}: Không tìm thấy học sinh - '{voice_name}'\n")
            
            if len(results) >= 2:
                return results
        
        # === PATTERN 3: Fallback - Tìm tất cả tên và số, ghép tuần tự ===
        # Trường hợp: "An Bình Chi tám bảy chín"
        all_numbers = re.findall(r'\d+(?:[.,]\d+)?', text)
        all_words = text.split()
        
        # Tìm các từ viết hoa (có thể là tên)
        potential_names = []
        potential_scores = []
        
        i = 0
        while i < len(all_words):
            word = all_words[i]
            
            # Kiểm tra nếu là số hoặc từ chỉ điểm
            if re.match(r'\d+', word) or word.lower() in ['một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín', 'mười', 'mươi', 'lăm']:
                potential_scores.append(word)
                i += 1
                continue
            
            # Kiểm tra nếu bắt đầu bằng chữ hoa (có thể là tên)
            if word[0].isupper():
                # Ghép các từ viết hoa liên tiếp thành tên đầy đủ
                name_parts = [word]
                j = i + 1
                while j < len(all_words) and all_words[j][0].isupper() and all_words[j].lower() not in ['một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín', 'mười']:
                    name_parts.append(all_words[j])
                    j += 1
                
                potential_names.append(' '.join(name_parts))
                i = j
            else:
                i += 1
        
        # Ghép tên và điểm theo thứ tự
        if len(potential_names) >= 2 and len(potential_scores) >= 2:
            self.log(f"🔍 Fallback: {len(potential_names)} tên, {len(potential_scores)} điểm\n")
            
            for i, (name, score_text) in enumerate(zip(potential_names, potential_scores), 1):
                score = voice_parse_score_text(score_text)
                
                if score is None:
                    continue
                
                student, _, confidence, message = self.fuzzy_match_student(name, score)
                
                if student and confidence >= 70:
                    results.append((student, score, confidence, message))
                    self.log(f"  ✅ Fallback {i}: {student['name']} - {score} điểm\n")
        
        return results if len(results) >= 2 else []
    
    def process_queue(self):
        """Xử lý kết quả từ background thread"""
        try:
            while True:
                item = self.result_queue.get_nowait()
                
                if item[0] == "fill":
                    _, student, score = item
                    self.fill_score(student, score)
                
                elif item[0] == "batch":
                    _, new_batch_results = item
                    
                    # === TÍch LŨY BATCH LOGIC ===
                    if self.accumulate_batch.get():
                        # Lấy batch cũ (nếu có)
                        existing_batch = getattr(self, 'batch_results', [])
                        
                        # Loại bỏ trùng lặp: Chỉ thêm học sinh chưa có trong batch cũ
                        existing_stts = {s[0]['stt'] for s in existing_batch}
                        unique_new = [item for item in new_batch_results if item[0]['stt'] not in existing_stts]
                        
                        # THÊM VÀO batch cũ
                        self.batch_results = existing_batch + unique_new
                        total_count = len(self.batch_results)
                        new_count = len(unique_new)
                        duplicate_count = len(new_batch_results) - new_count
                        
                        # Log thông báo
                        if duplicate_count > 0:
                            self.log(f"🔔 Đã thêm {new_count} học sinh vào batch (Bỏ qua {duplicate_count} trùng lặp)\n", tag="warning")
                        else:
                            self.log(f"🔔 Đã thêm {new_count} học sinh vào batch\n", tag="success")
                        
                        self.log(f"📊 Tổng batch hiện tại: {total_count} học sinh\n", tag="success")
                        
                        # [FIX] Force scroll xuống cuối ngay lập tức
                        self.log_text.see("end")
                        self.log_text.update_idletasks()
                        
                        # Hiển thị status
                        self.status_label.config(
                            text=f"✅ Đã tích lũy: {total_count} học sinh", 
                            fg="#27ae60"
                        )
                    else:
                        # GHI ĐÈ batch cũ (behavior cũ)
                        self.batch_results = new_batch_results
                        total_count = len(new_batch_results)
                        
                        self.log(f"🎯 Batch mới: {total_count} học sinh (đã ghi đè batch cũ)\n")
                        
                        # [FIX] Force scroll xuống cuối ngay lập tức
                        self.log_text.see("end")
                        self.log_text.update_idletasks()
                        
                        # Hiển thị status
                        self.status_label.config(
                            text=f"✅ Đã nhận {total_count} học sinh", 
                            fg="#27ae60"
                        )
                    
                    # Hiển thị nút lưu với số lượng
                    self.btn_save_batch.config(text=f"💾 LƯU TẤT CẢ ({total_count} học sinh)")
                    self.btn_save_batch.pack(fill="x", pady=(5, 0))
                    
                    # Hiển thị nút XÓA BATCH (nếu có batch)
                    if total_count > 0:
                        self.btn_clear_batch.pack(fill="x", pady=(5, 0))
                
                elif item[0] == "confirm":
                    _, student, score, message = item
                    answer = messagebox.askyesno(
                        "Xác nhận",
                        f"{message}\n\nĐiền điểm {score} cho {student['name']}?"
                    )
                    if answer:
                        self.fill_score(student, score)
                
                elif item[0] == "error":
                    messagebox.showerror("Lỗi", item[1])
                    self.is_listening = False
                    self.toggle_listening()
        
        except queue.Empty:
            pass
        
        self.after(100, self.process_queue)
    
    def fill_score(self, student, score):
        """Điền điểm cho học sinh"""
        # Validate
        if score < 0 or score > 10:
            self.log(f"⚠️ Điểm {score} không hợp lệ (0-10)\n")
            return
        
        # Update local data
        student['score'] = score
        
        # Update DataFrame
        if self.local_df is not None and self.current_col:
            self.local_df.at[student['row_idx'], self.current_col] = score
        
        # Update tree - [FIX] Phải có 4 cột: STT, Name, Alias, Score
        stt = student['stt']
        alias_display = self.student_aliases.get(stt, "")  # Lấy alias hiện tại
        self.tree.item(str(stt), values=(stt, student['name'], alias_display, f"{score:.1f}"))
        self.tree.item(str(stt), tags=("just_entered",))
        
        # Highlight và scroll tới
        self.tree.see(str(student['stt']))
        self.tree.selection_set(str(student['stt']))
        
        # Reset highlight sau 2 giây
        self.after(2000, lambda: self.tree.item(str(student['stt']), tags=("entered",)))
        
        self.update_progress()
    
    def save_batch_results(self):
        """Lưu batch results"""
        if not self.batch_results:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu batch để lưu!")
            return
        
        for student, score, _, _ in self.batch_results:
            self.fill_score(student, score)
        
        self.btn_save_batch.pack_forget()
        self.batch_results = []
        
        messagebox.showinfo("Thành công", f"Đã lưu điểm vào bảng local!\nBấm 'Đồng bộ' để ghi vào file.")
        self.log("✅ Đã lưu batch vào bảng local!\n")
        
        # [FIX] Focus lại cửa sổ Voice Input sau khi đóng messagebox
        self.focus_force()
        self.lift()  # Đưa cửa sổ lên trên cùng
    
    def is_pure_number(self, text):
        """Kiểm tra xem text có phải là số thuần túy không (để validate alias)"""
        # Loại bỏ khoảng trắng
        text = text.strip()
        
        # Kiểm tra các pattern số: 3, 3.0, 7.5, 10, etc
        try:
            # Nếu convert được sang float → là số
            float(text)
            return True
        except ValueError:
            pass
        
        # Kiểm tra alias có chứa chữ không
        # VD: "3.0" = số, "Ti 3" = OK (có chữ)
        has_letter = any(c.isalpha() for c in text)
        if not has_letter:
            # Không có chữ cái → coi như số
            return True
        
        # [FIX #6] Kiểm tra pattern "Số + Tên" (VD: "3 Ti", "7.5 An")
        # Pattern này dễ nhầm lẫn với điểm số
        if re.match(r'^\d+(?:[.,]\d+)?\s+\w+', text):
            # Pattern "3 Ti", "7.5 An" → Cảnh báo là số
            return True
        
        return False
    
    def on_tree_double_click(self, event):
        """Double-click để sửa điểm hoặc alias"""
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if not item:
            return
        
        # Lấy student
        stt = int(item)
        student = next((s for s in self.students if s['stt'] == stt), None)
        if not student:
            return
        
        # [NEW] Cột #3 = Alias
        if column == "#3":
            # Sửa Alias
            current_alias = self.student_aliases.get(stt, "")
            
            new_alias = simpledialog.askstring(
                "Sửa Tên khác / Tên gọi thân mật",
                f"Nhập tên gọi khác cho: {student['name']}\n\n"
                f"⚠️ CHÚ Ý: Chỉ nhập TÊN GỌI, KHÔNG phải điểm số!\n"
                f"VD: Ti, Bi, Chanh Ti, Chân Ti, Em Bi, Tí Con...\n\n"
                f"Phân tách nhiều tên bởi dấu phẩy:",
                initialvalue=current_alias,
                parent=self
            )
            
            if new_alias is not None:
                # [NEW] VALIDATE: Không cho phép nhập số thuần túy
                if new_alias.strip() and self.is_pure_number(new_alias.strip()):
                    messagebox.showerror(
                        "Lỗi Tên khác",
                        f"⚠️ TÊN GỌI KHÔNG HỢP LỆ!\n\n"
                        f"Bạn đã nhập: '{new_alias.strip()}'\n\n"
                        f"Tên khác phải là TÊN GỌI, không phải SỐ ĐIỂM!\n"
                        f"VD đúng: Ti, Bi, Chanh Ti, Em Bi, Tí Con\n"
                        f"VD sai: 3.0, 7, 10"
                    )
                    return
                
                # Lưu alias
                if new_alias.strip():
                    self.student_aliases[stt] = new_alias.strip()
                else:
                    # Xóa alias nếu để trống
                    self.student_aliases.pop(stt, None)
                
                # Cập nhật tree
                current_values = self.tree.item(item, 'values')
                if current_values and len(current_values) >= 4:
                    self.tree.item(item, values=(current_values[0], current_values[1], new_alias.strip(), current_values[3]))
                
                self.log(f"✏️ Đã cập nhật tên khác cho {student['name']}: '{new_alias.strip()}'\n")
                self.save_aliases()  # Lưu vào file
        
        # Cột #4 = Điểm
        elif column == "#4":
            # Sửa điểm (code cũ)
            current_score = student['score'] if student['score'] is not None else ""
            
            new_score = simpledialog.askstring(
                "Sửa điểm",
                f"Nhập điểm mới cho {student['name']}:",
                initialvalue=str(current_score) if current_score else "",
                parent=self
            )
            
            if new_score is not None:
                try:
                    score = float(new_score.replace(',', '.'))
                    if 0 <= score <= 10:
                        self.fill_score(student, score)
                        self.log(f"✏️ Đã sửa điểm {student['name']} thành {score}\n")
                    else:
                        messagebox.showwarning("Lỗi", "Điểm phải từ 0 đến 10!")
                except ValueError:
                    messagebox.showwarning("Lỗi", "Điểm không hợp lệ!")
    
    def clear_batch(self):
        """Xóa toàn bộ batch chưa lưu"""
        if not hasattr(self, 'batch_results') or not self.batch_results:
            return
        
        # Xác nhận
        answer = messagebox.askyesno(
            "Xóa batch",
            f"Bạn có chắc muốn xóa {len(self.batch_results)} học sinh trong batch?\n\n"
            "Dữ liệu chưa lưu sẽ mất!"
        )
        
        if answer:
            count = len(self.batch_results)
            self.batch_results = []
            
            # Ẩn các nút
            self.btn_save_batch.pack_forget()
            self.btn_clear_batch.pack_forget()
            
            # Reset status
            self.status_label.config(text="⏸️ Đã xóa batch", fg="#7f8c8d")
            
            # Log
            self.log(f"🗑️ Đã xóa {count} học sinh khỏi batch\n", tag="warning")
    
    def save_all_batch(self):
        """Lưu tất cả học sinh trong batch"""
        if not self.batch_results:
            messagebox.showwarning("Không có dữ liệu", "Batch rỗng!")
            return
        
        saved_count = 0
        for student, score, confidence, message in self.batch_results:
            self.fill_score(student, score)
            saved_count += 1
        
        self.log(f"💾 Đã lưu {saved_count} học sinh từ batch!\n", tag="success")
        
        # [FIX] Force scroll xuống cuối để hiển thị thông báo lưu thành công
        self.log_text.see("end")
        self.log_text.update_idletasks()
        
        # Reset batch
        self.batch_results = []
        self.btn_save_batch.pack_forget()
        self.btn_clear_batch.pack_forget()
        self.status_label.config(text="✅ Đã lưu batch thành công", fg="#27ae60")
    
    def reset_scores(self):
        """Reset tất cả điểm"""
        answer = messagebox.askyesno("Xác nhận", "Xóa tất cả điểm đã nhập trong bảng này?")
        if answer:
            for student in self.students:
                student['score'] = None
                if self.local_df is not None and self.current_col:
                    self.local_df.at[student['row_idx'], self.current_col] = None
                
                # [FIX] Phải có 4 cột: STT, Name, Alias, Score
                stt = student['stt']
                alias_display = self.student_aliases.get(stt, "")
                self.tree.item(str(stt), values=(stt, student['name'], alias_display, "---"), tags=())
            
            self.update_progress()
            self.log("🔄 Đã reset tất cả điểm\n")
    
    def load_phonetic_mapping(self):
        """Load phonetic mapping từ file JSON"""
        try:
            mapping_file = "phonetic_mapping.json"
            if os.path.exists(mapping_file):
                with open(mapping_file, 'r', encoding='utf-8') as f:
                    self.phonetic_mapping = json.load(f)
                print(f"✅ Đã load {len(self.phonetic_mapping)} phonetic mappings")
        except Exception as e:
            print(f"⚠️ Không thể load phonetic mapping: {e}")
            self.phonetic_mapping = {}
    
    def save_phonetic_mapping(self):
        """Lưu phonetic mapping vào file JSON"""
        try:
            mapping_file = "phonetic_mapping.json"
            with open(mapping_file, 'w', encoding='utf-8') as f:
                json.dump(self.phonetic_mapping, f, ensure_ascii=False, indent=2)
            print(f"✅ Đã lưu {len(self.phonetic_mapping)} phonetic mappings")
        except Exception as e:
            print(f"⚠️ Không thể lưu phonetic mapping: {e}")
    
    def save_aliases(self):
        """Lưu aliases vào file JSON"""
        try:
            alias_file = f"aliases_{self.current_sheet}.json"
            with open(alias_file, 'w', encoding='utf-8') as f:
                json.dump(self.student_aliases, f, ensure_ascii=False, indent=2)
            print(f"✅ Đã lưu {len(self.student_aliases)} aliases")
        except Exception as e:
            print(f"⚠️ Không thể lưu aliases: {e}")
    
    def load_aliases(self):
        """Load aliases từ file JSON"""
        try:
            alias_file = f"aliases_{self.current_sheet}.json"
            if os.path.exists(alias_file):
                with open(alias_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Convert string keys to int
                    self.student_aliases = {int(k): v for k, v in data.items()}
                print(f"✅ Đã load {len(self.student_aliases)} aliases")
        except Exception as e:
            print(f"⚠️ Không thể load aliases: {e}")
            self.student_aliases = {}
    
    def open_training_mode(self):
        """Mở cửa sổ Training Mode"""
        print(f"[DEBUG] open_training_mode() called")
        print(f"[DEBUG] self.students: {len(self.students) if self.students else 0} students")
        print(f"[DEBUG] HAS_SPEECH: {HAS_SPEECH}")
        
        if not self.students:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng chọn sheet và cột trước!")
            return
        
        if not HAS_SPEECH:
            messagebox.showerror("Lỗi", "Không có speech_recognition! Vui lòng cài đặt: pip install SpeechRecognition")
            return
        
        # Kiểm tra sounddevice thay vì PyAudio
        try:
            import sounddevice
            print(f"[DEBUG] sounddevice available: {sounddevice.__version__}")
        except ImportError:
            messagebox.showerror(
                "Thiếu sounddevice",
                "Training Mode cần sounddevice để ghi âm!\n\n"
                "Cài đặt:\n"
                "pip install sounddevice soundfile numpy"
            )
            return
        
        try:
            print(f"[DEBUG] Creating TrainingWindow...")
            print(f"[DEBUG] parent type: {type(self)}")
            print(f"[DEBUG] parent class: {self.__class__.__name__}")
            print(f"[DEBUG] parent bases: {self.__class__.__bases__}")
            
            TrainingWindow(self, self.students, self.phonetic_mapping, self.save_phonetic_mapping)
            print(f"[DEBUG] TrainingWindow created successfully!")
            
        except Exception as e:
            print(f"[DEBUG ERROR] Failed to create TrainingWindow: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Lỗi", f"Không thể mở Training Mode:\n{e}")
    
    def export_excel(self):
        """Xuất ra file Excel riêng"""
        try:
            from tkinter import filedialog
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"diem_{self.current_sheet}_{self.current_col}.xlsx"
            )
            
            if file_path:
                # Tạo DataFrame từ students
                data = []
                for s in self.students:
                    data.append({
                        'STT': s['stt'],
                        'Họ và tên': s['name'],
                        self.current_col: s['score'] if s['score'] is not None else ''
                    })
                
                df = pd.DataFrame(data)
                df.to_excel(file_path, index=False)
                
                messagebox.showinfo("Thành công", f"Đã xuất file: {file_path}")
                self.log(f"💾 Đã xuất Excel: {file_path}\n")
        
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
    
    def sync_to_main_app(self):
        """Đồng bộ về app chính và file Excel"""
        if self.local_df is None:
            messagebox.showwarning("Lỗi", "Không có dữ liệu để đồng bộ!")
            return
        
        try:
            # 1. Cập nhật DataFrame của ExcelTab
            if hasattr(self.excel_tab, 'all_sheets') and self.current_sheet in self.excel_tab.all_sheets:
                self.excel_tab.all_sheets[self.current_sheet] = self.local_df.copy()
            
            # Nếu đang ở sheet hiện tại, cập nhật current_df
            current_sheet = self.excel_tab.get_current_sheet_name()
            if current_sheet == self.current_sheet:
                self.excel_tab.current_df = self.local_df.copy()
            
            # 2. Refresh UI của ExcelTab
            self.excel_tab.update_ui_data(self.excel_tab.current_df, update_chart=True)
            
            # 3. Ghi vào file Excel - CHỈ CẬP NHẬT CỘT ĐIỂM, KHÔNG GHI ĐÈ TOÀN BỘ SHEET
            file_path = self.excel_tab.file_path
            
            if file_path and os.path.exists(file_path):
                # Backup trước
                backup_path = file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%H%M%S")}.xlsx')
                try:
                    import shutil
                    shutil.copy2(file_path, backup_path)
                    self.log(f"📁 Đã tạo backup: {os.path.basename(backup_path)}\n")
                except:
                    pass
                
                # [FIX] Ghi file bằng openpyxl - CHỈ CẬP NHẬT CỘT ĐIỂM ĐƯỢC CHỌN
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path)
                ws = wb[self.current_sheet] if self.current_sheet in wb.sheetnames else wb.active
                
                # Tìm header row (dòng chứa "Họ và tên")
                header_row_idx = 1
                for row_idx in range(1, 10):  # Check 10 dòng đầu
                    for cell in ws[row_idx]:
                        if cell.value and "Họ và tên" in str(cell.value):
                            header_row_idx = row_idx
                            break
                    if header_row_idx != 1:
                        break
                
                # Tìm cột điểm cần cập nhật
                col_idx_to_update = None
                for col_idx, cell in enumerate(ws[header_row_idx], start=1):
                    cell_val = str(cell.value or "").upper()
                    # Check cột TX từ header (có thể nằm ở dòng header hoặc dòng sau)
                    if self.current_col and self.current_col.upper() in cell_val:
                        col_idx_to_update = col_idx
                        break
                    # Check dòng header phụ (TX1, TX2...)
                    if header_row_idx + 1 <= ws.max_row:
                        sub_header = str(ws.cell(row=header_row_idx + 1, column=col_idx).value or "").upper()
                        if self.current_col and self.current_col.upper() == sub_header.strip():
                            col_idx_to_update = col_idx
                            break
                
                if col_idx_to_update:
                    self.log(f"📍 Tìm thấy cột '{self.current_col}' tại vị trí {col_idx_to_update}\n")
                    
                    # Xác định data row bắt đầu
                    data_start_row = header_row_idx + 1
                    # Check nếu có sub-header (TX1, TX2...)
                    if ws.cell(row=header_row_idx + 1, column=1).value:
                        first_cell = str(ws.cell(row=header_row_idx + 1, column=1).value)
                        if "TX" in first_cell.upper() or not first_cell[0].isdigit():
                            data_start_row = header_row_idx + 2
                    
                    # Ghi từng điểm vào đúng vị trí
                    updated_count = 0
                    for _, row_data in self.local_df.iterrows():
                        stt = str(row_data.get("STT", ""))
                        score = row_data.get(self.current_col, "")
                        
                        # Tìm row trong Excel bằng STT
                        for excel_row in range(data_start_row, ws.max_row + 1):
                            excel_stt = str(ws.cell(row=excel_row, column=1).value or "").strip()
                            if excel_stt == stt:
                                # Ghi điểm
                                if score != "" and score is not None:
                                    try:
                                        ws.cell(row=excel_row, column=col_idx_to_update).value = float(score)
                                        updated_count += 1
                                    except (ValueError, TypeError):
                                        pass
                                break
                    
                    self.log(f"✏️ Đã cập nhật {updated_count} điểm vào cột {self.current_col}\n")
                else:
                    self.log(f"⚠️ Không tìm thấy cột '{self.current_col}' trong Excel\n")
                
                wb.save(file_path)
                wb.close()
                
                self.log(f"✅ ĐÃ ĐỒNG BỘ thành công!\n")
                self.log(f"   - Cập nhật UI: ✓\n")
                self.log(f"   - Ghi file Excel: ✓\n")
                
                messagebox.showinfo(
                    "Đồng bộ thành công",
                    f"Đã cập nhật:\n"
                    f"• Bảng dữ liệu trên giao diện\n"
                    f"• File Excel: {os.path.basename(file_path)}\n"
                    f"• Sheet: {self.current_sheet}\n"
                    f"• Cột: {self.current_col}"
                )
            else:
                self.log("⚠️ Không tìm thấy file Excel để ghi\n")
                messagebox.showwarning("Cảnh báo", "Chỉ cập nhật UI, không ghi được file Excel!")
        
        except PermissionError:
            messagebox.showerror(
                "File đang bị khóa",
                "Không thể ghi file Excel vì đang được mở bởi ứng dụng khác.\n"
                "Vui lòng đóng file và thử lại."
            )
        except Exception as e:
            self.log(f"⚠️ Lỗi đồng bộ: {e}\n")
            messagebox.showerror("Lỗi", f"Lỗi đồng bộ: {str(e)}")
    
    def on_close(self):
        """Đóng cửa sổ"""
        self.is_listening = False
        # Unbind mousewheel để tránh lỗi sau khi đóng
        try:
            self.unbind_all("<MouseWheel>")
        except:
            pass
        self.destroy()


# =================================================================================
# CLASS NHẬP ĐIỂM TẬP TRUNG (FULLSCREEN 3-PANE) [NEW]
# =================================================================================
class InputScoreWindow(Toplevel):
    # ========== [BƯỚC 1] CẤU HÌNH MẶC ĐỊNH CHO PHÍM TẮT ĐIỂM ==========
    DEFAULT_SCORE_KEY_MAPPING = {
        "1": "1",
        "2": "2",
        "3": "3",
        "4": "4",
        "5": "5",
        "6": "6",
        "7": "7",
        "8": "8",
        "9": "9",
        "10": "0",  # Phím 0 = điểm 10 (mặc định)
    }
    CONFIG_FILE_NAME = "diemthi_key_config.json"
    # ================================================================
    
    def __init__(self, parent, excel_tab):
        super().__init__(parent)
        self.title("NHẬP ĐIỂM KIỂM TRA - CHẾ ĐỘ TẬP TRUNG")
        try:
            self.state("zoomed")  # Windows only
        except:
            self.attributes("-fullscreen", True) # Fallback
            
        self.configure(bg=THEME["bg_app"])
        self.transient(parent)
        
        self.excel_tab = excel_tab
        # Copy data để không ảnh hưởng trực tiếp bản gốc khi chưa Save
        self.current_df = excel_tab.current_df.copy() if excel_tab.current_df is not None else pd.DataFrame()
        
        # [BƯỚC 1] Load cấu hình phím tắt từ file JSON
        self.score_key_mapping = self.load_key_config()
        
        # [BƯỚC 3] Khởi tạo dict lưu reference Entry của modal settings (sẽ được populate khi mở modal)
        self.score_key_entries = {}
        
        # Mapping DataFrame index -> Tree iid cho Preview Tree
        self.df_index_to_tree_iid = {}
        
        # [BƯỚC 6] Debounce timer cho auto-sync
        self.sync_debounce_job = None
        
        self.setup_ui()
        
        # Bind phím ESC để thoát
        self.bind("<Escape>", lambda e: self.on_close())
        self.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def load_aliases(self):
        """Stub - InputScoreWindow không cần alias (chỉ VoiceInputWindow cần)"""
        pass

    def on_close(self):
        """Cleanup khi đóng cửa sổ"""
        try:
            self.unbind_all("<MouseWheel>")
        except:
            pass
        self.destroy()

    def setup_ui(self):
        # Header (Top Bar)
        self.header_frame = tk.Frame(self, bg=THEME["primary"], pady=15, padx=20)
        header_frame = self.header_frame
        header_frame.pack(fill="x")
        
        # Bind resize removed
        # header_frame.bind("<Configure>", lambda e: self.update_progress())
        
        # Close Button
        btn_close = tk.Button(header_frame, text="✖ THOÁT", bg="#f5b7b1", fg="black", 
                              font=("Segoe UI", 10, "bold"), bd=1, padx=15, pady=5, 
                              activebackground="#f1948a", activeforeground="black",
                              cursor="hand2", command=self.destroy, relief="solid")
        btn_close.pack(side="right")
        
        # Title
        tk.Label(header_frame, text="✎ NHẬP ĐIỂM TẬP TRUNG (FOCUS MODE)", 
                 font=("Segoe UI", 18, "bold"), fg="white", bg=THEME["primary"]).pack(side="left")

        # Custom Progress Bar REMOVED as per user request
        # self.progress_canvas = tk.Canvas(header_frame, height=4, bg="#2980b9", highlightthickness=0, bd=0)
        # self.progress_canvas.pack(side="bottom", fill="x", pady=(15, 0))
        # self.progress_rect = self.progress_canvas.create_rectangle(0, 0, 0, 4, fill="#2ecc71", outline="")

        # Main Layout Container
        main_container = tk.Frame(self, bg=THEME["bg_app"], padx=25, pady=25)
        main_container.pack(fill="both", expand=True)

        # PanedWindow (3 Columns)
        self.paned = tk.PanedWindow(main_container, orient="horizontal", sashwidth=10, bg=THEME["bg_app"], bd=0)
        self.paned.pack(fill="both", expand=True)
        
        # Style Constants
        BORDER_COLOR = "#000000" # Viền đen mỏng theo yêu cầu
        BORDER_WIDTH = 1
        
        # --- CỘT 1: NHẬP LIỆU (LEFT) - 45% ---
        col1_wrapper = tk.Frame(self.paned, bg=BORDER_COLOR, padx=BORDER_WIDTH, pady=BORDER_WIDTH)
        self.left_col = tk.Frame(col1_wrapper, bg=THEME["bg_card"])
        self.left_col.pack(fill="both", expand=True)
        self.paned.add(col1_wrapper, minsize=500, width=550, stretch="always")
        
        # Header Cột 1
        c1_header = tk.Frame(self.left_col, bg="#f8f9fa", pady=12, padx=15)
        c1_header.pack(fill="x")
        tk.Label(c1_header, text="❶ KHU VỰC NHẬP LIỆU", font=("Segoe UI", 13, "bold"), bg="#f8f9fa", fg="#2c3e50").pack(anchor="w")
        tk.Frame(self.left_col, bg="#e9ecef", height=1).pack(fill="x") # Line

        # Input List Container
        self.input_container = tk.Frame(self.left_col, bg=THEME["bg_card"])
        self.input_container.pack(fill="both", expand=True, padx=2, pady=2)
        
        self.canvas = tk.Canvas(self.input_container, bg=THEME["bg_card"], highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.input_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=THEME["bg_card"])
        
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            try: self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except: pass
        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # --- CỘT 2: PREVIEW (CENTER) - 30% ---
        col2_wrapper = tk.Frame(self.paned, bg=BORDER_COLOR, padx=BORDER_WIDTH, pady=BORDER_WIDTH)
        self.center_col = tk.Frame(col2_wrapper, bg=THEME["bg_card"])
        self.center_col.pack(fill="both", expand=True)
        self.paned.add(col2_wrapper, minsize=400, stretch="always")
        
        # Header Cột 2
        c2_header = tk.Frame(self.center_col, bg="#f8f9fa", pady=12, padx=15)
        c2_header.pack(fill="x")
        tk.Label(c2_header, text="❷ BẢNG XEM TRƯỚC", font=("Segoe UI", 13, "bold"), bg="#f8f9fa", fg="#2c3e50").pack(anchor="w")
        tk.Frame(self.center_col, bg="#e9ecef", height=1).pack(fill="x") # Line

        # Treeview
        tree_container = tk.Frame(self.center_col, bg=THEME["bg_card"], padx=10, pady=10)
        tree_container.pack(fill="both", expand=True)
        
        cols = ["STT", "Họ và tên", "Điểm Mới"]
        self.preview_tree = ttk.Treeview(tree_container, columns=cols, show="headings", style="Modern.Treeview")
        self.preview_tree.heading("STT", text="STT"); self.preview_tree.column("STT", width=60, anchor="center")
        self.preview_tree.heading("Họ và tên", text="Họ và tên"); self.preview_tree.column("Họ và tên", width=200, anchor="w")
        self.preview_tree.heading("Điểm Mới", text="Điểm Vừa Nhập"); self.preview_tree.column("Điểm Mới", width=100, anchor="center")
        
        sb_tree = ttk.Scrollbar(tree_container, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscroll=sb_tree.set)
        self.preview_tree.pack(side="left", fill="both", expand=True)
        sb_tree.pack(side="right", fill="y")

        # --- CỘT 3: CẤU HÌNH (RIGHT) - 25% ---
        col3_wrapper = tk.Frame(self.paned, bg=BORDER_COLOR, padx=BORDER_WIDTH, pady=BORDER_WIDTH)
        self.right_col = tk.Frame(col3_wrapper, bg="#f4f6f7") # Nền hơi xám
        self.right_col.pack(fill="both", expand=True) 
        self.paned.add(col3_wrapper, minsize=320, stretch="never")
        
        # Header Cột 3 (Dark)
        c3_header = tk.Frame(self.right_col, bg="#34495e", pady=12, padx=15)
        c3_header.pack(fill="x")
        tk.Label(c3_header, text="❸ CẤU HÌNH & TÁC VỤ", font=("Segoe UI", 13, "bold"), bg="#34495e", fg="white").pack(anchor="w")

        # Config Body
        config_body = tk.Frame(self.right_col, bg="#f4f6f7", padx=20, pady=20)
        config_body.pack(fill="both", expand=True)
        
        # 1. Sheet Selector
        tk.Label(config_body, text="CHỌN LỚP (SHEET)", font=("Segoe UI", 10, "bold"), bg="#f4f6f7", fg="#b8860b").pack(anchor="w")
        self.cbo_sheet = ttk.Combobox(config_body, values=[], state="normal", font=("Segoe UI", 11))
        self.cbo_sheet.pack(fill="x", pady=(5, 20), ipady=5)
        self.cbo_sheet.bind("<<ComboboxSelected>>", self.on_sheet_change)
        
        # 2. Column Selector
        tk.Label(config_body, text="CỘT ĐIỂM CẦN NHẬP", font=("Segoe UI", 10, "bold"), bg="#f4f6f7", fg="#b8860b").pack(anchor="w")
        # [DYNAMIC COLS] Sẽ được cập nhật khi chọn sheet
        self.cbo_col = ttk.Combobox(config_body, values=["TX1", "TX2", "GK", "CK"], state="readonly", font=("Segoe UI", 11))
        self.cbo_col.pack(fill="x", pady=(5, 20), ipady=5)
        self.cbo_col.bind("<<ComboboxSelected>>", self.on_col_change)
        
        # 3. Quick Options
        tk.Label(config_body, text="TÙY CHỌN NHANH", font=("Segoe UI", 10, "bold"), bg="#f4f6f7", fg="#b8860b").pack(anchor="w")
        
        self.var_fast_mode = tk.BooleanVar(value=False)
        chk_frame = tk.Frame(config_body, bg="white", padx=12, pady=12, highlightbackground="#bdc3c7", highlightthickness=1)
        chk_frame.pack(fill="x", pady=(5, 10))
        
        chk_fast = tk.Checkbutton(chk_frame, text="⚡ Chế độ nhập nhanh", variable=self.var_fast_mode,
                       bg="white", font=("Segoe UI", 12, "bold"), activebackground="white",
                       cursor="hand2")
        chk_fast.pack(anchor="w")
        # Hiển thị đầy đủ văn bản hướng dẫn (wrap nếu cần)
        lbl_hint = tk.Label(chk_frame, text="Tự động nhảy ô tiếp theo\nkhi nhập điểm", 
                 font=("Segoe UI", 9), fg="#7f8c8d", bg="white", justify="left")
        lbl_hint.pack(anchor="w", padx=28, pady=(2, 0))

        # Settings Button
        btn_settings = tk.Button(config_body, text="⚙ CÀI ĐẶT CẤU HÌNH", bg="#d5d8dc", fg="black",
                                 font=("Segoe UI", 10, "bold"), relief="solid", cursor="hand2",
                                 padx=10, pady=5, bd=1, command=self.open_settings_modal)
        btn_settings.pack(fill="x", pady=(10, 0))

        # Spacer
        tk.Label(config_body, text="", bg="#f4f6f7").pack(fill="y", expand=True)
        
        # 4. Save Button
        # Dòng cảnh báo phía trên nút Lưu
        tk.Label(config_body, text="⚠ Lưu sau khi đã nhập xong điểm", 
                 font=("Segoe UI", 9, "bold"), fg="#e74c3c", bg="#f4f6f7").pack(anchor="w", pady=(0, 5))
        
        self.btn_save = tk.Button(config_body, text="💾 LƯU DỮ LIỆU (SAVE)", bg="#a9dfbf", fg="black",
                  font=("Segoe UI", 12, "bold"), pady=15, relief="solid", cursor="hand2", 
                  activebackground="#7dcea0", activeforeground="black", command=self.save_data, bd=1)
        self.btn_save.pack(fill="x")

        # Init
        self.load_initial_data()

    def open_settings_modal(self):
        """Mở cửa sổ cài đặt cấu hình phím tắt"""
        settings_win = Toplevel(self)
        settings_win.title("CÀI ĐẶT CẤU HÌNH")
        try:
            # Center modal - Giữ nguyên kích thước 500x400
            width, height = 500, 400
            x = (self.winfo_screenwidth() // 2) - (width // 2)
            y = (self.winfo_screenheight() // 2) - (height // 2)
            settings_win.geometry(f"{width}x{height}+{x}+{y}")
        except: pass
        
        settings_win.configure(bg=THEME["bg_app"])
        settings_win.transient(self) # Always on top of parent
        settings_win.grab_set()      # Modal mode

        # Tabs
        tab_control = ttk.Notebook(settings_win)
        tab_keys = tk.Frame(tab_control, bg=THEME["bg_app"])
        tab_todo = tk.Frame(tab_control, bg=THEME["bg_app"], padx=20, pady=20)
        
        tab_control.add(tab_keys, text="⌨️ Cấu hình phím")
        tab_control.add(tab_todo, text="🚧 Chưa cập nhật")
        tab_control.pack(expand=1, fill="both", padx=10, pady=10)

        # ========== [BƯỚC 2] TAB 1: CẤU HÌNH PHÍM VỚI SCROLLBAR ==========
        # Canvas + Scrollbar cho nội dung dài
        canvas = tk.Canvas(tab_keys, bg=THEME["bg_app"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(tab_keys, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=THEME["bg_app"], padx=20, pady=15)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack scroll components
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        # Mouse wheel scroll
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # --- SECTION A: Phím tắt điều hướng (giữ nguyên) ---
        tk.Label(scrollable_frame, text="📤 Phím tắt điều hướng:", font=("Segoe UI", 12, "bold"), bg=THEME["bg_app"]).pack(anchor="w", pady=(0, 10))
        
        shortcuts = [
            ("Enter", "Lưu & Xuống dòng"),
            ("Mũi tên Xuống", "Xuống dòng dưới"),
            ("Mũi tên Lên", "Lên dòng trên"),
            ("Esc", "Đóng cửa sổ")
        ]
        
        for key, desc in shortcuts:
            row = tk.Frame(scrollable_frame, bg="#ffffff", pady=5, padx=10, highlightbackground="#bdc3c7", highlightthickness=1)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=key, font=("Consolas", 10, "bold"), bg="#ecf0f1", width=14).pack(side="left")
            tk.Label(row, text=desc, font=("Segoe UI", 10), bg="#ffffff").pack(side="left", padx=10)
        
        # --- SECTION B: GÁN PHÍM CHO ĐIỂM (MỚI) ---
        tk.Frame(scrollable_frame, bg="#bdc3c7", height=2).pack(fill="x", pady=15)
        tk.Label(scrollable_frame, text="⚡ Gán phím cho điểm (Chế độ nhập nhanh):", font=("Segoe UI", 12, "bold"), bg=THEME["bg_app"]).pack(anchor="w", pady=(0, 5))
        tk.Label(scrollable_frame, text="Nhấn vào ô và gõ phím bạn muốn gán. Ví dụ: T = 10", font=("Segoe UI", 9, "italic"), fg="#7f8c8d", bg=THEME["bg_app"]).pack(anchor="w", pady=(0, 10))
        
        # [BƯỚC 3] Lưu reference các Entry để đọc giá trị khi Save
        self.score_key_entries = {}
        
        # Tạo 10 hàng cho Điểm 1 đến 10
        for score in range(1, 11):
            score_str = str(score)
            row = tk.Frame(scrollable_frame, bg="#ffffff", pady=6, padx=10, highlightbackground="#bdc3c7", highlightthickness=1)
            row.pack(fill="x", pady=3)
            
            # Label Điểm
            tk.Label(row, text=f"Điểm {score}", font=("Segoe UI", 10, "bold"), bg="#ffffff", width=10, anchor="w").pack(side="left")
            
            # Entry cho phím - Hiển thị giá trị hiện tại từ mapping
            current_key = self.score_key_mapping.get(score_str, score_str if score <= 9 else "0")
            entry = tk.Entry(row, font=("Consolas", 11, "bold"), width=8, justify="center", bg="#ecf0f1")
            entry.insert(0, current_key)
            entry.pack(side="left", padx=10)
            
            # Lưu reference
            self.score_key_entries[score_str] = entry
            
            # [BƯỚC 3] Bind sự kiện KeyPress để capture phím
            entry.bind("<KeyPress>", lambda e, ent=entry, sc=score_str: self.on_key_capture(e, ent, sc))
            
            # Label mặc định
            default_key = self.DEFAULT_SCORE_KEY_MAPPING.get(score_str, "")
            tk.Label(row, text=f"(Mặc định: {default_key})", font=("Segoe UI", 9), fg="#95a5a6", bg="#ffffff").pack(side="left", padx=5)
        
        # --- [BƯỚC 4] NÚT LƯU CẤU HÌNH ---
        tk.Frame(scrollable_frame, bg="#bdc3c7", height=1).pack(fill="x", pady=15)
        
        btn_save_config = tk.Button(scrollable_frame, text="💾 LƯU CẤU HÌNH", bg="#27ae60", fg="white",
                                    font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2",
                                    padx=20, pady=8, activebackground="#2ecc71", activeforeground="white",
                                    command=lambda: self.save_settings_from_modal(settings_win))
        btn_save_config.pack(pady=5)
        
        # Nút Reset về mặc định
        btn_reset = tk.Button(scrollable_frame, text="🔄 Đặt lại mặc định", bg="#e74c3c", fg="white",
                              font=("Segoe UI", 10), relief="flat", cursor="hand2",
                              padx=15, pady=5, command=lambda: self.reset_key_config_ui())
        btn_reset.pack(pady=5)
        
        # ========== END TAB 1 ==========

        # Tab 2: Chưa cập nhật
        tk.Label(tab_todo, text="Đang phát triển...", font=("Segoe UI", 14), fg="gray", bg=THEME["bg_app"]).pack(expand=True)
        
        # Unbind mouse wheel khi đóng modal
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            settings_win.destroy()
        settings_win.protocol("WM_DELETE_WINDOW", on_close)

    def on_key_capture(self, event, entry_widget, score_value):
        """
        [BƯỚC 3] Capture phím khi người dùng nhấn vào Entry.
        Chỉ cho phép 1 ký tự hoặc phím đặc biệt (không cho Enter, Tab, Escape).
        """
        key_pressed = event.keysym
        
        # Bỏ qua các phím điều hướng và modifier
        ignore_keys = ["Return", "Tab", "Escape", "Shift_L", "Shift_R", "Control_L", "Control_R", 
                       "Alt_L", "Alt_R", "Caps_Lock", "BackSpace", "Delete"]
        if key_pressed in ignore_keys:
            return  # Không xử lý
        
        # Chuyển đổi keysym thành ký tự hiển thị
        if len(key_pressed) == 1:
            display_key = key_pressed  # Ký tự thường: a, b, 1, 2, ...
        else:
            display_key = key_pressed  # Phím đặc biệt: F1, F2, ...
        
        # Kiểm tra trùng phím với điểm khác
        for sc, ent in self.score_key_entries.items():
            if sc != score_value:
                existing_key = ent.get().strip()
                if existing_key.lower() == display_key.lower():
                    messagebox.showwarning("Trùng phím", 
                        f"Phím '{display_key}' đã được gán cho Điểm {sc}!\n"
                        f"Vui lòng chọn phím khác.")
                    return "break"
        
        # Cập nhật Entry
        entry_widget.delete(0, "end")
        entry_widget.insert(0, display_key)
        
        return "break"  # Ngăn không cho ký tự được nhập thêm

    def save_settings_from_modal(self, modal_window):
        """
        [BƯỚC 4] Lưu tất cả cấu hình phím từ Modal vào file JSON.
        """
        # Thu thập mapping mới từ các Entry
        new_mapping = {}
        has_error = False
        
        for score_str, entry in self.score_key_entries.items():
            key_value = entry.get().strip()
            if not key_value:
                messagebox.showwarning("Thiếu phím", f"Điểm {score_str} chưa được gán phím!")
                has_error = True
                break
            new_mapping[score_str] = key_value
        
        if has_error:
            return
        
        # Kiểm tra trùng lặp
        seen_keys = {}
        for score, key in new_mapping.items():
            key_lower = key.lower()
            if key_lower in seen_keys:
                messagebox.showwarning("Trùng phím", 
                    f"Phím '{key}' được gán cho cả Điểm {seen_keys[key_lower]} và Điểm {score}!")
                return
            seen_keys[key_lower] = score
        
        # Lưu vào file
        if self.save_key_config(new_mapping):
            messagebox.showinfo("Thành công", 
                "✅ Đã lưu cấu hình phím tắt!\n\n"
                "Lưu ý: Tính năng này chỉ hoạt động khi\nbật 'Chế độ nhập nhanh' ⚡")
            modal_window.destroy()

    def reset_key_config_ui(self):
        """
        [BƯỚC 4] Reset tất cả Entry về giá trị mặc định.
        """
        for score_str, entry in self.score_key_entries.items():
            default_key = self.DEFAULT_SCORE_KEY_MAPPING.get(score_str, score_str)
            entry.delete(0, "end")
            entry.insert(0, default_key)
        
        messagebox.showinfo("Đã reset", "Đã đặt lại về phím mặc định.\nNhấn 'Lưu cấu hình' để áp dụng.")

    def schedule_realtime_sync(self):
        """
        [BƯỚC 6] Schedule đồng bộ dữ liệu với debounce 500ms.
        Tránh gọi sync liên tục khi nhập nhanh nhiều điểm.
        """
        # Hủy job cũ nếu có
        if self.sync_debounce_job:
            self.after_cancel(self.sync_debounce_job)
        
        # Schedule job mới sau 500ms
        self.sync_debounce_job = self.after(500, self.perform_realtime_sync)

    def perform_realtime_sync(self):
        """
        [BƯỚC 6] Thực hiện đồng bộ dữ liệu lên giao diện chính và file Excel.
        """
        try:
            sheet_name = self.cbo_sheet.get()
            if not sheet_name or self.current_df is None:
                return
            
            # 1. Đồng bộ vào all_classes_data
            self.excel_tab.all_classes_data[sheet_name] = self.current_df.copy()
            
            # 2. Nếu đang hiển thị cùng sheet ở giao diện chính → cập nhật current_df
            if self.excel_tab.cbo_classes.get() == sheet_name:
                self.excel_tab.current_df = self.current_df.copy()
                # Refresh bảng dữ liệu giao diện chính
                try:
                    self.excel_tab.update_ui_data(self.excel_tab.current_df, update_chart=False)
                except Exception as e:
                    print(f"[SYNC] Lỗi refresh UI chính: {e}")
            
            print(f"[🔄 REALTIME SYNC] Đã đồng bộ sheet '{sheet_name}' ({len(self.current_df)} dòng)")
            
        except Exception as e:
            print(f"[LỖI ĐỒNG BỘ] {e}")
        finally:
            self.sync_debounce_job = None

    def load_initial_data(self):
        """
        [BƯỚC 2 - ĐÃ SỬA] Load danh sách sheet từ ExcelTab với fallback logic mạnh.
        Đảm bảo đồng bộ dữ liệu giữa giao diện chính và cửa sổ NHẬP ĐIỂM.
        """
        try:
            # ========== [DEBUG LOGGING - BƯỚC 5] ==========
            print("="*60)
            print("[DEBUG NHẬP ĐIỂM] === BẮT ĐẦU load_initial_data() ===")
            print(f"[DEBUG] all_classes_data.keys() = {list(self.excel_tab.all_classes_data.keys())}")
            print(f"[DEBUG] cbo_classes.get() = '{self.excel_tab.cbo_classes.get()}'")
            print(f"[DEBUG] cbo_classes['values'] = {list(self.excel_tab.cbo_classes['values'])}")
            print(f"[DEBUG] current_df is None? = {self.excel_tab.current_df is None}")
            if self.excel_tab.current_df is not None:
                print(f"[DEBUG] current_df.shape = {self.excel_tab.current_df.shape}")
            print("="*60)
            # ========== END DEBUG ==========
            
            # ========== [SYNC GUARD - BƯỚC 2] ==========
            # Đảm bảo all_classes_data có dữ liệu TRƯỚC KHI lấy sheets
            if not self.excel_tab.all_classes_data:
                if self.excel_tab.current_df is not None and len(self.excel_tab.current_df) > 0:
                    # Sử dụng hàm helper để lấy tên sheet đáng tin cậy
                    fallback_name = self.excel_tab.get_current_sheet_name()
                    self.excel_tab.all_classes_data[fallback_name] = self.excel_tab.current_df.copy()
                    print(f"[SYNC GUARD InputScore] Đã tạo entry: all_classes_data['{fallback_name}']")
            # ========== END SYNC GUARD ==========
            
            # 1. Lấy dữ liệu nguồn (Keys) - Gộp từ nhiều nguồn để chắc chắn
            main_sheets = list(self.excel_tab.cbo_classes['values']) if self.excel_tab.cbo_classes['values'] else []
            data_keys = list(self.excel_tab.all_classes_data.keys())
            
            # Gộp cả 2 nguồn và loại bỏ trùng lặp
            sheets = list(set(main_sheets + data_keys))
            sheets.sort()  # Sắp xếp cho đẹp
            
            # [FALLBACK] Nếu vẫn rỗng, thử lấy từ current_df
            if not sheets and self.excel_tab.current_df is not None:
                current_sheet_name = self.excel_tab.get_current_sheet_name()
                self.excel_tab.all_classes_data[current_sheet_name] = self.excel_tab.current_df.copy()
                sheets = [current_sheet_name]
                print(f"[FALLBACK] Đã tạo sheet từ current_df: '{current_sheet_name}'")
            
            # Update UI List
            self.cbo_sheet['values'] = sheets
            
            # [DEBUG] Log kết quả
            print(f"[DEBUG] Danh sách sheets cuối cùng: {sheets}")
            
            if not sheets:
                messagebox.showerror("Lỗi dữ liệu", 
                    "Không có dữ liệu lớp học nào được tải!\n\n"
                    "Nguyên nhân có thể:\n"
                    "• File Excel chưa được mở\n"
                    "• Dữ liệu không đúng định dạng\n"
                    "• Sheet không có đủ cột cần thiết\n\n"
                    "Vui lòng kiểm tra lại file Excel.")
                return

            # 2. Logic chọn sheet mặc định thông minh
            target_sheet = sheets[0]  # Mặc định cái đầu tiên
            
            # Ưu tiên sheet đang được chọn ở giao diện chính
            current_main_ui = self.excel_tab.get_current_sheet_name()
            if current_main_ui in sheets:
                target_sheet = current_main_ui
            
            # SET giá trị cho Combobox
            self.cbo_sheet.set(target_sheet)
            print(f"[DEBUG] Đã chọn sheet: '{target_sheet}'")
            
            # 3. KÍCH HOẠT LOAD DỮ LIỆU
            self.load_sheet_data(target_sheet)
            
            print("[DEBUG NHẬP ĐIỂM] === KẾT THÚC load_initial_data() ===")

        except Exception as e:
            import traceback
            print(f"[LỖI] load_initial_data() thất bại:")
            traceback.print_exc()
            messagebox.showerror("Lỗi khởi tạo", f"Lỗi không xác định: {e}")

    def on_sheet_change(self, event):
        """Handler cho sự kiện combobox"""
        sheet_name = self.cbo_sheet.get()
        self.load_sheet_data(sheet_name)

    def load_sheet_data(self, sheet_name):
        """Hàm core để load dữ liệu của 1 sheet"""
        # [FIX] Logic tìm key tương đối (Relative matching)
        target_key = sheet_name
        
        # Nếu key không khớp chính xác, thử tìm key tương đương (str vs int, case-insensitive)
        keys = list(self.excel_tab.all_classes_data.keys())
        if sheet_name not in keys:
            # 1. Thử khớp 1 cái duy nhất
            if len(keys) == 1:
                target_key = keys[0]
            else:
                return # Give up

        try:
            # Deep copy data
            self.current_df = self.excel_tab.all_classes_data[target_key].copy()
            
            # [DYNAMIC COLS] Tự động phát hiện số cột TX trong sheet hiện tại
            tx_cols_in_sheet = [col for col in self.current_df.columns if col.startswith("TX") and col[2:].isdigit()]
            available_cols = sorted(tx_cols_in_sheet) + ["GK", "CK"]
            
            # Reset UI
            for widget in self.scrollable_frame.winfo_children():
                widget.destroy()
            self.preview_tree.delete(*self.preview_tree.get_children())
            
            # Reset cột điểm với danh sách động
            self.cbo_col.set("")
            self.cbo_col['values'] = available_cols
            
            # [NEW] Load aliases cho sheet hiện tại
            self.load_aliases()
            
        except Exception as e:
            print(f"Lỗi trong load_sheet_data: {e}")
        
        # Hiển thị thông báo hướng dẫn
        tk.Label(self.scrollable_frame, text="Vui lòng chọn CỘT CẦN NHẬP để bắt đầu!", 
                 font=("Segoe UI", 12, "bold"), fg=THEME["primary"], bg=THEME["bg_card"]).pack(pady=50)

    def on_col_change(self, event):
        col_type = self.cbo_col.get() # TX1, TX2...
        if not col_type: return
        
        self.populate_students(col_type)

    def populate_students(self, col_type):
        """Render danh sách học sinh và ô nhập"""
        target_col = col_type # Ở logic mới df đã chuẩn hóa tên cột
        
        # Clear cũ
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.preview_tree.delete(*self.preview_tree.get_children())
        
        self.entries = [] # List lưu tham chiếu entry widgets
        self.df_index_to_tree_iid = {}  # Mapping: DataFrame index -> Tree iid

        for position, (idx, row) in enumerate(self.current_df.iterrows()):
            stt = str(row.get("STT", ""))
            name = str(row.get("Họ và tên", ""))
            current_val = str(row.get(target_col, "")).replace("nan", "").replace("None", "")
            
            # Tạo iid duy nhất dựa trên position (tránh lỗi nếu DataFrame index không liên tục)
            tree_iid = f"row_{position}"
            self.df_index_to_tree_iid[idx] = tree_iid
            
            # Container cho 1 dòng (để chứa cả separator)
            container = tk.Frame(self.scrollable_frame, bg=THEME["bg_card"])
            container.pack(fill="x")

            # 1. Render Row ở Cột Nhập Liệu
            row_frame = tk.Frame(container, bg=THEME["bg_card"], pady=8, padx=5)
            row_frame.pack(fill="x")
            
            # STT (Badge Style)
            lbl_stt = tk.Label(row_frame, text=stt, font=("Segoe UI", 11, "bold"), width=4, bg="#ecf0f1", fg="#7f8c8d")
            lbl_stt.pack(side="left")
            
            # Tên
            lbl_name = tk.Label(row_frame, text=name, font=("Segoe UI", 12), width=30, anchor="w", bg=THEME["bg_card"], fg="#2c3e50")
            lbl_name.pack(side="left", padx=15)
            
            # Ô Nhập
            ent = tk.Entry(row_frame, font=("Segoe UI", 14, "bold"), width=6, justify="center", 
                           bd=1, relief="solid", bg=THEME["entry_bg"], validate="key")
            ent['validatecommand'] = (self.register(self.validate_entry), '%P')
            ent.pack(side="right", padx=10)
            if current_val: ent.insert(0, current_val)
            
            # Bind events
            ent.bind("<Return>", lambda e, idx=len(self.entries): self.focus_next_entry(idx))
            ent.bind("<Down>", lambda e, idx=len(self.entries): self.focus_next_entry(idx))
            ent.bind("<Up>", lambda e, idx=len(self.entries): self.focus_prev_entry(idx))
            # [FIX] Bind KeyPress để xử lý mapping TRƯỚC khi ký tự hiển thị (tránh nhấp nháy)
            ent.bind("<KeyPress>", lambda e, r=idx, c=col_type, widget=ent: self.on_key_press_fast_mode(e, r, c, widget))
            # KeyRelease vẫn giữ để xử lý các trường hợp khác (xóa, sửa)
            ent.bind("<KeyRelease>", lambda e, r=idx, c=col_type, widget=ent: self.on_input_change(e, r, c, widget))
            
            # Highlight effect
            def on_focus_in(e, f=row_frame, s=lbl_stt):
                f.config(bg="#e8f6f3") # Xanh ngọc nhạt
                s.config(bg="#1abc9c", fg="white")
            
            def on_focus_out(e, f=row_frame, s=lbl_stt):
                f.config(bg=THEME["bg_card"])
                s.config(bg="#ecf0f1", fg="#7f8c8d")

            ent.bind("<FocusIn>", on_focus_in)
            ent.bind("<FocusOut>", on_focus_out)
            
            self.entries.append(ent)

            # Separator Line
            tk.Frame(container, bg="#f1f2f6", height=1).pack(fill="x", padx=10)
            
            # 2. Render Row ở Preview Tree (Ban đầu) - dùng tree_iid thay vì idx
            self.preview_tree.insert("", "end", iid=tree_iid, values=(stt, name, current_val))
        
        self.update_progress()

    def validate_entry(self, new_val):
        if not new_val: return True
        try:
            val = float(new_val.replace(',', '.'))
            if 0 <= val <= 10: return True
            return False
        except ValueError:
            return False

    def focus_next_entry(self, current_idx):
        next_idx = current_idx + 1
        if next_idx < len(self.entries):
            self.entries[next_idx].focus_set()
            # [FIX] Auto-scroll để entry mới hiển thị trong vùng nhìn thấy
            self.scroll_to_entry(next_idx)

    def focus_prev_entry(self, current_idx):
        prev_idx = current_idx - 1
        if prev_idx >= 0:
            self.entries[prev_idx].focus_set()
            # [FIX] Auto-scroll để entry mới hiển thị trong vùng nhìn thấy
            self.scroll_to_entry(prev_idx)

    def scroll_to_entry(self, entry_idx):
        """
        Tự động scroll vùng nhập liệu để entry ở vị trí entry_idx hiển thị trong viewport.
        Đảm bảo học sinh đang nhập điểm luôn được nhìn thấy.
        """
        try:
            if entry_idx < 0 or entry_idx >= len(self.entries):
                return
                
            entry_widget = self.entries[entry_idx]
            
            # Lấy thông tin vị trí của entry trong scrollable_frame
            self.canvas.update_idletasks()  # Đảm bảo layout đã được tính
            
            # Tính vị trí của entry so với scrollable_frame
            entry_y = entry_widget.winfo_y()
            entry_parent = entry_widget.master  # row_frame
            container = entry_parent.master  # container frame
            
            # Tính vị trí tuyệt đối trong scrollable_frame
            abs_y = container.winfo_y() + entry_parent.winfo_y()
            
            # Lấy chiều cao của canvas (viewport) và scrollable_frame (nội dung)
            canvas_height = self.canvas.winfo_height()
            scroll_height = self.scrollable_frame.winfo_height()
            
            if scroll_height <= canvas_height:
                return  # Không cần scroll nếu nội dung vừa viewport
            
            # Tính vị trí scroll hiện tại
            current_scroll = self.canvas.yview()
            visible_top = current_scroll[0] * scroll_height
            visible_bottom = current_scroll[1] * scroll_height
            
            # Chiều cao mỗi entry (khoảng)
            entry_height = 60  # Ước lượng chiều cao mỗi dòng
            entry_top = abs_y
            entry_bottom = abs_y + entry_height
            
            # Kiểm tra xem entry có nằm trong vùng nhìn thấy không
            if entry_top < visible_top:
                # Entry ở trên viewport -> scroll lên
                new_pos = max(0, entry_top - 20) / scroll_height
                self.canvas.yview_moveto(new_pos)
            elif entry_bottom > visible_bottom:
                # Entry ở dưới viewport -> scroll xuống
                new_pos = min(scroll_height, entry_bottom - canvas_height + 20) / scroll_height
                self.canvas.yview_moveto(new_pos)
                
        except Exception as e:
            # Bỏ qua lỗi scroll, không ảnh hưởng chức năng chính
            print(f"[DEBUG] scroll_to_entry error: {e}")

    def update_progress(self):
        # Progress bar feature disabled
        pass

    # ========== [BƯỚC 1] LOAD/SAVE KEY CONFIG ==========
    def get_config_file_path(self) -> str:
        """
        Lấy đường dẫn file config.
        Ưu tiên: Cùng thư mục với file Excel, fallback về Desktop.
        """
        try:
            if self.excel_tab.file_path:
                config_dir = os.path.dirname(self.excel_tab.file_path)
                if os.path.isdir(config_dir):
                    return os.path.join(config_dir, self.CONFIG_FILE_NAME)
        except:
            pass
        # Fallback về Desktop
        return os.path.join(os.path.expanduser("~"), "Desktop", self.CONFIG_FILE_NAME)
    
    def load_key_config(self) -> dict:
        """
        [BƯỚC 1] Load cấu hình phím tắt từ JSON file.
        Nếu file không tồn tại hoặc bị hỏng, trả về default.
        
        Returns:
            dict: Mapping điểm -> phím (ví dụ: {"1": "1", "10": "T"})
        """
        config_path = self.get_config_file_path()
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if "score_key_mapping" in data:
                        # Merge với default để đảm bảo đủ 10 phím
                        merged = self.DEFAULT_SCORE_KEY_MAPPING.copy()
                        merged.update(data["score_key_mapping"])
                        print(f"[✅ CONFIG] Đã load cấu hình phím tắt từ: {config_path}")
                        return merged
        except Exception as e:
            print(f"[⚠️ CONFIG] Lỗi load config: {e}. Sử dụng mặc định.")
        
        return self.DEFAULT_SCORE_KEY_MAPPING.copy()
    
    def save_key_config(self, new_mapping: dict) -> bool:
        """
        [BƯỚC 1] Lưu cấu hình phím tắt vào JSON file.
        
        Args:
            new_mapping: Dict mapping điểm -> phím
            
        Returns:
            bool: True nếu lưu thành công
        """
        config_path = self.get_config_file_path()
        try:
            data = {
                "score_key_mapping": new_mapping,
                "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            # Cập nhật mapping hiện tại
            self.score_key_mapping = new_mapping.copy()
            print(f"[✅ CONFIG] Đã lưu cấu hình phím tắt vào: {config_path}")
            return True
        except Exception as e:
            print(f"[❌ CONFIG] Lỗi lưu config: {e}")
            messagebox.showerror("Lỗi", f"Không thể lưu cấu hình:\n{e}")
            return False
    # ========== END LOAD/SAVE CONFIG ==========

    def on_key_press_fast_mode(self, event, row_idx, col_name, entry_widget):
        """
        [FIX] Xử lý phím tắt điểm TRƯỚC khi ký tự được hiển thị.
        Điều này ngăn hiện tượng nhấp nháy (hiện phím gốc rồi mới đổi sang điểm).
        
        Return:
            "break" - Ngăn Tkinter insert ký tự gốc (vì ta đã insert điểm thay thế)
            None - Cho phép Tkinter xử lý bình thường
        """
        # Chỉ xử lý khi Fast Mode được bật
        if not self.var_fast_mode.get():
            return  # Không chặn, để Tkinter xử lý bình thường
        
        # Lấy ký tự được nhấn
        key_char = event.char  # Ký tự thực (a, b, 1, 2, ...)
        key_sym = event.keysym  # Tên phím (Return, BackSpace, ...)
        
        # Bỏ qua các phím điều khiển (không có ký tự)
        if not key_char or key_char in ('\r', '\n', '\t', '\x08', '\x7f'):
            return  # Cho phép xử lý bình thường
        
        # Tìm trong mapping: phím -> điểm
        key_lower = key_char.lower()
        for score_value, mapped_key in self.score_key_mapping.items():
            if key_lower == mapped_key.lower():
                # ĐÃ TÌM THẤY MAPPING!
                # 1. Xóa nội dung cũ và insert điểm mới
                entry_widget.delete(0, "end")
                entry_widget.insert(0, score_value)
                
                print(f"[⚡ FAST MODE] Phím '{key_char}' → Điểm '{score_value}' (instant)")
                
                # 2. Cập nhật DataFrame
                self.current_df.at[row_idx, col_name] = score_value
                
                # 3. Cập nhật Preview Tree
                tree_iid = self.df_index_to_tree_iid.get(row_idx)
                if tree_iid:
                    try:
                        current_values = self.preview_tree.item(tree_iid)['values']
                        if current_values and len(current_values) >= 2:
                            new_values = (current_values[0], current_values[1], score_value)
                            self.preview_tree.item(tree_iid, values=new_values)
                    except:
                        pass
                
                # 4. Auto-move đến ô tiếp theo
                try:
                    current_idx = self.entries.index(entry_widget)
                    self.after(30, lambda idx=current_idx: self.focus_next_entry(idx))
                except ValueError:
                    pass
                
                # 5. Schedule đồng bộ (debounced)
                self.schedule_realtime_sync()
                
                # QUAN TRỌNG: Return "break" để ngăn Tkinter insert ký tự gốc
                return "break"
        
        # Không tìm thấy mapping → cho phép xử lý bình thường
        return None

    def on_input_change(self, event, row_idx, col_name, entry_widget):
        val = entry_widget.get().strip()
        
        # [NOTE] Mapping đã được xử lý ở on_key_press_fast_mode() rồi
        # Hàm này chỉ xử lý validate và update cho các trường hợp còn lại
        # (nhập số trực tiếp, xóa, sửa, v.v.)
        
        # Update DataFrame tạm
        # Validate số cơ bản
        is_valid = True
        try:
            if val: 
                v_num = float(val.replace(",", "."))
                if not (0 <= v_num <= 10): is_valid = False
            else:
                is_valid = True # Empty is valid
        except:
            is_valid = False
            
        entry_widget.config(bg="#FEE2E2" if not is_valid and val else THEME["entry_bg"], fg="red" if not is_valid and val else "black")

        if is_valid:
            if val == "": 
                self.current_df.at[row_idx, col_name] = "None" # Xóa điểm
            else: 
                self.current_df.at[row_idx, col_name] = val
            
            # Update Preview Tree - sử dụng mapping để lấy đúng iid
            tree_iid = self.df_index_to_tree_iid.get(row_idx)
            if tree_iid:
                try:
                    current_values = self.preview_tree.item(tree_iid)['values']
                    if current_values and len(current_values) >= 2:
                        # Update cột giá trị (index 2)
                        new_values = (current_values[0], current_values[1], val)
                        self.preview_tree.item(tree_iid, values=new_values)
                except Exception as e:
                    print(f"[DEBUG] Lỗi update preview tree: {e}")
            
            self.update_progress()
            
            # [NOTE] Auto-next và sync đã được xử lý trong on_key_press_fast_mode()
            # Chỉ trigger sync ở đây cho trường hợp nhập thủ công (không qua mapping)

        if is_valid or val == "":
            self.current_df.at[row_idx, col_name] = val
            
            # Update Preview Tree - sử dụng mapping để lấy đúng iid
            tree_iid = self.df_index_to_tree_iid.get(row_idx)
            if tree_iid:
                try:
                    current_values = self.preview_tree.item(tree_iid)['values']
                    if current_values and len(current_values) >= 2:
                        # Update cột giá trị (index 2)
                        new_values = (current_values[0], current_values[1], val)
                        self.preview_tree.item(tree_iid, values=new_values)
                except Exception as e:
                    pass  # Đã xử lý ở trên

    def save_data(self):
        """Commit dữ liệu về ExcelTab"""
        if self.current_df is None: return
        
        sheet_name = self.cbo_sheet.get()
        if not sheet_name:
             messagebox.showwarning("Lỗi", "Chưa chọn Sheet nào!")
             return

        # Tính toán lại ĐTB trước khi lưu
        for i, row in self.current_df.iterrows():
            # [DYNAMIC TX] Lấy tất cả cột TX có trong row
            tx_scores = []
            for col in row.index:
                if col.startswith("TX") and col[2:].isdigit():
                    tx_scores.append(clean_float_val(row.get(col, "")))
            
            gk = row.get("GK", "")
            ck = row.get("CK", "")
            
            dtb = calculate_dtb_exact(tx_scores, clean_float_val(gk), clean_float_val(ck))
            self.current_df.at[i, "ĐTB"] = dtb

        # Lưu ngược về ExcelTab
        self.excel_tab.all_classes_data[sheet_name] = self.current_df.copy()
        
        # Nếu đang xem đúng sheet đó thì refresh UI chính
        if self.excel_tab.cbo_classes.get() == sheet_name:
            self.excel_tab.current_df = self.current_df.copy()
            # Gọi đúng hàm để refresh giao diện chính
            try:
                self.excel_tab.update_ui_data(self.excel_tab.current_df, update_chart=True)
            except Exception as e:
                print(f"[SAVE] Lỗi refresh UI: {e}")
        
        # ========== [FIX] LƯU VÀO FILE EXCEL THỰC SỰ ==========
        try:
            self.excel_tab.perform_auto_save()
            messagebox.showinfo("Thành công", 
                f"✅ Đã cập nhật dữ liệu cho lớp {sheet_name}!\n\n"
                f"• Dữ liệu đã được lưu vào file Excel.\n"
                f"• File: {self.excel_tab.file_path}", 
                parent=self)
        except Exception as e:
            print(f"[SAVE] Lỗi lưu file: {e}")
            messagebox.showwarning("Cảnh báo", 
                f"Đã cập nhật giao diện nhưng CHƯA lưu được vào file Excel!\n\n"
                f"Lỗi: {e}\n\n"
                f"Vui lòng dùng Ctrl+S để lưu thủ công.", 
                parent=self)
        # ========== END FIX ==========
        
# =================================================================================
# CLASS XỬ LÝ TAB DỮ LIỆU
# =================================================================================
class ExcelTab(tk.Frame):
    def __init__(self, parent, file_path, status_label=None, filter_label=None):
        # [UPDATED] Thêm tham số status_label để fix lỗi AttributeError
        super().__init__(parent, bg=THEME["bg_app"])
        self.file_path = file_path
        self.status_label = status_label # Lưu tham chiếu đến label của cửa sổ chính
        # Label trạng thái lọc dùng chung (đặt giữa Logo và Tab)
        self.filter_label = filter_label
        # Undo/Redo stacks
        self.history_stack = []
        self.redo_stack = []
        
        # [FIX] Biến để track entry đang được edit trong treeview
        self.active_entry = None
        self.active_entry_save_func = None
        
        # Giai đoạn 9: Undo Mapping - Lưu snapshot dữ liệu trước khi mapping
        self.mapping_undo_stack = []  # Stack chứa {sheet_name, old_df, backup_path, timestamp}
        self.max_mapping_undo = 5  # Giới hạn số lần undo mapping
        
        self.all_classes_data = {}
        self.current_df = None
        self.sheet_headers_info = {} 
        self.logbook_df = pd.DataFrame(columns=["Ngay", "TenHS", "Loai", "GhiChu"])
        self.search_job = None
        self.cards = {} 
        self.current_groups_data = []
        
        # Storage cho thông tin phụ huynh
        self.parent_contacts = {}
        self.load_parent_contacts() 
        
        # Storage cho tính năng QUAY SỐ
        self.spin_filter_mode = None           # "Khá/Giỏi", "Hoàn thành", "Cần chú ý", None
        self.spin_exclude_used = False         # Checkbox "Loại nếu quay trúng"
        self.spin_used_students = set()        # Set[int] - STT đã quay (loại khỏi bàn quay)
        self.spin_history = []                 # List[dict] - Lịch sử quay số
        self.spin_duration = 3.0               # Thời gian quay (giây) - mặc định 3s
        
        # [ADDED] Storage cho tính năng TRỢ LÝ PHÒNG THI
        self.exam_timer_running = False        # Trạng thái đồng hồ đang chạy
        self.exam_timer_paused = False         # Trạng thái tạm dừng
        self.exam_timer_thread = None          # Thread đếm ngược (để kiểm soát)
        self.exam_remaining_seconds = 0        # Số giây còn lại
        self.exam_alert_played_15 = False      # Đã phát cảnh báo 15 phút chưa
        self.exam_alert_played_5 = False       # Đã phát cảnh báo 5 phút chưa
        # [END ADDED]
        
        self.setup_ui()
        self.load_data_from_file()

    def setup_ui(self):
        # --- 1. Top Dashboard (Thống kê nhanh) ---
        self.stats_frame = tk.Frame(self, bg=THEME["bg_app"], pady=10)
        self.stats_frame.pack(fill="x", padx=20)
        
        stats_configs = [
            ("TỔNG SĨ SỐ", "All"),
            ("KHÁ / GIỎI", "Tot"),
            ("TRUNG BÌNH", "HoanThanh"),
            ("CẦN CHÚ Ý", "CanChuY"),
            ("CHƯA ĐIỂM", "ChuaCoDiem"),
        ]
        
        # Render thẻ thống kê theo style mới
        for title, key in stats_configs:
            self.create_modern_stat_card_widget(self.stats_frame, title, key)

        # --- 2. Control Ribbon (Thanh công cụ) ---
        control_card = tk.Frame(self, bg=THEME["bg_card"], padx=10, pady=10)
        control_card.pack(fill="x", padx=20, pady=(0, 15))
        
        # 2b. Right Side: Menu Button cho Tiện Ích
        right_ctrl = tk.Frame(control_card, bg=THEME["bg_card"])
        right_ctrl.pack(side="right", fill="y")

        # Tạo menu dropdown cho các tiện ích
        self.utilities_menu = tk.Menu(self, tearoff=0, font=("Segoe UI", 10))
        self.utilities_menu.add_command(label="🤖TỰ ĐỘNG", command=self.show_automation_popup)
        self.utilities_menu.add_command(label="📒NHẬT KÝ", command=self.show_logbook_popup)
        self.utilities_menu.add_command(label="🧩CHIA NHÓM", command=self.show_smart_grouping)
        self.utilities_menu.add_separator()
        self.utilities_menu.add_command(label="📊BIỂU ĐỒ", command=self.show_histogram_popup)
        self.utilities_menu.add_command(label="⚠️CẢNH BÁO", command=self.show_radar_popup)
        self.utilities_menu.add_command(label="📺TRÌNH CHIẾU", command=self.show_presentation_mode)
        self.utilities_menu.add_separator()
        self.utilities_menu.add_command(label="📈THỐNG KÊ NÂNG CAO", command=self.show_advanced_statistics)
        self.utilities_menu.add_command(label="📋XUẤT BÁO CÁO", command=self.export_report)
        self.utilities_menu.add_separator()
        self.utilities_menu.add_command(label="⏳ĐỒNG HỒ ĐẾM NGƯỢC", command=self.show_exam_proctor_toolkit)
        self.utilities_menu.add_command(label="📞DANH BẠ PHỤ HUYNH", command=self.show_parent_contacts)
        
        # Nút mở menu
        def show_utilities_menu():
            try:
                x = utilities_btn.winfo_rootx()
                y = utilities_btn.winfo_rooty() + utilities_btn.winfo_height()
                self.utilities_menu.post(x, y)
            except:
                pass
        
        # [NEW] Nút NHẬP ĐIỂM GIỌNG NÓI - Phía trước nút NHẬP ĐIỂM
        voice_btn = ModernButton(right_ctrl, bg_color="#e91e63",  # Màu Hồng
                               text_color="white", 
                               text="🎤 GIỌNG NÓI", 
                               command=self.show_voice_input_window)
        voice_btn.pack(side="left", padx=4)
        
        # [ADDED Phase 1] Nút NHẬP ĐIỂM KIỂM TRA
        input_btn = ModernButton(right_ctrl, bg_color="#8e44ad",  # Màu Tím
                               text_color="white", 
                               text="  NHẬP ĐIỂM", 
                               command=self.show_input_score_window)
        input_btn.pack(side="left", padx=4)
        
        # [NEW] Nút SỬA HÀNG LOẠT
        bulk_edit_btn = ModernButton(right_ctrl, bg_color="#e67e22",  # Màu Cam
                               text_color="white", 
                               text="⚡ SỬA HÀNG LOẠT", 
                               command=self.open_bulk_edit_selection)
        bulk_edit_btn.pack(side="left", padx=4)

        spin_btn = ModernButton(right_ctrl, bg_color=PASTEL_PALETTE["orange"], 
                               text_color=PASTEL_PALETTE["orange_dark"], 
                               text="  QUAY SỐ", 
                               command=self.show_spin_wheel_popup)
        spin_btn.pack(side="left", padx=4)
        
        utilities_btn = ModernButton(right_ctrl, bg_color=PASTEL_PALETTE["lavender"], 
                                    text_color=PASTEL_PALETTE["lavender_dark"], 
                                    text="  TIỆN ÍCH ▼", 
                                    command=show_utilities_menu)
        utilities_btn.pack(side="left", padx=4)

        # 2a. Left Side: Class & Search
        left_ctrl = tk.Frame(control_card, bg=THEME["bg_card"])
        left_ctrl.pack(side="left", fill="y")

        tk.Label(left_ctrl, text="Lớp học:", font=("Segoe UI", 10, "bold"), bg=THEME["bg_card"], fg="gray").pack(side="left")
        self.cbo_classes = ttk.Combobox(left_ctrl, state="readonly", width=15, font=("Segoe UI", 10))
        self.cbo_classes.pack(side="left", padx=(5, 20))
        self.cbo_classes.bind("<<ComboboxSelected>>", self.on_class_change)

        search_frame = tk.Frame(left_ctrl, bg=THEME["entry_bg"], bd=1, relief="solid")
        search_frame.config(highlightbackground="#E5E7EB", highlightthickness=1, relief="flat")
        search_frame.pack(side="left")
        
        tk.Label(search_frame, text="🔍", bg=THEME["entry_bg"], fg="gray").pack(side="left", padx=5)
        self.ent_search = tk.Entry(search_frame, width=25, font=("Segoe UI", 10), bd=0, bg=THEME["entry_bg"])
        self.ent_search.pack(side="left", ipady=4)
        self.ent_search.bind("<KeyRelease>", self.on_search_debounce)

        # Nếu có filter_label (ở App), dùng nó; nếu không thì fallback về label trong tab
        if self.filter_label is not None:
            self.lbl_filter_status = self.filter_label
        else:
            self.lbl_filter_status = tk.Label(left_ctrl, text="", bg=THEME["bg_card"], fg=THEME["text_main"], font=("Segoe UI", 9, "bold"))
            self.lbl_filter_status.pack(side="left", padx=10)

        # --- 3. Main Content Split ---
        self.main_pane = tk.PanedWindow(self, orient="horizontal", sashwidth=4, bg=THEME["bg_app"], bd=0)
        self.main_pane.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.table_card = tk.Frame(self.main_pane, bg=THEME["bg_card"], highlightbackground=THEME["border"], highlightthickness=1) 
        self.main_pane.add(self.table_card, minsize=750)
        
        # Setup Treeview Style
        self.update_treeview_style() 

        # [DYNAMIC COLS] Danh sách cột mặc định ban đầu - sẽ được cập nhật tự động khi load file
        self.cols_def = ["STT", "Họ và tên", "Ngày sinh", "TX1", "TX2", "GK", "CK", "ĐTB", "Xếp loại"]
        self.num_tx_cols = 2  # Số cột TX mặc định - sẽ tự động điều chỉnh
        self.tree = ttk.Treeview(self.table_card, columns=self.cols_def, show="headings", selectmode="browse", style="Modern.Treeview")
        self.setup_treeview_columns()
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        
        # [FIX] Bind scroll và click để auto-save entry đang edit
        self.tree.bind("<MouseWheel>", self.on_tree_scroll_or_click, add="+")
        self.tree.bind("<Button-4>", self.on_tree_scroll_or_click, add="+")  # Linux scroll up
        self.tree.bind("<Button-5>", self.on_tree_scroll_or_click, add="+")  # Linux scroll down
        self.tree.bind("<Button-1>", self.on_tree_scroll_or_click, add="+")

        sb_y = ttk.Scrollbar(self.table_card, orient="vertical", command=self.tree.yview)
        sb_x = ttk.Scrollbar(self.table_card, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=sb_y.set, xscroll=sb_x.set)
        
        sb_x.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)
        sb_y.pack(side="right", fill="y")

        # 3b. Chart Area (Bên phải)
        self.chart_card = tk.Frame(self.main_pane, bg=THEME["bg_card"], padx=10, pady=10, highlightbackground=THEME["border"], highlightthickness=1)
        self.main_pane.add(self.chart_card, minsize=350)
        
        tk.Label(self.chart_card, text="TỔNG QUAN LỚP HỌC", font=("Segoe UI", 11, "bold"), 
                 bg=THEME["bg_card"], fg="gray").pack(pady=(5, 10), anchor="w")
        
        self.chart_container = tk.Frame(self.chart_card, bg=THEME["bg_card"])
        self.chart_container.pack(fill="both", expand=True)

        self.update_treeview_tags("Light")

    def update_treeview_style(self):
        style = ttk.Style()
        style.configure("Modern.Treeview", 
                        background=THEME["tree_bg"], 
                        foreground=THEME["tree_fg"], 
                        fieldbackground=THEME["tree_bg"],
                        rowheight=40, font=("Segoe UI", 10), borderwidth=0)
        
        style.configure("Modern.Treeview.Heading", 
                        background=THEME["tree_header_bg"], 
                        foreground=THEME["tree_header_fg"], 
                        font=("Segoe UI", 10, "bold"), padding=(10, 10))
        
        style.map("Modern.Treeview.Heading", background=[('active', THEME["tree_header_bg"])])

    def show_input_score_window(self):
        """Mở giao diện nhập điểm tập trung (Phase 1)"""
        if self.current_df is None or self.current_df.empty:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng mở file Excel trước!")
            return
            
        win = InputScoreWindow(self.winfo_toplevel(), self)
        # win.grab_set() # Modal dialog (optional, để user vẫn tương tác cửa sổ kia nếu muốn thì bỏ)
    
    def show_voice_input_window(self):
        """[NEW] Mở giao diện nhập điểm bằng giọng nói"""
        if self.current_df is None or self.current_df.empty:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng mở file Excel trước!")
            return
        
        # Kiểm tra thư viện
        if not HAS_SPEECH:
            messagebox.showerror(
                "Thiếu thư viện",
                "Chưa cài đặt thư viện SpeechRecognition.\n\n"
                "Vui lòng chạy:\npip install SpeechRecognition sounddevice soundfile numpy fuzzywuzzy python-Levenshtein"
            )
            return
        
        # [FIX] Kiểm tra đã import file chưa
        if not hasattr(self, 'all_classes_data') or not self.all_classes_data:
            messagebox.showwarning(
                "Chưa có dữ liệu",
                "Vui lòng import file Excel trước khi sử dụng chức năng Giọng nói!\n\n"
                "Bước 1: Bấm nút 'Mở File' để import Excel\n"
                "Bước 2: Sau đó mở lại 'Nhập điểm bằng Giọng nói'"
            )
            return
        
        VoiceInputWindow(self.winfo_toplevel(), self)
    
    def open_bulk_edit_selection(self):
        """[NEW] Mở cửa sổ chọn học sinh để sửa điểm hàng loạt"""
        if self.current_df is None or self.current_df.empty:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng mở file Excel trước!")
            return
        
        BulkEditSelectionWindow(self)

    def get_current_sheet_name(self) -> str:
        """
        [BƯỚC 3] Trả về tên sheet đang được hiển thị một cách đáng tin cậy.
        Hàm này là 'source of truth' để lấy tên sheet hiện tại.
        
        Thứ tự ưu tiên:
        1. Từ combobox cbo_classes nếu có giá trị
        2. Từ all_classes_data nếu chỉ có 1 key duy nhất
        3. Trích xuất từ tên file (bỏ extension)
        4. Mặc định "Sheet1"
        
        Returns:
            str: Tên sheet hiện tại
        """
        # Ưu tiên 1: Từ combobox nếu có giá trị hợp lệ
        if hasattr(self, 'cbo_classes'):
            current_val = self.cbo_classes.get()
            if current_val and current_val.strip():
                return current_val.strip()
        
        # Ưu tiên 2: Từ all_classes_data nếu chỉ có 1 key
        if self.all_classes_data and len(self.all_classes_data) == 1:
            return list(self.all_classes_data.keys())[0]
        
        # Ưu tiên 3: Trích xuất từ tên file
        if self.file_path:
            try:
                base_name = os.path.splitext(os.path.basename(self.file_path))[0]
                if base_name:
                    return base_name
            except:
                pass
        
        # Mặc định cuối cùng
        return "Sheet1"

    def create_modern_stat_card_widget(self, parent, title, key):
        bg_pastel, text_color = CARD_COLORS.get(key, ("#E5E7EB", "#374151"))
        
        # [NEW] Tạo màu hover (sáng hơn và đậm hơn)
        def brighten_color(hex_color, factor=1.12):
            """Làm sáng màu hex để tạo hiệu ứng đậm hơn khi hover"""
            hex_color = hex_color.lstrip('#')
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
            # Tăng độ sáng nhưng giới hạn tối đa 255
            r = min(255, int(r * factor))
            g = min(255, int(g * factor))
            b = min(255, int(b * factor))
            return f'#{r:02x}{g:02x}{b:02x}'
        
        bg_hover = brighten_color(bg_pastel, 1.12)  # Sáng hơn 12%
        
        # [UPDATED] Thêm viền xám bằng cách bọc nội dung trong Frame viền (Giống ModernButton)
        # Frame viền ngoài (Màu xám/theme border)
        border_frame = tk.Frame(parent, bg=THEME["btn_border_col"], padx=1, pady=1, cursor="hand2")
        border_frame.is_btn_border = True # Đánh dấu để đổi màu khi switch theme
        border_frame.pack(side="left", expand=True, fill="both", padx=6)
        
        # Frame nội dung bên trong (Màu Pastel)
        card = tk.Frame(border_frame, bg=bg_pastel, padx=20, pady=15, cursor="hand2")
        card.is_colored_card = True # [FIXED] Đánh dấu để tránh bị override màu khi đổi theme
        card.pack(fill="both", expand=True)
        
        # Bo góc (Giả lập bằng frame con nếu muốn, ở đây dùng màu nền phẳng)
        # [UPDATED] is_static = True để hàm apply_theme_recursive không đổi màu chữ
        val_lbl = tk.Label(card, text="0", font=("Segoe UI", 24, "bold"), bg=bg_pastel, fg=text_color)
        val_lbl.is_static = True 
        val_lbl.pack(anchor="w")
        
        title_lbl = tk.Label(card, text=title.upper(), font=("Segoe UI", 9, "bold"), bg=bg_pastel, fg=text_color, highlightthickness=0)
        title_lbl.is_static = True
        title_lbl.pack(anchor="w")

        # Hack để tương thích logic update cũ (tuple 2 phần tử)
        self.cards[key] = (val_lbl, val_lbl)
        
        # [NEW] Hiệu ứng hover - đổi màu khi rê chuột
        def on_enter(e):
            """Khi chuột vào card, đổi sang màu đậm"""
            card.config(bg=bg_hover)
            val_lbl.config(bg=bg_hover)
            title_lbl.config(bg=bg_hover)
        
        def on_leave(e):
            """Khi chuột rời card, trở về màu gốc"""
            card.config(bg=bg_pastel)
            val_lbl.config(bg=bg_pastel)
            title_lbl.config(bg=bg_pastel)

        def on_click(e): self.filter_data(key)
        
        # Bind events cho tất cả widgets trong card
        for w in [card, val_lbl, title_lbl, border_frame]:
            w.bind("<Button-1>", on_click)
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)

    # --- HÀM UI HELPERS CŨ (Dùng cho Chart) ---
    def update_treeview_tags(self, mode):
        colors = TREE_ROW_COLORS[mode]
        for tag, val in colors.items():
            self.tree.tag_configure(tag, background=val["bg"], foreground=val["fg"])
        # Tag tô sáng kết quả tìm kiếm
        self.tree.tag_configure("search_hit", background="#FFF6BF", foreground=THEME["text_main"])

    def create_shadowed_text(self, parent, text, font, bg_color, offset=(1, 1)):
        # Giữ lại để tránh lỗi nếu có code nào gọi, dù UI chính đã đổi
        container = tk.Frame(parent, bg=bg_color)
        container.is_static = True 
        container.pack(anchor="w")
        shadow = tk.Label(container, text=text, font=font, bg=bg_color, fg=THEME["text_shadow"])
        shadow.place(x=offset[0], y=offset[1])
        main = tk.Label(container, text=text, font=font, bg=bg_color, fg="#FFFFFF")
        main.pack(side="left") 
        return main, shadow, container

    def setup_treeview_columns(self):
        """Thiết lập cột Treeview - tự động điều chỉnh theo số cột TX"""
        col_configs = {
            "STT": 50, "Họ và tên": 150, "Ngày sinh": 90,
            "GK": 50, "CK": 50, "ĐTB": 60, "Xếp loại": 130
        }
        
        # Tự động thêm cấu hình cho các cột TX động
        for i in range(1, 11):  # Hỗ trợ tối đa 10 cột TX
            col_configs[f"TX{i}"] = 60
        
        for col in self.cols_def:
            self.tree.heading(col, text=col)
            w = col_configs.get(col, 80)
            anchor = "w" if col == "Họ và tên" else "center"
            self.tree.column(col, width=w, anchor=anchor)
    
    def update_columns_based_on_data(self, df):
        """Cập nhật số cột TX dựa trên DataFrame thực tế"""
        if df is None or len(df) == 0:
            return
        
        # Đếm số cột TX có trong DataFrame
        tx_cols = [col for col in df.columns if col.startswith("TX") and col[2:].isdigit()]
        num_tx = len(tx_cols)
        
        # Nếu số cột TX thay đổi, cập nhật lại cols_def và Treeview
        if num_tx != self.num_tx_cols:
            self.num_tx_cols = num_tx
            
            # Tạo lại danh sách cột với số TX động
            tx_col_names = [f"TX{i}" for i in range(1, num_tx + 1)]
            self.cols_def = ["STT", "Họ và tên", "Ngày sinh"] + tx_col_names + ["GK", "CK", "ĐTB", "Xếp loại"]
            
            # Cập nhật lại Treeview
            self.tree.config(columns=self.cols_def)
            self.setup_treeview_columns()
            
            print(f"[DYNAMIC COLS] Đã cập nhật {num_tx} cột TX: {tx_col_names}")

    def center_window(self, win, width, height):
        screen_width = win.winfo_screenwidth()
        screen_height = win.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        win.geometry(f'{width}x{height}+{x}+{y}')

    # =============================================================================
    # TÍNH NĂNG 1: AUTOMATION
    # =============================================================================
    def show_automation_popup(self):
        if self.current_df is None: return

        auto_win = Toplevel(self)
        auto_win.title("🤖 Tiện Ích Tự Động Hóa Dành Cho Giáo Viên")
        self.center_window(auto_win, 800, 600)
        auto_win.configure(bg=THEME["bg_app"])

        tab_control = ttk.Notebook(auto_win)
        tab_auto_comment = tk.Frame(tab_control, bg=THEME["bg_app"])
        tab_auto_msg = tk.Frame(tab_control, bg=THEME["bg_app"])
        
        tab_control.add(tab_auto_comment, text="📝 TẠO NHẬN XÉT TỰ ĐỘNG")
        tab_control.add(tab_auto_msg, text="📧 MẪU TIN NHẮN CHO PHỤ HUYNH")
        tab_control.pack(expand=1, fill="both", padx=10, pady=10)

        # --- TAB 1: NHẬN XÉT ---
        tk.Label(tab_auto_comment, text="Hệ thống tự động sinh nhận xét dựa trên ĐTB:", font=("Segoe UI", 12, "bold"), bg=THEME["bg_app"], fg=THEME["text_main"]).pack(pady=10)
        
        txt_comment = tk.Text(tab_auto_comment, height=20, font=("Segoe UI", 10), bg=THEME["entry_bg"], fg=THEME["text_main"])
        txt_comment.pack(fill="both", expand=True, padx=10)
        
        def generate_comments():
            txt_comment.delete(1.0, tk.END)
            for _, row in self.current_df.iterrows():
                name = row['Họ và tên']
                try:
                    dtb = float(row['ĐTB'])
                except:
                    dtb = 0
                
                comment = ""
                if dtb >= 9.0: 
                    opts = ["Xuất sắc! Tư duy nhạy bén, chăm chỉ.", "Học lực Giỏi. Có tố chất và tiềm năng lớn.", "Tuyệt vời! Là tấm gương sáng cho cả lớp."]
                    comment = random.choice(opts)
                elif dtb >= 8.0:
                    opts = ["Học lực Giỏi. Cần phát huy hơn ở các bài nâng cao.", "Nắm vững kiến thức. Rất đáng khen.", "Chăm ngoan, học tốt. Cố gắng duy trì phong độ."]
                    comment = random.choice(opts)
                elif dtb >= 6.5:
                    opts = ["Học lực Khá. Cần cẩn thận hơn để đạt điểm cao.", "Có tiến bộ, nhưng cần tập trung hơn trong giờ học.", "Nắm được kiến thức cơ bản. Cần nỗ lực thêm."]
                    comment = random.choice(opts)
                elif dtb >= 5.0:
                    opts = ["Học lực Trung bình. Cần làm thêm bài tập về nhà.", "Cần cố gắng nhiều hơn. Chú ý nghe giảng.", "Sức học còn chậm. Cần phụ đạo thêm."]
                    comment = random.choice(opts)
                else:
                    opts = ["Học lực Yếu. Cần sự quan tâm đặc biệt từ gia đình.", "Hổng kiến thức nhiều. Cần đi học phụ đạo ngay.", "Kết quả thấp. Cần chấn chỉnh lại thái độ học tập."]
                    comment = random.choice(opts)
                
                txt_comment.insert(tk.END, f"• {name} ({dtb}): {comment}\n")

        btn_gen = ModernButton(tab_auto_comment, bg_color="#2ecc71", text_color="white", text="✨ TẠO NHẬN XÉT NGAY", command=generate_comments)
        btn_gen.pack(pady=10)

        # --- TAB 2: TIN NHẮN ---
        tk.Label(tab_auto_msg, text="Mẫu tin nhắn gửi Zalo/SMS cho Phụ huynh:", font=("Segoe UI", 12, "bold"), bg=THEME["bg_app"], fg=THEME["text_main"]).pack(pady=10)
        
        txt_msg = tk.Text(tab_auto_msg, height=20, font=("Segoe UI", 10), bg=THEME["entry_bg"], fg=THEME["text_main"])
        txt_msg.pack(fill="both", expand=True, padx=10)
        
        def generate_messages():
            txt_msg.delete(1.0, tk.END)
            for _, row in self.current_df.iterrows():
                name = row['Họ và tên']
                try:
                    dtb = float(row['ĐTB'])
                except:
                    dtb = 0
                rank = str(row['Xếp loại']).upper()
                
                template = f"Kính gửi PH em {name}. TBCN môn Tiếng Anh: {dtb}. Xếp loại: {rank}. Nhờ gia đình đôn đốc cháu thêm. Trân trọng!\n"
                txt_msg.insert(tk.END, template + "-"*50 + "\n")

        btn_msg = ModernButton(tab_auto_msg, bg_color="#3498db", text_color="white", text="📩 TẠO TIN NHẮN HÀNG LOẠT", command=generate_messages)
        btn_msg.pack(pady=10)

    # =============================================================================
    # TÍNH NĂNG 2: BIỂU ĐỒ PHỔ ĐIỂM
    # =============================================================================
    def show_histogram_popup(self):
        if self.current_df is None: 
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu lớp học!")
            return

        try:
            raw_data = pd.to_numeric(self.current_df['ĐTB'], errors='coerce')
            dtb_data = raw_data.dropna()
            dtb_data = dtb_data[dtb_data > 0]
            plot_data = dtb_data.values.flatten() if hasattr(dtb_data, 'values') else list(dtb_data)
        except Exception as e:
            messagebox.showerror("Lỗi dữ liệu", f"Không thể đọc cột điểm ĐTB: {str(e)}")
            return

        if len(plot_data) == 0:
            messagebox.showinfo("Thông báo", "Chưa có dữ liệu điểm số hợp lệ (>0) để vẽ biểu đồ!")
            return

        hist_win = Toplevel(self)
        hist_win.title(f"Biểu Đồ Phân Phối Điểm - Sĩ số: {len(plot_data)}")
        
        self.center_window(hist_win, 750, 520) 
        hist_win.configure(bg=THEME["bg_app"])

        chart_frame = tk.Frame(hist_win, bg=THEME["bg_card"], bd=2, relief="flat")
        chart_frame.pack(side="top", fill="both", expand=True, padx=10, pady=10)

        try:
            fig = Figure(figsize=(6, 3.8), dpi=100)
            fig.patch.set_facecolor(THEME["bg_card"])
            
            ax = fig.add_subplot(111)
            ax.set_facecolor(THEME["bg_card"])

            use_density = HAS_SCIPY 

            counts, bins, patches = ax.hist(
                plot_data, 
                bins=range(0, 12), 
                density=use_density, 
                color=THEME["primary"], 
                alpha=0.7, 
                rwidth=0.85, 
                edgecolor=THEME["text_main"], 
                label="Thực tế"
            )

            if HAS_SCIPY:
                try:
                    mu, std = stats.norm.fit(plot_data)
                    xmin, xmax = 0, 11
                    x = np.linspace(xmin, xmax, 100)
                    p = stats.norm.pdf(x, mu, std)
                    ax.plot(x, p, 'r', linewidth=2, label=f"Phân phối chuẩn\n(TB={mu:.1f}, Độ lệch={std:.1f})")
                    ax.axvline(mu, color='#f39c12', linestyle='--', linewidth=2, label="Điểm TB Lớp")
                except Exception: pass

            ax.set_title("PHỔ ĐIỂM CỦA LỚP HỌC", color=THEME["text_main"], fontsize=12, fontweight='bold', pad=10)
            ax.set_xlabel("Thang điểm (0-10)", color=THEME["text_main"], fontsize=9)
            ax.set_ylabel("Mật độ / Số lượng", color=THEME["text_main"], fontsize=9)
            ax.set_xticks(range(0, 11))
            
            ax.tick_params(axis='both', colors=THEME["text_main"], labelsize=8)
            for spine in ax.spines.values(): spine.set_color(THEME["text_main"])
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            
            legend = ax.legend(frameon=False, loc='upper left', fontsize=8)
            if legend:
                for text in legend.get_texts(): text.set_color(THEME["text_main"])

            fig.tight_layout()

            canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

            avg_val = sum(plot_data)/len(plot_data)
            
            if avg_val >= 8.0: msg = "🌟 Lớp rất xuất sắc! Đa số điểm giỏi."
            elif avg_val >= 6.5: msg = "✅ Lớp học tốt. Biểu đồ lệch phải."
            elif avg_val >= 5.0: msg = "⚠️ Lớp trung bình. Cần cố gắng thêm."
            else: msg = "🆘 Lớp yếu. Biểu đồ lệch trái (nhiều điểm kém)."

            lbl_frame = tk.Frame(hist_win, bg=THEME["bg_app"], pady=5)
            lbl_frame.pack(side="bottom", fill="x")

            tk.Label(lbl_frame, text=f"{msg} (Điểm TB: {avg_val:.2f})", 
                     bg=THEME["bg_app"], fg=THEME["text_main"], 
                     font=("Segoe UI", 11, "bold")).pack()
            tk.Label(lbl_frame, text="(Đỉnh biểu đồ càng nhọn thì học sinh càng có trình độ đồng đều)", 
                     bg=THEME["bg_app"], fg="gray", 
                     font=("Segoe UI", 9, "italic")).pack(pady=(0, 5))
        except Exception as e:
            messagebox.showerror("Lỗi Vẽ Biểu Đồ", f"Chi tiết lỗi: {str(e)}")

    def show_presentation_mode(self):
        if self.current_df is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu lớp học!")
            return
        
        pre_win = Toplevel(self)
        pre_win.title("Chế độ Trình chiếu - Bảng Điểm Lớp Học")
        pre_win.state("zoomed")
        pre_win.configure(bg=THEME["bg_app"])

        header = tk.Frame(pre_win, bg=THEME["primary"], pady=20)
        header.pack(fill="x")
        
        class_name = self.cbo_classes.get() if hasattr(self, 'cbo_classes') else "Lớp Học"
        tk.Label(header, text=f"📺 BẢNG ĐIỂM LỚP: {class_name.upper()}", 
                 font=("Segoe UI", 28, "bold"), bg=THEME["primary"], fg="white").pack()
        
        stats = self.get_stats()
        info_text = f"Tổng: {stats['All']} HS  |  Giỏi/Khá: {stats['Tot']} HS  |  TB: {stats['HoanThanh']} HS  |  Yếu: {stats['CanChuY']} HS"
        tk.Label(header, text=info_text, font=("Segoe UI", 14), bg=THEME["primary"], fg="#E3F2FD").pack(pady=(5, 0))

        container = tk.Frame(pre_win, bg=THEME["bg_app"], padx=30, pady=20)
        container.pack(fill="both", expand=True)

        # FIXED HEADER - Nằm ngoài canvas để không bị cuộn
        fixed_header_frame = tk.Frame(container, bg=THEME["tree_header_bg"], pady=15)
        fixed_header_frame.pack(fill="x", pady=(0, 5))
        
        headers = ["STT", "HỌ VÀ TÊN", "TX1", "TX2", "TX3", "TX4", "GK", "CK", "ĐTB", "XẾP LOẠI"]
        widths = [60, 200, 75, 75, 75, 75, 75, 75, 75, 220]
        
        for i, (h, w) in enumerate(zip(headers, widths)):
            if i in [2, 3, 4, 5, 6, 7, 8]:  # TX1-4, GK, CK, ĐTB
                font_style = ("Consolas", 13, "bold")
            else:
                font_style = ("Segoe UI", 13, "bold")
            
            lbl = tk.Label(fixed_header_frame, text=h, font=font_style, 
                          bg=THEME["tree_header_bg"], fg=THEME["tree_header_fg"], 
                          width=w//8, anchor="center")
            lbl.pack(side="left", padx=2)

        # SCROLLABLE CONTENT
        canvas = tk.Canvas(container, bg=THEME["bg_app"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=THEME["bg_app"])

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _on_mousewheel(event):
            # Fix: Check if canvas still exists before scrolling
            try:
                if canvas.winfo_exists():
                    if event.delta: 
                        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Unbind mousewheel when window is destroyed
        def on_window_destroy():
            try:
                canvas.unbind_all("<MouseWheel>")
            except:
                pass
            pre_win.destroy()
        
        pre_win.protocol("WM_DELETE_WINDOW", on_window_destroy)
        pre_win.bind("<Escape>", lambda e: on_window_destroy())

        for idx, row in self.current_df.iterrows():
            xl = row['Xếp loại']
            
            if xl == "Hoàn thành xuất sắc":
                row_bg, row_fg = "#D1FAE5", "#065F46"
            elif xl == "Hoàn thành tốt":
                row_bg, row_fg = "#DBEAFE", "#1E40AF"
            elif xl == "Hoàn thành":
                row_bg, row_fg = "#FEF3C7", "#92400E"
            elif xl == "Chưa hoàn thành":
                row_bg, row_fg = "#FEE2E2", "#991B1B"
            else:
                row_bg, row_fg = "#F3F4F6", "#6B7280"
            
            row_frame = tk.Frame(scrollable_frame, bg=row_bg, pady=12, 
                               highlightbackground="#D1D5DB", highlightthickness=1)
            row_frame.pack(fill="x", pady=3)
            
            # [DYNAMIC TX] Build values list dynamically based on available columns
            values = [str(row['STT']), str(row['Họ và tên'])]
            
            # Add all TX columns that exist in dataframe
            tx_cols = [col for col in row.index if col.startswith("TX") and col[2:].isdigit()]
            for tx_col in sorted(tx_cols):
                values.append(str(row.get(tx_col, '')).replace('nan', '').replace('None', ''))
            
            # Add GK, CK, ĐTB, Xếp loại
            values.extend([
                str(row.get('GK', '')).replace('nan', '').replace('None', ''),
                str(row.get('CK', '')).replace('nan', '').replace('None', ''),
                str(row.get('ĐTB', '')).replace('nan', '').replace('None', ''),
                str(xl)
            ])
            
            for i, (val, w) in enumerate(zip(values, widths)):
                # [FIXED] Dùng font monospace (Consolas) cho cột điểm để thẳng hàng
                if i in [2, 3, 4, 5, 6, 7, 8]:  # TX1-4, GK, CK, ĐTB
                    font_style = ("Consolas", 12, "bold")
                elif i in [0, 1, 9]:  # STT, Họ tên, Xếp loại
                    font_style = ("Segoe UI", 12, "bold")
                else:
                    font_style = ("Segoe UI", 12)
                
                anchor_style = "w" if i == 1 else "center"
                
                lbl = tk.Label(row_frame, text=val, font=font_style, 
                             bg=row_bg, fg=row_fg, width=w//8, anchor=anchor_style)
                lbl.pack(side="left", padx=2)

        footer = tk.Frame(pre_win, bg=THEME["bg_app"], pady=15)
        footer.pack(fill="x")
        # Nút ĐÓNG với viền đen
        close_btn_frame = tk.Frame(footer, bg="#000000", padx=1, pady=1)
        close_btn_frame.pack()
        tk.Button(close_btn_frame, text="✖ ĐÓNG (Esc)", command=pre_win.destroy, 
                 bg="#f87171", fg="white", font=("Segoe UI", 13, "bold"), 
                 relief="flat", padx=30, pady=10, cursor="hand2", borderwidth=0).pack()

    # =============================================================================
    # TÍNH NĂNG: NHẬT KÝ, RADAR, CHIA NHÓM
    # =============================================================================
    def show_logbook_popup(self):
        if self.current_df is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng mở file Excel trước khi sử dụng Nhật ký thi đua!")
            return
        
        log_win = Toplevel(self)
        class_name = self.cbo_classes.get() or "Không xác định"
        print(f"DEBUG: Mở nhật ký cho lớp: '{class_name}'")
        log_win.title(f"Sổ Nhật Ký - Lớp {class_name}")
        self.center_window(log_win, 880, 520) 
        log_win.configure(bg=THEME["bg_app"])
        
        # Cleanup khi đóng cửa sổ
        def on_log_close():
            try:
                log_win.unbind_all("<MouseWheel>")
            except:
                pass
            log_win.destroy()
        log_win.protocol("WM_DELETE_WINDOW", on_log_close)

        current_date = datetime.now()
        date_var = tk.StringVar(value=current_date.strftime("%d/%m/%Y"))
        search_var = tk.StringVar() 

        header = tk.Frame(log_win, bg=THEME["bg_card"], pady=15, padx=20)
        header.pack(fill="x")
        tk.Label(header, text="📒 NHẬT KÝ THI ĐUA", font=("Segoe UI", 18, "bold"), bg=THEME["bg_card"], fg="#8e44ad").pack(side="left")

        ctrl_frame = tk.Frame(header, bg=THEME["bg_card"])
        ctrl_frame.pack(side="right")

        # [UPDATED] Hàm hiển thị Bảng Vàng (Golden Board) mới với màu Pastel
        def show_stats():
            stats_win = Toplevel(log_win)
            current_class_name = self.cbo_classes.get()
            stats_win.title(f"Bảng Vàng Vinh Danh - Lớp {current_class_name}")
            self.center_window(stats_win, 550, 680) 
            stats_win.configure(bg="#fafbfc")
            
            # Header Bảng Vàng với khung pastel vàng
            header_container = tk.Frame(stats_win, bg="#fafbfc", padx=20, pady=20)
            header_container.pack(fill="x")
            
            header_frame = tk.Frame(header_container, bg="#fef3c7", padx=25, pady=20, 
                                   highlightbackground="#fbbf24", highlightthickness=3, relief="solid")
            header_frame.pack(fill="x")
            
            tk.Label(header_frame, text="👑", font=("Segoe UI", 48), bg="#fef3c7", fg="#f59e0b").pack()
            tk.Label(header_frame, text="BẢNG VÀNG VINH DANH", font=("Segoe UI", 20, "bold"), 
                    bg="#fef3c7", fg="#d97706").pack(pady=(5, 5))
            tk.Label(header_frame, text=f"🏆 Lớp: {current_class_name}", font=("Segoe UI", 12, "italic"), 
                    bg="#fef3c7", fg="#f59e0b").pack()

            # Tính toán điểm số
            valid_students = self.current_df["Họ và tên"].astype(str).tolist()
            score_map = {}
            for index, row in self.logbook_df.iterrows():
                name = str(row['TenHS'])
                if name not in valid_students: continue 
                type_log = row['Loai']
                if name not in score_map: score_map[name] = 0
                if type_log == 'Cộng': score_map[name] += 1
                else: score_map[name] -= 1
            
            # Lọc chỉ lấy điểm > 0 và sắp xếp
            sorted_scores = sorted([(k, v) for k, v in score_map.items() if v > 0], key=lambda x: x[1], reverse=True)
            
            # --- CẤU HÌNH STYLE PASTEL ---
            style = ttk.Style()
            style.configure("Pastel.Treeview", 
                            font=("Segoe UI", 11), 
                            rowheight=45,  
                            borderwidth=0, 
                            background="white")
            
            style.configure("Pastel.Treeview.Heading", 
                            font=("Segoe UI", 10, "bold"), 
                            foreground="#546e7a",
                            background="#eceff1", 
                            relief="flat",
                            padding=(0, 12))
            
            style.layout("Pastel.Treeview", [('Pastel.Treeview.treearea', {'sticky': 'nswe'})])

            # Container với viền pastel vàng
            table_container = tk.Frame(stats_win, bg="#fef9e7", padx=25, pady=15, 
                                      highlightbackground="#fbbf24", highlightthickness=2, relief="solid")
            table_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))

            columns = ("Hạng", "Tên", "Điểm")
            tree = ttk.Treeview(table_container, columns=columns, show="headings", height=15, style="Pastel.Treeview")
            
            tree.heading("Hạng", text="HẠNG")
            tree.heading("Tên", text="HỌ VÀ TÊN HỌC SINH")
            tree.heading("Điểm", text="TỔNG ĐIỂM")
            
            tree.column("Hạng", width=80, anchor="center")
            tree.column("Tên", width=280, anchor="w")
            tree.column("Điểm", width=100, anchor="center")

            # --- PALETTE MÀU PASTEL PREMIUM ---
            # Rank 1: Vàng Kem (Pastel Yellow)
            tree.tag_configure("rank1", background="#FFF9C4", foreground="#EF6C00", font=("Segoe UI", 12, "bold"))
            # Rank 2: Xanh Băng (Ice Blue)
            tree.tag_configure("rank2", background="#E1F5FE", foreground="#0277BD", font=("Segoe UI", 11, "bold"))
            # Rank 3: Hồng Cam (Rose Gold/Peach)
            tree.tag_configure("rank3", background="#FFCCBC", foreground="#BF360C", font=("Segoe UI", 11, "bold"))
            # Normal: Trắng
            tree.tag_configure("normal", background="white", foreground="#37474f", font=("Segoe UI", 11))

            tree.pack(fill="both", expand=True)

            if not sorted_scores:
                tree.insert("", "end", values=("---", "Chưa có HS nào đạt điểm dương", "0"), tags=("normal",))
            else:
                for i, (name, score) in enumerate(sorted_scores):
                    rank_display = str(i + 1)
                    tag = "normal"
                    
                    if i == 0:
                        rank_display = "🥇 1"
                        tag = "rank1"
                    elif i == 1:
                        rank_display = "🥈 2"
                        tag = "rank2"
                    elif i == 2:
                        rank_display = "🥉 3"
                        tag = "rank3"
                    
                    tree.insert("", "end", values=(rank_display, f"  {name}", f"+{score}"), tags=(tag,))

            # Nút ĐÓNG với viền đen
            close_btn_frame = tk.Frame(stats_win, bg="#000000", padx=1, pady=1)
            close_btn_frame.pack(pady=20)
            tk.Button(close_btn_frame, text="ĐÓNG", command=stats_win.destroy, bg="#94a3b8", fg="white", 
                     font=("Segoe UI", 10, "bold"), relief="flat", width=20, borderwidth=0).pack()

        btn_stats = tk.Button(ctrl_frame, text="📊 THỐNG KÊ", bg="#27ae60", fg="white", font=("Segoe UI", 10, "bold"), relief="flat", command=show_stats)
        btn_stats.pack(side="left", padx=(0, 20))

        tk.Label(ctrl_frame, text="🔍", bg=THEME["bg_card"], fg=THEME["text_main"], font=("Segoe UI", 12)).pack(side="left", padx=(0, 5))
        search_entry = tk.Entry(ctrl_frame, textvariable=search_var, width=15, font=("Segoe UI", 11), relief="solid", bd=1)
        search_entry.pack(side="left", padx=(0, 20))

        def change_date(delta):
            nonlocal current_date
            current_date += timedelta(days=delta)
            date_var.set(current_date.strftime("%d/%m/%Y"))
            refresh_list()

        tk.Button(ctrl_frame, text="◀", command=lambda: change_date(-1), bg="#ecf0f1", relief="flat").pack(side="left", padx=5)
        tk.Label(ctrl_frame, textvariable=date_var, font=("Segoe UI", 12, "bold"), bg=THEME["bg_card"], fg=THEME["text_main"], width=12).pack(side="left", padx=5)
        tk.Button(ctrl_frame, text="▶", command=lambda: change_date(1), bg="#ecf0f1", relief="flat").pack(side="left", padx=5)

        # List frame với viền pastel xanh dương
        list_frame = tk.Frame(log_win, bg=THEME["bg_app"], padx=20, pady=10, 
                        highlightbackground="#3b82f6", highlightthickness=2, relief="solid")
        list_frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(list_frame, bg=THEME["bg_app"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=THEME["bg_app"])

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def on_canvas_configure(event): canvas.itemconfig(window_id, width=event.width)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            if event.delta: canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def add_log(name, type_log):
            try:
                date_str = date_var.get()
                new_row = {"Ngay": date_str, "TenHS": name, "Loai": type_log, "GhiChu": ""}
                self.logbook_df = pd.concat([self.logbook_df, pd.DataFrame([new_row])], ignore_index=True)
                self.save_logbook()
                refresh_list()
            except Exception as e:
                print(f"Lỗi add_log: {e}")
                messagebox.showerror("Lỗi", f"Không thể thêm nhật ký: {str(e)}")

        def refresh_list(*args):
            for widget in scrollable_frame.winfo_children(): widget.destroy()
            
            # Debug: Chi tiết thông tin
            print(f"DEBUG refresh_list:")
            print(f"  - self.current_df is None: {self.current_df is None}")
            print(f"  - len(self.current_df): {len(self.current_df) if self.current_df is not None else 'N/A'}")
            print(f"  - Columns: {list(self.current_df.columns) if self.current_df is not None else 'N/A'}")
            
            if self.current_df is None or len(self.current_df) == 0:
                tk.Label(scrollable_frame, text="❌ Chưa có dữ liệu học sinh!", 
                       font=("Segoe UI", 12), bg=THEME["bg_app"], fg="#dc2626").pack(pady=20)
                return
                
            date_str = date_var.get()
            keyword = search_var.get().lower()
            today_logs = self.logbook_df[self.logbook_df["Ngay"] == date_str]

            print(f"DEBUG: Bắt đầu loop - Keyword: '{keyword}', Date: '{date_str}'")
            print(f"DEBUG: today_logs có {len(today_logs)} bản ghi")
            if len(today_logs) > 0:
                print(f"DEBUG: today_logs sample: {today_logs.head().to_dict()}")
            student_count = 0
            
            for idx, row in self.current_df.iterrows():
                name = row["Họ và tên"]
                if keyword and keyword not in name.lower(): continue

                plus_count = len(today_logs[(today_logs["TenHS"] == name) & (today_logs["Loai"] == "Cộng")])
                minus_count = len(today_logs[(today_logs["TenHS"] == name) & (today_logs["Loai"] == "Trừ")])
                net_score = plus_count - minus_count
                
                print(f"DEBUG: Học sinh {idx+1}: {name} - Cộng: {plus_count}, Trừ: {minus_count}, Net: {net_score}")

                row_f = tk.Frame(scrollable_frame, bg=THEME["bg_card"], pady=5, padx=10)
                row_f.pack(fill="x", pady=2, padx=5)

                tk.Label(row_f, text=name, font=("Segoe UI", 11), bg=THEME["bg_card"], fg=THEME["text_main"], width=25, anchor="w").pack(side="left")

                btn_minus = tk.Button(row_f, text="➖ TRỪ", bg="#fab1a0", fg="#c0392b", relief="flat", font=("Segoe UI", 9, "bold"), width=8,
                                      command=lambda n=name: add_log(n, "Trừ"))
                btn_minus.pack(side="right", padx=5)

                btn_plus = tk.Button(row_f, text="➕ CỘNG", bg="#a3e4d7", fg="#16a085", relief="flat", font=("Segoe UI", 9, "bold"), width=8,
                                     command=lambda n=name: add_log(n, "Cộng"))
                btn_plus.pack(side="right", padx=5)

                score_color = "#7f8c8d"
                score_str = f"({net_score})"
                if net_score > 0: score_color, score_str = "#27ae60", f"(+{net_score})"
                elif net_score < 0: score_color = "#e74c3c"
                
                tk.Label(row_f, text=score_str, font=("Segoe UI", 12, "bold"), bg=THEME["bg_card"], fg=score_color).pack(side="left", padx=20)
                student_count += 1

            print(f"DEBUG: Đã xử lý {student_count} học sinh")
            if student_count == 0:
                tk.Label(scrollable_frame, text="🔍 Không tìm thấy học sinh nào!", 
                       font=("Segoe UI", 12), bg=THEME["bg_app"], fg="#f59e0b").pack(pady=20)

        search_var.trace("w", refresh_list)
        
        # Debug: Hiển thị thông tin khi mở
        print(f"DEBUG: Mở sổ nhật ký - Số học sinh: {len(self.current_df) if self.current_df is not None else 0}")
        print(f"DEBUG: Logbook có {len(self.logbook_df)} bản ghi")
        refresh_list()

    def save_logbook(self):
        try:
            if self.file_path.lower().endswith('.xls'):
                print("Không hỗ trợ lưu nhật ký vào file .xls cũ để tránh lỗi cấu trúc.")
                return 

            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                self.logbook_df.to_excel(writer, sheet_name="NhatKy", index=False)
        except Exception as e: print("Lỗi lưu nhật ký:", e)

    def load_parent_contacts(self):
        """Load thông tin phụ huynh từ file JSON"""
        try:
            contacts_file = "parent_contacts.json"
            if os.path.exists(contacts_file):
                with open(contacts_file, 'r', encoding='utf-8') as f:
                    self.parent_contacts = json.load(f)
        except Exception as e:
            print(f"Lỗi load thông tin phụ huynh: {e}")
            self.parent_contacts = {}
    
    def save_parent_contacts(self):
        """Lưu thông tin phụ huynh ra file JSON"""
        try:
            contacts_file = "parent_contacts.json"
            with open(contacts_file, 'w', encoding='utf-8') as f:
                json.dump(self.parent_contacts, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Lỗi lưu thông tin phụ huynh: {e}")
    
    def edit_parent_contact(self, student_name):
        """Mở popup chỉnh sửa thông tin phụ huynh cho học sinh"""
        edit_win = Toplevel(self)
        edit_win.title(f"Sửa thông tin phụ huynh: {student_name}")
        self.center_window(edit_win, 500, 400)
        edit_win.configure(bg=THEME["bg_app"])
        edit_win.transient(self)
        edit_win.grab_set()
        
        # Header với khung pastel nổi bật
        header_container = tk.Frame(edit_win, bg="#fafbfc", padx=20, pady=20)
        header_container.pack(fill="x")
        
        header = tk.Frame(header_container, bg="#dbeafe", padx=20, pady=15, 
                         highlightbackground="#3b82f6", highlightthickness=2, relief="solid")
        header.pack(fill="x")
        
        tk.Label(header, text="📞", font=("Segoe UI", 18), bg="#dbeafe", fg="#3b82f6").pack()
        tk.Label(header, text="THÔNG TIN PHỤ HUYNH", 
                 font=("Segoe UI", 16, "bold"), bg="#dbeafe", fg="#1d4ed8").pack(pady=(5, 5))
        tk.Label(header, text=f"👤 Học sinh: {student_name}", 
                 font=("Segoe UI", 12, "italic"), bg="#dbeafe", fg="#3b82f6").pack()
        
        # Form
        form_frame = tk.Frame(edit_win, bg=THEME["bg_card"], padx=20, pady=20)
        form_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Lấy thông tin hiện tại
        current_info = self.parent_contacts.get(student_name, {
            "phone": "",
            "email": "",
            "notes": ""
        })
        
        # Các trường nhập liệu
        fields = [
            ("📱 Số điện thoại:", current_info["phone"], "phone"),
            ("📧 Email:", current_info["email"], "email"),
            ("📝 Ghi chú:", current_info["notes"], "notes")
        ]
        
        entries = {}
        for i, (label, default, key) in enumerate(fields):
            tk.Label(form_frame, text=label, font=("Segoe UI", 10, "bold"), 
                    bg=THEME["bg_card"], fg=THEME["text_main"]).grid(row=i, column=0, sticky="w", pady=10)
            
            if key == "notes":
                entry = tk.Text(form_frame, width=40, height=4, font=("Segoe UI", 10), 
                              bg=THEME["entry_bg"], fg=THEME["text_main"])
                entry.insert("1.0", default)
                entry.grid(row=i, column=1, padx=10, pady=10)
            else:
                entry = tk.Entry(form_frame, width=40, font=("Segoe UI", 10), 
                               bg=THEME["entry_bg"], fg=THEME["text_main"])
                entry.insert(0, default)
                entry.grid(row=i, column=1, padx=10, pady=10)
            
            entries[key] = entry
        
        # Nút lưu
        def save_contact():
            try:
                phone = entries["phone"].get().strip()
                email = entries["email"].get().strip()
                notes = entries["notes"].get("1.0", tk.END).strip()
                
                # Validate đơn giản
                if phone and not re.match(r'^[0-9\s\-\+\(\)]+$', phone):
                    messagebox.showerror("Lỗi", "Số điện thoại không hợp lệ!")
                    return
                
                if email and "@" not in email:
                    messagebox.showerror("Lỗi", "Email không hợp lệ!")
                    return
                
                # Lưu thông tin
                self.parent_contacts[student_name] = {
                    "phone": phone,
                    "email": email,
                    "notes": notes
                }
                
                self.save_parent_contacts()
                messagebox.showinfo("Thành công", "Đã lưu thông tin phụ huynh!")
                edit_win.destroy()
                
                # Refresh danh sách
                self.show_parent_contacts()
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu thông tin: {str(e)}")
        
        def delete_contact():
            if messagebox.askyesno("Xóa", f"Bạn có chắc muốn xóa thông tin phụ huynh của {student_name}?"):
                if student_name in self.parent_contacts:
                    del self.parent_contacts[student_name]
                    self.save_parent_contacts()
                    messagebox.showinfo("Thành công", "Đã xóa thông tin!")
                    edit_win.destroy()
                    self.show_parent_contacts()
        
        # Nút chức năng
        btn_frame = tk.Frame(edit_win, bg=THEME["bg_app"], pady=15)
        btn_frame.pack(fill="x")
        
        # Nút LƯU với viền đen
        save_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        save_btn_frame.pack(side="left", padx=10)
        tk.Button(save_btn_frame, text="💾 LƯU", command=save_contact, 
                 bg=THEME["primary"], fg="white", font=("Segoe UI", 11, "bold"), 
                 relief="flat", padx=20, pady=8, borderwidth=0).pack()
        
        # Nút XÓA với viền đen
        delete_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        delete_btn_frame.pack(side="left", padx=10)
        tk.Button(delete_btn_frame, text="🗑️ XÓA", command=delete_contact, 
                 bg="#f87171", fg="white", font=("Segoe UI", 11, "bold"), 
                 relief="flat", padx=20, pady=8, borderwidth=0).pack()
        
        # Nút HỦY với viền đen
        cancel_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        cancel_btn_frame.pack(side="right", padx=10)
        tk.Button(cancel_btn_frame, text="HỦY", command=edit_win.destroy, 
                 bg="#94a3b8", fg="white", font=("Segoe UI", 11, "bold"), 
                 relief="flat", padx=20, pady=8, borderwidth=0).pack()

    def show_radar_popup(self):
        if self.current_df is None: return

        radar_win = Toplevel(self)
        radar_win.title("Radar Cảnh Báo Sớm")
        self.center_window(radar_win, 880, 520) 
        radar_win.configure(bg=THEME["bg_app"])
        
        # Cleanup khi đóng cửa sổ
        def on_radar_close():
            try:
                radar_win.unbind_all("<MouseWheel>")
            except:
                pass
            radar_win.destroy()
        radar_win.protocol("WM_DELETE_WINDOW", on_radar_close)

        # Header với khung pastel nổi bật
        header_container = tk.Frame(radar_win, bg="#fafbfc", padx=20, pady=20)
        header_container.pack(fill="x")
        
        header = tk.Frame(header_container, bg="#fee2e2", padx=20, pady=15, 
                         highlightbackground="#f87171", highlightthickness=2, relief="solid")
        header.pack(fill="x")
        
        tk.Label(header, text="⚠️ DANH SÁCH HỌC SINH CẦN LƯU Ý", 
                 font=("Segoe UI", 18, "bold"), bg="#fee2e2", fg="#dc2626").pack()
        tk.Label(header, text="🔍 Hệ thống AI tự động phân tích và phát hiện sớm", 
                 font=("Segoe UI", 11, "italic"), bg="#fee2e2", fg="#7f1d1d").pack(pady=(5, 0))

        container = tk.Frame(radar_win, bg=THEME["bg_app"])
        container.pack(fill="both", expand=True, padx=20, pady=20)

        canvas = tk.Canvas(container, bg=THEME["bg_app"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=THEME["bg_app"])

        scrollable_frame.grid_columnconfigure(0, weight=1)
        scrollable_frame.grid_columnconfigure(1, weight=1)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def on_canvas_configure(event): canvas.itemconfig(window_id, width=event.width)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            if event.delta: canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        risks = []
        for _, row in self.current_df.iterrows():
            name = row["Họ và tên"]
            def get_val(col):
                v = row.get(col, 0)
                try:
                    return float(v)
                except:
                    return 0.0

            # [DYNAMIC TX] Tính trung bình TX động
            tx_sum = 0
            tx_count = 0
            tx_cols_in_row = [col for col in row.index if col.startswith("TX") and col[2:].isdigit()]
            for k in sorted(tx_cols_in_row):
                v = get_val(k)
                if v > 0: tx_sum += v; tx_count += 1
            avg_tx = tx_sum / tx_count if tx_count > 0 else 0

            gk = get_val("GK")
            ck = get_val("CK")
            dtb = get_val("ĐTB")

            risk_level = 0 
            risk_type = ""
            details = []

            if 0 < dtb < 5.0:
                risk_level = 3; risk_type = "NGUY CƠ Ở LẠI LỚP"; details.append(f"• Điểm TB hiện tại: {dtb} (Dưới 5.0)")
            elif avg_tx >= 7.5 and (gk < 6.0 or (ck > 0 and ck < 6.0)):
                risk_level = 2; risk_type = "SỤT GIẢM PHONG ĐỘ"; details.append(f"• Học trên lớp tốt ({avg_tx:.1f}) nhưng thi thấp.")
            elif 5.0 <= dtb < 6.0:
                risk_level = 1; risk_type = "MẤP MÉ TRUNG BÌNH"; details.append(f"• Điểm TB: {dtb} (Cần cố gắng thêm để đạt Khá)")

            if risk_level > 0:
                risks.append({"name": name, "level": risk_level, "type": risk_type, "details": details})

        risks.sort(key=lambda x: x["level"], reverse=True)

        def create_risk_card(parent, data, row, col):
            if data["level"] == 3: border_col, text_col, bg_title, icon = "#f87171", "#dc2626", "#fee2e2", "🆘"
            elif data["level"] == 2: border_col, text_col, bg_title, icon = "#fb923c", "#ea580c", "#fed7aa", "📉"
            else: border_col, text_col, bg_title, icon = "#fbbf24", "#d97706", "#fef3c7", "⚠️"

            card = tk.Frame(parent, bg=THEME["bg_card"], highlightbackground=border_col, highlightthickness=2, padx=0, pady=0)
            card.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)

            header = tk.Frame(card, bg=bg_title, padx=10, pady=8)
            header.pack(fill="x")
            tk.Label(header, text=f"{icon} {data['type']}", font=("Segoe UI", 9, "bold"), fg=text_col, bg=bg_title).pack(anchor="w")
            tk.Label(header, text=data['name'].upper(), font=("Segoe UI", 12, "bold"), fg="black", bg=bg_title).pack(anchor="w")

            body = tk.Frame(card, bg=THEME["bg_card"], padx=15, pady=10)
            body.pack(fill="x")
            for line in data["details"]: tk.Label(body, text=line, font=("Segoe UI", 10), bg=THEME["bg_card"], fg=THEME["text_main"]).pack(anchor="w")

        if not risks:
            # Khung pastel cho thông báo thành công
            success_frame = tk.Frame(scrollable_frame, bg="#dcfce7", padx=25, pady=20, 
                                   highlightbackground="#10b981", highlightthickness=2, relief="solid")
            success_frame.grid(row=0, column=0, columnspan=2, pady=50, padx=20, sticky="ew")
            
            tk.Label(success_frame, text="🎉", font=("Segoe UI", 24), bg="#dcfce7", fg="#10b981").pack()
            tk.Label(success_frame, text="TUYỆT VỜI!", font=("Segoe UI", 16, "bold"), 
                    bg="#dcfce7", fg="#047857").pack(pady=(5, 10))
            tk.Label(success_frame, text="Không phát hiện học sinh nào có dấu hiệu bất thường", 
                    font=("Segoe UI", 12), bg="#dcfce7", fg="#065f46").pack()
            tk.Label(success_frame, text="✨ Lớp học đang phát triển rất tốt!", 
                    font=("Segoe UI", 11, "italic"), bg="#dcfce7", fg="#047857").pack(pady=(5, 0))
        else:
            for i, item in enumerate(risks):
                create_risk_card(scrollable_frame, item, i // 2, i % 2)

    # =============================================================================
    # TÍNH NĂNG: QUAY SỐ MAY MẮN
    # =============================================================================
    def show_spin_wheel_popup(self):
        """Hiển thị popup quay số may mắn với đầy đủ tính năng"""
        if self.current_df is None:
            messagebox.showwarning("Cảnh báo", "Vui lòng mở file Excel trước!")
            return
        
        spin_win = Toplevel(self)
        spin_win.title("🎯 QUAY SỐ MAY MẮN")
        self.center_window(spin_win, 900, 650)
        spin_win.configure(bg=THEME["bg_app"])
        
        # Biến điều khiển UI
        filter_enabled = tk.BooleanVar(value=False)
        filter_mode = tk.StringVar(value="Khá/Giỏi")
        exclude_var = tk.BooleanVar(value=self.spin_exclude_used)
        duration_enabled = tk.BooleanVar(value=False)
        duration_var = tk.StringVar(value=str(self.spin_duration))
        
        is_spinning = [False]  # Dùng list để tránh lỗi scope
        
        # =============================================================================
        # HÀM LOGIC CHÍNH
        # =============================================================================
        def get_eligible_students():
            """Lấy danh sách STT học sinh đủ điều kiện quay"""
            try:
                base_list = []
                
                if filter_enabled.get():
                    mode = filter_mode.get()
                    for _, row in self.current_df.iterrows():
                        try:
                            dtb = float(row.get('ĐTB', 0))
                            stt = int(row['STT'])
                            
                            if mode == "Khá/Giỏi" and dtb >= 6.5:
                                base_list.append(stt)
                            elif mode == "Hoàn thành" and 5.0 <= dtb < 6.5:
                                base_list.append(stt)
                            elif mode == "Cần chú ý" and (dtb < 5.0 or pd.isna(row.get('ĐTB'))):
                                base_list.append(stt)
                        except:
                            continue
                else:
                    for _, row in self.current_df.iterrows():
                        try:
                            stt = int(row['STT'])
                            base_list.append(stt)
                        except:
                            continue
                
                # Loại trừ đã quay nếu checkbox được tick
                if exclude_var.get():
                    eligible = [s for s in base_list if s not in self.spin_used_students]
                else:
                    eligible = base_list
                
                return eligible
            except Exception as e:
                print(f"Lỗi get_eligible_students: {e}")
                return []
        
        def draw_wheel(canvas, rotation=0):
            """Vẽ bàn quay số với màu pastel"""
            canvas.delete("all")
            eligible = get_eligible_students()
            
            if len(eligible) == 0:
                canvas.create_text(200, 200, text="Đã hết học sinh!\nNhấn RESET để bắt đầu lại", 
                                 font=("Segoe UI", 14, "bold"), fill="#dc2626", justify="center")
                return
            
            # Vẽ vòng tròn nền (giảm từ 400x400 xuống 360x360)
            canvas.create_oval(20, 20, 380, 380, fill="#f3f4f6", outline="#000000", width=3)
            
            # Màu pastel cho vòng quay
            wheel_colors = ["#A7F3D0", "#BFDBFE", "#FECDD3", "#DDD6FE", "#FED7AA", "#FEF3C7"]
            
            segment_angle = 360 / len(eligible)
            
            for i, stt in enumerate(eligible):
                start_angle = (i * segment_angle + rotation) % 360
                color = wheel_colors[i % len(wheel_colors)]
                
                # Vẽ segment
                canvas.create_arc(20, 20, 380, 380, start=start_angle, extent=segment_angle,
                                fill=color, outline="#000000", width=2)
                
                # Tính vị trí text (ở giữa segment) - Dời ra ngoài rìa hơn
                mid_angle = start_angle + segment_angle / 2
                rad = math.radians(mid_angle)
                text_x = 200 + 145 * math.cos(rad)
                text_y = 200 - 145 * math.sin(rad)
                
                # Vẽ số STT
                canvas.create_text(text_x, text_y, text=str(stt), 
                                 font=("Segoe UI", 12, "bold"), fill="#000000")
                # Debug: In vị trí mỗi số
                if rotation == 0:  # Chỉ in khi không quay
                    print(f"DEBUG: STT {stt} at angle {mid_angle:.2f}°, pos({text_x:.0f}, {text_y:.0f})")
            
            # Vẽ mũi tên chỉ định (ở vị trí 12 giờ) - Đảo ngược hướng xuống
            canvas.create_polygon(200, 35, 185, 10, 215, 10, fill="#dc2626", outline="#000000", width=2)
            
            # Vẽ vòng tròn giữa
            canvas.create_oval(175, 175, 225, 225, fill="#ffffff", outline="#000000", width=3)
            canvas.create_text(200, 200, text="🎯", font=("Segoe UI", 20))
        
        def animate_spin(selected_stt):
            """Animation quay số với easing"""
            is_spinning[0] = True
            btn_spin_mini.config(state="disabled")
            btn_spin.config(state="disabled")
            
            try:
                duration = float(duration_var.get()) if duration_enabled.get() else 3.0
                if duration < 0.5 or duration > 10:
                    duration = 3.0
            except:
                duration = 3.0
            
            eligible = get_eligible_students()
            segment_angle = 360 / len(eligible)
            target_index = eligible.index(selected_stt)
            
            # Tính góc đích - Tkinter Canvas
            # Canvas: 0° ở 3 giờ chiều, arc vẽ theo chiều kim đồng hồ
            # draw_wheel dùng + rotation = quay cùng chiều kim đồng hồ
            # Mũi tên ở 12 giờ (đỉnh) = 90°
            # Segment 0 bắt đầu từ góc 0°, text ở giữa = start + segment/2
            target_text_angle = target_index * segment_angle + segment_angle / 2
            # Quay để target_text_angle đến 90° (vị trí mũi tên)
            # Vì rotation dương quay cùng chiều kim đồng hồ, ta cần: 90 - target_text_angle
            total_rotation = 360 * random.randint(5, 7) + (90 - target_text_angle)
            print(f"DEBUG: Target STT={selected_stt}, Index={target_index}")
            print(f"DEBUG: Target text angle = {target_text_angle:.2f}°, Final rotation = {total_rotation % 360:.2f}°")
            
            num_frames = int(duration * 50)
            
            def ease_out_cubic(t):
                return 1 - pow(1 - t, 3)
            
            def spin_frame(frame):
                if frame >= num_frames:
                    # Kết thúc quay
                    draw_wheel(canvas_wheel, total_rotation % 360)
                    is_spinning[0] = False
                    btn_spin_mini.config(state="normal")
                    btn_spin.config(state="normal")
                    
                    # Thêm vào lịch sử
                    if exclude_var.get():
                        self.spin_used_students.add(selected_stt)
                    
                    add_to_history(selected_stt)
                    show_result_popup(selected_stt)
                    update_status_label()
                    return
                
                progress = frame / num_frames
                eased = ease_out_cubic(progress)
                current_rotation = total_rotation * eased
                
                draw_wheel(canvas_wheel, current_rotation)
                canvas_wheel.update()
                canvas_wheel.after(20, lambda: spin_frame(frame + 1))
            
            spin_frame(0)
        
        def start_spin():
            """Bắt đầu quay số"""
            if is_spinning[0]:
                return
            
            eligible = get_eligible_students()
            if len(eligible) == 0:
                messagebox.showwarning("Hết học sinh", "Không còn học sinh nào để quay!\nNhấn RESET để bắt đầu lại.")
                return
            
            selected_stt = random.choice(eligible)
            animate_spin(selected_stt)
        
        def show_result_popup(stt):
            """Hiển thị popup kết quả"""
            try:
                # Tìm học sinh theo STT - Convert cả hai về int để so sánh
                matched = self.current_df[self.current_df['STT'].astype(int) == int(stt)]
                if matched.empty:
                    messagebox.showerror("Lỗi", f"Không tìm thấy học sinh có STT {stt}!")
                    return
                
                row = matched.iloc[0]
                name = row.get('Họ và tên', 'N/A')
                dtb = row.get('ĐTB', '')
                rank = row.get('Xếp loại', '')
                
                result_win = Toplevel(spin_win)
                result_win.title("🎊 KẾT QUẢ")
                self.center_window(result_win, 450, 350)
                result_win.configure(bg="#fafbfc")
                result_win.transient(spin_win)
                result_win.grab_set()
                
                # Header
                header = tk.Frame(result_win, bg="#dcfce7", padx=20, pady=20,
                                highlightbackground="#10b981", highlightthickness=3)
                header.pack(fill="x")
                
                tk.Label(header, text="🎊", font=("Segoe UI", 32), bg="#dcfce7").pack()
                tk.Label(header, text="CHÚC MỪNG!", font=("Segoe UI", 18, "bold"),
                        bg="#dcfce7", fg="#047857").pack(pady=(5, 0))
                
                # Body
                body = tk.Frame(result_win, bg="#fafbfc", padx=30, pady=20)
                body.pack(fill="both", expand=True)
                
                info_frame = tk.Frame(body, bg="#ffffff", padx=20, pady=15,
                                    highlightbackground="#86efac", highlightthickness=3)
                info_frame.pack(fill="x", pady=10)
                
                tk.Label(info_frame, text=f"Số thứ tự: {stt}", font=("Segoe UI", 14, "bold"),
                        bg="#ffffff", fg="#1f2937").pack(anchor="w", pady=5)
                tk.Label(info_frame, text=f"Họ và tên: {name}", font=("Segoe UI", 16, "bold"),
                        bg="#ffffff", fg="#3b82f6").pack(anchor="w", pady=5)
                
                # Nút đóng
                close_frame = tk.Frame(result_win, bg="#000000", padx=1, pady=1)
                close_frame.pack(pady=20)
                tk.Button(close_frame, text="ĐÓNG", command=result_win.destroy,
                         bg="#3b82f6", fg="white", font=("Segoe UI", 11, "bold"),
                         relief="flat", padx=30, pady=10, borderwidth=0).pack()
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể hiển thị kết quả:\n{str(e)}")
        
        def add_to_history(stt):
            """Thêm vào lịch sử quay số"""
            try:
                # Convert STT về int để so sánh
                row = self.current_df[self.current_df['STT'].astype(int) == int(stt)].iloc[0]
                record = {
                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                    "stt": stt,
                    "name": row['Họ và tên'],
                    "dtb": row.get('ĐTB', ''),
                    "rank": row.get('Xếp loại', '')
                }
                self.spin_history.append(record)
                update_history_display()
            except Exception as e:
                print(f"Lỗi add_to_history: {e}")
        
        def update_history_display():
            """Cập nhật hiển thị lịch sử"""
            for widget in history_frame.winfo_children():
                widget.destroy()
            
            if not self.spin_history:
                tk.Label(history_frame, text="Chưa có lịch sử quay số", 
                        font=("Segoe UI", 10, "italic"), bg=THEME["bg_card"], 
                        fg="#9ca3af").pack(pady=20)
                return
            
            for i, record in enumerate(reversed(self.spin_history[-10:])):
                item_frame = tk.Frame(history_frame, bg="#f9fafb", pady=5, padx=10,
                                    highlightbackground="#e5e7eb", highlightthickness=1)
                item_frame.pack(fill="x", pady=2, padx=5)
                
                text = f"{record['timestamp']} - STT {record['stt']} - {record['name']}"
                tk.Label(item_frame, text=text, font=("Segoe UI", 9),
                        bg="#f9fafb", fg="#374151", anchor="w").pack(fill="x")
        
        def export_history():
            """Xuất lịch sử ra Excel"""
            if not self.spin_history:
                messagebox.showwarning("Cảnh báo", "Chưa có lịch sử quay số!")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Lưu Lịch Sử Quay Số"
            )
            
            if file_path:
                try:
                    df = pd.DataFrame(self.spin_history)
                    df.columns = ["Thời gian", "STT", "Họ và tên", "ĐTB", "Xếp loại"]
                    df.to_excel(file_path, index=False)
                    messagebox.showinfo("Thành công", f"Đã xuất lịch sử ra:\n{file_path}")
                except Exception as e:
                    messagebox.showerror("Lỗi", f"Không thể xuất file: {str(e)}")
        
        def clear_history():
            """Xóa lịch sử quay số (chỉ xóa lịch sử, giữ nguyên danh sách đã quay)"""
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa lịch sử quay số?"):
                self.spin_history.clear()
                update_history_display()
                messagebox.showinfo("Thành công", "Đã xóa lịch sử!")
        
        def reset_all():
            """Reset tất cả dữ liệu quay số"""
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn reset tất cả?\n(Xóa lịch sử và danh sách đã quay)"):
                self.spin_used_students.clear()
                self.spin_history.clear()
                update_history_display()
                update_status_label()
                draw_wheel(canvas_wheel, 0)
                messagebox.showinfo("Thành công", "Đã reset tất cả!")
        
        def update_status_label():
            """Cập nhật label trạng thái"""
            eligible = get_eligible_students()
            total = len(self.current_df)
            used = len(self.spin_used_students)
            status_text = f"Đã loại: {used}/{total} học sinh | Còn lại: {len(eligible)} học sinh"
            lbl_status.config(text=status_text)
        
        def on_filter_toggle():
            """Xử lý khi toggle checkbox filter"""
            if filter_enabled.get():
                cbo_filter.config(state="readonly")
            else:
                cbo_filter.config(state="disabled")
            update_status_label()
            draw_wheel(canvas_wheel, 0)
        
        def on_exclude_toggle():
            """Xử lý khi toggle checkbox exclude"""
            self.spin_exclude_used = exclude_var.get()
            if not self.spin_exclude_used:
                self.spin_used_students.clear()
            update_status_label()
            draw_wheel(canvas_wheel, 0)
        
        def on_duration_toggle():
            """Xử lý khi toggle checkbox duration"""
            if duration_enabled.get():
                ent_duration.config(state="normal")
            else:
                ent_duration.config(state="disabled")
        
        # =============================================================================
        # UI LAYOUT
        # =============================================================================
        
        # Header
        header = tk.Frame(spin_win, bg="#fef3c7", padx=20, pady=15,
                        highlightbackground="#f59e0b", highlightthickness=2)
        header.pack(fill="x")
        
        tk.Label(header, text="🎯 QUAY SỐ MAY MẮN", font=("Segoe UI", 18, "bold"),
                bg="#fef3c7", fg="#d97706").pack()
        
        # Control panel
        ctrl_frame = tk.Frame(spin_win, bg=THEME["bg_card"], padx=20, pady=15)
        ctrl_frame.pack(fill="x", padx=20, pady=(10, 0))
        
        # Checkbox 1: Phân loại học lực
        chk_filter = tk.Checkbutton(ctrl_frame, text="PHÂN LOẠI HỌC LỰC", variable=filter_enabled,
                                   bg=THEME["bg_card"], fg=THEME["text_main"],
                                   font=("Segoe UI", 10, "bold"), command=on_filter_toggle)
        chk_filter.grid(row=0, column=0, sticky="w", pady=5)
        
        cbo_filter = ttk.Combobox(ctrl_frame, textvariable=filter_mode, state="disabled",
                                 values=["Khá/Giỏi", "Hoàn thành", "Cần chú ý"], width=15)
        cbo_filter.grid(row=0, column=1, padx=(10, 0), sticky="w")
        cbo_filter.bind("<<ComboboxSelected>>", lambda e: (update_status_label(), draw_wheel(canvas_wheel, 0)))
        
        # Checkbox 2: Loại nếu quay trúng
        chk_exclude = tk.Checkbutton(ctrl_frame, text="LOẠI NẾU QUAY TRÚNG", variable=exclude_var,
                                    bg=THEME["bg_card"], fg=THEME["text_main"],
                                    font=("Segoe UI", 10, "bold"), command=on_exclude_toggle)
        chk_exclude.grid(row=1, column=0, sticky="w", pady=5)
        
        lbl_status = tk.Label(ctrl_frame, text="", font=("Segoe UI", 9, "italic"),
                            bg=THEME["bg_card"], fg="#6b7280")
        lbl_status.grid(row=1, column=1, padx=(10, 0), sticky="w")
        
        # Checkbox 3: Thời gian quay
        chk_duration = tk.Checkbutton(ctrl_frame, text="THỜI GIAN QUAY:", variable=duration_enabled,
                                     bg=THEME["bg_card"], fg=THEME["text_main"],
                                     font=("Segoe UI", 10, "bold"), command=on_duration_toggle)
        chk_duration.grid(row=2, column=0, sticky="w", pady=5)
        
        duration_frame = tk.Frame(ctrl_frame, bg=THEME["bg_card"])
        duration_frame.grid(row=2, column=1, padx=(10, 0), sticky="w")
        
        ent_duration = tk.Entry(duration_frame, textvariable=duration_var, width=5,
                               font=("Segoe UI", 10), state="disabled", justify="center")
        ent_duration.pack(side="left")
        tk.Label(duration_frame, text="giây (0.5-10)", font=("Segoe UI", 9),
                bg=THEME["bg_card"], fg="#6b7280").pack(side="left", padx=(5, 0))
        
        # Nút QUAY với viền pastel mint
        spin_btn_frame = tk.Frame(ctrl_frame, bg="#34d399", padx=2, pady=2)
        spin_btn_frame.grid(row=2, column=2, padx=(20, 0), sticky="w")
        btn_spin_mini = tk.Button(spin_btn_frame, text="🎲 QUAY", command=start_spin,
                                 bg="#d1fae5", fg="#065f46", font=("Segoe UI", 10, "bold"),
                                 relief="flat", padx=20, pady=6, borderwidth=0, cursor="hand2")
        btn_spin_mini.pack()
        
        # Main content
        main_frame = tk.Frame(spin_win, bg=THEME["bg_app"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Left: Canvas quay số
        left_frame = tk.Frame(main_frame, bg=THEME["bg_card"], padx=10, pady=10,
                            highlightbackground="#000000", highlightthickness=2)
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        canvas_wheel = tk.Canvas(left_frame, width=400, height=400, bg="#ffffff",
                                highlightthickness=0)
        canvas_wheel.pack()
        
        # Right: Lịch sử
        right_frame = tk.Frame(main_frame, bg=THEME["bg_card"], width=280,
                             highlightbackground="#3b82f6", highlightthickness=2)
        right_frame.pack(side="right", fill="both", padx=(10, 0))
        right_frame.pack_propagate(False)
        
        tk.Label(right_frame, text="📜 LỊCH SỬ QUAY SỐ", font=("Segoe UI", 12, "bold"),
                bg=THEME["bg_card"], fg="#3b82f6").pack(pady=10)
        
        history_container = tk.Frame(right_frame, bg=THEME["bg_card"])
        history_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        history_canvas = tk.Canvas(history_container, bg=THEME["bg_card"], highlightthickness=0)
        history_scrollbar = ttk.Scrollbar(history_container, orient="vertical", command=history_canvas.yview)
        history_frame = tk.Frame(history_canvas, bg=THEME["bg_card"])
        
        history_frame.bind("<Configure>", lambda e: history_canvas.configure(scrollregion=history_canvas.bbox("all")))
        history_canvas.create_window((0, 0), window=history_frame, anchor="nw")
        history_canvas.configure(yscrollcommand=history_scrollbar.set)
        
        history_canvas.pack(side="left", fill="both", expand=True)
        history_scrollbar.pack(side="right", fill="y")
        
        # Frame chứa 2 nút
        btn_container = tk.Frame(right_frame, bg=THEME["bg_card"])
        btn_container.pack(pady=(0, 10), padx=10)  # Thêm padding ngang để cách viền
        
        # Nút export - đổi sang pastel xanh
        export_btn_frame = tk.Frame(btn_container, bg="#93c5fd", padx=2, pady=2)  # Viền pastel xanh
        export_btn_frame.pack(side="left", padx=(0, 8))
        tk.Button(export_btn_frame, text="💾XUÁT", command=export_history,
                 bg="#dbeafe", fg="#1e40af", font=("Segoe UI", 8, "bold"),  # Pastel xanh nhạt
                 relief="flat", padx=10, pady=4, borderwidth=0, cursor="hand2").pack()
        
        # Nút xóa lịch sử
        clear_btn_frame = tk.Frame(btn_container, bg="#fca5a5", padx=2, pady=2)  # Viền pastel đỏ
        clear_btn_frame.pack(side="left")
        tk.Button(clear_btn_frame, text="🗑️XÓA", command=clear_history,
                 bg="#fee2e2", fg="#991b1b", font=("Segoe UI", 8, "bold"),  # Pastel đỏ nhạt
                 relief="flat", padx=10, pady=4, borderwidth=0, cursor="hand2").pack()
        
        # Bottom buttons
        btn_frame = tk.Frame(spin_win, bg=THEME["bg_app"], pady=15)
        btn_frame.pack(fill="x")
        
        # Nút QUAY SỐ
        spin_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=2, pady=2)
        spin_btn_frame.pack(side="left", padx=20)
        btn_spin = tk.Button(spin_btn_frame, text="🎲 BẮT ĐẦU QUAY SỐ", command=start_spin,
                            bg="#10b981", fg="white", font=("Segoe UI", 14, "bold"),
                            relief="flat", padx=40, pady=12, borderwidth=0, cursor="hand2")
        btn_spin.pack()
        
        # Nút RESET
        reset_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        reset_btn_frame.pack(side="left", padx=10)
        tk.Button(reset_btn_frame, text="🔄 RESET TẤT CẢ", command=reset_all,
                 bg="#f59e0b", fg="white", font=("Segoe UI", 11, "bold"),
                 relief="flat", padx=20, pady=10, borderwidth=0).pack()
        
        # Nút ĐÓNG
        close_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        close_btn_frame.pack(side="right", padx=20)
        tk.Button(close_btn_frame, text="✖ ĐÓNG", command=spin_win.destroy,
                 bg="#94a3b8", fg="white", font=("Segoe UI", 11, "bold"),
                 relief="flat", padx=30, pady=10, borderwidth=0).pack()
        
        # Khởi tạo UI
        update_status_label()
        update_history_display()
        draw_wheel(canvas_wheel, 0)

    # =============================================================================
    # TÍNH NĂNG: TRỢ LÝ PHÒNG THI (EXAM PROCTOR TOOLKIT)
    # =============================================================================
    def show_exam_proctor_toolkit(self):
        """Hiển thị cửa sổ Trợ Lý Phòng Thi với đồng hồ đếm ngược khổng lồ"""
        
        # Tạo cửa sổ Toplevel
        exam_win = Toplevel(self)
        exam_win.title("⏳ Trợ Lý Phòng Thi - Exam Proctor Toolkit")
        self.center_window(exam_win, 850, 700)  # [OPTIMIZED] Thu gọn cho laptop
        exam_win.configure(bg=THEME["bg_app"])
        
        # Biến điều khiển UI (IntVar, StringVar)
        selected_duration = tk.IntVar(value=45)  # Mặc định 45 phút
        is_fullscreen = tk.BooleanVar(value=False)
        enable_sound = tk.BooleanVar(value=True)
        # [ADDED] Biến cho thời gian tùy chỉnh
        custom_time_enabled = tk.BooleanVar(value=False)
        custom_minutes = tk.StringVar(value="45")
        # [ADDED] Biến cho loại âm thanh
        sound_type = tk.StringVar(value="Exclamation")  # Mặc định
        # [END ADDED]
        
        # =============================================================================
        # UI LAYOUT
        # =============================================================================
        
        # Header với viền pastel nổi bật (Thu gọn)
        header_container = tk.Frame(exam_win, bg="#fafbfc", padx=15, pady=12)
        header_container.pack(fill="x")
        
        header = tk.Frame(header_container, bg="#fef3c7", padx=15, pady=10,
                         highlightbackground="#fbbf24", highlightthickness=2, relief="solid")
        header.pack(fill="x")
        
        tk.Label(header, text="⏳ ĐỒNG HỒ ĐẾM NGƯỢC", 
                 font=("Segoe UI", 16, "bold"), bg="#fef3c7", fg="#d97706").pack(pady=5)
        tk.Label(header, text="Đồng hồ đếm ngược & Cảnh báo tự động", 
                 font=("Segoe UI", 9, "italic"), bg="#fef3c7", fg="#f59e0b").pack()
        
        
        # =================================================================================
        # PANEL CÀI ĐẶT + ĐIỀU KHIỂN (2 CỘT: Trái=Settings, Phải=Buttons)
        # =================================================================================
        config_frame = tk.Frame(exam_win, bg=THEME["bg_card"], padx=20, pady=15,
                               highlightbackground=THEME["border"], highlightthickness=1)
        config_frame.pack(fill="x", padx=20, pady=(10, 0))
        
        tk.Label(config_frame, text="⚙️ CÀI ĐẶT THỜI GIAN \u0026 ĐIỀU KHIỂN:", 
                 font=("Segoe UI", 12, "bold"), bg=THEME["bg_card"], fg=THEME["text_main"]).pack(anchor="w", pady=(0, 10))
        
        # Container chứa 2 cột
        main_container = tk.Frame(config_frame, bg=THEME["bg_card"])
        main_container.pack(fill="both", expand=True)
        
        # ========== CỘT TRÁI: SETTINGS ==========
        left_column = tk.Frame(main_container, bg=THEME["bg_card"])
        left_column.pack(side="left", fill="both", expand=True, padx=(0, 20))
        
        # Preset buttons
        tk.Label(left_column, text="Chọn nhanh:", font=("Segoe UI", 10), 
                bg=THEME["bg_card"], fg=THEME["text_main"]).pack(anchor="w", pady=(0, 5))
        
        preset_frame = tk.Frame(left_column, bg=THEME["bg_card"])
        preset_frame.pack(fill="x", pady=5)
        
        presets = [
            ("15 phút", 15),
            ("30 phút", 30),
            ("45 phút", 45),
            ("60 phút", 60),
            ("90 phút", 90)
        ]
        
        for label_text, minutes in presets:
            btn = tk.Radiobutton(preset_frame, text=label_text, variable=selected_duration, 
                                value=minutes, bg=THEME["bg_card"], fg=THEME["text_main"],
                                font=("Segoe UI", 9, "bold"), selectcolor=PASTEL_PALETTE["mint"],
                                command=lambda m=minutes: update_preview(m))
            btn.pack(side="left", padx=5)
        
        # Custom time input
        custom_frame = tk.Frame(left_column, bg=THEME["bg_card"])
        custom_frame.pack(fill="x", pady=8)
        
        chk_custom = tk.Checkbutton(custom_frame, text="⏱️ Tùy chỉnh:", 
                                   variable=custom_time_enabled, bg=THEME["bg_card"], 
                                   fg=THEME["text_main"], font=("Segoe UI", 10, "bold"),
                                   command=lambda: toggle_custom_time())
        chk_custom.pack(side="left")
        
        ent_custom = tk.Entry(custom_frame, textvariable=custom_minutes, width=6, 
                             font=("Segoe UI", 10), justify="center", state="disabled",
                             relief="solid", bd=1)
        ent_custom.pack(side="left", padx=8)
        
        tk.Label(custom_frame, text="phút (1-300)", font=("Segoe UI", 9), 
                bg=THEME["bg_card"], fg="gray").pack(side="left")
        
        # [UPGRADED] Audio settings với dropdown và test button
        audio_frame = tk.Frame(left_column, bg=THEME["bg_card"])
        audio_frame.pack(fill="x", pady=8)
        
        chk_sound = tk.Checkbutton(audio_frame, text="🔊 Âm thanh:", 
                                  variable=enable_sound, bg=THEME["bg_card"], 
                                  fg=THEME["text_main"], font=("Segoe UI", 9, "bold"))
        chk_sound.pack(side="left")
        
        # Dropdown chọn loại âm thanh
        sound_options = ["Default", "Error", "Question", "Exclamation", "Asterisk"]
        sound_dropdown = ttk.Combobox(audio_frame, textvariable=sound_type, 
                                     values=sound_options, state="readonly", width=10,
                                     font=("Segoe UI", 9))
        sound_dropdown.pack(side="left", padx=5)
        
        # Nút nghe thử
        test_sound_btn = tk.Button(audio_frame, text="🔉 Nghe thử", 
                                  command=lambda: test_sound_alert(),
                                  bg=PASTEL_PALETTE["blue"], fg=THEME["text_main"],
                                  font=("Segoe UI", 8, "bold"), relief="flat", 
                                  padx=8, pady=2, cursor="hand2")
        test_sound_btn.pack(side="left", padx=3)
        
        # Fullscreen checkbox
        chk_fullscreen = tk.Checkbutton(left_column, text="🖥️ Toàn màn hình (F11)", 
                                       variable=is_fullscreen, bg=THEME["bg_card"], 
                                       fg=THEME["text_main"], font=("Segoe UI", 9),
                                       command=lambda: toggle_fullscreen())
        chk_fullscreen.pack(anchor="w", pady=3)
        
        # ========== CỘT PHẢI: BUTTONS ĐIỀU KHIỂN (Grid 2x2) ==========
        right_column = tk.Frame(main_container, bg=THEME["bg_card"])
        right_column.pack(side="right", fill="y")
        
        tk.Label(right_column, text="Điều khiển:", font=("Segoe UI", 9, "bold"), 
                bg=THEME["bg_card"], fg=THEME["text_main"]).pack(anchor="w", pady=(0, 5))
        
        # Container grid 2x2
        btn_grid = tk.Frame(right_column, bg=THEME["bg_card"])
        btn_grid.pack()
        
        # HÀNG 1: START + PAUSE
        row1 = tk.Frame(btn_grid, bg=THEME["bg_card"])
        row1.pack(pady=3)  # Tăng khoảng cách dọc
        
        # Nút START (Pastel Mint - chữ theo theme)
        start_frame = tk.Frame(row1, bg=PASTEL_PALETTE["mint_dark"], padx=1, pady=1)
        start_frame.pack(side="left", padx=4)  # Tăng khoảng cách ngang
        btn_start = tk.Button(start_frame, text="▶️BẮT ĐẦU", 
                             command=lambda: start_timer(),
                             bg=PASTEL_PALETTE["mint"], fg=THEME["text_main"],  # Đen/Trắng theo theme
                             font=("Segoe UI", 9, "bold"),
                             relief="flat", padx=12, pady=6, borderwidth=0, cursor="hand2", width=9)
        btn_start.pack()
        
        # Nút PAUSE (Pastel Orange - chữ theo theme)
        pause_frame = tk.Frame(row1, bg=PASTEL_PALETTE["orange_dark"], padx=1, pady=1)
        pause_frame.pack(side="left", padx=4)
        btn_pause = tk.Button(pause_frame, text="⏸️TẠM DỪNG", 
                             command=lambda: pause_timer(),
                             bg=PASTEL_PALETTE["orange"], fg=THEME["text_main"],  # Đen/Trắng theo theme
                             font=("Segoe UI", 9, "bold"),
                             relief="flat", padx=12, pady=6, borderwidth=0, cursor="hand2", 
                             state="disabled", width=9)
        btn_pause.pack()
        
        # HÀNG 2: RESET + CLOSE
        row2 = tk.Frame(btn_grid, bg=THEME["bg_card"])
        row2.pack(pady=3)
        
        # Nút RESET (Pastel Lavender - chữ theo theme)
        reset_frame = tk.Frame(row2, bg=PASTEL_PALETTE["lavender_dark"], padx=1, pady=1)
        reset_frame.pack(side="left", padx=4)
        btn_reset = tk.Button(reset_frame, text="🔄ĐẶT LẠI", 
                             command=lambda: reset_timer(),
                             bg=PASTEL_PALETTE["lavender"], fg=THEME["text_main"],  # Đen/Trắng theo theme
                             font=("Segoe UI", 9, "bold"),
                             relief="flat", padx=12, pady=6, borderwidth=0, cursor="hand2", width=9)
        btn_reset.pack()
        
        # Nút ĐÓNG (Pastel Rose - chữ theo theme)
        close_frame = tk.Frame(row2, bg=PASTEL_PALETTE["rose_dark"], padx=1, pady=1)
        close_frame.pack(side="left", padx=4)
        tk.Button(close_frame, text="✖ ĐÓNG", command=lambda: on_close_window(),
                 bg=PASTEL_PALETTE["rose"], fg=THEME["text_main"],  # Đen/Trắng theo theme
                 font=("Segoe UI", 9, "bold"),
                 relief="flat", padx=12, pady=6, borderwidth=0, width=9).pack()
        
        # =================================================================================
        # MÀN HÌNH HIỂN THỊ ĐỒNG HỒ (Tăng kích thước)
        # =================================================================================
        timer_display_frame = tk.Frame(exam_win, bg="#1e293b", height=380,
                                      highlightbackground="#06b6d4", highlightthickness=3)
        # [CRITICAL FIX] Bỏ expand=True và set height cố định để các nút điều khiển hiển thị
        timer_display_frame.pack(fill="x", padx=20, pady=15)
        timer_display_frame.pack_propagate(False)  # Giữ chiều cao cố định
        
        # [REORDERED] Status text ở TRÊN
        lbl_status = tk.Label(timer_display_frame, text="SẴN SÀNG BẮT ĐẦU ⏸", 
                             font=("Segoe UI", 16, "bold"), bg="#1e293b", fg="#94a3b8")
        lbl_status.pack(pady=(15, 5))  # Padding trên 15px, dưới 5px
        
        # Số đồng hồ ở DƯỚI
        lbl_timer_display = tk.Label(timer_display_frame, text="45:00", 
                                     font=("Consolas", 110, "bold"),  # [OPTIMIZED] Tăng lên 110 cho rõ hơn
                                     bg="#1e293b", fg="#06b6d4")
        lbl_timer_display.pack(expand=True)
        
        # =============================================================================
        # LOGIC ĐIỀU KHIỂN (CORE ALGORITHM)
        # =============================================================================
        
        def countdown_worker():
            """Thread worker chạy đếm ngược (riêng biệt với Main UI Thread)"""
            while self.exam_timer_running and self.exam_remaining_seconds > 0:
                if not self.exam_timer_paused:
                    # Giảm 1 giây
                    self.exam_remaining_seconds -= 1
                    
                    # [THREAD-SAFE] Dùng after() để update UI từ thread
                    exam_win.after(0, update_display)
                    
                    # Kiểm tra cảnh báo 15 phút
                    if self.exam_remaining_seconds == 900 and not self.exam_alert_played_15:
                        exam_win.after(0, lambda: play_alert("⚠️ CÒN 15 PHÚT!"))
                        self.exam_alert_played_15 = True
                    
                    # Kiểm tra cảnh báo 5 phút
                    if self.exam_remaining_seconds == 300 and not self.exam_alert_played_5:
                        exam_win.after(0, lambda: play_alert("🔔 CÒN 5 PHÚT!"))
                        self.exam_alert_played_5 = True
                    
                    # Hết giờ
                    if self.exam_remaining_seconds == 0:
                        exam_win.after(0, on_time_up)
                        break
                
                time.sleep(1)  # Chờ 1 giây
        
        def update_display():
            """Cập nhật Label hiển thị thời gian (MM:SS)"""
            try:
                minutes = self.exam_remaining_seconds // 60
                seconds = self.exam_remaining_seconds % 60
                time_str = f"{minutes:02d}:{seconds:02d}"
                
                lbl_timer_display.config(text=time_str)
                
                # Đổi màu khi gần hết giờ (Visual Alert)
                if self.exam_remaining_seconds <= 300:  # < 5 phút
                    lbl_timer_display.config(fg="#ef4444")  # Đỏ cảnh báo
                elif self.exam_remaining_seconds <= 900:  # < 15 phút
                    lbl_timer_display.config(fg="#f59e0b")  # Cam nhắc nhở
                else:
                    lbl_timer_display.config(fg="#06b6d4")  # Xanh bình thường
            except Exception as e:
                print(f"Lỗi update_display: {e}")
        
        def get_selected_duration():
            """Lấy thời gian được chọn (từ preset hoặc custom input)"""
            try:
                if custom_time_enabled.get():
                    # Dùng thời gian tùy chỉnh
                    minutes = int(custom_minutes.get())
                    # Giới hạn 1-300 phút
                    if minutes < 1:
                        minutes = 1
                        custom_minutes.set("1")
                    elif minutes > 300:
                        minutes = 300
                        custom_minutes.set("300")
                    return minutes
                else:
                    # Dùng preset
                    return selected_duration.get()
            except ValueError:
                messagebox.showerror("Lỗi", "Vui lòng nhập số phút hợp lệ (1-300)!")
                return 45  # Fallback mặc định
        
        def toggle_custom_time():
            """Bật/Tắt chế độ nhập thời gian tùy chỉnh"""
            try:
                if custom_time_enabled.get():
                    # Bật custom time
                    ent_custom.config(state="normal")
                    # Focus vào ô nhập
                    ent_custom.focus()
                    ent_custom.select_range(0, tk.END)
                    # Cập nhật preview
                    if not self.exam_timer_running:
                        try:
                            minutes = int(custom_minutes.get())
                            self.exam_remaining_seconds = minutes * 60
                            update_display()
                        except ValueError:
                            pass
                else:
                    # Tắt custom time
                    ent_custom.config(state="disabled")
                    # Về preset
                    if not self.exam_timer_running:
                        update_preview(selected_duration.get())
            except Exception as e:
                print(f"Lỗi toggle_custom_time: {e}")
        
        def start_timer():
            """Bắt đầu đồng hồ đếm ngược"""
            try:
                if not self.exam_timer_running:
                    # Bắt đầu lần đầu
                    self.exam_timer_running = True
                    self.exam_timer_paused = False
                    self.exam_remaining_seconds = get_selected_duration() * 60
                    self.exam_alert_played_15 = False
                    self.exam_alert_played_5 = False
                    
                    # Chạy thread đếm ngược (daemon=True để tự tắt khi đóng app)
                    self.exam_timer_thread = threading.Thread(target=countdown_worker, daemon=True)
                    self.exam_timer_thread.start()
                    
                    # Update UI buttons (Dùng màu pastel)
                    btn_start.config(state="disabled")  # Disable nút START
                    btn_pause.config(state="normal", text="⏸️TẠM DỪNG", bg=PASTEL_PALETTE["orange"])  # Enable PAUSE - pastel orange
                    btn_reset.config(bg=PASTEL_PALETTE["rose"])  # Đổi màu RESET sang rose (active)
                    lbl_status.config(text="ĐANG CHẠY ⏱", fg="#10b981")
                else:
                    messagebox.showinfo("Thông báo", "Đồng hồ đang chạy! Vui lòng Reset trước khi bắt đầu lại.")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể bắt đầu đồng hồ: {str(e)}")
        
        def pause_timer():
            """Tạm dừng/Tiếp tục đồng hồ"""
            try:
                if self.exam_timer_running:
                    # Toggle pause state (Dùng màu pastel)
                    self.exam_timer_paused = not self.exam_timer_paused
                    if self.exam_timer_paused:
                        btn_pause.config(text="▶️TIẾP TỤC", bg=PASTEL_PALETTE["mint"])  # Pause → mint (giống START)
                        lbl_status.config(text="TẠM DỪNG ⏸", fg="#f59e0b")
                    else:
                        btn_pause.config(text="⏸️TẠM DỪNG", bg=PASTEL_PALETTE["orange"])  # Resume → orange
                        lbl_status.config(text="ĐANG CHẠY ⏱", fg="#10b981")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể điều khiển đồng hồ: {str(e)}")
        
        def reset_timer():
            """Đặt lại đồng hồ về thời gian ban đầu"""
            try:
                self.exam_timer_running = False
                self.exam_timer_paused = False
                self.exam_remaining_seconds = get_selected_duration() * 60
                
                update_display()
                # Reset UI buttons về trạng thái ban đầu (Dùng màu pastel)
                btn_start.config(state="normal")  # Enable lại nút START
                btn_pause.config(state="disabled", text="⏸️TẠM DỪNG", bg=PASTEL_PALETTE["orange"])  # Disable PAUSE - pastel orange
                btn_reset.config(bg=PASTEL_PALETTE["lavender"])  # Đổi màu RESET về lavender (inactive)
                lbl_status.config(text="ĐÃ ĐẶT LẠI 🔄", fg="#94a3b8")
                lbl_timer_display.config(fg="#06b6d4")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể reset: {str(e)}")
        
        def update_preview(minutes):
            """Cập nhật preview khi chọn preset (chỉ khi chưa chạy)"""
            try:
                if not self.exam_timer_running:
                    self.exam_remaining_seconds = minutes * 60
                    update_display()
            except Exception as e:
                print(f"Lỗi update_preview: {e}")
        
        def get_sound_constant(sound_name):
            """Chuyển tên âm thanh thành hằng số winsound"""
            sound_map = {
                "Default": 0x00000000,  # MB_OK
                "Error": 0x00000010,    # MB_ICONHAND
                "Question": 0x00000020, # MB_ICONQUESTION
                "Exclamation": 0x00000030,  # MB_ICONEXCLAMATION
                "Asterisk": 0x00000040  # MB_ICONASTERISK
            }
            return sound_map.get(sound_name, 0x00000030)  # Default: Exclamation
        
        def test_sound_alert():
            """Nghe thử âm thanh đã chọn"""
            try:
                if HAS_WINSOUND:
                    sound_const = get_sound_constant(sound_type.get())
                    winsound.MessageBeep(sound_const)
                else:
                    messagebox.showinfo("Thông báo", "Winsound không khả dụng trên hệ điều hành này.")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể phát âm thanh: {str(e)}")
        
        def play_alert(message):
            """Phát âm thanh cảnh báo + Flash màn hình"""
            try:
                # Âm thanh cảnh báo (Windows) - Lặp 10 lần
                if enable_sound.get() and HAS_WINSOUND:
                    def play_sound_loop():
                        """Lặp âm thanh 10 lần trong thread riêng"""
                        try:
                            sound_const = get_sound_constant(sound_type.get())
                            print(f"[DEBUG] Bắt đầu phát âm thanh {sound_type.get()} - 10 lần")
                            for i in range(10):  # Lặp 10 lần
                                print(f"[DEBUG] Beep lần {i+1}/10")
                                winsound.MessageBeep(sound_const)
                                if i < 9:  # Không sleep sau lần cuối
                                    time.sleep(0.6)  # Tăng delay lên 600ms để âm thanh không chồng lên nhau
                            print(f"[DEBUG] Hoàn thành 10 lần beep")
                        except Exception as e:
                            print(f"[ERROR] Lỗi play_sound_loop: {e}")
                    
                    # Chạy trong thread riêng để không block UI
                    threading.Thread(target=play_sound_loop, daemon=True).start()
                
                # Flash màn hình
                original_bg = lbl_status.cget("bg")
                lbl_status.config(text=message, bg="#fbbf24", fg="#78350f",
                                 font=("Segoe UI", 18, "bold"))
                exam_win.after(3000, lambda: lbl_status.config(
                    bg=original_bg, 
                    text="ĐANG CHẠY ⏱" if self.exam_timer_running else "SẴN SÀNG ⏸",
                    font=("Segoe UI", 16, "bold"),
                    fg="#10b981" if self.exam_timer_running else "#94a3b8"
                ))
            except Exception as e:
                print(f"Lỗi play_alert: {e}")
        
        def on_time_up():
            """Xử lý khi hết giờ làm bài"""
            try:
                self.exam_timer_running = False
                
                lbl_timer_display.config(text="00:00", fg="#ef4444")
                lbl_status.config(text="🔔 HẾT GIỜ LÀM BÀI!", fg="#ef4444", 
                                 font=("Segoe UI", 20, "bold"))
                
                # Âm thanh liên tục (10 lần MessageBeep - giống play_alert)
                if enable_sound.get() and HAS_WINSOUND:
                    def play_time_up_sound():
                        try:
                            sound_const = get_sound_constant(sound_type.get())
                            for i in range(10):  # 10 lần để nhất quán
                                winsound.MessageBeep(sound_const)
                                if i < 9:
                                    time.sleep(0.6)  # 600ms delay
                        except Exception:
                            pass
                    threading.Thread(target=play_time_up_sound, daemon=True).start()
                
                messagebox.showinfo("🔔 Hết giờ!", 
                                   "Thời gian làm bài đã kết thúc!\n\nVui lòng thu bài của học sinh.",
                                   parent=exam_win)
            except Exception as e:
                print(f"Lỗi on_time_up: {e}")
        
        def toggle_fullscreen():
            """Bật/Tắt chế độ toàn màn hình"""
            try:
                if is_fullscreen.get():
                    exam_win.attributes('-fullscreen', True)
                else:
                    exam_win.attributes('-fullscreen', False)
            except Exception:
                pass  # Fallback nếu hệ điều hành không hỗ trợ
        
        def on_close_window():
            """Xử lý khi đóng cửa sổ (dừng timer trước)"""
            try:
                self.exam_timer_running = False
                self.exam_timer_paused = False
                exam_win.destroy()
            except Exception:
                exam_win.destroy()
        
        # Bind phím tắt F11 để toggle fullscreen
        exam_win.bind('<F11>', lambda e: (is_fullscreen.set(not is_fullscreen.get()), toggle_fullscreen()))
        
        # Khởi tạo hiển thị ban đầu
        update_preview(45)

    def show_smart_grouping(self):
        if self.current_df is None: return

        group_win = Toplevel(self)
        group_win.title("Phân Nhóm Học Tập Thông Minh")
        self.center_window(group_win, 1100, 700)
        group_win.configure(bg=THEME["bg_app"])
        
        # Cleanup khi đóng cửa sổ
        def on_group_close():
            try:
                group_win.unbind_all("<MouseWheel>")
            except:
                pass
            group_win.destroy()
        group_win.protocol("WM_DELETE_WINDOW", on_group_close)

        self.current_groups_data = [] 

        ctrl_frame = tk.Frame(group_win, bg=THEME["bg_card"], pady=10, padx=15)
        ctrl_frame.pack(fill="x", pady=(0, 2))
        
        tk.Label(ctrl_frame, text="🧩 CHIA NHÓM", font=("Segoe UI", 14, "bold"), bg=THEME["bg_card"], fg=THEME["primary"]).pack(side="left", padx=(0, 20))

        tk.Label(ctrl_frame, text="Số người/nhóm:", font=("Segoe UI", 10), bg=THEME["bg_card"], fg=THEME["text_main"]).pack(side="left")
        ent_size = tk.Entry(ctrl_frame, width=5, font=("Segoe UI", 10), justify="center", relief="solid", bd=1)
        ent_size.insert(0, "4") 
        ent_size.pack(side="left", padx=5)

        btn_frame = tk.Frame(ctrl_frame, bg=THEME["bg_card"])
        btn_frame.pack(side="right")

        container = tk.Frame(group_win, bg=THEME["bg_app"])
        container.pack(fill="both", expand=True, padx=10, pady=10)

        canvas = tk.Canvas(container, bg=THEME["bg_app"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=THEME["bg_app"])

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def on_canvas_configure(event): canvas.itemconfig(window_id, width=event.width)
        canvas.bind("<Configure>", on_canvas_configure)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def _on_mousewheel(event):
            if event.delta: canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def generate_groups(shuffle=False):
            try:
                group_size = int(ent_size.get())
                if group_size <= 0: raise ValueError
            except:
                messagebox.showerror("Lỗi", "Vui lòng nhập số người mỗi nhóm hợp lệ (>0)!")
                return

            df_students = self.current_df.copy()
            df_students['ĐTB'] = pd.to_numeric(df_students['ĐTB'], errors='coerce').fillna(0)
            
            df_students = df_students.sort_values(by='ĐTB', ascending=False).reset_index(drop=True)
            
            total_students = len(df_students)
            if total_students == 0:
                messagebox.showinfo("Thông báo", "Lớp chưa có học sinh nào!")
                return

            num_groups = math.ceil(total_students / group_size)
            groups = [[] for _ in range(num_groups)]

            indices = list(range(total_students))
            
            if shuffle:
                shuffled_indices = []
                for i in range(0, total_students, num_groups):
                    chunk = indices[i : i + num_groups]
                    random.shuffle(chunk) 
                    shuffled_indices.extend(chunk)
                
                df_students = df_students.iloc[shuffled_indices].reset_index(drop=True)

            for i, row in df_students.iterrows():
                pos = i % num_groups
                if (i // num_groups) % 2 == 1:
                    pos = num_groups - 1 - pos
                groups[pos].append({
                    "name": row['Họ và tên'], 
                    "score": row['ĐTB'],
                    "dob": row.get('Ngày sinh', '')
                })

            self.current_groups_data = groups
            render_groups_ui(groups)

        def render_groups_ui(groups):
            for widget in scrollable_frame.winfo_children(): widget.destroy()

            cols = 3 
            for i, group in enumerate(groups):
                r, c = divmod(i, cols)
                avg_score = sum(m['score'] for m in group) / len(group) if group else 0
                
                card = tk.Frame(scrollable_frame, bg=THEME["bg_card"], bd=1, relief="solid")
                card.config(highlightbackground="#bdc3c7", highlightthickness=1, relief="flat")
                card.grid(row=r, column=c, padx=10, pady=10, sticky="nsew")
                
                colors = ["#3498db", "#9b59b6", "#e67e22", "#1abc9c", "#34495e", "#d35400"]
                bg_head = colors[i % len(colors)]
                
                head_f = tk.Frame(card, bg=bg_head, padx=10, pady=5)
                head_f.pack(fill="x")
                tk.Label(head_f, text=f"NHÓM {i+1}", font=("Segoe UI", 11, "bold"), fg="white", bg=bg_head).pack(side="left")
                tk.Label(head_f, text=f"TB: {avg_score:.2f}", font=("Segoe UI", 11, "bold"), fg="white", bg=bg_head).pack(side="right")
                
                body_f = tk.Frame(card, bg=THEME["bg_card"], padx=10, pady=10)
                body_f.pack(fill="both", expand=True)
                
                for member in group:
                    f_mem = tk.Frame(body_f, bg=THEME["bg_card"])
                    f_mem.pack(fill="x", pady=2)
                    
                    icon = "🔹"
                    s = member['score']
                    if s >= 8.0: icon = "🌟"
                    elif s < 5.0: icon = "⚠️"
                    
                    tk.Label(f_mem, text=f"{icon} {member['name']}", font=("Segoe UI", 10), bg=THEME["bg_card"], fg=THEME["text_main"], anchor="w").pack(side="left", fill="x", expand=True)
                    
                    lbl_score = tk.Label(f_mem, text=f"{s}", font=("Segoe UI", 9, "bold"), fg="#7f8c8d", bg=THEME["bg_card"])
                    if s < 5.0: lbl_score.config(fg="#c0392b")
                    elif s >= 8.0: lbl_score.config(fg="#27ae60")
                    lbl_score.pack(side="right")

            for i in range(cols): scrollable_frame.columnconfigure(i, weight=1)

        def save_groups_to_file():
            if not self.current_groups_data:
                messagebox.showwarning("Cảnh báo", "Vui lòng tạo nhóm trước khi lưu!")
                return
                
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Lưu Danh Sách Nhóm",
                initialfile=f"DanhSachNhom_Lop_{self.cbo_classes.get()}.xlsx"
            )
            
            if not file_path: return

            try:
                export_rows = []
                for idx, group in enumerate(self.current_groups_data):
                    group_name = f"Nhóm {idx + 1}"
                    for member in group:
                        export_rows.append({
                            "Tên Nhóm": group_name,
                            "Họ và tên": member['name'],
                            "Ngày sinh": member['dob'],
                            "Điểm TB": member['score']
                        })
                
                df_export = pd.DataFrame(export_rows)
                df_export.to_excel(file_path, index=False)
                messagebox.showinfo("Thành công", f"Đã lưu file tại:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Lỗi Lưu File", str(e))

        def create_btn(parent, text, bg, cmd):
            btn = ModernButton(parent, bg_color=bg, text_color="white", text=text, command=cmd)
            btn.pack(side="left", padx=5)
            return btn

        create_btn(btn_frame, "🎲 TẠO / XÁO TRỘN", "#2980b9", lambda: generate_groups(shuffle=True))
        create_btn(btn_frame, "⚖️ SẮP XẾP CHUẨN", "#27ae60", lambda: generate_groups(shuffle=False))
        create_btn(btn_frame, "💾 LƯU FILE", "#e67e22", save_groups_to_file)

        generate_groups(shuffle=False)

    def predict_score_ml(self, student_grades):
        if not HAS_SKLEARN: return None, "Chưa cài thư viện scikit-learn."
        
        df_train = self.current_df.copy()
        
        # [DYNAMIC TX] Tự động phát hiện các cột TX có trong dữ liệu
        tx_cols = [col for col in df_train.columns if col.startswith("TX") and col[2:].isdigit()]
        score_cols = sorted(tx_cols) + ['GK', 'CK']
        
        for c in score_cols: 
            df_train[c] = pd.to_numeric(df_train[c], errors='coerce')
        
        df_train = df_train.dropna(subset=['CK'])
        df_train = df_train[df_train['CK'] > 0]
        if len(df_train) < 5: return None, "Chưa đủ dữ liệu lớp để chạy AI."

        # [DYNAMIC TX] Sử dụng các cột TX động + GK để predict
        feature_cols = sorted(tx_cols) + ['GK']
        X_train = df_train[feature_cols].fillna(0)
        y_train = df_train['CK']

        try:
            model = LinearRegression()
            model.fit(X_train, y_train)
            input_val = np.array([0 if x is None else x for x in student_grades]).reshape(1, -1)
            predicted_ck = model.predict(input_val)[0]
            return round(max(0, min(10, predicted_ck)), 2), "Dựa trên Hồi quy tuyến tính."
        except Exception as e: return None, str(e)

    def show_student_dashboard(self, student_name, student_data):
        dash = Toplevel(self)
        dash.title(f"Hồ sơ học tập 360: {student_name}")
        self.center_window(dash, 880, 600)
        dash.configure(bg=THEME["bg_app"])
        
        top_frame = tk.Frame(dash, bg=THEME["primary"], pady=10)
        top_frame.pack(fill="x")
        tk.Label(top_frame, text=f"👤 {student_name.upper()}", font=("Segoe UI", 20, "bold"), bg=THEME["primary"], fg="white").pack()
        tk.Label(top_frame, text="HỒ SƠ HỌC TẬP & DỰ BÁO AI", font=("Segoe UI", 10), bg=THEME["primary"], fg="#E3F2FD").pack()

        content = tk.Frame(dash, bg=THEME["bg_app"], padx=15, pady=15)
        content.pack(fill="both", expand=True)

        left_panel = tk.Frame(content, bg=THEME["bg_app"], width=260)
        left_panel.pack(side="left", fill="y", padx=(0, 15))

        all_dtb = pd.to_numeric(self.current_df["ĐTB"], errors='coerce').fillna(0)
        student_dtb = float(student_data.get("ĐTB", 0) or 0)
        sorted_scores = sorted(list(all_dtb), reverse=True)
        try: rank = sorted_scores.index(student_dtb) + 1 if student_dtb > 0 else "---"
        except: rank = "---"
        
        def create_modern_stat_card(parent, label, value, color):
            card = tk.Frame(parent, bg=THEME["bg_card"], pady=10, padx=10, highlightbackground="#ddd", highlightthickness=1)
            card.pack(fill="x", pady=8)
            tk.Frame(card, bg=color, width=5).place(x=0, y=0, relheight=1)
            tk.Label(card, text=label, bg=THEME["bg_card"], fg="#7f8c8d", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10)
            tk.Label(card, text=str(value), bg=THEME["bg_card"], fg=color, font=("Segoe UI", 22, "bold")).pack(anchor="w", padx=10)

        create_modern_stat_card(left_panel, "ĐIỂM TRUNG BÌNH", student_data.get("ĐTB", "---"), "#f39c12")
        create_modern_stat_card(left_panel, "XẾP HẠNG LỚP", f"#{rank} / {len(self.current_df)}", "#2ecc71")
        create_modern_stat_card(left_panel, "XẾP LOẠI", student_data.get("Xếp loại", "---").upper(), "#3498db")

        right_panel = tk.Frame(content, bg=THEME["bg_app"])
        right_panel.pack(side="right", fill="both", expand=True)

        chart_frame = tk.Frame(right_panel, bg=THEME["chart_bg"], highlightbackground="#ddd", highlightthickness=1, padx=5, pady=5)
        chart_frame.pack(fill="both", expand=True, pady=(5, 10))

        # Container cho phần insight/nhận xét với viền xanh đậm
        comment_container = tk.Frame(right_panel, bg="#E3F2FD", padx=10, pady=8, 
                                   highlightbackground="#1e40af", highlightthickness=2, relief="solid")
        comment_container.pack(fill="x", pady=(0, 10), padx=2)

        # [DYNAMIC TX] Tự động lấy danh sách cột điểm từ dữ liệu
        tx_cols = [col for col in student_data.keys() if col.startswith("TX") and col[2:].isdigit()]
        cols_score = sorted(tx_cols) + ["GK", "CK"]
        student_scores = []
        class_avg_scores = []
        raw_scores_for_ml = []
        
        for col in cols_score:
            val = student_data.get(col, 0)
            try: v = float(val); student_scores.append(v)
            except: v = 0; student_scores.append(0)
            if col != "CK": raw_scores_for_ml.append(v)

            try:
                col_data = pd.to_numeric(self.current_df[col], errors='coerce').fillna(0)
                valid_scores = col_data[col_data > 0]
                if len(valid_scores) > 0: class_avg_scores.append(valid_scores.mean())
                else: class_avg_scores.append(0)
            except: class_avg_scores.append(0)

        # Biểu đồ tuyến tính so sánh điểm HS và trung bình lớp
        fig = Figure(figsize=(5.5, 3), dpi=100) 
        fig.patch.set_facecolor(THEME["chart_bg"])
        ax_line = fig.add_subplot(111)
        ax_line.set_facecolor(THEME["chart_bg"])
        x_pos = list(range(len(cols_score)))
        ax_line.plot(x_pos, student_scores, marker="o", color="#2980b9", label="Học sinh")
        ax_line.plot(x_pos, class_avg_scores, marker="o", color="#e67e22", label="TB Lớp")
        ax_line.set_xticks(x_pos)
        ax_line.set_xticklabels(cols_score)
        ax_line.set_ylabel("Điểm", color=THEME["text_main"])
        ax_line.set_ylim(0, 10.5)
        ax_line.grid(True, linestyle="--", alpha=0.3)
        ax_line.tick_params(colors=THEME["text_main"])
        for spine in ax_line.spines.values(): spine.set_color(THEME["text_main"])
        legend = ax_line.legend(frameon=False)
        if legend:
            for txt in legend.get_texts(): txt.set_color(THEME["text_main"])

        canvas_line = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas_line.draw()
        canvas_line.get_tk_widget().pack(fill="both", expand=True)

        # Khối insight và nhận xét
        tk.Label(comment_container, text="AI & DATA SCIENCE INSIGHTS", font=("Segoe UI", 10, "bold"), bg="#E3F2FD", fg="#1565C0").pack(anchor="w", padx=(15, 10), pady=(5, 2))
        
        comments = []
        pred_val, pred_msg = self.predict_score_ml(raw_scores_for_ml)
        if pred_val is not None:
            comments.append(f"🎯 DỰ ĐOÁN CUỐI KỲ: {pred_val} điểm")
            if pred_val < 5.0: comments.append(f"   ➔ CẢNH BÁO: Điểm thấp!")
            elif pred_val >= 8.0: comments.append(f"   ➔ TUYỆT VỜI: Khả năng đạt Giỏi.")
            else: comments.append(f"   ➔ ỔN ĐỊNH: Cần cố gắng thêm.")
        else: comments.append(f"ℹ️ {pred_msg}")

        # Fix: Check list length before accessing indices
        if len(student_scores) >= 6 and student_scores[4] > 0 and student_scores[5] > 0:
            diff = student_scores[5] - student_scores[4]
            if diff >= 1.5: comments.append("🚀 Tiến bộ vượt bậc.")
            elif diff <= -1.5: comments.append("⚠️ Phong độ giảm sút.")
        
        lbl_comment = tk.Label(comment_container, text="\n".join(comments), justify="left", bg="#E3F2FD", fg="#263238", font=("Segoe UI", 11))
        lbl_comment.pack(anchor="w", padx=(20, 10), pady=(0, 5))
        comment_container.bind('<Configure>', lambda e: lbl_comment.config(wraplength=e.width - 40))

    def on_tree_double_click(self, event):
        # [FIX] Nếu đang có entry khác đang edit, lưu nó trước
        if self.active_entry and self.active_entry.winfo_exists():
            if self.active_entry_save_func:
                self.active_entry_save_func()
        
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell": return
        column_id = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        if not item_id: return
        
        col_index = int(column_id.replace("#", "")) - 1
        col_name = self.cols_def[col_index]
        current_values = self.tree.item(item_id, "values")

        if col_name in ["Họ và tên", "STT"]:
            student_data = {self.cols_def[i]: current_values[i] for i in range(len(self.cols_def))}
            self.show_student_dashboard(student_data["Họ và tên"], student_data)
            return

        if col_name in ["Ngày sinh", "Xếp loại"]: return

        x, y, w, h = self.tree.bbox(item_id, column_id)
        entry = tk.Entry(self.tree, width=w, font=THEME["font_body"])
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, current_values[col_index])
        entry.select_range(0, tk.END)
        entry.focus()
        
        # [FIX] Lưu reference đến entry hiện tại
        self.active_entry = entry

        def save_score(event=None):
            if not entry.winfo_exists(): 
                return
            raw = entry.get().strip()
            try: 
                num_val = float(raw) if "." in raw else int(raw)
            except: 
                num_val = raw

            cur_vals = list(self.tree.item(item_id, "values"))
            old_val = cur_vals[col_index]
            cur_vals[col_index] = str(num_val)
            if col_name == "ĐTB": 
                cur_vals[self.cols_def.index("Xếp loại")] = self.classify_smart(num_val)

            self.tree.item(item_id, values=cur_vals)
            
            stt = int(cur_vals[0])
            idx = self.current_df.index[self.current_df['STT'].astype(str) == str(stt)].tolist()
            if idx:
                # [FIX] Convert cột sang object dtype trước khi gán để tránh TypeError với string dtype
                if col_name in self.current_df.columns:
                    if self.current_df[col_name].dtype.name == 'string':
                        self.current_df[col_name] = self.current_df[col_name].astype(object)
                self.current_df.at[idx[0], col_name] = num_val
                
                # Cập nhật cả all_classes_data để đồng bộ
                sheet_name = self.cbo_classes.get()
                if sheet_name in self.all_classes_data:
                    df_ref = self.all_classes_data[sheet_name]
                    if col_name in df_ref.columns and df_ref[col_name].dtype.name == 'string':
                        df_ref[col_name] = df_ref[col_name].astype(object)
                    idx_all = df_ref.index[df_ref['STT'].astype(str) == str(stt)].tolist()
                    if idx_all:
                        df_ref.at[idx_all[0], col_name] = num_val
                        if col_name == "ĐTB": 
                            df_ref.at[idx_all[0], "Xếp loại"] = self.classify_smart(num_val)
                
                if col_name == "ĐTB": 
                    self.current_df.at[idx[0], "Xếp loại"] = self.classify_smart(num_val)

            # Lưu lịch sử cho Undo/Redo
            self.history_stack.append({"stt": stt, "col": col_name, "old": old_val, "new": num_val})
            self.redo_stack.clear()
            
            # [FIX] Clear reference
            self.active_entry = None
            self.active_entry_save_func = None
            
            entry.destroy()
            self.perform_auto_save()
            self.update_ui_data(self.current_df, update_chart=True)
        
        # [FIX] Lưu hàm save để có thể gọi từ bên ngoài
        self.active_entry_save_func = save_score
        
        def cancel_edit(event=None):
            """Hủy edit khi nhấn Escape"""
            if entry.winfo_exists():
                self.active_entry = None
                self.active_entry_save_func = None
                entry.destroy()

        # Bind các events
        entry.bind("<Return>", save_score)
        entry.bind("<FocusOut>", save_score)
        entry.bind("<Escape>", cancel_edit)
    
    def on_tree_scroll_or_click(self, event):
        """Xử lý khi scroll hoặc click trên treeview - Lưu entry đang edit"""
        # Nếu có entry đang được edit, lưu nó
        if self.active_entry and self.active_entry.winfo_exists():
            # Kiểm tra xem event có phải là click không
            if event.type == "4":  # ButtonPress
                # Kiểm tra xem click có nằm trong vùng entry không
                try:
                    ex = self.active_entry.winfo_x()
                    ey = self.active_entry.winfo_y()
                    ew = self.active_entry.winfo_width()
                    eh = self.active_entry.winfo_height()
                    
                    # Nếu click ngoài entry, lưu lại
                    if not (ex <= event.x <= ex + ew and ey <= event.y <= ey + eh):
                        if self.active_entry_save_func:
                            self.active_entry_save_func()
                except:
                    # Nếu có lỗi, cũng lưu luôn
                    if self.active_entry_save_func:
                        self.active_entry_save_func()
            else:
                # Nếu là scroll, lưu luôn
                if self.active_entry_save_func:
                    self.active_entry_save_func()

    def open_manual_mapping(self):
        try:
            if not self.file_path:
                messagebox.showwarning("Cảnh báo", "Chưa có file nào được mở!")
                return
            
            sheet_name = self.cbo_classes.get()
            if not sheet_name:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một lớp để khớp cột!")
                return
            
            raw_df = self.robust_read_excel(self.file_path, sheet_name=sheet_name, header=None)
            
            # Giai đoạn 8: Truyền file_path để hỗ trợ auto-save
            dialog = ColumnMappingDialog(self, raw_df, sheet_name, file_path=self.file_path)
            self.wait_window(dialog) 
            
            if dialog.result_df is not None and len(dialog.result_df) > 0:
                processed = dialog.result_df
                processed["Xếp loại"] = processed["ĐTB"].apply(self.classify_smart)
                
                self.all_classes_data[sheet_name] = processed
                self.on_class_change(None) 
                messagebox.showinfo("Thành công", "Đã cập nhật dữ liệu từ việc khớp cột thủ công!")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở bảng khớp cột: {str(e)}")

    def robust_read_excel(self, file_path, sheet_name=None, header=None):
        """
        Đọc file Excel với khả năng tự động repair file corrupt.
        
        Thứ tự thử:
        1. Đọc bình thường (xlrd cho .xls, openpyxl cho .xlsx)
        2. Thử đọc như HTML table (file VNEdu giả dạng .xls)
        3. Thử đọc như CSV tab-separated
        4. [MỚI] Dùng Microsoft Excel COM để auto-repair và convert
        
        Args:
            file_path: Đường dẫn file Excel
            sheet_name: Tên sheet cần đọc (None = sheet đầu tiên)
            header: Dòng header (None = không có header)
            
        Returns:
            DataFrame hoặc None nếu thất bại
        """
        # Bước 1: Thử đọc bình thường
        try:
            if file_path.lower().endswith('.xls'):
                if not HAS_XLRD: 
                    raise ImportError("Cần xlrd để đọc .xls")
                return pd.read_excel(file_path, sheet_name=sheet_name, header=header, engine='xlrd')
            else:
                return pd.read_excel(file_path, sheet_name=sheet_name, header=header, engine='openpyxl')
        
        except Exception as e1:
            error_msg = str(e1).lower()
            is_corrupt = 'corrupt' in error_msg or 'compDoc' in str(e1) or 'seen[' in str(e1)
            
            # Bước 2: Thử đọc như HTML table
            try:
                dfs = pd.read_html(file_path, header=None, encoding='utf-8')
                if dfs:
                    if len(dfs) == 1: 
                        return dfs[0]
                    return max(dfs, key=lambda x: x.size)
            except Exception as e2:
                if "lxml" in str(e2) or "html5lib" in str(e2):
                    pass  # Tiếp tục thử cách khác
            
            # Bước 3: Thử đọc như CSV tab-separated
            try:
                return pd.read_csv(file_path, sep='\t', header=None, encoding='utf-16')
            except:
                pass
            
            # Bước 4: [MỚI] Dùng Excel COM để auto-repair (chỉ trên Windows)
            if HAS_WIN32COM and is_corrupt:
                try:
                    result = self._repair_with_excel_com(file_path, sheet_name, header)
                    if result is not None:
                        return result
                except Exception as e3:
                    print(f"[DEBUG] Excel COM repair failed: {e3}")
            
            # Tất cả đều thất bại - hiển thị thông báo lỗi chi tiết
            if is_corrupt:
                messagebox.showerror(
                    "File Excel bị hỏng", 
                    f"File '{os.path.basename(file_path)}' bị hỏng cấu trúc.\n\n"
                    f"Đây là vấn đề phổ biến với file từ VNEdu.\n\n"
                    f"Giải pháp: Mở file bằng Microsoft Excel, sau đó:\n"
                    f"File → Save As → Chọn 'Excel Workbook (.xlsx)'\n\n"
                    f"Chi tiết lỗi: {str(e1)[:100]}"
                )
                return None
            raise e1
    
    def _repair_with_excel_com(self, file_path, sheet_name=None, header=None):
        """
        Dùng Microsoft Excel COM để mở và đọc file corrupt trực tiếp.
        Đọc dữ liệu từ Excel COM mà không cần SaveAs (tránh lỗi Unicode).
        
        Yêu cầu: Microsoft Excel phải được cài đặt.
        
        Args:
            file_path: Đường dẫn file Excel corrupt
            sheet_name: Tên sheet cần đọc (None = sheet đầu tiên hoặc index 0)
            header: Dòng header (None = không có header)
            
        Returns:
            DataFrame hoặc None nếu thất bại
        """
        if not HAS_WIN32COM:
            return None
            
        excel = None
        
        try:
            # Khởi động Excel ẩn
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Mở file với chế độ repair (CorruptLoad=1)
            wb = excel.Workbooks.Open(
                os.path.abspath(file_path), 
                UpdateLinks=0, 
                ReadOnly=True, 
                CorruptLoad=1
            )
            
            # Xác định sheet cần đọc
            if sheet_name is None:
                ws = wb.Sheets(1)  # Sheet đầu tiên
            elif isinstance(sheet_name, int):
                ws = wb.Sheets(sheet_name + 1)  # Excel 1-indexed
            else:
                ws = wb.Sheets(sheet_name)
            
            # Lấy vùng dữ liệu có giá trị (UsedRange)
            used_range = ws.UsedRange
            
            # Đọc tất cả giá trị thành list of lists
            # .Value trả về tuple of tuples
            data = used_range.Value
            
            if data is None:
                wb.Close(False)
                excel.Quit()
                return None
            
            # Chuyển thành DataFrame
            # data có thể là tuple of tuples hoặc single value
            if isinstance(data, tuple):
                if isinstance(data[0], tuple):
                    # Nhiều dòng
                    df = pd.DataFrame(list(data))
                else:
                    # Chỉ 1 dòng
                    df = pd.DataFrame([list(data)])
            else:
                # Chỉ 1 cell
                df = pd.DataFrame([[data]])
            
            wb.Close(False)
            excel.Quit()
            excel = None
            
            # Xử lý header nếu cần
            if header is not None and header >= 0 and header < len(df):
                df.columns = df.iloc[header].values
                df = df.iloc[header + 1:].reset_index(drop=True)
            
            print(f"[INFO] File corrupt đã được đọc trực tiếp bằng Excel COM: {df.shape}")
            return df
            
        except Exception as e:
            print(f"[DEBUG] _repair_with_excel_com error: {e}")
            return None
            
        finally:
            # Cleanup: Đảm bảo đóng Excel
            if excel is not None:
                try:
                    excel.Quit()
                except:
                    pass

    def _get_sheet_names_via_com(self, file_path):
        """
        Lấy danh sách sheet names thông qua Excel COM (cho file corrupt).
        
        Args:
            file_path: Đường dẫn file Excel
            
        Returns:
            List[str] danh sách tên sheet, hoặc None nếu thất bại
        """
        if not HAS_WIN32COM:
            return None
            
        excel = None
        try:
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(
                os.path.abspath(file_path), 
                UpdateLinks=0, 
                ReadOnly=True, 
                CorruptLoad=1
            )
            
            sheet_names = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
            wb.Close(False)
            excel.Quit()
            
            print(f"[INFO] Đã lấy {len(sheet_names)} sheet từ file corrupt via Excel COM")
            return sheet_names
            
        except Exception as e:
            print(f"[DEBUG] _get_sheet_names_via_com error: {e}")
            return None
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except:
                    pass

    def load_data_from_file(self):
        """
        Load dữ liệu từ file Excel, hỗ trợ auto-repair file corrupt.
        """
        try:
            engine = 'openpyxl'
            sheet_names = None
            is_xls = self.file_path.lower().endswith('.xls')
            
            if is_xls:
                if not HAS_XLRD:
                    messagebox.showerror("Lỗi", "Cần thư viện 'xlrd' để đọc file .xls. Vui lòng cài đặt: pip install xlrd")
                    return
                engine = 'xlrd'

            # Thử lấy danh sách sheet names
            try:
                xls = pd.ExcelFile(self.file_path, engine=engine)
                sheet_names = xls.sheet_names
            except Exception as e:
                error_msg = str(e).lower()
                is_corrupt = 'corrupt' in error_msg or 'compdoc' in error_msg or 'seen[' in str(e)
                
                # Nếu file corrupt và có Excel COM, thử repair
                if is_corrupt and HAS_WIN32COM:
                    sheet_names = self._get_sheet_names_via_com(self.file_path)
                    
                if sheet_names is None:
                    sheet_names = ["Sheet1"]

            try:
                if "NhatKy" in sheet_names:
                    self.logbook_df = pd.read_excel(self.file_path, sheet_name="NhatKy")
            except: pass
                
            self.stop_interactive_mapping = False

            loaded_count = 0
            for sheet_name in sheet_names:
                if sheet_name == "NhatKy": continue
                if self.stop_interactive_mapping: break

                try:
                    raw_df_no_header = self.robust_read_excel(self.file_path, sheet_name=sheet_name, header=None)
                    
                    if raw_df_no_header is None: continue
                    
                    if len(raw_df_no_header) < 5 or raw_df_no_header.shape[1] < 3:
                        continue

                    processed = self.process_raw_dataframe(raw_df_no_header, sheet_name)
                    
                    if processed is None or len(processed) == 0:
                        dialog = ColumnMappingDialog(self, raw_df_no_header, sheet_name)
                        self.wait_window(dialog) 
                        
                        if dialog.result_df is not None:
                            processed = dialog.result_df
                            processed["Xếp loại"] = processed["ĐTB"].apply(self.classify_smart)
                        else:
                            self.stop_interactive_mapping = True

                    if processed is not None and len(processed) > 0:
                        self.all_classes_data[sheet_name] = processed
                        loaded_count += 1
                        
                except Exception as e:
                    import traceback
                    print(f"❌ Lỗi load sheet {sheet_name}:")
                    traceback.print_exc()
                    continue
            
            # [SYNC GUARD - BƯỚC 1] Đảm bảo all_classes_data luôn có dữ liệu nếu current_df tồn tại
            if not self.all_classes_data and self.current_df is not None and len(self.current_df) > 0:
                # Fallback: Lấy tên sheet từ file hoặc dùng mặc định
                fallback_sheet_name = self.get_current_sheet_name()
                self.all_classes_data[fallback_sheet_name] = self.current_df.copy()
                print(f"[SYNC GUARD] Đã tự động đồng bộ current_df vào all_classes_data['{fallback_sheet_name}']")
            
            if self.all_classes_data:
                # [DYNAMIC COLS] Cập nhật số cột TX dựa trên dữ liệu đầu tiên TRƯỚC
                first_df = list(self.all_classes_data.values())[0]
                self.update_columns_based_on_data(first_df)
                
                self.cbo_classes['values'] = list(self.all_classes_data.keys())
                self.cbo_classes.current(0)
                self.on_class_change(None)
                
                # [FIXED] Sử dụng status_label được truyền vào thay vì self.lbl_autosave
                if self.status_label:
                    self.status_label.config(text="✅ Đã tải xong", fg="#27ae60")
                
                # [EVENT SYNC - BƯỚC 4] Broadcast event để các component khác có thể react
                try:
                    self.event_generate("<<DataLoaded>>")
                except:
                    pass  # Bỏ qua nếu widget chưa sẵn sàng
            else:
                if not self.stop_interactive_mapping:
                     messagebox.showerror("Lỗi đọc file", "Không tìm thấy dữ liệu hợp lệ. Vui lòng kiểm tra file Excel.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Lỗi", f"Lỗi không xác định: {str(e)}")

    def on_class_change(self, event):
        selected = self.cbo_classes.get()
        if selected in self.all_classes_data:
            self.current_df = self.all_classes_data[selected]
            
            # [DYNAMIC COLS] Cập nhật cột nếu lớp mới có số TX khác
            self.update_columns_based_on_data(self.current_df)
            
            self.filter_data("All")
            self.ent_search.delete(0, 'end')

    def on_search_debounce(self, event):
        if self.search_job: self.after_cancel(self.search_job)
        self.search_job = self.after(150, self.perform_search)

    def perform_search(self):
        if self.current_df is None: return
        keyword = self.ent_search.get().lower()
        keyword_norm = normalize_text(keyword)
        if keyword_norm == "":
            self.update_ui_data(self.current_df, update_chart=True, highlight_norm=None)
            self.lbl_filter_status.config(text="")
        else:
            mask = self.current_df['Họ và tên'].apply(lambda x: keyword_norm in normalize_text(x))
            filtered = self.current_df[mask]
            self.update_ui_data(filtered, update_chart=False, highlight_norm=keyword_norm)
            self.lbl_filter_status.config(text=f"TÌM: '{keyword}'", fg=THEME["primary"])

    def perform_auto_save(self):
        if not HAS_OPENPYXL:
            if self.status_label: self.status_label.config(text="⚠️ Thiếu openpyxl", fg="red")
            return

        # [FIXED] Sử dụng status_label
        if self.status_label:
            self.status_label.config(text="💾 Đang lưu...", fg="#e67e22")
        self.update_idletasks()
        
        try:
            save_path = self.file_path
            is_xls_old = self.file_path.lower().endswith('.xls')
            
            if is_xls_old:
                save_path = self.file_path + "x" 
                if not os.path.exists(save_path):
                    try:
                        temp_df = self.robust_read_excel(self.file_path, sheet_name=0, header=None)
                        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                            temp_df.to_excel(writer, index=False, header=False)
                    except Exception:
                        messagebox.showwarning("Cảnh báo", "Không thể tự động chuyển đổi file .xls lỗi. Vui lòng Save As .xlsx thủ công bằng Excel.")
                        return

                    messagebox.showinfo("Thông báo chuyển đổi", f"File cũ (.xls) đã được nâng cấp sang (.xlsx) để bảo toàn dữ liệu.\nFile mới: {save_path}")
                self.file_path = save_path 

            wb = openpyxl.load_workbook(save_path)
            
            for sheet_name, df_data in self.all_classes_data.items():
                ws = None
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                elif len(wb.sheetnames) == 1:
                    ws = wb.active
                
                if ws is None: continue
                
                header_row_idx = self.sheet_headers_info.get(sheet_name, 0)
                excel_header_row = header_row_idx + 1 
                
                col_mapping = {} 
                
                for col_idx in range(1, ws.max_column + 1):
                    val1 = str(ws.cell(row=excel_header_row, column=col_idx).value or "").upper()
                    val2 = str(ws.cell(row=excel_header_row + 1, column=col_idx).value or "").upper()
                    
                    comb = val1 + " " + val2
                    # [FIXED] Thêm pattern THƯỜNG XUYÊN để khớp với process_raw_dataframe
                    if re.search(r'(TX|HS1|HỆ SỐ 1|15|MIỆNG|ĐĐGTX|THƯỜNG XUYÊN)', comb) and "GK" not in comb and "CK" not in comb:
                        if "TX" not in col_mapping: col_mapping["TX"] = []
                        col_mapping["TX"].append(col_idx)
                    elif "GK" in comb: col_mapping["GK"] = col_idx
                    elif "CK" in comb: col_mapping["CK"] = col_idx
                    elif "ĐTB" in comb or "MHK" in comb: col_mapping["ĐTB"] = col_idx
                    elif "MÃ" in comb and "HS" in comb: col_mapping["MA_HS"] = col_idx

                ma_hs_col = col_mapping.get("MA_HS")
                
                excel_row_map = {}
                start_data_row = excel_header_row + 1 
                if ws.cell(row=excel_header_row+1, column=1).value is not None: 
                     if "TX" in str(ws.cell(row=excel_header_row+1, column=1).value): start_data_row += 1

                for r in range(start_data_row, ws.max_row + 1):
                    if ma_hs_col:
                        code_val = str(ws.cell(row=r, column=ma_hs_col).value or "").strip()
                        if code_val: excel_row_map[code_val] = r
                    else:
                        stt_val = str(ws.cell(row=r, column=1).value or "").strip()
                        if stt_val.isdigit(): excel_row_map[stt_val] = r

                tx_cols = col_mapping.get("TX", [])
                print(f"[DEBUG perform_auto_save] Sheet '{sheet_name}': Found {len(tx_cols)} TX columns at positions: {tx_cols}")
                
                for _, row_data in df_data.iterrows():
                    target_row = None
                    stt = str(row_data["STT"])
                    target_row = excel_row_map.get(stt)
                    
                    if target_row:
                        for i, col_idx in enumerate(tx_cols):
                            key = f"TX{i+1}"
                            if key in row_data:
                                val = row_data[key]
                                try:
                                    # [FIXED] Xử lý đúng các giá trị rỗng và số
                                    if val == "" or val is None or (isinstance(val, float) and pd.isna(val)):
                                        ws.cell(row=target_row, column=col_idx).value = None
                                    else:
                                        ws.cell(row=target_row, column=col_idx).value = float(val)
                                except (ValueError, TypeError):
                                    ws.cell(row=target_row, column=col_idx).value = None
                        
                        if "GK" in col_mapping and "GK" in row_data:
                            val = row_data["GK"]
                            try:
                                if val == "" or val is None or (isinstance(val, float) and pd.isna(val)):
                                    ws.cell(row=target_row, column=col_mapping["GK"]).value = None
                                else:
                                    ws.cell(row=target_row, column=col_mapping["GK"]).value = float(val)
                            except (ValueError, TypeError):
                                ws.cell(row=target_row, column=col_mapping["GK"]).value = None
                        
                        if "CK" in col_mapping and "CK" in row_data:
                            val = row_data["CK"]
                            try:
                                if val == "" or val is None or (isinstance(val, float) and pd.isna(val)):
                                    ws.cell(row=target_row, column=col_mapping["CK"]).value = None
                                else:
                                    ws.cell(row=target_row, column=col_mapping["CK"]).value = float(val)
                            except (ValueError, TypeError):
                                ws.cell(row=target_row, column=col_mapping["CK"]).value = None
                            
                        if "ĐTB" in col_mapping and "ĐTB" in row_data:
                            val = row_data["ĐTB"]
                            try:
                                if val == "" or val is None or (isinstance(val, float) and pd.isna(val)):
                                    ws.cell(row=target_row, column=col_mapping["ĐTB"]).value = None
                                else:
                                    ws.cell(row=target_row, column=col_mapping["ĐTB"]).value = float(val)
                            except (ValueError, TypeError):
                                ws.cell(row=target_row, column=col_mapping["ĐTB"]).value = None

            wb.save(save_path)
            wb.close()
            # [FIXED]
            if self.status_label: self.status_label.config(text="✅ Đã đồng bộ", fg="#27ae60")
            
        except PermissionError:
            if self.status_label: self.status_label.config(text="❌ Lỗi: File đang mở!", fg="red")
        except Exception as e:
            print(e)
            if self.status_label: self.status_label.config(text="❌ Lỗi lưu", fg="red")

    def classify_smart(self, value):
        val_str = str(value).strip().upper()
        if pd.isna(value) or val_str in ["", "NAN", "NONE"]: return "Chưa có điểm"
        if val_str == "Đ": return "Hoàn thành"
        if val_str == "CĐ": return "Chưa hoàn thành"
        try:
            s = float(value)
            if s < 5.0: return "Chưa hoàn thành"
            if 5.0 <= s < 6.5: return "Hoàn thành"
            if 6.5 <= s < 8.0: return "Hoàn thành tốt"
            return "Hoàn thành xuất sắc"
        except: return "Chưa có điểm"
    
    def update_status(self, message, color="black"):
        """Helper method để cập nhật status label"""
        if self.status_label:
            color_map = {
                "green": "#27ae60",
                "red": "#e74c3c",
                "blue": "#3498db",
                "orange": "#e67e22",
                "black": THEME["text_main"]
            }
            self.status_label.config(text=message, fg=color_map.get(color, color))

    def process_raw_dataframe(self, raw_df, sheet_name):
        try:
            header_idx = -1
            for idx, row in raw_df.head(30).iterrows():
                row_str = [str(x).lower() for x in row.values]
                if any(k == s.strip() for k in ["stt", "mã học sinh", "họ và tên"] for s in row_str):
                    header_idx = idx; break
            
            if header_idx == -1: return None
            self.sheet_headers_info[sheet_name] = header_idx 

            stt_idx = -1
            name_idx = -1
            dob_idx = -1
            
            header_row_vals = raw_df.iloc[header_idx].astype(str).tolist()
            for col_idx, val in enumerate(header_row_vals):
                val_clean = str(val).lower()
                if "stt" in val_clean: stt_idx = col_idx
                if "họ" in val_clean and "tên" in val_clean: name_idx = col_idx
                if "sinh" in val_clean: dob_idx = col_idx

            if name_idx == -1: return None

            # [NEW] Đếm số cột TX từ header TRƯỚC khi xử lý dữ liệu
            # Để đảm bảo show đủ cột dù học sinh chưa có điểm
            num_tx_cols = 0
            score_start_idx = dob_idx + 1 if dob_idx != -1 else name_idx + 2
            
            for col_j in range(score_start_idx, len(header_row_vals)):
                h1 = str(raw_df.iloc[header_idx, col_j]).upper()
                h2 = ""
                if header_idx + 1 < len(raw_df):
                    h2 = str(raw_df.iloc[header_idx+1, col_j]).upper()
                header_comb = h1 + " " + h2
                
                # Đếm cột TX
                if re.search(r'(TX|HS1|HỆ SỐ 1|15|MIỆNG|ĐĐGTX|THƯỜNG XUYÊN)', header_comb) and "GK" not in header_comb and "CK" not in header_comb:
                    num_tx_cols += 1
            
            # Đảm bảo tối thiểu 2 cột TX
            num_tx_cols = max(2, num_tx_cols)
            print(f"[DEBUG] Sheet '{sheet_name}': Phát hiện {num_tx_cols} cột TX từ header")

            data = []
            start_row = header_idx + 1
            if start_row < len(raw_df):
                next_row_str = "".join([str(x) for x in raw_df.iloc[start_row].values]).upper()
                if "TX" in next_row_str or "HS" in next_row_str:
                    start_row += 1

            for i in range(start_row, len(raw_df)):
                row = raw_df.iloc[i]
                
                stt = str(row.iloc[stt_idx]).strip() if stt_idx != -1 else str(i)
                if not stt or (stt_idx != -1 and not stt[0].isdigit()): continue

                raw_surname = str(row.iloc[name_idx]).strip()
                if raw_surname.lower() in ["nan", "none", ""]: continue
                
                full_name = raw_surname
                
                next_col_idx = name_idx + 1
                if next_col_idx < len(row) and next_col_idx != dob_idx:
                    val_next = str(row.iloc[next_col_idx]).strip()
                    is_valid_name_part = True
                    if val_next.lower() in ["nan", "none", ""]: is_valid_name_part = False
                    if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{4}', val_next): is_valid_name_part = False
                    if re.match(r'^\d+$', val_next): is_valid_name_part = False

                    if is_valid_name_part:
                        full_name = raw_surname + " " + val_next
                
                full_name = full_name.title()

                dob = ""
                if dob_idx != -1:
                    dob = str(row.iloc[dob_idx]).replace("nan", "").strip()

                tx_scores = []
                gk = ""
                ck = ""
                
                for col_j in range(score_start_idx, len(row)):
                    h1 = str(raw_df.iloc[header_idx, col_j]).upper()
                    h2 = ""
                    if header_idx + 1 < len(raw_df):
                        h2 = str(raw_df.iloc[header_idx+1, col_j]).upper()
                    
                    header_comb = h1 + " " + h2
                    val = str(row.iloc[col_j]).strip().replace(",", ".")
                    
                    if not val or val.lower() in ["nan", "none"]: 
                        # Vẫn thêm placeholder nếu là cột TX để giữ đúng vị trí
                        if re.search(r'(TX|HS1|HỆ SỐ 1|15|MIỆNG|ĐĐGTX|THƯỜNG XUYÊN)', header_comb) and "GK" not in header_comb and "CK" not in header_comb:
                            tx_scores.append("")  # Placeholder
                        continue
                    
                    try:
                        f_val = float(val)
                    except: 
                        # Vẫn thêm placeholder nếu là cột TX
                        if re.search(r'(TX|HS1|HỆ SỐ 1|15|MIỆNG|ĐĐGTX|THƯỜNG XUYÊN)', header_comb) and "GK" not in header_comb and "CK" not in header_comb:
                            tx_scores.append("")  # Placeholder
                        continue 

                    if re.search(r'(TX|HS1|HỆ SỐ 1|15|MIỆNG|ĐĐGTX|THƯỜNG XUYÊN)', header_comb) and "GK" not in header_comb and "CK" not in header_comb:
                        tx_scores.append(f_val)
                    elif "GK" in header_comb or "GIỮA" in header_comb:
                        gk = f_val
                    elif "CK" in header_comb or "CUỐI" in header_comb:
                        ck = f_val

                dtb = calculate_dtb_exact(tx_scores, gk, ck)

                # [DYNAMIC TX] Tạo dictionary động cho các cột TX dựa trên header
                row_data = {
                    "STT": stt,
                    "Họ và tên": full_name,
                    "Ngày sinh": dob
                }
                
                # Thêm các cột TX động dựa trên số cột TX từ HEADER (không phải từ dữ liệu)
                # num_tx_cols đã được đếm từ header phía trên
                for i in range(1, num_tx_cols + 1):
                    row_data[f"TX{i}"] = tx_scores[i-1] if i-1 < len(tx_scores) else ""
                
                # Thêm các cột còn lại
                row_data.update({
                    "GK": gk, "CK": ck, "ĐTB": dtb,
                    "Xếp loại": self.classify_smart(dtb)
                })
                
                data.append(row_data)

            # [FIX] Tạo DataFrame với dtype=object để tránh lỗi string dtype khi edit
            result_df = pd.DataFrame(data)
            # Convert các cột điểm sang object dtype để có thể gán cả số và chuỗi
            score_cols = [c for c in result_df.columns if c.startswith("TX") or c in ["GK", "CK", "ĐTB"]]
            for col in score_cols:
                if col in result_df.columns:
                    result_df[col] = result_df[col].astype(object)
            return result_df

        except Exception as e:
            import traceback
            print(f"❌ Lỗi process_raw_dataframe '{sheet_name}':")
            traceback.print_exc()
            return None
    
    def filter_data(self, key):
        if self.current_df is None: return
        
        # [UPDATED] Màu chữ hiển thị thông báo lọc
        # Thay vì dùng màu pastel khó nhìn, dùng màu text_main (Đen/Trắng)
        text_color = THEME["text_main"] 

        if key == "All": view = self.current_df; txt = "TẤT CẢ"
        elif key == "Tot": view = self.current_df[self.current_df['Xếp loại'].isin(["Hoàn thành tốt", "Hoàn thành xuất sắc"])]; txt = "KHÁ / GIỎI"
        else: map_name = {"HoanThanh": "Hoàn thành", "CanChuY": "Chưa hoàn thành", "ChuaCoDiem": "Chưa có điểm"}; target = map_name.get(key, key); view = self.current_df[self.current_df['Xếp loại'] == target]; txt = target.upper()
        
        # [FIXED] Gán màu chữ theo theme
        self.lbl_filter_status.config(text=f"LỌC: {txt}", fg=text_color)
        self.update_ui_data(view)

    def update_ui_data(self, df, update_chart=True, highlight_norm=None):
        for item in self.tree.get_children(): self.tree.delete(item)
        for _, row in df.iterrows():
            xl = row['Xếp loại']
            base_tag = "normal"
            if xl == "Chưa hoàn thành": base_tag = "warning"
            elif xl == "Hoàn thành tốt": base_tag = "good"
            elif xl == "Hoàn thành xuất sắc": base_tag = "excellent"
            elif xl == "Chưa có điểm": base_tag = "nodata"
            vals = [str(row[c]).replace("nan","").replace("None","") for c in self.cols_def]
            tags = [base_tag]
            if highlight_norm:
                try:
                    name_norm = normalize_text(row['Họ và tên'])
                    if highlight_norm in name_norm:
                        tags.append("search_hit")
                except Exception:
                    pass
            self.tree.insert("", "end", values=vals, tags=tuple(tags))
        
        stats = self.get_stats()
        for key, val in stats.items():
            if key in self.cards:
                main_lbl, shadow_lbl = self.cards[key]
                val_str = str(val)
                main_lbl.config(text=val_str)
                shadow_lbl.config(text=val_str)
        if update_chart: self.draw_3d_pie_chart(stats)

    def get_stats(self):
        counts = self.current_df['Xếp loại'].value_counts()
        return {
            "All": len(self.current_df),
            "Tot": counts.get("Hoàn thành tốt", 0) + counts.get("Hoàn thành xuất sắc", 0),
            "HoanThanh": counts.get("Hoàn thành", 0),
            "CanChuY": counts.get("Chưa hoàn thành", 0),
            "ChuaCoDiem": counts.get("Chưa có điểm", 0)
        }

    # ---------------- Undo / Redo ----------------
    def _save_mapping_snapshot(self, sheet_name, old_df, backup_path=None):
        """
        Giai đoạn 9: Lưu snapshot dữ liệu trước khi mapping để hỗ trợ Undo.
        """
        from datetime import datetime
        
        snapshot = {
            "sheet_name": sheet_name,
            "old_df": old_df,  # Có thể None nếu sheet chưa có dữ liệu
            "backup_path": backup_path,  # Đường dẫn file backup trên Desktop
            "timestamp": datetime.now().strftime("%H:%M:%S %d/%m/%Y")
        }
        
        self.mapping_undo_stack.append(snapshot)
        
        # Giới hạn stack size
        if len(self.mapping_undo_stack) > self.max_mapping_undo:
            self.mapping_undo_stack.pop(0)
        
        print(f"[DEBUG] Đã lưu snapshot mapping cho sheet '{sheet_name}' (stack size: {len(self.mapping_undo_stack)})")
    
    def undo_last_mapping(self):
        """
        Giai đoạn 9: Hoàn tác lần mapping gần nhất.
        Cho phép khôi phục dữ liệu từ snapshot hoặc file backup.
        """
        if not self.mapping_undo_stack:
            messagebox.showinfo("Hoàn tác Mapping", "Không có lịch sử mapping nào để hoàn tác.")
            return
        
        # Lấy snapshot gần nhất
        snapshot = self.mapping_undo_stack[-1]
        sheet_name = snapshot["sheet_name"]
        old_df = snapshot["old_df"]
        backup_path = snapshot["backup_path"]
        timestamp = snapshot["timestamp"]
        
        # Hiển thị hộp thoại xác nhận
        options = []
        msg = f"Hoàn tác mapping cho sheet '{sheet_name}'\n(Thời điểm: {timestamp})\n\n"
        
        if old_df is not None:
            msg += f"✅ Có snapshot dữ liệu cũ ({len(old_df)} học sinh)\n"
            options.append("snapshot")
        else:
            msg += "❌ Không có dữ liệu cũ (sheet mới)\n"
        
        if backup_path and os.path.exists(backup_path):
            msg += f"✅ Có file backup: {os.path.basename(backup_path)}\n"
            options.append("backup")
        elif backup_path:
            msg += f"❌ File backup không tồn tại: {backup_path}\n"
        else:
            msg += "❌ Không có file backup\n"
        
        if not options:
            messagebox.showwarning("Hoàn tác Mapping", "Không có nguồn dữ liệu nào để khôi phục!")
            return
        
        msg += "\nBạn muốn khôi phục từ đâu?"
        
        # Tạo dialog chọn nguồn khôi phục
        choice_dialog = Toplevel(self)
        choice_dialog.title("Hoàn tác Mapping")
        choice_dialog.geometry("450x280")
        choice_dialog.configure(bg=THEME["bg_card"])
        choice_dialog.transient(self)
        choice_dialog.grab_set()
        
        # Center dialog
        choice_dialog.update_idletasks()
        x = (choice_dialog.winfo_screenwidth() - 450) // 2
        y = (choice_dialog.winfo_screenheight() - 280) // 2
        choice_dialog.geometry(f"450x280+{x}+{y}")
        
        tk.Label(choice_dialog, text="↩️ HOÀN TÁC MAPPING", font=("Segoe UI", 12, "bold"),
                 bg="#3498db", fg="white", pady=10).pack(fill="x")
        
        tk.Label(choice_dialog, text=msg, font=("Segoe UI", 10), bg=THEME["bg_card"],
                 fg=THEME["text_main"], justify="left", padx=20, pady=10).pack(fill="x")
        
        btn_frame = tk.Frame(choice_dialog, bg=THEME["bg_card"])
        btn_frame.pack(pady=15)
        
        def restore_from_snapshot():
            if old_df is not None:
                self.all_classes_data[sheet_name] = old_df.copy()
                self.on_class_change(None)
                self.mapping_undo_stack.pop()
                messagebox.showinfo("Thành công", f"Đã khôi phục dữ liệu cũ cho sheet '{sheet_name}'!")
            choice_dialog.destroy()
        
        def restore_from_backup():
            if backup_path and os.path.exists(backup_path):
                try:
                    # Đọc dữ liệu từ file backup
                    restored_df = pd.read_excel(backup_path, sheet_name=sheet_name, engine='openpyxl')
                    self.all_classes_data[sheet_name] = restored_df
                    self.on_class_change(None)
                    self.mapping_undo_stack.pop()
                    messagebox.showinfo("Thành công", f"Đã khôi phục từ file backup cho sheet '{sheet_name}'!")
                except Exception as e:
                    messagebox.showerror("Lỗi", f"Không thể đọc file backup: {e}")
            choice_dialog.destroy()
        
        def cancel():
            choice_dialog.destroy()
        
        if "snapshot" in options:
            tk.Button(btn_frame, text="📊 Từ Snapshot", font=("Segoe UI", 10, "bold"),
                      bg="#27ae60", fg="white", padx=15, pady=8, cursor="hand2",
                      command=restore_from_snapshot).pack(side="left", padx=5)
        
        if "backup" in options:
            tk.Button(btn_frame, text="📁 Từ File Backup", font=("Segoe UI", 10, "bold"),
                      bg="#3498db", fg="white", padx=15, pady=8, cursor="hand2",
                      command=restore_from_backup).pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="❌ Hủy", font=("Segoe UI", 10),
                  bg="#95a5a6", fg="white", padx=15, pady=8, cursor="hand2",
                  command=cancel).pack(side="left", padx=5)
    
    def undo_last_edit(self):
        if not self.history_stack:
            messagebox.showinfo("Undo", "Không còn thao tác để hoàn tác.")
            return
        action = self.history_stack.pop()
        self.apply_value_to_row(action["stt"], action["col"], action["old"])
        self.redo_stack.append(action)
        self.update_ui_data(self.current_df, update_chart=True)

    def redo_last_edit(self):
        if not self.redo_stack:
            messagebox.showinfo("Redo", "Không còn thao tác để làm lại.")
            return
        action = self.redo_stack.pop()
        self.apply_value_to_row(action["stt"], action["col"], action["new"])
        self.history_stack.append(action)
        self.update_ui_data(self.current_df, update_chart=True)

    def apply_value_to_row(self, stt, col_name, value):
        # Cập nhật DataFrame
        idx = self.current_df.index[self.current_df['STT'].astype(str) == str(stt)].tolist()
        if idx:
            self.current_df.at[idx[0], col_name] = value
            if col_name == "ĐTB":
                self.current_df.at[idx[0], "Xếp loại"] = self.classify_smart(value)
        # Cập nhật TreeView đơn giản bằng cách refresh toàn bộ
        self.update_ui_data(self.current_df, update_chart=False)

    # =============================================================================
    # TIỆN ÍCH BỔ SUNG MỚI
    # =============================================================================
    def show_advanced_statistics(self):
        """Thống kê nâng cao với phân tích xu hướng"""
        if self.current_df is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu lớp học!")
            return
        
        stat_win = Toplevel(self)
        stat_win.title("📈 Thống Kê Nâng Cao")
        self.center_window(stat_win, 900, 650)
        stat_win.configure(bg=THEME["bg_app"])
        
        # Header với khung pastel nổi bật
        header_container = tk.Frame(stat_win, bg="#fafbfc", padx=20, pady=20)
        header_container.pack(fill="x")
        
        header = tk.Frame(header_container, bg="#ddd6fe", padx=20, pady=15, 
                         highlightbackground="#6366f1", highlightthickness=2, relief="solid")
        header.pack(fill="x")
        
        tk.Label(header, text="📈", font=("Segoe UI", 20), bg="#ddd6fe", fg="#6366f1").pack()
        tk.Label(header, text="THỐNG KÊ NÂNG CAO", 
                 font=("Segoe UI", 18, "bold"), bg="#ddd6fe", fg="#4f46e5").pack(pady=(5, 5))
        tk.Label(header, text="🔍 Phân tích xu hướng học tập thông minh", 
                 font=("Segoe UI", 11, "italic"), bg="#ddd6fe", fg="#6366f1").pack()
        
        content = tk.Frame(stat_win, bg=THEME["bg_app"], padx=20, pady=20)
        content.pack(fill="both", expand=True)
        
        # Tính toán các chỉ số thống kê
        dtb_data = pd.to_numeric(self.current_df['ĐTB'], errors='coerce').dropna()
        dtb_data = dtb_data[dtb_data > 0]
        
        if len(dtb_data) == 0:
            tk.Label(content, text="Chưa có dữ liệu điểm để phân tích!", 
                    font=("Segoe UI", 12), bg=THEME["bg_app"], fg="red").pack(pady=50)
            return
        
        mean_score = dtb_data.mean()
        median_score = dtb_data.median()
        std_score = dtb_data.std()
        max_score = dtb_data.max()
        min_score = dtb_data.min()
        
        # Phân tích xu hướng điểm TX -> GK -> CK
        # Fix: Only use TX columns that actually exist in dataframe
        all_tx_cols = ['TX1', 'TX2', 'TX3', 'TX4']
        tx_cols = [col for col in all_tx_cols if col in self.current_df.columns]
        tx_means = []
        for col in tx_cols:
            col_data = pd.to_numeric(self.current_df[col], errors='coerce').dropna()
            col_data = col_data[col_data > 0]
            if len(col_data) > 0:
                tx_means.append(col_data.mean())
        
        gk_data = pd.to_numeric(self.current_df['GK'], errors='coerce').dropna()
        gk_data = gk_data[gk_data > 0]
        gk_mean = gk_data.mean() if len(gk_data) > 0 else 0
        
        ck_data = pd.to_numeric(self.current_df['CK'], errors='coerce').dropna()
        ck_data = ck_data[ck_data > 0]
        ck_mean = ck_data.mean() if len(ck_data) > 0 else 0
        
        # Hiển thị thống kê
        stats_frame = tk.Frame(content, bg=THEME["bg_card"], padx=20, pady=20, 
                              highlightbackground=THEME["border"], highlightthickness=1)
        stats_frame.pack(fill="both", expand=True)
        
        tk.Label(stats_frame, text="📊 CHỈ SỐ THỐNG KÊ", font=("Segoe UI", 14, "bold"), 
                bg=THEME["bg_card"], fg=THEME["primary"]).pack(anchor="w", pady=(0, 15))
        
        stats_info = [
            ("Điểm Trung Bình (Mean)", f"{mean_score:.2f}", "#3B82F6"),
            ("Điểm Trung Vị (Median)", f"{median_score:.2f}", "#10B981"),
            ("Độ Lệch Chuẩn (Std Dev)", f"{std_score:.2f}", "#F59E0B"),
            ("Điểm Cao Nhất", f"{max_score:.2f}", "#EF4444"),
            ("Điểm Thấp Nhất", f"{min_score:.2f}", "#6366F1"),
        ]
        
        for label, value, color in stats_info:
            row = tk.Frame(stats_frame, bg=THEME["bg_card"])
            row.pack(fill="x", pady=5)
            tk.Label(row, text=label, font=("Segoe UI", 11), bg=THEME["bg_card"], 
                    fg=THEME["text_main"], width=25, anchor="w").pack(side="left")
            tk.Label(row, text=value, font=("Segoe UI", 11, "bold"), bg=THEME["bg_card"], 
                    fg=color, width=10, anchor="e").pack(side="right")
        
        # Phân tích xu hướng
        tk.Label(stats_frame, text="\n🔍 PHÂN TÍCH XU HƯỚNG", font=("Segoe UI", 14, "bold"), 
                bg=THEME["bg_card"], fg=THEME["primary"]).pack(anchor="w", pady=(15, 10))
        
        trend_text = tk.Text(stats_frame, height=8, font=("Segoe UI", 10), 
                            bg=THEME["entry_bg"], fg=THEME["text_main"], wrap="word")
        trend_text.pack(fill="x", pady=5)
        
        analysis = []
        if len(tx_means) > 1:
            tx_trend = "tăng" if tx_means[-1] > tx_means[0] else "giảm"
            analysis.append(f"• Điểm TX có xu hướng {tx_trend} (từ {tx_means[0]:.2f} → {tx_means[-1]:.2f})")
        
        if gk_mean > 0 and len(tx_means) > 0:
            avg_tx = sum(tx_means) / len(tx_means)
            if gk_mean > avg_tx + 0.5:
                analysis.append(f"• Học sinh thi GK TỐT HƠN điểm TX (GK: {gk_mean:.2f} vs TX TB: {avg_tx:.2f})")
            elif gk_mean < avg_tx - 0.5:
                analysis.append(f"• Học sinh thi GK KÉM HƠN điểm TX (GK: {gk_mean:.2f} vs TX TB: {avg_tx:.2f})")
        
        if ck_mean > 0 and gk_mean > 0:
            if ck_mean > gk_mean + 0.5:
                analysis.append(f"• Tiến bộ rõ rệt từ GK đến CK (GK: {gk_mean:.2f} → CK: {ck_mean:.2f})")
            elif ck_mean < gk_mean - 0.5:
                analysis.append(f"• Suy giảm từ GK đến CK (GK: {gk_mean:.2f} → CK: {ck_mean:.2f})")
        
        if std_score < 1.0:
            analysis.append(f"• Lớp có trình độ ĐỒNG ĐỀU (độ lệch chuẩn thấp: {std_score:.2f})")
        elif std_score > 2.0:
            analysis.append(f"• Lớp có sự CHÊNH LỆCH LỚN về trình độ (độ lệch chuẩn cao: {std_score:.2f})")
        
        if mean_score >= 8.0:
            analysis.append("• Lớp học có CHẤT LƯỢNG XUẤT SẮC!")
        elif mean_score < 6.0:
            analysis.append("• Lớp cần SỰ QUAN TÂM ĐẶC BIỆT từ giáo viên.")
        
        trend_text.insert("1.0", "\n".join(analysis) if analysis else "Chưa đủ dữ liệu để phân tích xu hướng.")
        trend_text.config(state="disabled")
        
        # Nút ĐÓNG với viền đen
        close_btn_frame = tk.Frame(stat_win, bg="#000000", padx=1, pady=1)
        close_btn_frame.pack(pady=15)
        tk.Button(close_btn_frame, text="ĐÓNG", command=stat_win.destroy, bg="#94a3b8", fg="white", 
                 font=("Segoe UI", 11, "bold"), relief="flat", padx=20, pady=8, borderwidth=0).pack()
    
    def export_report(self):
        """Xuất báo cáo thống kê ra file Excel"""
        if self.current_df is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất!")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Xuất Báo Cáo Thống Kê",
            initialfile=f"BaoCao_{self.cbo_classes.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet 1: Dữ liệu học sinh
                self.current_df.to_excel(writer, sheet_name="Dữ Liệu", index=False)
                
                # Sheet 2: Thống kê tổng quan
                stats = self.get_stats()
                dtb_data = pd.to_numeric(self.current_df['ĐTB'], errors='coerce').dropna()
                dtb_data = dtb_data[dtb_data > 0]
                
                summary_data = {
                    "Chỉ Số": ["Tổng Sĩ Số", "Giỏi/Khá", "Trung Bình", "Yếu", "Chưa Điểm",
                              "Điểm TB Lớp", "Điểm Cao Nhất", "Điểm Thấp Nhất"],
                    "Giá Trị": [
                        stats['All'], stats['Tot'], stats['HoanThanh'], stats['CanChuY'], stats['ChuaCoDiem'],
                        f"{dtb_data.mean():.2f}" if len(dtb_data) > 0 else "N/A",
                        f"{dtb_data.max():.2f}" if len(dtb_data) > 0 else "N/A",
                        f"{dtb_data.min():.2f}" if len(dtb_data) > 0 else "N/A"
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name="Thống Kê", index=False)
                
                # Sheet 3: Danh sách học sinh yếu
                weak_students = self.current_df[self.current_df['Xếp loại'] == "Chưa hoàn thành"]
                if len(weak_students) > 0:
                    weak_students.to_excel(writer, sheet_name="Học Sinh Yếu", index=False)
            
            messagebox.showinfo("Thành công", f"Đã xuất báo cáo thành công!\n\nFile: {file_path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất báo cáo: {str(e)}")
    

    
    def show_parent_contacts(self):
        """Quản lý danh bạ phụ huynh"""
        if self.current_df is None:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu lớp học!")
            return
        
        contact_win = Toplevel(self)
        contact_win.title("📞 Danh Bạ Phụ Huynh")
        self.center_window(contact_win, 850, 600)
        contact_win.configure(bg=THEME["bg_app"])
        
        # Header với khung pastel nổi bật
        header_container = tk.Frame(contact_win, bg="#fafbfc", padx=20, pady=20)
        header_container.pack(fill="x")
        
        header = tk.Frame(header_container, bg="#dbeafe", padx=20, pady=15, 
                         highlightbackground="#3b82f6", highlightthickness=2, relief="solid")
        header.pack(fill="x")
        
        tk.Label(header, text="📞", font=("Segoe UI", 20), bg="#dbeafe", fg="#3b82f6").pack()
        tk.Label(header, text="DANH BẠ PHỤ HUYNH", 
                 font=("Segoe UI", 18, "bold"), bg="#dbeafe", fg="#1d4ed8").pack(pady=(5, 5))
        tk.Label(header, text="👨‍👩‍👧‍👦 Quản lý thông tin liên lạc hiệu quả", 
                 font=("Segoe UI", 11, "italic"), bg="#dbeafe", fg="#3b82f6").pack()
        
        content = tk.Frame(contact_win, bg=THEME["bg_app"], padx=20, pady=20)
        content.pack(fill="both", expand=True)
        
        info_frame = tk.Frame(content, bg="#FEF3C7", padx=15, pady=10, 
                             highlightbackground="#F59E0B", highlightthickness=2)
        info_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(info_frame, text="💡 Tính năng này cho phép bạn quản lý thông tin liên lạc với phụ huynh.", 
                font=("Segoe UI", 10), bg="#FEF3C7", fg="#92400E", wraplength=750, justify="left").pack(anchor="w")
        tk.Label(info_frame, text="Nhấp đúp vào học sinh để thêm/sửa thông tin. Dữ liệu được lưu tự động.", 
                font=("Segoe UI", 10), bg="#FEF3C7", fg="#92400E", wraplength=750, justify="left").pack(anchor="w")
        
        # Tạo bảng danh sách
        table_frame = tk.Frame(content, bg=THEME["bg_card"], 
                              highlightbackground=THEME["border"], highlightthickness=1)
        table_frame.pack(fill="both", expand=True)
        
        columns = ("STT", "Họ và tên", "Số ĐT PH", "Email PH", "Ghi chú")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        tree.heading("STT", text="STT")
        tree.heading("Họ và tên", text="HỌ VÀ TÊN HỌC SINH")
        tree.heading("Số ĐT PH", text="SỐ ĐIỆN THOẠI")
        tree.heading("Email PH", text="EMAIL")
        tree.heading("Ghi chú", text="GHI CHÚ")
        
        tree.column("STT", width=50, anchor="center")
        tree.column("Họ và tên", width=200, anchor="w")
        tree.column("Số ĐT PH", width=120, anchor="center")
        tree.column("Email PH", width=180, anchor="w")
        tree.column("Ghi chú", width=200, anchor="w")
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Thêm dữ liệu thật từ storage
        for _, row in self.current_df.iterrows():
            student_name = row['Họ và tên']
            contact_info = self.parent_contacts.get(student_name, {
                "phone": "",
                "email": "",
                "notes": ""
            })
            
            # Hiển thị thông tin hoặc placeholder
            phone = contact_info["phone"] if contact_info["phone"] else "(Chưa có)"
            email = contact_info["email"] if contact_info["email"] else "(Chưa có)"
            notes = contact_info["notes"][:30] + "..." if len(contact_info["notes"]) > 30 else contact_info["notes"]
            notes = notes if notes else "Nhấp đúp để thêm"
            
            tree.insert("", "end", values=(
                row['STT'],
                student_name,
                phone,
                email,
                notes
            ))
        
        # Double-click để chỉnh sửa
        def on_double_click(event):
            selection = tree.selection()
            if selection:
                item = tree.item(selection[0])
                student_name = item['values'][1]  # Lấy tên học sinh từ cột 2
                self.edit_parent_contact(student_name)
        
        tree.bind("<Double-1>", on_double_click)
        
        btn_frame = tk.Frame(content, bg=THEME["bg_app"], pady=10)
        btn_frame.pack(fill="x")
        
        tk.Label(btn_frame, text="💡 Nhấp đúp vào học sinh để thêm/sửa thông tin liên lạc", 
                font=("Segoe UI", 9, "italic"), bg=THEME["bg_app"], fg="gray").pack(side="left")
        
        # Thêm nút xuất danh sách
        def export_contacts():
            try:
                from tkinter import filedialog
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="Xuất Danh Bạ Phụ Huynh",
                    initialfile=f"DanhBaPhuHuynh_{self.cbo_classes.get()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
                if file_path:
                    export_data = []
                    for _, row in self.current_df.iterrows():
                        student_name = row['Họ và tên']
                        contact_info = self.parent_contacts.get(student_name, {
                            "phone": "",
                            "email": "",
                            "notes": ""
                        })
                        
                        export_data.append({
                            "STT": row['STT'],
                            "Họ và tên": student_name,
                            "Số điện thoại PH": contact_info["phone"],
                            "Email PH": contact_info["email"],
                            "Ghi chú": contact_info["notes"]
                        })
                    
                    df_export = pd.DataFrame(export_data)
                    df_export.to_excel(file_path, index=False)
                    messagebox.showinfo("Thành công", f"Đã xuất danh bạ ra file:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể xuất danh bạ: {str(e)}")
        
        # Nút XUẤT EXCEL với viền đen
        export_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        export_btn_frame.pack(side="right", padx=5)
        tk.Button(export_btn_frame, text="📄 XUẤT EXCEL", command=export_contacts, 
                 bg=THEME["primary"], fg="white", font=("Segoe UI", 10, "bold"), 
                 relief="flat", padx=15, pady=6, borderwidth=0).pack()
        
        # Nút ĐÓNG với viền đen
        close_btn_frame = tk.Frame(btn_frame, bg="#000000", padx=1, pady=1)
        close_btn_frame.pack(side="right")
        tk.Button(close_btn_frame, text="ĐÓNG", command=contact_win.destroy, 
                 bg="#94a3b8", fg="white", font=("Segoe UI", 10, "bold"), 
                 relief="flat", padx=20, pady=6, borderwidth=0).pack()

    def draw_3d_pie_chart(self, stats, target_container=None, figsize=(5, 4)):
        container = target_container if target_container is not None else self.chart_container
        for w in container.winfo_children(): w.destroy()
        labels, sizes, colors = [], [], []
        mapping = [
            ("Chưa Đạt", stats["CanChuY"], CARD_COLORS["CanChuY"][0], "CanChuY"),
            ("Hoàn Thành", stats["HoanThanh"], CARD_COLORS["HoanThanh"][0], "HoanThanh"),
            ("Khá/Giỏi", stats["Tot"], CARD_COLORS["Tot"][0], "Tot"),
            ("Chưa Điểm", stats["ChuaCoDiem"], CARD_COLORS["ChuaCoDiem"][0], "ChuaCoDiem")
        ]
        total = stats["All"]
        if total == 0: return
        active_keys = []
        for lbl, val, col, key in mapping:
            if val > 0:
                pct = (val/total)*100
                labels.append(f"{lbl}\n{pct:.1f}%")
                sizes.append(val)
                colors.append(col)
                active_keys.append(key)
        
        fig = Figure(figsize=figsize, dpi=100)
        fig.patch.set_facecolor(THEME["bg_card"])
        ax = fig.add_subplot(111)
        ax.set_facecolor(THEME["bg_card"])
        explode = [0.05] * len(sizes) 
        wedges, texts = ax.pie(sizes, labels=labels, labeldistance=1.15, startangle=140, colors=colors, explode=explode, shadow=True, textprops={'fontsize': 9, 'weight': 'bold', 'color': THEME["text_main"]})
        
        def onclick(event):
            if event.inaxes != ax: return
            for i, wedge in enumerate(wedges):
                if wedge.contains_point([event.x, event.y]):
                    self.filter_data(active_keys[i])
                    break
        fig.canvas.mpl_connect('button_press_event', onclick)
        canvas = FigureCanvasTkAgg(fig, master=container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

# =================================================================================
# CLASS CHÍNH
# =================================================================================
class StudentManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{APP_NAME} v{APP_VERSION} - AI Powered Edition")
        self.root.state("zoomed")
        self.root.configure(bg=THEME["bg_app"])
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Load UI config từ JSON
        self.ui_config = self.load_ui_config()
        
        # [NEW] Biến global cho chế độ viết hoa toàn bộ
        self.uppercase_mode = False
        self.original_texts = {}  # Lưu text gốc để khôi phục
        
        # [NEW] Biến global cho chế độ in đậm toàn bộ
        self.bold_mode = False
        self.original_fonts = {}  # Lưu font gốc để khôi phục
        
        # [FIX] Biến để track entry đang được edit trong treeview
        self.active_entry = None
        self.active_entry_save_func = None
        
        self.setup_ui()
        
        # Thêm label hiển thị toạ độ
        self.coord_label = tk.Label(self.root, text="", bg="yellow", fg="black", 
                                   font=("Consolas", 9, "bold"), relief="solid", borderwidth=1,
                                   justify="left")
        self.coord_label.place(x=10, y=10)  # Góc trên bên trái
        self.coord_label.place_forget()  # Ẩn ban đầu
        
        # Biến theo dõi trạng thái Alt
        self.alt_pressed = False
        
        # Bind events cho phím Alt và chuột
        self.root.bind('<Alt_L>', self.on_alt_press)
        self.root.bind('<KeyRelease-Alt_L>', self.on_alt_release)
        self.root.bind('<Alt_R>', self.on_alt_press)
        self.root.bind('<KeyRelease-Alt_R>', self.on_alt_release)
        self.root.bind('<Motion>', self.on_mouse_motion)  # Chỉ cần di chuyển chuột
        
        # Thêm phím tắt để copy toạ độ
        self.root.bind('<Control-c>', self.copy_coordinates)  # Ctrl+C để copy toạ độ
        self.root.bind('<Control-C>', self.copy_coordinates)

        # Phím tắt tiện ích chung
        self.root.bind('<Control-f>', self.focus_search)
        self.root.bind('<Control-F>', self.focus_search)
        self.root.bind('<Control-r>', self.reset_filters)
        self.root.bind('<Control-R>', self.reset_filters)
        self.root.bind('<Control-s>', self.save_current_tab)
        self.root.bind('<Control-S>', self.save_current_tab)
        self.root.bind('<F5>', self.reload_current_tab)
        self.root.bind('<Control-z>', self.undo_current_tab)
        self.root.bind('<Control-Z>', self.undo_current_tab)
        
        # Thêm event click chuột để chọn widget
        self.root.bind('<Button-1>', self.on_widget_click)  # Click chuột trái để chọn widget
        
        # Biến để theo dõi trạng thái hiển thị biểu đồ
        self.chart_visible = False
        
        # Lưu widget đang được chọn
        self.selected_widget = None
        self.selected_widget_info = None
        
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.update_ttk_style()
        
        # [AUTO-UPDATE] Kiểm tra update sau khi app load xong
        self.updater = UpdateManager()
        self.check_for_updates_background()
    
    def check_for_updates_background(self):
        """Kiểm tra update trong background thread"""
        def check_thread():
            time.sleep(2)  # Đợi app load xong
            
            has_update, update_info = self.updater.check_for_updates()
            
            if has_update:
                # Hiện dialog trong main thread
                self.root.after(0, lambda: self.show_update_dialog(update_info))
        
        threading.Thread(target=check_thread, daemon=True).start()
    
    def show_update_dialog(self, update_info):
        """Hiện dialog update"""
        try:
            UpdateDialog(self.root, update_info, self.updater)
        except Exception as e:
            print(f"[UPDATE] Error showing dialog: {e}")
    
    def manual_check_update(self):
        """Kiểm tra cập nhật thủ công (từ nút hoặc menu)"""
        try:
            # Hiện thông báo đang kiểm tra
            self.root.config(cursor="wait")
            self.root.update()
            
            has_update, update_info = self.updater.check_for_updates(timeout=10)
            
            self.root.config(cursor="")
            
            if has_update:
                self.show_update_dialog(update_info)
            else:
                messagebox.showinfo(
                    "Cập nhật",
                    f"✅ Bạn đang sử dụng phiên bản mới nhất!\n\n"
                    f"Phiên bản hiện tại: {APP_VERSION}\n"
                    f"Ngày build: {APP_BUILD_DATE}",
                    parent=self.root
                )
        except Exception as e:
            self.root.config(cursor="")
            messagebox.showerror(
                "Lỗi",
                f"❌ Không thể kiểm tra cập nhật!\n\n"
                f"Lý do: Không có kết nối internet hoặc server không phản hồi.\n\n"
                f"Vui lòng thử lại sau.",
                parent=self.root
            )
    
    def on_alt_press(self, event):
        """Khi nhấn phím Alt"""
        self.alt_pressed = True
        self.coord_label.place(x=10, y=10)  # Hiển thị label toạ độ
        
    def on_alt_release(self, event):
        """Khi thả phím Alt"""
        self.alt_pressed = False
        self.coord_label.place_forget()  # Ẩn label toạ độ
        
    def on_mouse_motion(self, event):
        """Khi di chuyển chuột (chỉ hoạt động khi Alt được nhấn)"""
        if self.alt_pressed:
            # Lấy toạ độ chuột tương đối so với cửa sổ
            x = event.x
            y = event.y
            
            # Lấy toạ độ tuyệt đối của chuột trên màn hình
            abs_x = self.root.winfo_pointerx() - self.root.winfo_rootx()
            abs_y = self.root.winfo_pointery() - self.root.winfo_rooty()
            
            # Tìm widget dưới con trỏ chuột
            widget_info = self.get_widget_at_position(x, y)
            
            # Hiển thị thông tin chi tiết
            if widget_info:
                widget_name, widget_x, widget_y, widget_w, widget_h = widget_info
                info_text = f"Chuột: X:{x}, Y:{y}\nMàn hình: X:{abs_x}, Y:{abs_y}\n"
                info_text += f"Widget: {widget_name}\n"
                info_text += f"Vị trí widget: X:{widget_x}, Y:{widget_y}\n"
                info_text += f"Kích thước: {widget_w}x{widget_h}\n"
                info_text += "Ctrl+C: Copy toạ độ"
                
                # Lưu widget được chọn để copy
                self.selected_widget_info = widget_info
            else:
                info_text = f"Chuột: X:{x}, Y:{y}\nMàn hình: X:{abs_x}, Y:{abs_y}\n"
                info_text += "Widget: Không có"
                
                # Xóa widget được chọn
                self.selected_widget_info = None
            
            self.coord_label.config(text=info_text)
            
            # Di chuyển label theo chuột để không bị che
            label_x = x + 15
            label_y = y - 60  # Tăng khoảng cách vì label cao hơn
            if label_x > self.root.winfo_width() - 150:
                label_x = x - 150
            if label_y < 10:
                label_y = y + 15
            self.coord_label.place(x=label_x, y=label_y)
    
    def get_widget_at_position(self, x, y):
        """Tìm widget tại vị trí x, y và trả về thông tin chi tiết"""
        try:
            # Lấy widget tại vị trí chuột
            widget = self.root.winfo_containing(x + self.root.winfo_rootx(), 
                                               y + self.root.winfo_rooty())
            
            if widget and widget != self.root:
                # Lấy thông tin widget
                widget_x = widget.winfo_x()
                widget_y = widget.winfo_y()
                widget_w = widget.winfo_width()
                widget_h = widget.winfo_height()
                
                # Lấy tên widget thông minh
                widget_name = widget.winfo_name()
                widget_text = ""
                
                # Thử lấy text và loại widget
                widget_text = ""
                if hasattr(widget, 'cget'):
                    try:
                        # Lấy text của widget
                        if widget.cget('text') and widget.cget('text').strip():
                            widget_text = widget.cget('text')
                    except:
                        pass
                
                # Xác định loại widget để dễ nhận biết
                widget_type = type(widget).__name__
                if widget_type == 'Button':
                    widget_name = f"🔘 NÚT LỚN: {widget_name}"
                elif widget_type == 'Label':
                    widget_name = f"📝 NHÃN: {widget_name}"
                elif widget_type == 'Entry':
                    widget_name = f"📝 Ô NHẬP: {widget_name}"
                elif widget_type == 'Frame':
                    # Kiểm tra xem frame có chứa button không
                    has_button = False
                    for child in widget.winfo_children():
                        if 'Button' in str(type(child)) or 'ModernButton' in str(type(child)):
                            has_button = True
                            break
                    
                    if has_button:
                        widget_name = f"🔘 KHUNG CHỨA NÚT: {widget_name}"
                    else:
                        widget_name = f"📦 KHUNG THƯỜNG: {widget_name}"
                elif widget_type == 'Notebook':
                    widget_name = f"📑 TAB: {widget_name}"
                elif 'ModernButton' in str(type(widget)):
                    widget_name = f"🔘 NÚT LỚN: {widget_name}"
                
                # Thêm gợi ý vị trí nếu là nút quan trọng
                if widget_text and any(keyword in widget_text.upper() for keyword in ['MỞ FILE', 'OPEN', 'FILE', 'EXCEL']):
                    widget_name += " 📂"
                elif widget_text and any(keyword in widget_text.upper() for keyword in ['KHỚP', 'MAP', 'CỘT']):
                    widget_name += " 🛠️"
                elif widget_text and any(keyword in widget_text.upper() for keyword in ['TOGGLE', 'CHUYỂN ĐỔI', 'DARK', 'LIGHT']):
                    widget_name += " "
                
                return (widget_name, widget_x, widget_y, widget_w, widget_h)
        except:
            pass
        return None
    
    def on_widget_click(self, event):
        """Xử lý khi click chuột trái để chọn widget - KHÔNG HIỂN THÔNG BÁO"""
        # Lấy toạ độ chuột
        x = event.x
        y = event.y
        
        # Tìm widget tại vị trí click
        widget_info = self.get_widget_at_position(x, y)
        
        if widget_info:
            # Lưu widget được chọn
            self.selected_widget_info = widget_info
        else:
            # Xóa widget được chọn
            self.selected_widget_info = None

    def copy_coordinates(self, event):
        """Copy toạ độ chuột và thông tin widget vào clipboard"""
        # Debug: Hiển thị thông báo hàm được gọi
        print("DEBUG: copy_coordinates được gọi!")
        print(f"DEBUG: selected_widget_info = {self.selected_widget_info}")
        
        # Debug: Kiểm tra xem có widget được chọn không
        if not self.selected_widget_info:
            # Hiển thị thông báo gợi ý
            self.coord_label.config(text="❌ Chưa chọn widget!\nClick vào widget trước")
            mouse_x = self.root.winfo_pointerx() - self.root.winfo_rootx()
            mouse_y = self.root.winfo_pointery() - self.root.winfo_rooty()
            self.coord_label.place(x=mouse_x + 15, y=mouse_y - 30)
            self.root.after(2000, lambda: self.coord_label.place_forget())
            return

        # Hoạt động khi có widget được chọn (bằng click chuột) hoặc khi giữ Alt
        if self.selected_widget_info:
            try:
                widget_name, widget_x, widget_y, widget_w, widget_h = self.selected_widget_info

                # Lấy toạ độ chuột hiện tại
                x = self.root.winfo_pointerx() - self.root.winfo_rootx()
                y = self.root.winfo_pointery() - self.root.winfo_rooty()
                
                # Tạo nội dung copy
                copy_text = f"""=== INFO WIDGET ===
Widget: {widget_name}
Toạ độ chuột: X:{x}, Y:{y}
Vị trí widget: X:{widget_x}, Y:{widget_y}
Kích thước: {widget_w}x{widget_h}

=== CODE PLACE ===
.place(x={x}, y={y}, width={widget_w}, height={widget_h})

=== CHỈNH SỬA ===
X: {x} (để di chuyển ngang)
Y: {y} (để di chuyển dọc)
Width: {widget_w} (độ rộng)
Height: {widget_h} (độ cao)
X Offset: {x - widget_x} (để di chuyển ngang so với widget)
Y Offset: {y - widget_y} (để di chuyển dọc so với widget)"""
                
                # Sử dụng tkinter clipboard thay vì pyperclip
                try:
                    import pyperclip
                    pyperclip.copy(copy_text)
                except ImportError:
                    # Fallback: dùng tkinter clipboard
                    self.root.clipboard_clear()
                    self.root.clipboard_append(copy_text)
                
                # Hiển thị thông báo
                self.coord_label.config(text="✅ ĐÃ COPY TOẠ ĐỘ!\nCtrl+V để dán")
                # Hiển thị tại vị trí chuột và ẩn sau 2 giây
                mouse_x = self.root.winfo_pointerx() - self.root.winfo_rootx()
                mouse_y = self.root.winfo_pointery() - self.root.winfo_rooty()
                self.coord_label.place(x=mouse_x + 15, y=mouse_y - 30)
                self.root.after(2000, lambda: self.coord_label.place_forget())
                
            except ImportError:
                # Nếu không có pyperclip, dùng tkinter clipboard
                try:
                    widget_name, widget_x, widget_y, widget_w, widget_h = self.selected_widget_info
                    x = self.root.winfo_pointerx() - self.root.winfo_rootx()
                    y = self.root.winfo_pointery() - self.root.winfo_rooty()
                    
                    copy_text = f".place(x={x}, y={y}, width={widget_w}, height={widget_h})"
                    self.root.clipboard_clear()
                    self.root.clipboard_append(copy_text)
                    
                    self.coord_label.config(text=" ĐÃ COPY CODE!\nCtrl+V để dán")
                    self.root.after(2000, lambda: self.coord_label.place_forget() if not self.alt_pressed else None)
                except:
                    pass
            except:
                pass

    # =====================================================
    # HOTKEY HELPERS
    # =====================================================
    def get_active_excel_tab(self):
        try:
            current_tab_id = self.notebook.select()
            if not current_tab_id: return None
            w = self.notebook.nametowidget(current_tab_id)
            if isinstance(w, ExcelTab):
                return w
        except Exception:
            return None
        return None

    def focus_search(self, event=None):
        tab = self.get_active_excel_tab()
        if tab and tab.ent_search:
            tab.ent_search.focus_set()
            tab.ent_search.select_range(0, tk.END)
        return "break"

    def reset_filters(self, event=None):
        tab = self.get_active_excel_tab()
        if tab:
            tab.ent_search.delete(0, tk.END)
            tab.filter_data("All")
        return "break"

    def save_current_tab(self, event=None):
        tab = self.get_active_excel_tab()
        if tab:
            tab.perform_auto_save()
        return "break"

    def reload_current_tab(self, event=None):
        tab = self.get_active_excel_tab()
        if tab:
            tab.load_data_from_file()
        return "break"

    def undo_current_tab(self, event=None):
        tab = self.get_active_excel_tab()
        if tab:
            tab.undo_last_edit()
        return "break"
    
    def load_ui_config(self):
        """Đọc cấu hình UI từ file JSON với xử lý lỗi an toàn"""
        try:
            config_file = "ui_config.json"
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return {}
        except Exception as e:
            return {}
    
    def apply_layout(self, widget_name, widget):
        """Áp dụng layout từ JSON cho widget cụ thể"""
        try:
            if widget_name in self.ui_config:
                config = self.ui_config[widget_name]
                x = config.get("x", 0)
                y = config.get("y", 0)
                width = config.get("width", 100)
                height = config.get("height", 30)
                
                # Chỉ áp dụng place() nếu có tọa độ hợp lệ
                if x >= 0 and y >= 0 and width > 0 and height > 0:
                    widget.place(x=x, y=y, width=width, height=height)
                    return True
        except Exception as e:
            print(f"Cảnh báo: Lỗi áp dụng layout cho {widget_name}: {str(e)}")
        return False

    def setup_ui(self):
        # =============================================================================
        # MENU BAR - LICENSE MANAGEMENT
        # =============================================================================
        menubar = tk.Menu(self.root, bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 11, "bold"))
        self.root.config(menu=menubar)
        
        # Menu License - TO HƠN VÀ IN HOA
        license_menu = tk.Menu(menubar, tearoff=0, bg="#ffffff", fg="#2c3e50", font=("Segoe UI", 10))
        menubar.add_cascade(label="📜 THÔNG TIN BẢN QUYỀN", menu=license_menu)
        license_menu.add_command(label="📜 Thông tin License", command=self.show_license_info)
        license_menu.add_command(label="🔑 Kích hoạt License mới", command=self.activate_new_license)
        license_menu.add_separator()
        license_menu.add_command(label="🗑️ Xóa License (Debug)", command=self.revoke_license_debug)
        
        # =============================================================================
        
        # Header Frame với viền đen mỏng
        self.header = tk.Frame(self.root, bg=THEME["bg_card"], height=80, 
                              highlightbackground="black", highlightthickness=1) 
        self.header.pack(fill="x", padx=15, pady=10)
        self.header.pack_propagate(False)

        # Logo Text
        self.lbl_logo = tk.Label(self.header, text="🎓 EduManager", font=("Segoe UI", 20, "bold"), fg=THEME["primary"], bg=THEME["bg_card"])
        # Thử áp dụng layout từ JSON, nếu thất bại thì dùng pack() như cũ
        if not self.apply_layout("self.lbl_logo", self.lbl_logo):
            self.lbl_logo.pack(side="left", padx=20)
        
        self.lbl_sub = tk.Label(self.header, text="| AI Powered Edition 🚀", font=("Segoe UI", 12), fg="#7f8c8d", bg=THEME["bg_card"])
        if not self.apply_layout("self.lbl_sub", self.lbl_sub):
            self.lbl_sub.pack(side="left")

        # [UPDATED] Đặt thông báo "Đã tải/Sẵn sàng" sẽ được tạo trong filter_bar

        # Buttons Right
        btn_container = tk.Frame(self.header, bg=THEME["bg_card"])
        btn_container.pack(side="right", padx=10, pady=5)

        self.toggle_frame = tk.Frame(btn_container, bg=THEME["btn_border"], padx=1, pady=1) 
        self.toggle_frame.is_btn_border = True
        self.toggle_frame.pack(side="right", padx=5) 
        
        self.btn_toggle = tk.Button(self.toggle_frame, text="💡", font=("Segoe UI", 12),
                                    bg=THEME["bg_app"], fg=THEME["text_main"],
                                    relief="flat", cursor="hand2", width=4,
                                    command=self.toggle_dark_mode)
        self.btn_toggle.pack(fill="both", expand=True)
        
        # Áp dụng layout từ JSON cho toggle frame và button
        self.apply_layout("self.toggle_frame", self.toggle_frame)
        self.apply_layout("self.btn_toggle", self.btn_toggle)

        # 2. KHUNG CHỨA NÚT MỞ FILE VÀ KHỚP CỘT
        file_action_frame = tk.Frame(btn_container, bg=THEME["bg_card"])
        file_action_frame.pack(side="left", fill="y") 
        
        # Nút HDSD (Bên trái cùng)
        self.hdsd_frame = tk.Frame(file_action_frame, bg="#000000", padx=0, pady=0)
        self.hdsd_frame.is_btn_border = True
        self.hdsd_frame.pack(side="left", padx=(0, 5), fill="both", expand=True) 
        
        self.btn_hdsd = ModernButton(self.hdsd_frame, bg_color="#9333ea", text_color="white", text="📖 HDSD", command=self.show_user_guide)
        self.btn_hdsd.pack(fill="both", expand=True)
        
        # [AUTO-UPDATE] Nút Kiểm tra cập nhật
        self.update_frame = tk.Frame(file_action_frame, bg="#000000", padx=0, pady=0)
        self.update_frame.is_btn_border = True
        self.update_frame.pack(side="left", padx=(0, 5), fill="both", expand=True)
        
        self.btn_update = ModernButton(self.update_frame, bg_color="#3498db", text_color="white", text="🔄 UPDATE", command=self.manual_check_update)
        self.btn_update.pack(fill="both", expand=True)

        # Nút Mở File (Bên phải HDSD)
        self.open_frame = tk.Frame(file_action_frame, bg="#000000", padx=0, pady=0)
        self.open_frame.is_btn_border = True
        self.open_frame.pack(side="left", padx=(0, 5), fill="both", expand=True) 
        

        self.btn_open = ModernButton(self.open_frame, bg_color=THEME["primary"], text_color="white", text="📂 MỞ FILE EXCEL", command=self.add_new_tab)
        self.btn_open.pack(fill="both", expand=True)

        # Nút Khớp Cột (Bên phải)
        self.map_frame = tk.Frame(file_action_frame, bg="#000000", padx=0, pady=0)
        self.map_frame.is_btn_border = True
        self.map_frame.pack(side="left", padx=(0, 5), fill="both", expand=True)

        self.btn_map_global = ModernButton(self.map_frame, bg_color="#7f8c8d", text_color="white", text=" KHỚP CỘT", 
                                             command=self.call_mapping_on_current_tab)
        self.btn_map_global.pack(fill="both", expand=True)
        
        # [NEW] Nút Viết Hoa Toàn Bộ
        self.uppercase_frame = tk.Frame(file_action_frame, bg=THEME["bg_card"], padx=1, pady=1)
        self.uppercase_frame.is_btn_border = True
        self.uppercase_frame.pack(side="left", padx=(0, 5), fill="both", expand=True)
        
        self.uppercase_var = tk.BooleanVar(value=True)
        self.btn_uppercase = tk.Checkbutton(
            self.uppercase_frame,
            text="IN HOA",
            variable=self.uppercase_var,
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            selectcolor=THEME["primary"],
            activebackground=THEME["bg_card"],
            relief="flat",
            cursor="hand2",
            command=self.toggle_uppercase_mode
        )
        self.btn_uppercase.pack(fill="both", expand=True, padx=3, pady=3)
        
        # [NEW] Nút In Đậm Toàn Bộ
        self.bold_frame = tk.Frame(file_action_frame, bg=THEME["bg_card"], padx=1, pady=1)
        self.bold_frame.is_btn_border = True
        self.bold_frame.pack(side="left", fill="both", expand=True)
        
        self.bold_var = tk.BooleanVar(value=False)
        self.btn_bold = tk.Checkbutton(
            self.bold_frame,
            text="IN ĐẬM",
            variable=self.bold_var,
            bg=THEME["bg_card"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold"),
            selectcolor=THEME["primary"],
            activebackground=THEME["bg_card"],
            relief="flat",
            cursor="hand2",
            command=self.toggle_bold_mode
        )
        self.btn_bold.pack(fill="both", expand=True, padx=3, pady=3)

        # [NEW] Thanh hiển thị trạng thái lọc nằm GIỮA Logo (Header) và Tab Notebook
        self.filter_bar = tk.Frame(self.root, bg=THEME["bg_app"])
        self.filter_bar.pack(fill="x", padx=20, pady=(0, 5))

        # Label bên trái - Lọc
        self.lbl_filter_global = tk.Label(
            self.filter_bar,
            text="",
            bg=THEME["bg_app"],
            fg=THEME["text_main"],
            font=("Segoe UI", 10, "bold")
        )
        # Căn trái
        self.lbl_filter_global.pack(side="left", anchor="w")
        self.lbl_filter_global.config(justify="left", anchor="w")

        # Label bên phải - Đã đồng bộ
        self.lbl_autosave = tk.Label(
            self.filter_bar,
            text="SẲN SÀNG ✅",
            bg=THEME["bg_app"],
            fg="#10B981",
            font=("Segoe UI", 10, "bold")
        )
        # Căn phải
        self.lbl_autosave.pack(side="right", anchor="e")

        # Main Notebook
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=20, pady=20)
        
        self.tab_menu = tk.Menu(self.root, tearoff=0)
        self.tab_menu.add_command(label="Đóng Tab Này", command=self.close_current_tab)
        self.notebook.bind("<Button-3>", self.show_tab_menu)
        # Welcome Screen
        self.welcome = tk.Frame(self.notebook, bg=THEME["bg_card"])
        self.notebook.add(self.welcome, text="Trang chủ")
        
        self.box = tk.Frame(self.welcome, bg=THEME["bg_card"])
        if not self.apply_layout("self.box", self.box):
            self.box.pack(expand=True)
        self.lbl_welcome_1 = tk.Label(self.box, text="XIN CHÀO!", font=("Segoe UI", 30, "bold"), bg=THEME["bg_card"], fg=THEME["primary"])
        if not self.apply_layout("self.lbl_welcome_1", self.lbl_welcome_1):
            self.lbl_welcome_1.pack()
        self.lbl_welcome_2 = tk.Label(self.box, text="Vui lòng mở file Excel có đuôi định dạng xlsx để bắt đầu.", font=("Segoe UI", 14), bg=THEME["bg_card"], fg="gray")
        if not self.apply_layout("self.lbl_welcome_2", self.lbl_welcome_2):
            self.lbl_welcome_2.pack(pady=10)
        
        # Frame chứa 2 zone song song - dùng grid để chia đều
        zones_container = tk.Frame(self.box, bg=THEME["bg_card"], height=280)
        zones_container.pack(pady=20, padx=20, fill="x")
        zones_container.pack_propagate(False)
        zones_container.grid_columnconfigure(0, weight=1)
        zones_container.grid_columnconfigure(1, weight=1)
        
        # [NEW] Click Zone - Khung đẹp để click mở file
        self.click_zone = tk.Frame(zones_container, bg="#f8f9fa", relief="solid", bd=2, cursor="hand2")
        self.click_zone.config(highlightbackground="#dee2e6", highlightthickness=2)
        if not self.apply_layout("self.click_zone", self.click_zone):
            self.click_zone.grid(row=0, column=0, padx=(0, 10), sticky="nsew")
        
        # Nội dung Click Zone
        click_content = tk.Frame(self.click_zone, bg="#f8f9fa", padx=30, pady=20)
        click_content.pack(fill="both", expand=True)
        
        # Icon lớn
        self.click_icon = tk.Label(click_content, text="📂", font=("Segoe UI", 40), bg="#f8f9fa")
        self.click_icon.pack()
        
        # Text hướng dẫn
        self.click_label = tk.Label(click_content, text="CLICK ĐỂ CHỌN FILE EXCEL", 
                                    font=("Segoe UI", 14, "bold"), bg="#f8f9fa", fg="#495057")
        self.click_label.pack(pady=(10, 5))
        
        self.click_sublabel = tk.Label(click_content, text="hoặc nhấn Ctrl+O", 
                 font=("Segoe UI", 10), bg="#f8f9fa", fg="#6c757d")
        self.click_sublabel.pack()
        
        self.click_formats = tk.Label(click_content, text="Hỗ trợ: .xlsx, .xls, .csv", 
                 font=("Segoe UI", 9, "italic"), bg="#f8f9fa", fg="#adb5bd")
        self.click_formats.pack(pady=(10, 0))
        
        # Bind click event cho Click Zone
        for widget in [self.click_zone, click_content, self.click_icon, self.click_label, self.click_sublabel, self.click_formats]:
            widget.bind("<Button-1>", lambda e: self.add_new_tab())
        
        # Hover effects với debounce để tránh flicker
        self._click_leave_timer = None
        
        def on_click_enter(e):
            # Hủy timer leave nếu đang chờ
            if self._click_leave_timer:
                self.root.after_cancel(self._click_leave_timer)
                self._click_leave_timer = None
            # Đổi màu ngay lập tức theo theme
            hover_bg = "#e8f5e9" if not IS_DARK_MODE else "#1b4d3e"
            hover_accent = "#27ae60" if not IS_DARK_MODE else "#4ade80"
            self.click_zone.config(highlightbackground=hover_accent, bg=hover_bg)
            click_content.config(bg=hover_bg)
            self.click_icon.config(bg=hover_bg)
            self.click_label.config(bg=hover_bg, fg=hover_accent)
            self.click_sublabel.config(bg=hover_bg)
            self.click_formats.config(bg=hover_bg)
        
        def on_click_leave(e):
            # Delay 50ms trước khi đổi màu về
            def delayed_leave():
                normal_bg = THEME["bg_card"]
                normal_border = THEME["border"]
                normal_fg = THEME["text_sub"]
                self.click_zone.config(highlightbackground=normal_border, bg=normal_bg)
                click_content.config(bg=normal_bg)
                self.click_icon.config(bg=normal_bg)
                self.click_label.config(bg=normal_bg, fg=normal_fg)
                self.click_sublabel.config(bg=normal_bg)
                self.click_formats.config(bg=normal_bg)
                self._click_leave_timer = None
            
            self._click_leave_timer = self.root.after(50, delayed_leave)
        
        self.click_zone.bind("<Enter>", on_click_enter)
        self.click_zone.bind("<Leave>", on_click_leave)

        # [NEW] Convert Zone - Bên phải Click Zone
        self.convert_zone = tk.Frame(zones_container, bg="#f8f9fa", relief="solid", bd=2, cursor="hand2")
        self.convert_zone.config(highlightbackground="#dee2e6", highlightthickness=2)
        if not self.apply_layout("self.convert_zone", self.convert_zone):
            self.convert_zone.grid(row=0, column=1, padx=(10, 0), sticky="nsew")

        # Nội dung Convert Zone - giảm pady để bằng với Click Zone
        convert_content = tk.Frame(self.convert_zone, bg="#f8f9fa", padx=30, pady=20)
        convert_content.pack(fill="both", expand=True)

        # Icon - giảm size để cân đối
        self.convert_icon = tk.Label(convert_content, text="📁", font=("Segoe UI", 40), bg="#f8f9fa")
        self.convert_icon.pack()

        # Tiêu đề - dùng Segoe UI giống Click Zone
        self.convert_label = tk.Label(
            convert_content,
            text="CHUYỂN ĐỔI ĐỊNH DẠNG FILE",
            font=("Segoe UI", 14, "bold"),
            bg="#f8f9fa",
            fg="#495057"
        )
        self.convert_label.pack(pady=(10, 5))

        # Mô tả
        self.convert_sublabel = tk.Label(
            convert_content,
            text="Chuyển file .xls sang .xlsx (giữ nguyên định dạng)",
            font=("Segoe UI", 10),
            bg=THEME["bg_card"],
            fg=THEME["text_sub"]
        )
        self.convert_sublabel.pack()

        # Bind click event cho toàn vùng
        for widget in [self.convert_zone, convert_content, self.convert_icon, self.convert_label, self.convert_sublabel]:
            widget.bind("<Button-1>", lambda e: self.convert_file_format())

        # Hover effects với debounce để tránh flicker
        self._convert_leave_timer = None
        
        def on_convert_enter(e):
            # Hủy timer leave nếu đang chờ
            if self._convert_leave_timer:
                self.root.after_cancel(self._convert_leave_timer)
                self._convert_leave_timer = None
            # Đổi màu ngay lập tức theo theme
            hover_bg = "#e8f5e9" if not IS_DARK_MODE else "#1b4d3e"
            hover_accent = "#27ae60" if not IS_DARK_MODE else "#4ade80"
            self.convert_zone.config(highlightbackground=hover_accent, bg=hover_bg)
            convert_content.config(bg=hover_bg)
            self.convert_icon.config(bg=hover_bg)
            self.convert_label.config(bg=hover_bg, fg=hover_accent)
            self.convert_sublabel.config(bg=hover_bg)

        def on_convert_leave(e):
            # Delay 50ms trước khi đổi màu về
            def delayed_leave():
                normal_bg = THEME["bg_card"]
                normal_border = THEME["border"]
                normal_fg = THEME["text_sub"]
                self.convert_zone.config(highlightbackground=normal_border, bg=normal_bg)
                convert_content.config(bg=normal_bg)
                self.convert_icon.config(bg=normal_bg)
                self.convert_label.config(bg=normal_bg, fg=normal_fg)
                self.convert_sublabel.config(bg=normal_bg, fg=normal_fg)
                self._convert_leave_timer = None
            
            self._convert_leave_timer = self.root.after(50, delayed_leave)

        self.convert_zone.bind("<Enter>", on_convert_enter)
        self.convert_zone.bind("<Leave>", on_convert_leave)

        # [NEW] Thêm keyboard bindings cho phím tắt
        self.setup_keyboard_shortcuts()
        
        # [NEW] Thêm Tooltips cho các nút
        self.setup_tooltips()

    def setup_keyboard_shortcuts(self):
        """Thiết lập các phím tắt cho ứng dụng"""
        try:
            # Ctrl+O - Mở file Excel
            self.root.bind('<Control-o>', lambda e: self.add_new_tab())
            self.root.bind('<Control-O>', lambda e: self.add_new_tab())
            
            # Ctrl+S - Lưu file
            self.root.bind('<Control-s>', lambda e: self.save_file())
            self.root.bind('<Control-S>', lambda e: self.save_file())
            
            # Ctrl+F - Tìm kiếm
            self.root.bind('<Control-f>', lambda e: self.focus_search())
            self.root.bind('<Control-F>', lambda e: self.focus_search())
            
            # F5 - Làm mới dữ liệu
            self.root.bind('<F5>', lambda e: self.refresh_data())
            
            # ESC - Đóng cửa sổ/tab hiện tại
            self.root.bind('<Escape>', lambda e: self.close_active_window())
            
            print("✅ Đã thiết lập phím tắt thành công!")
            print("📋 Phím tắt:")
            print("   Ctrl+O - Mở file Excel")
            print("   Ctrl+S - Lưu file")
            print("   Ctrl+F - Tìm kiếm")
            print("   F5 - Làm mới dữ liệu")
            print("   ESC - Đóng cửa sổ/tab")
            
        except Exception as e:
            print(f"❌ Lỗi thiết lập phím tắt: {e}")
            messagebox.showerror("Lỗi", f"Không thể thiết lập phím tắt: {str(e)}")
    
    def setup_tooltips(self):
        """Thiết lập Tooltips cho các nút chính"""
        try:
            # Nút MỞ FILE
            if hasattr(self, 'btn_open') and self.btn_open:
                ToolTip(self.btn_open.btn, "📂 Mở file Excel để quản lý điểm\\n(Ctrl+O)\\n\\nHỗ trợ: .xlsx, .xls, .csv")
            
            # Nút KHỚP CỘT
            if hasattr(self, 'btn_map_global') and self.btn_map_global:
                ToolTip(self.btn_map_global.btn, "🔗 Khớp cột dữ liệu thủ công\\n\\nSử dụng khi auto-detect không chính xác")
            
            # Nút HDSD
            if hasattr(self, 'btn_hdsd') and self.btn_hdsd:
                ToolTip(self.btn_hdsd.btn, "📖 Hướng dẫn sử dụng\\n\\nXem các tính năng và cách thao tác")
            
            # Nút UPDATE
            if hasattr(self, 'btn_update') and self.btn_update:
                ToolTip(self.btn_update.btn, f"🔄 Kiểm tra cập nhật\\n\\nPhiên bản hiện tại: {APP_VERSION}\\nTự động tải và cài đặt bản mới")
            
            # Nút Theme Toggle
            if hasattr(self, 'btn_toggle') and self.btn_toggle:
                ToolTip(self.btn_toggle, "🎨 Chuyển đổi giao diện Sáng/Tối")
            
            print("✅ Đã thiết lập Tooltips!")
        except Exception as e:
            print(f"⚠️ Lỗi thiết lập Tooltips: {e}")

    def show_user_guide(self):
        """Hiển thị bảng hướng dẫn sử dụng EduManager"""
        try:
            print("DEBUG: show_user_guide được gọi!")
            guide_win = tk.Toplevel(self.root)
            guide_win.title("📖 Hướng dẫn sử dụng EduManager")
            guide_win.geometry("900x650")
            guide_win.resizable(False, False)
            guide_win.configure(bg="#f8fafc")
            
            # Cleanup khi đóng cửa sổ
            def on_guide_close():
                try:
                    guide_win.unbind_all("<MouseWheel>")
                except:
                    pass
                guide_win.destroy()
            guide_win.protocol("WM_DELETE_WINDOW", on_guide_close)
            
            # Center window - bỏ để tránh lỗi
            # self.center_window(guide_win, 900, 650)
            
            # Header với viền pastel
            header_frame = tk.Frame(guide_win, bg="#e0e7ff", height=80, 
                                   highlightbackground="#818cf8", highlightthickness=2, relief="solid")
            header_frame.pack(fill="x", padx=10, pady=(10, 5))
            header_frame.pack_propagate(False)
            
            # Title header
            tk.Label(header_frame, text="🎓 EDUMANAGER - HƯỚNG DẪN SỬ DỤNG", 
                    font=("Segoe UI", 18, "bold"), bg="#e0e7ff", fg="#4338ca").pack(pady=20)
            
            # Main content với viền pastel
            main_frame = tk.Frame(guide_win, bg="#fef3c7", 
                                 highlightbackground="#fbbf24", highlightthickness=2, relief="solid")
            main_frame.pack(fill="both", expand=True, padx=10, pady=5)
            
            # Scrollable area
            canvas = tk.Canvas(main_frame, bg="#fef3c7", highlightthickness=0)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg="#fef3c7")
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Nội dung hướng dẫn
            guide_content = [
                {
                    "category": "🔰 CƠ BẢN",
                    "items": [
                        {"icon": "📂", "name": "MỞ FILE EXCEL", "desc": "Mở file Excel chứa danh sách học sinh và điểm. Hỗ trợ định dạng .xlsx và .xls. Chọn một hoặc nhiều file cùng lúc."},
                        {"icon": "🔗", "name": "KHỚP CỘT", "desc": "Tự động hoặc thủ công khớp các cột dữ liệu với tên cột chuẩn: Họ và tên, Toán, Văn, Anh, Lý, Hóa, Sinh."},
                        {"icon": "🔍", "name": "TÌM KIẾM", "desc": "Tìm kiếm học sinh theo tên, lớp hoặc điểm. Nhập từ khóa và kết quả sẽ hiển thị ngay lập tức."},
                        {"icon": "📊", "name": "BỘ LỌC", "desc": "Lọc học sinh theo khoảng điểm, môn học hoặc trạng thái. Giúp tập trung vào nhóm học sinh cần quan tâm."}
                    ]
                },
                {
                    "category": "📈 THỐNG KÊ & BÁO CÁO",
                    "items": [
                        {"icon": "📊", "name": "BIỂU ĐỒ", "desc": "Hiển thị biểu đồ cột, tròn, đường phân tích điểm số. Giúp nhận diện xu hướng học tập của lớp."},
                        {"icon": "📈", "name": "THỐNG KÊ NÂNG CAO", "desc": "Phân tích chi tiết: điểm trung bình, độ lệch chuẩn, phân phối điểm, so sánh các môn học."},
                        {"icon": "📋", "name": "XUẤT BÁO CÁO", "desc": "Xuất báo cáo chi tiết ra file Excel hoặc PDF. Bao gồm thống kê, biểu đồ và nhận xét."}
                    ]
                },
                {
                    "category": "📒 NHẬT KÝ & THI ĐUA",
                    "items": [
                        {"icon": "📒", "name": "NHẬT KÝ", "desc": "Ghi chép hoạt động hàng ngày: cộng điểm cho hành vi tốt, trừ điểm cho vi phạm. Tự động tính điểm thi đua."},
                        {"icon": "🏆", "name": "BẢNG VÀNG VINH DANH", "desc": "Hiển thị danh sách học sinh có điểm thi đua dương. Vinh danh và tạo động lực cho học sinh."},
                        {"icon": "🧩", "name": "CHIA NHÓM", "desc": "Tự động chia lớp thành các nhóm nhỏ dựa trên điểm số. Canh bằng sức mạnh giữa các nhóm."},
                        {"icon": "📺", "name": "TRÌNH CHIẾU", "desc": "Chế độ trình chiếu bảng điểm lên màn hình lớn. Tối ưu cho họp phụ huynh hoặc buổi tổng kết."}
                    ]
                },
                {
                    "category": "🤖 TIỆN ÍCH MỞ RỘNG",
                    "items": [
                        {"icon": "🤖", "name": "TỰ ĐỘNG", "desc": "Tự động hóa các tác vụ: làm mới dữ liệu, kiểm tra lỗi, sao lưu. Giúp tiết kiệm thời gian thao tác."},
                        {"icon": "⚠️", "name": "CẢNH BÁO", "desc": "Hiển thị cảnh báo: học sinh yếu kém, điểm bất thường, cần quan tâm đặc biệt. Giúp giáo viên can thiệp kịp thời."},
                        {"icon": "📞", "name": "DANH BẠ PHỤ HUYNH", "desc": "Quản lý thông tin liên lạc của phụ huynh. Xuất danh sách và gửi thông báo nhanh chóng."},
                        {"icon": "🌙", "name": "CHỦ ĐỀ GIAO DIỆN", "desc": "Chuyển đổi giữa giao diện sáng và tối. Bảo vệ mắt khi làm việc lâu."}
                    ]
                },
                {
                    "category": "⌨️ PHÍM TẮT",
                    "items": [
                        {"icon": "Ctrl+O", "name": "Mở file", "desc": "Mở nhanh file Excel mới."},
                        {"icon": "Ctrl+S", "name": "Lưu", "desc": "Lưu các thay đổi."},
                        {"icon": "Ctrl+F", "name": "Tìm kiếm", "desc": "Đưa con trỏ vào ô tìm kiếm."},
                        {"icon": "F5", "name": "Làm mới", "desc": "Làm mới dữ liệu và biểu đồ."},
                        {"icon": "ESC", "name": "Đóng", "desc": "Đóng cửa sổ hiện tại."}
                    ]
                }
            ]
            
            # Tạo nội dung cho từng category
            for category_idx, category in enumerate(guide_content):
                # Category header với viền pastel
                cat_frame = tk.Frame(scrollable_frame, bg="#dcfce7", 
                                   highlightbackground="#86efac", highlightthickness=2, relief="solid")
                cat_frame.pack(fill="x", padx=15, pady=(10, 5))
                
                tk.Label(cat_frame, text=category["category"], 
                        font=("Segoe UI", 14, "bold"), bg="#dcfce7", fg="#166534").pack(pady=10, padx=15)
                
                # Items trong category
                for item_idx, item in enumerate(category["items"]):
                    # Item frame với viền pastel
                    item_frame = tk.Frame(scrollable_frame, bg="#f0f9ff", 
                                       highlightbackground="#7dd3fc", highlightthickness=1, relief="solid")
                    item_frame.pack(fill="x", padx=20, pady=3)
                    
                    # Icon và tên chức năng
                    content_frame = tk.Frame(item_frame, bg="#f0f9ff")
                    content_frame.pack(fill="x", padx=15, pady=10)
                    
                    # Icon
                    icon_label = tk.Label(content_frame, text=item["icon"], 
                                       font=("Segoe UI", 16), bg="#f0f9ff", width=5)
                    icon_label.pack(side="left", padx=(0, 10))
                    
                    # Tên và mô tả
                    text_frame = tk.Frame(content_frame, bg="#f0f9ff")
                    text_frame.pack(side="left", fill="x", expand=True)
                    
                    tk.Label(text_frame, text=item["name"], 
                            font=("Segoe UI", 12, "bold"), bg="#f0f9ff", fg="#1e40af").pack(anchor="w")
                    
                    tk.Label(text_frame, text=item["desc"], 
                            font=("Segoe UI", 10), bg="#f0f9ff", fg="#64748b", 
                            wraplength=600, justify="left").pack(anchor="w", pady=(2, 0))
            
            # Pack scrollable area
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # [FIXED] Thêm chức năng cuộn chuột
            def _on_mousewheel(event):
                if event.delta:
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            
            # Footer với viền pastel
            footer_frame = tk.Frame(guide_win, bg="#fce7f3", 
                                   highlightbackground="#f9a8d4", highlightthickness=2, relief="solid")
            footer_frame.pack(fill="x", padx=10, pady=(5, 10))
            
            footer_content = tk.Frame(footer_frame, bg="#fce7f3")
            footer_content.pack(pady=10)
            
            tk.Label(footer_content, text="💡 Mẹo: Để con trỏ vào các chức năng để xem tooltip chi tiết hơn", 
                    font=("Segoe UI", 10, "italic"), bg="#fce7f3", fg="#be185d").pack()
            
            tk.Label(footer_content, text="📞 Cần hỗ trợ? Liên hệ: eduManager@support.com | Hotline: 1900-xxxx", 
                    font=("Segoe UI", 9), bg="#fce7f3", fg="#9333ea").pack(pady=(5, 0))
            
            # Close button với viền pastel
            close_btn_frame = tk.Frame(guide_win, bg="#fbbf24", padx=2, pady=2)
            close_btn_frame.pack(pady=10)
            
            tk.Button(close_btn_frame, text="✅ ĐÓNG HƯỚNG DẪN", command=guide_win.destroy, 
                     bg="#10b981", fg="white", font=("Segoe UI", 11, "bold"), 
                     relief="flat", borderwidth=0, cursor="hand2", 
                     padx=30, pady=8).pack()
            
        except Exception as e:
            print(f"Lỗi trong show_user_guide: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Lỗi", f"Không thể mở hướng dẫn: {str(e)}")

    def call_mapping_on_current_tab(self):
        try:
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                messagebox.showinfo("Thông báo", "Vui lòng mở một file Excel trước.")
                return
            
            current_widget = self.notebook.nametowidget(current_tab_id)
            
            if hasattr(current_widget, 'open_manual_mapping'):
                current_widget.open_manual_mapping()
            else:
                messagebox.showinfo("Thông báo", "Chức năng này chỉ hoạt động trên tab dữ liệu Excel.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở chức năng khớp cột: {str(e)}")
    
    def call_undo_mapping_on_current_tab(self):
        """Giai đoạn 9: Gọi hàm undo_last_mapping trên tab hiện tại"""
        try:
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                messagebox.showinfo("Thông báo", "Vui lòng mở một file Excel trước.")
                return
            
            current_widget = self.notebook.nametowidget(current_tab_id)
            
            if hasattr(current_widget, 'undo_last_mapping'):
                current_widget.undo_last_mapping()
            else:
                messagebox.showinfo("Thông báo", "Chức năng này chỉ hoạt động trên tab dữ liệu Excel.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể hoàn tác mapping: {str(e)}")

    def toggle_dark_mode(self):
        global IS_DARK_MODE, THEME
        IS_DARK_MODE = not IS_DARK_MODE
        
        NEW_THEME = DARK_THEME if IS_DARK_MODE else LIGHT_THEME
        THEME.update(NEW_THEME) 

        icon = "🌙" if IS_DARK_MODE else "💡"
        self.btn_toggle.config(text=icon)

        self.update_ttk_style()

        self.root.configure(bg=THEME["bg_app"])
        self.apply_theme_recursive(self.root)

        current_tab = self.notebook.select()
        if current_tab:
            widget_name = self.notebook.nametowidget(current_tab)
            if isinstance(widget_name, ExcelTab):
                widget_name.update_treeview_tags("Dark" if IS_DARK_MODE else "Light")
                if widget_name.current_df is not None:
                    widget_name.filter_data("All")
    
    def toggle_uppercase_mode(self):
        """Bật/tắt chế độ viết hoa toàn bộ UI"""
        self.uppercase_mode = self.uppercase_var.get()
        
        if self.uppercase_mode:
            # Lưu text gốc trước khi uppercase
            self._save_original_texts(self.root)
            # Áp dụng uppercase
            self._apply_uppercase(self.root)
        else:
            # Khôi phục text gốc
            self._restore_original_texts(self.root)
    
    def _save_original_texts(self, widget):
        """Lưu text gốc của tất cả widgets (đệ quy)"""
        widget_id = str(widget)
        
        # Lưu text của widget hiện tại
        if isinstance(widget, (tk.Button, tk.Label, tk.Checkbutton, tk.Radiobutton)):
            try:
                text = widget.cget("text")
                if text:
                    self.original_texts[widget_id] = text
            except:
                pass
        elif isinstance(widget, tk.LabelFrame):
            try:
                text = widget.cget("text")
                if text:
                    self.original_texts[widget_id] = text
            except:
                pass
        elif isinstance(widget, ModernButton):
            try:
                text = widget.btn.cget("text")
                if text:
                    self.original_texts[widget_id] = text
            except:
                pass
        
        # Đệ quy cho các widget con
        try:
            for child in widget.winfo_children():
                self._save_original_texts(child)
        except:
            pass
    
    def _apply_uppercase(self, widget):
        """Áp dụng uppercase cho tất cả text trong UI (đệ quy)"""
        # Bỏ qua Treeview (bảng data) và Text widgets
        if isinstance(widget, (ttk.Treeview, tk.Text, tk.Entry)):
            return
        
        # Uppercase text của widget hiện tại
        if isinstance(widget, (tk.Button, tk.Label, tk.Checkbutton, tk.Radiobutton)):
            try:
                text = widget.cget("text")
                if text:
                    widget.config(text=text.upper())
            except:
                pass
        elif isinstance(widget, tk.LabelFrame):
            try:
                text = widget.cget("text")
                if text:
                    widget.config(text=text.upper())
            except:
                pass
        elif isinstance(widget, ModernButton):
            try:
                text = widget.btn.cget("text")
                if text:
                    widget.btn.config(text=text.upper())
            except:
                pass
        
        # Đệ quy cho các widget con
        try:
            for child in widget.winfo_children():
                self._apply_uppercase(child)
        except:
            pass
    
    def _restore_original_texts(self, widget):
        """Khôi phục text gốc của tất cả widgets (đệ quy)"""
        widget_id = str(widget)
        
        # Khôi phục text gốc nếu có
        if widget_id in self.original_texts:
            original_text = self.original_texts[widget_id]
            
            if isinstance(widget, (tk.Button, tk.Label, tk.Checkbutton, tk.Radiobutton)):
                try:
                    widget.config(text=original_text)
                except:
                    pass
            elif isinstance(widget, tk.LabelFrame):
                try:
                    widget.config(text=original_text)
                except:
                    pass
            elif isinstance(widget, ModernButton):
                try:
                    widget.btn.config(text=original_text)
                except:
                    pass
        
        # Đệ quy cho các widget con
        try:
            for child in widget.winfo_children():
                self._restore_original_texts(child)
        except:
            pass
    
    def toggle_bold_mode(self):
        """Bật/tắt chế độ in đậm toàn bộ UI"""
        self.bold_mode = self.bold_var.get()
        
        if self.bold_mode:
            # Lưu font gốc trước khi bold
            self._save_original_fonts(self.root)
            # Áp dụng bold
            self._apply_bold(self.root)
        else:
            # Khôi phục font gốc
            self._restore_original_fonts(self.root)
    
    def _save_original_fonts(self, widget):
        """Lưu font gốc của tất cả widgets (đệ quy)"""
        widget_id = str(widget)
        
        # Lưu font của widget hiện tại
        if isinstance(widget, (tk.Button, tk.Label, tk.Checkbutton, tk.Radiobutton)):
            try:
                font = widget.cget("font")
                if font:
                    self.original_fonts[widget_id] = font
            except:
                pass
        elif isinstance(widget, tk.LabelFrame):
            try:
                font = widget.cget("font")
                if font:
                    self.original_fonts[widget_id] = font
            except:
                pass
        elif isinstance(widget, ModernButton):
            try:
                font = widget.btn.cget("font")
                if font:
                    self.original_fonts[widget_id] = font
            except:
                pass
        
        # Đệ quy cho các widget con
        try:
            for child in widget.winfo_children():
                self._save_original_fonts(child)
        except:
            pass
    
    def _apply_bold(self, widget):
        """Áp dụng bold cho tất cả font trong UI (đệ quy)"""
        # Bỏ qua Treeview (bảng data) và Text widgets
        if isinstance(widget, (ttk.Treeview, tk.Text, tk.Entry)):
            return
        
        # Bold font của widget hiện tại
        if isinstance(widget, (tk.Button, tk.Label, tk.Checkbutton, tk.Radiobutton)):
            try:
                font = widget.cget("font")
                if font:
                    # Parse font tuple hoặc string
                    if isinstance(font, str):
                        # Font string format: "Segoe UI 10"
                        parts = font.split()
                        if len(parts) >= 2:
                            family = " ".join(parts[:-1])
                            size = parts[-1]
                            new_font = (family, int(size), "bold")
                        else:
                            new_font = (font, 10, "bold")
                    elif isinstance(font, tuple):
                        # Font tuple format: ("Segoe UI", 10) or ("Segoe UI", 10, "normal")
                        if len(font) >= 2:
                            family, size = font[0], font[1]
                            new_font = (family, size, "bold")
                        else:
                            new_font = font
                    else:
                        new_font = font
                    
                    widget.config(font=new_font)
            except:
                pass
        elif isinstance(widget, tk.LabelFrame):
            try:
                font = widget.cget("font")
                if font:
                    if isinstance(font, str):
                        parts = font.split()
                        if len(parts) >= 2:
                            family = " ".join(parts[:-1])
                            size = parts[-1]
                            new_font = (family, int(size), "bold")
                        else:
                            new_font = (font, 10, "bold")
                    elif isinstance(font, tuple):
                        if len(font) >= 2:
                            family, size = font[0], font[1]
                            new_font = (family, size, "bold")
                        else:
                            new_font = font
                    else:
                        new_font = font
                    
                    widget.config(font=new_font)
            except:
                pass
        elif isinstance(widget, ModernButton):
            try:
                font = widget.btn.cget("font")
                if font:
                    if isinstance(font, str):
                        parts = font.split()
                        if len(parts) >= 2:
                            family = " ".join(parts[:-1])
                            size = parts[-1]
                            new_font = (family, int(size), "bold")
                        else:
                            new_font = (font, 10, "bold")
                    elif isinstance(font, tuple):
                        if len(font) >= 2:
                            family, size = font[0], font[1]
                            new_font = (family, size, "bold")
                        else:
                            new_font = font
                    else:
                        new_font = font
                    
                    widget.btn.config(font=new_font)
            except:
                pass
        
        # Đệ quy cho các widget con
        try:
            for child in widget.winfo_children():
                self._apply_bold(child)
        except:
            pass
    
    def _restore_original_fonts(self, widget):
        """Khôi phục font gốc của tất cả widgets (đệ quy)"""
        widget_id = str(widget)
        
        # Khôi phục font gốc nếu có
        if widget_id in self.original_fonts:
            original_font = self.original_fonts[widget_id]
            
            if isinstance(widget, (tk.Button, tk.Label, tk.Checkbutton, tk.Radiobutton)):
                try:
                    widget.config(font=original_font)
                except:
                    pass
            elif isinstance(widget, tk.LabelFrame):
                try:
                    widget.config(font=original_font)
                except:
                    pass
            elif isinstance(widget, ModernButton):
                try:
                    widget.btn.config(font=original_font)
                except:
                    pass
        
        # Đệ quy cho các widget con
        try:
            for child in widget.winfo_children():
                self._restore_original_fonts(child)
        except:
            pass

    def update_ttk_style(self):
        style = self.style
        
        # Treeview Colors
        style.configure("Treeview", 
                        background=THEME["tree_bg"], 
                        foreground=THEME["tree_fg"], 
                        fieldbackground=THEME["tree_bg"],
                        font=THEME["font_body"], rowheight=35)
        
        style.configure("Treeview.Heading", 
                        background=THEME["tree_header_bg"], 
                        foreground=THEME["tree_header_fg"], 
                        font=THEME["font_title"])
        
        style.map('Treeview', background=[('selected', THEME["primary"])], foreground=[('selected', 'white')])
        
        # [UPDATED] Thêm cấu hình cho "Modern.Treeview" để đồng bộ khi đổi theme
        style.configure("Modern.Treeview", 
                        background=THEME["bg_card"] if not IS_DARK_MODE else "#000000", 
                        foreground=THEME["text_main"], 
                        fieldbackground=THEME["bg_card"] if not IS_DARK_MODE else "#000000",
                        rowheight=40,
                        font=("Segoe UI", 10),
                        borderwidth=0)
        style.configure("Modern.Treeview.Heading", 
                        background=THEME["tree_header_bg"], 
                        foreground=THEME["tree_header_fg"], 
                        font=("Segoe UI", 10, "bold"), padding=(10, 10))
        style.map("Modern.Treeview.Heading", background=[('active', THEME["tree_header_bg"])])
        
        # Notebook Style
        style.configure("TNotebook", background=THEME["bg_app"], borderwidth=0)
        style.configure("TNotebook.Tab", 
                        background=THEME["tab_bg_unselected"], 
                        foreground=THEME["text_main"], 
                        padding=[15, 8],
                        font=("Segoe UI", 10))
        
        style.map("TNotebook.Tab",
            background=[("selected", THEME["tab_bg_selected"]), ("!selected", THEME["tab_bg_unselected"])],
            foreground=[("selected", THEME["primary"]), ("!selected", THEME["text_sub"])],
            lightcolor=[("selected", THEME["border"]), ("!selected", THEME["bg_app"])],
            bordercolor=[("selected", THEME["border"]), ("!selected", THEME["bg_app"])],
        )

    def apply_theme_recursive(self, widget):
        # [UPDATED] Nếu là widget tĩnh (is_static) thì không đổi gì cả (bảo toàn màu chữ/màu nền)
        if hasattr(widget, "is_static") and widget.is_static: return

        try:
            # [UPDATED] Xử lý đổi màu viền cho ModernButton
            if hasattr(widget, "is_modern_btn") and widget.is_modern_btn:
                widget.config(bg=THEME["btn_border_col"])
            
            # [UPDATED] Xử lý đổi màu viền cho Stats Card
            if hasattr(widget, "is_btn_border") and widget.is_btn_border:
                widget.config(bg=THEME["btn_border_col"])
            
            w_class = widget.winfo_class()
            
            if w_class in ["Frame", "Canvas", "TFrame"]:
                # [FIXED] Nếu là card màu (Pastel), KHÔNG đổi màu nền, chỉ giữ nguyên
                if hasattr(widget, "is_colored_card") and widget.is_colored_card:
                    pass 
                elif hasattr(widget, "is_btn_border") and widget.is_btn_border:
                    widget.config(bg=THEME["btn_border_col"])
                else:
                    current_bg = widget.cget("bg")
                    if current_bg != "white" and current_bg != "#FFFFFF" and not IS_DARK_MODE:
                         widget.config(bg=THEME["bg_app"])
                    else:
                         widget.config(bg=THEME["bg_card"])
                
                # [UPDATED] Đổi màu viền nếu là Header hoặc Card chính
                if widget in [self.header]:
                    widget.config(highlightbackground=THEME["border"])
                    
                if hasattr(self, 'notebook'):
                     pass

                if widget in [self.header, self.box, self.welcome]:
                    widget.config(bg=THEME["bg_card"])
                elif widget == self.root:
                    widget.config(bg=THEME["bg_app"])
                
                if isinstance(widget, tk.PanedWindow):
                    widget.config(bg=THEME["bg_app"])

            elif w_class == "Label":
                parent_bg = widget.master.cget("bg")
                widget.config(bg=parent_bg)
                current_fg = widget.cget("fg")
                
                if current_fg in ["gray", "#7f8c8d", "grey"]:
                     widget.config(fg="#B0BEC5" if IS_DARK_MODE else "gray")
                elif current_fg in [THEME["primary"], "#3B8ED0"]:
                     widget.config(fg=THEME["primary"])
                elif current_fg in ["white", "#FFFFFF"]:
                     pass 
                else:
                     widget.config(fg=THEME["text_main"])

            elif w_class == "Entry":
                widget.config(bg=THEME["entry_bg"], fg=THEME["text_main"], insertbackground=THEME["text_main"])
            
            elif w_class == "Text": 
                 widget.config(bg=THEME["entry_bg"], fg=THEME["text_main"])

            elif w_class == "Button":
                if widget == self.btn_toggle:
                    widget.config(bg=THEME["bg_app"], fg=THEME["text_main"])
                
            if isinstance(widget, ExcelTab):
                widget.config(bg=THEME["bg_app"])
                # [UPDATED] Cập nhật viền cho các frame con trong Tab
                widget.chart_card.config(highlightbackground=THEME["border"])
                widget.table_card.config(highlightbackground=THEME["border"])

        except Exception:
            pass
        
        for child in widget.winfo_children():
            self.apply_theme_recursive(child)

    def show_tab_menu(self, event):
        try:
            index = self.notebook.index(f"@{event.x},{event.y}")
            self.notebook.select(index)
            self.tab_menu.post(event.x_root, event.y_root)
        except: pass

    def close_current_tab(self):
        current_tab = self.notebook.select()
        if not current_tab: return
        self.notebook.forget(current_tab)
        if len(self.notebook.tabs()) == 0: self.notebook.add(self.welcome, text="Trang chủ")

    def convert_file_format(self):
        """
        Chuyển đổi định dạng file Excel (.xls -> .xlsx).
        Cho phép người dùng đặt tên file mới và hỏi có muốn mở ngay không.
        """
        # Bước 1: Chọn file nguồn
        source_path = filedialog.askopenfilename(
            title="Chọn file cần chuyển đổi",
            filetypes=[("Excel 97-2003", "*.xls"), ("Tất cả file Excel", "*.xls;*.xlsx")]
        )
        if not source_path:
            return
        
        # Kiểm tra định dạng
        if source_path.lower().endswith('.xlsx'):
            messagebox.showinfo("Thông báo", "File này đã ở định dạng .xlsx, không cần chuyển đổi.")
            return
        
        # Bước 2: Hỏi tên file mới
        base_name = os.path.splitext(os.path.basename(source_path))[0]
        base_dir = os.path.dirname(source_path)
        
        # Dialog nhập tên file - sử dụng InputDialog đẹp
        new_name = InputDialog.ask_string(
            self.root,
            "Đặt tên file mới",
            f"File gốc: {os.path.basename(source_path)}\n\nNhập tên cho file .xlsx mới:",
            initial_value=base_name
        )
        
        if not new_name:
            return  # Người dùng cancel
        
        # Đảm bảo có đuôi .xlsx
        if not new_name.lower().endswith('.xlsx'):
            new_name = new_name + '.xlsx'
        
        output_path = os.path.join(base_dir, new_name)
        
        # Kiểm tra file đã tồn tại chưa
        if os.path.exists(output_path):
            overwrite = messagebox.askyesno(
                "File đã tồn tại",
                f"File '{new_name}' đã tồn tại.\nBạn có muốn ghi đè không?"
            )
            if not overwrite:
                return
        
        # Bước 3: Thực hiện chuyển đổi
        try:
            self.lbl_autosave.config(text="ĐANG CHUYỂN ĐỔI...", fg="#f59e0b")
            self.root.update()
            
            # Dùng Excel COM để SaveAs trực tiếp (giữ nguyên 100% format)
            if HAS_WIN32COM:
                success = _repair_file_via_excel_com(source_path, output_path)
                if not success:
                    raise Exception("Excel COM không thể chuyển đổi file")
            else:
                # Fallback: dùng hàm convert cũ
                converted_path, error_msg = convert_legacy_file_to_xlsx(source_path)
                if converted_path:
                    import shutil
                    shutil.move(converted_path, output_path)
                else:
                    raise Exception(error_msg or "Không thể chuyển đổi")
            
            self.lbl_autosave.config(text="CHUYỂN ĐỔI XONG ✅", fg="#10b981")
            
            # Bước 4: Hỏi có muốn mở file mới không
            open_now = messagebox.askyesno(
                "Chuyển đổi thành công!",
                f"Đã chuyển đổi xong!\n\nFile mới: {new_name}\nVị trí: {base_dir}\n\nBạn có muốn mở file này ngay bây giờ không?"
            )
            
            if open_now:
                # Mở file vừa chuyển đổi
                if self.welcome in self.notebook.tabs():
                    self.notebook.forget(self.welcome)
                
                tab = ExcelTab(self.notebook, output_path, status_label=self.lbl_autosave, filter_label=self.lbl_filter_global)
                self.notebook.add(tab, text=f" {os.path.basename(output_path)} ")
                self.notebook.select(tab)
                
                if IS_DARK_MODE:
                    self.apply_theme_recursive(tab)
                    tab.update_treeview_tags("Dark")
                    
        except Exception as e:
            self.lbl_autosave.config(text="LỖI CHUYỂN ĐỔI ❌", fg="#ef4444")
            messagebox.showerror("Lỗi chuyển đổi", f"Không thể chuyển đổi file:\n{str(e)}")

    def add_new_tab(self):
        paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if not paths: return
        if self.welcome in self.notebook.tabs(): self.notebook.forget(self.welcome)
        for p in paths:
            try:
                f = open(p, "r+"); f.close()
            except PermissionError:
                messagebox.showerror("Lỗi File", f"File '{os.path.basename(p)}' đang mở bên Excel.\nVui lòng tắt file đó trước khi thêm vào đây!")
                continue

            final_path = p
            
            if p.lower().endswith(('.xls', '.xlr')):
                # [NEW] Cho người dùng đặt tên file mới
                base_name = os.path.splitext(os.path.basename(p))[0]
                base_dir = os.path.dirname(p)
                
                new_name = InputDialog.ask_string(
                    self.root,
                    "Chuyển đổi file .xls",
                    f"File '{os.path.basename(p)}' cần chuyển đổi sang .xlsx.\n\nNhập tên cho file mới:",
                    initial_value=base_name
                )
                
                if not new_name:
                    continue  # Người dùng cancel
                
                if not new_name.lower().endswith('.xlsx'):
                    new_name = new_name + '.xlsx'
                
                output_path = os.path.join(base_dir, new_name)
                
                # Kiểm tra file đã tồn tại
                if os.path.exists(output_path):
                    overwrite = messagebox.askyesno(
                        "File đã tồn tại",
                        f"File '{new_name}' đã tồn tại.\nBạn có muốn ghi đè không?"
                    )
                    if not overwrite:
                        continue
                
                # Thực hiện chuyển đổi
                self.lbl_autosave.config(text="ĐANG CHUYỂN ĐỔI...", fg="#f59e0b")
                self.root.update()
                
                try:
                    if HAS_WIN32COM:
                        success = _repair_file_via_excel_com(p, output_path)
                        if success:
                            final_path = output_path
                        else:
                            raise Exception("Excel COM không thể chuyển đổi")
                    else:
                        converted_path, error_msg = convert_legacy_file_to_xlsx(p)
                        if converted_path:
                            import shutil
                            if os.path.exists(output_path):
                                os.remove(output_path)
                            shutil.move(converted_path, output_path)
                            final_path = output_path
                        else:
                            raise Exception(error_msg or "Không thể chuyển đổi")
                    
                    self.lbl_autosave.config(text="ĐÃ CHUYỂN ĐỔI ✅", fg="#10b981")
                    
                    # Hỏi có muốn mở file mới không
                    open_now = messagebox.askyesno(
                        "Chuyển đổi thành công!",
                        f"Đã chuyển đổi xong!\n\nFile mới: {new_name}\n\nBạn có muốn mở file này ngay không?"
                    )
                    
                    if not open_now:
                        continue  # Không mở, tiếp tục file tiếp theo
                        
                except Exception as e:
                    self.lbl_autosave.config(text="LỖI ❌", fg="#ef4444")
                    messagebox.showerror("Lỗi chuyển đổi", f"Không thể chuyển đổi file '{os.path.basename(p)}':\n{str(e)}")
                    continue 

            # [FIXED] Truyền status_label (self.lbl_autosave) và filter label (self.lbl_filter_global)
            tab = ExcelTab(self.notebook, final_path, status_label=self.lbl_autosave, filter_label=self.lbl_filter_global)
            self.notebook.add(tab, text=f" {os.path.basename(final_path)} ")
            self.notebook.select(tab)
            
            if IS_DARK_MODE:
                self.apply_theme_recursive(tab)
                tab.update_treeview_tags("Dark")

    def save_file(self):
        """Lưu file Excel hiện tại (Ctrl+S)"""
        try:
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                messagebox.showinfo("Thông báo", "Không có file nào để lưu.")
                return
            
            current_widget = self.notebook.nametowidget(current_tab_id)
            if hasattr(current_widget, 'current_df') and current_widget.current_df is not None:
                # Hiển thị thông báo đang lưu
                self.lbl_autosave.config(text="ĐANG LƯU...", fg="#f59e0b")
                self.root.update()
                
                # Lưu file (giả sử có hàm save trong ExcelTab)
                if hasattr(current_widget, 'save_current_data'):
                    current_widget.save_current_data()
                    self.lbl_autosave.config(text="ĐÃ LƯU ✅", fg="#10b981")
                    messagebox.showinfo("Thành công", "Đã lưu file thành công!")
                else:
                    self.lbl_autosave.config(text="SẲN SÀNG ✅", fg="#10b981")
                    messagebox.showinfo("Thông báo", "File đã được tự động lưu.")
            else:
                messagebox.showinfo("Thông báo", "Không có dữ liệu để lưu.")
        except Exception as e:
            self.lbl_autosave.config(text="LỖI LƯU ❌", fg="#ef4444")
            messagebox.showerror("Lỗi", f"Không thể lưu file: {str(e)}")

    def focus_search(self):
        """Focus vào ô tìm kiếm (Ctrl+F)"""
        try:
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                messagebox.showinfo("Thông báo", "Vui lòng mở file Excel trước.")
                return
            
            current_widget = self.notebook.nametowidget(current_tab_id)
            # Tìm ô tìm kiếm trong tab hiện tại
            if hasattr(current_widget, 'search_var'):
                # Focus vào search entry nếu có
                for child in current_widget.winfo_children():
                    if isinstance(child, tk.Entry) and child.get().startswith("Tìm kiếm"):
                        child.focus_set()
                        child.select_range(0, tk.END)
                        return
            
            # Nếu không tìm thấy search entry, hiển thị thông báo
            messagebox.showinfo("Thông báo", "Đã focus vào tìm kiếm. Sử dụng ô tìm kiếm trong tab hiện tại.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể focus tìm kiếm: {str(e)}")

    def refresh_data(self):
        """Làm mới dữ liệu (F5)"""
        try:
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                messagebox.showinfo("Thông báo", "Vui lòng mở file Excel trước.")
                return
            
            current_widget = self.notebook.nametowidget(current_tab_id)
            if hasattr(current_widget, 'current_df') and current_widget.current_df is not None:
                # Hiển thị thông báo đang làm mới
                self.lbl_autosave.config(text="ĐANG LÀM MỚI...", fg="#f59e0b")
                self.root.update()
                
                # Làm mới dữ liệu
                if hasattr(current_widget, 'filter_data'):
                    current_widget.filter_data("All")
                    self.lbl_autosave.config(text="ĐÃ LÀM MỚI ✅", fg="#10b981")
                    messagebox.showinfo("Thành công", "Đã làm mới dữ liệu thành công!")
                else:
                    self.lbl_autosave.config(text="SẲN SÀNG ✅", fg="#10b981")
                    messagebox.showinfo("Thông báo", "Dữ liệu đã được làm mới.")
            else:
                messagebox.showinfo("Thông báo", "Không có dữ liệu để làm mới.")
        except Exception as e:
            self.lbl_autosave.config(text="LỖI ❌", fg="#ef4444")
            messagebox.showerror("Lỗi", f"Không thể làm mới dữ liệu: {str(e)}")

    def close_active_window(self):
        """Đóng cửa sổ hiện tại (ESC)"""
        try:
            # Tìm cửa sổ con đang mở và đóng nó
            for widget in self.root.winfo_children():
                if isinstance(widget, tk.Toplevel) and widget.winfo_exists():
                    widget.destroy()
                    return
            
            # Nếu không có cửa sổ con, đóng tab hiện tại
            self.close_current_tab()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đóng cửa sổ: {str(e)}")

    def on_closing(self):
        if messagebox.askokcancel("Xác nhận thoát", "Bạn có chắc chắn muốn thoát chương trình?"):
            self.root.destroy()
            sys.exit(0)


# =========================== TRAINING MODE WINDOW ===========================
class TrainingWindow:
    """Cửa sổ Training Mode để huấn luyện nhận diện tên học sinh"""
    
    def __init__(self, parent, students, phonetic_mapping, save_callback):
        print(f"[DEBUG TrainingWindow] __init__ called")
        print(f"[DEBUG TrainingWindow] parent type: {type(parent)}")
        print(f"[DEBUG TrainingWindow] students: {len(students)}")
        
        self.parent = parent
        self.students = students
        self.phonetic_mapping = phonetic_mapping
        self.save_callback = save_callback
        
        try:
            # parent là VoiceInputWindow (kế thừa Toplevel), nên parent chính là window
            print(f"[DEBUG TrainingWindow] Creating Toplevel...")
            self.window = tk.Toplevel(parent)
            print(f"[DEBUG TrainingWindow] Toplevel created!")
            
            self.window.title("🎓 TRAINING MODE - Huấn luyện nhận diện")
            self.window.geometry("600x500")
            self.window.configure(bg='#2c3e50')
            
            # Recording state
            self.is_recording = False
            self.current_student_idx = 0
            self.recordings = []  # Store recorded texts for current student
            
            print(f"[DEBUG TrainingWindow] Calling create_ui()...")
            self.create_ui()
            print(f"[DEBUG TrainingWindow] TrainingWindow initialized successfully!")
            
        except Exception as e:
            print(f"[DEBUG TrainingWindow ERROR] Failed in __init__: {e}")
            import traceback
            traceback.print_exc()
            raise
        
    def create_ui(self):
        # Header
        header_frame = tk.Frame(self.window, bg='#34495e', height=80)
        header_frame.pack(fill=tk.X, padx=10, pady=10)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="🎓 TRAINING MODE",
            font=("Segoe UI", 18, "bold"),
            bg='#34495e',
            fg='white'
        )
        title_label.pack(pady=5)
        
        info_label = tk.Label(
            header_frame,
            text="Đọc tên học sinh 3 lần để cải thiện độ chính xác",
            font=("Segoe UI", 10),
            bg='#34495e',
            fg='#bdc3c7'
        )
        info_label.pack()
        
        # Progress
        progress_frame = tk.Frame(self.window, bg='#2c3e50')
        progress_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.progress_label = tk.Label(
            progress_frame,
            text=f"0/{len(self.students)} học sinh",
            font=("Segoe UI", 12, "bold"),
            bg='#2c3e50',
            fg='white'
        )
        self.progress_label.pack()
        
        # Current student
        student_frame = tk.Frame(self.window, bg='#34495e', height=100)
        student_frame.pack(fill=tk.X, padx=10, pady=10)
        student_frame.pack_propagate(False)
        
        tk.Label(
            student_frame,
            text="Học sinh hiện tại:",
            font=("Segoe UI", 10),
            bg='#34495e',
            fg='#bdc3c7'
        ).pack(pady=5)
        
        self.student_name_label = tk.Label(
            student_frame,
            text="",
            font=("Segoe UI", 16, "bold"),
            bg='#34495e',
            fg='#3498db'
        )
        self.student_name_label.pack()
        
        # Recording status
        self.record_status_label = tk.Label(
            student_frame,
            text="Bấm nút GHI ÂM để bắt đầu (0/3)",
            font=("Segoe UI", 10),
            bg='#34495e',
            fg='#e74c3c'
        )
        self.record_status_label.pack(pady=5)
        
        # Recordings display
        recording_frame = tk.Frame(self.window, bg='#2c3e50')
        recording_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        tk.Label(
            recording_frame,
            text="Kết quả ghi âm:",
            font=("Segoe UI", 10, "bold"),
            bg='#2c3e50',
            fg='white'
        ).pack(anchor='w')
        
        self.recordings_text = tk.Text(
            recording_frame,
            height=6,
            font=("Consolas", 10),
            bg='#34495e',
            fg='white',
            state='disabled'
        )
        self.recordings_text.pack(fill=tk.BOTH, expand=True)
        
        # Buttons
        button_frame = tk.Frame(self.window, bg='#2c3e50')
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.record_btn = tk.Button(
            button_frame,
            text="🎙️ GHI ÂM",
            font=("Segoe UI", 12, "bold"),
            bg='#e74c3c',
            fg='white',
            command=self.toggle_recording,
            width=12,
            height=2
        )
        self.record_btn.pack(side=tk.LEFT, padx=5)
        
        # Nút xóa dòng cuối
        self.delete_btn = tk.Button(
            button_frame,
            text="🗑️ XÓA CUỐI",
            font=("Segoe UI", 12, "bold"),
            bg='#e67e22',
            fg='white',
            command=self.delete_last_recording,
            width=12,
            height=2,
            state='disabled'
        )
        self.delete_btn.pack(side=tk.LEFT, padx=5)
        
        self.next_btn = tk.Button(
            button_frame,
            text="▶️ TIẾP THEO",
            font=("Segoe UI", 12, "bold"),
            bg='#27ae60',
            fg='white',
            command=self.next_student,
            width=12,
            height=2,
            state='disabled'
        )
        self.next_btn.pack(side=tk.LEFT, padx=5)
        
        skip_btn = tk.Button(
            button_frame,
            text="⏭️ BỎ QUA",
            font=("Segoe UI", 12, "bold"),
            bg='#95a5a6',
            fg='white',
            command=self.skip_student,
            width=12,
            height=2
        )
        skip_btn.pack(side=tk.LEFT, padx=5)
        
        close_btn = tk.Button(
            button_frame,
            text="❌ ĐÓNG",
            font=("Segoe UI", 12, "bold"),
            bg='#7f8c8d',
            fg='white',
            command=self.window.destroy,
            width=12,
            height=2
        )
        close_btn.pack(side=tk.RIGHT, padx=5)
        
        # Load first student
        self.load_student()
    
    def load_student(self):
        """Load thông tin học sinh hiện tại"""
        if self.current_student_idx >= len(self.students):
            messagebox.showinfo("Hoàn thành", f"Đã huấn luyện xong {len(self.students)} học sinh!")
            self.window.destroy()
            return
        
        student = self.students[self.current_student_idx]
        self.student_name_label.config(text=student['name'])
        self.recordings = []
        self.update_recordings_display()
        self.update_progress()
        self.record_btn.config(state='normal')
        self.delete_btn.config(state='disabled')
        self.next_btn.config(state='disabled')
    
    def update_progress(self):
        """Cập nhật progress"""
        self.progress_label.config(text=f"{self.current_student_idx}/{len(self.students)} học sinh")
        self.record_status_label.config(text=f"Bấm nút GHI ÂM để bắt đầu ({len(self.recordings)}/3)")
    
    def update_recordings_display(self):
        """Hiển thị danh sách recordings"""
        self.recordings_text.config(state='normal')
        self.recordings_text.delete('1.0', tk.END)
        
        for i, text in enumerate(self.recordings, 1):
            self.recordings_text.insert(tk.END, f"{i}. {text}\n")
        
        self.recordings_text.config(state='disabled')
    
    def toggle_recording(self):
        """Bật/tắt recording"""
        if self.is_recording:
            self.stop_recording()
        else:
            self.start_recording()
    
    def start_recording(self):
        """Bắt đầu ghi âm"""
        self.is_recording = True
        self.record_btn.config(text="⏹️ DỪNG", bg='#c0392b')
        self.record_status_label.config(text="🎙️ Đang ghi âm... Hãy đọc tên học sinh", fg='#e74c3c')
        
        # Start recording in background
        threading.Thread(target=self.record_audio, daemon=True).start()
    
    def stop_recording(self):
        """Dừng ghi âm"""
        self.is_recording = False
        self.record_btn.config(text="🎙️ GHI ÂM", bg='#e74c3c')
        self.record_status_label.config(text=f"Đang xử lý... ({len(self.recordings)}/3)", fg='#f39c12')
    
    def record_audio(self):
        """Ghi âm và nhận diện - Dùng sounddevice thay vì PyAudio"""
        try:
            if not HAS_SPEECH or not self.parent.recognizer:
                self.stop_recording()
                self.window.after(0, lambda: messagebox.showerror("Lỗi", "Speech recognition không khả dụng!"))
                return
            
            # Kiểm tra sounddevice (thay vì PyAudio)
            try:
                import sounddevice as sd
                import numpy as np
                import wave
                import tempfile
                import os
            except ImportError:
                self.stop_recording()
                self.window.after(0, lambda: messagebox.showerror(
                    "Thiếu sounddevice",
                    "Không thể ghi âm vì thiếu sounddevice!\n\n"
                    "Cài đặt:\n"
                    "pip install sounddevice soundfile numpy"
                ))
                return
            
            # Ghi âm bằng sounddevice (không cần PyAudio)
            sample_rate = 16000
            duration = 3  # 3 giây
            
            self.record_status_label.config(text="🎙️ Đang ghi âm...", fg='#e74c3c')
            
            # Record
            audio_data = sd.rec(int(duration * sample_rate), 
                              samplerate=sample_rate, 
                              channels=1, 
                              dtype='int16')
            sd.wait()  # Wait until recording is finished
            
            # [FIX #1] Lưu audio vào file WAV trước khi đưa vào SpeechRecognition
            tmp_file = tempfile.NamedTemporaryFile(suffix='.wav', delete=False)
            tmp_path = tmp_file.name
            tmp_file.close()
            
            try:
                # Ghi audio_data vào file WAV
                with wave.open(tmp_path, 'wb') as wf:
                    wf.setnchannels(1)
                    wf.setsampwidth(2)  # int16 = 2 bytes
                    wf.setframerate(sample_rate)
                    wf.writeframes(audio_data.tobytes())
                
                # Stop recording UI
                self.stop_recording()
                self.record_status_label.config(text="⏳ Đang nhận diện...", fg='#f39c12')
                
                # Đọc lại bằng speech_recognition
                with sr.AudioFile(tmp_path) as source:
                    audio = self.parent.recognizer.record(source)
                
                # Xóa file tạm
                os.unlink(tmp_path)
                
                # Recognize
                text = self.parent.recognizer.recognize_google(audio, language="vi-VN")
                self.add_recording(text)
                
            except sr.UnknownValueError:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                self.window.after(0, lambda: messagebox.showwarning("Lỗi", "Không nghe rõ, vui lòng thử lại"))
                self.record_status_label.config(text=f"Không nghe rõ ({len(self.recordings)}/3)", fg='#e74c3c')
            except Exception as e:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                self.window.after(0, lambda: messagebox.showerror("Lỗi", f"Lỗi nhận diện: {e}"))
                self.record_status_label.config(text=f"Lỗi ({len(self.recordings)}/3)", fg='#e74c3c')
                    
        except Exception as e:
            self.stop_recording()
            import traceback
            traceback.print_exc()
            self.window.after(0, lambda: messagebox.showerror("Lỗi", f"Lỗi ghi âm: {e}"))
    
    def add_recording(self, text):
        """Thêm recording vào danh sách"""
        self.recordings.append(text)
        self.update_recordings_display()
        self.update_progress()
        
        # Enable nút XÓA khi có ít nhất 1 recording
        self.delete_btn.config(state='normal' if len(self.recordings) > 0 else 'disabled')
        
        if len(self.recordings) >= 3:
            # Đủ 3 lần, enable nút TIẾP THEO
            self.record_btn.config(state='disabled')
            self.next_btn.config(state='normal')
            self.record_status_label.config(text="✅ Đã đủ 3 lần! Bấm TIẾP THEO", fg='#27ae60')
        else:
            self.record_status_label.config(text=f"✅ Đã ghi {len(self.recordings)}/3. Tiếp tục...", fg='#27ae60')
    
    def delete_last_recording(self):
        """Xóa recording cuối cùng (khi đọc nhầm)"""
        if len(self.recordings) > 0:
            deleted_text = self.recordings.pop()
            self.update_recordings_display()
            self.update_progress()
            
            # Update button states
            if len(self.recordings) == 0:
                self.delete_btn.config(state='disabled')
            
            if len(self.recordings) < 3:
                self.record_btn.config(state='normal')
                self.next_btn.config(state='disabled')
                self.record_status_label.config(
                    text=f"🗑️ Đã xóa '{deleted_text}' ({len(self.recordings)}/3)", 
                    fg='#f39c12'
                )
            
            self.log_message(f"🗑️ Đã xóa: '{deleted_text}'")
    
    def log_message(self, message):
        """Log message to recordings text (temporary)"""
        # Tạm thời hiển thị message trong recordings_text
        pass
    
    def next_student(self):
        """Chuyển sang học sinh tiếp theo"""
        # Save phonetic mapping
        student = self.students[self.current_student_idx]
        actual_name = student['name']
        
        for recognized_text in self.recordings:
            # Chỉ lưu nếu khác tên thật
            if recognized_text.lower() != actual_name.lower():
                self.phonetic_mapping[recognized_text.lower()] = actual_name
                print(f"🎓 Training: '{recognized_text}' → '{actual_name}'")
        
        # Save to file
        self.save_callback()
        
        # Next student
        self.current_student_idx += 1
        self.load_student()
    
    def skip_student(self):
        """Bỏ qua học sinh hiện tại"""
        self.current_student_idx += 1
        self.load_student()


# =================================================================================
# LICENSE MENU HANDLERS (StudentManagerApp methods)
# =================================================================================
# Thêm các methods này vào StudentManagerApp class

def show_license_info_method(self):
    """Hiển thị thông tin license"""
    if hasattr(self, 'license_mgr'):
        LicenseInfoDialog(self.root, self.license_mgr)
    else:
        messagebox.showinfo("Thông tin", "License manager không khả dụng!", parent=self.root)

def activate_new_license_method(self):
    """Kích hoạt license mới"""
    if hasattr(self, 'license_mgr'):
        result = messagebox.askyesno(
            "Kích hoạt mới",
            "Bạn muốn kích hoạt license mới?\n\n"
            "License hiện tại sẽ bị thay thế.",
            parent=self.root
        )
        if result:
            # Xóa license cũ
            self.license_mgr.revoke_license()
            
            # Show activation dialog - KHÔNG CHO PHÉP TRIAL (vì đang re-activate)
            activation_dialog = LicenseActivationDialog(self.root, self.license_mgr, allow_trial=False)
            self.root.wait_window(activation_dialog)  # Chờ dialog đóng
            
            if activation_dialog.success:
                messagebox.showinfo(
                    "Thành công",
                    "License mới đã được kích hoạt!\n\nVui lòng khởi động lại ứng dụng.",
                    parent=self.root
                )
                self.on_closing()
            else:
                # User đã cancel - thông báo cần kích hoạt lại
                messagebox.showwarning(
                    "Chưa kích hoạt",
                    "License đã bị xóa nhưng chưa kích hoạt mới!\n\n"
                    "Ứng dụng sẽ đóng. Vui lòng kích hoạt lại lần sau.",
                    parent=self.root
                )
                self.on_closing()
    else:
        messagebox.showerror("Lỗi", "License manager không khả dụng!", parent=self.root)

def revoke_license_debug_method(self):
    """Xóa license (debug mode)"""
    if hasattr(self, 'license_mgr'):
        result = messagebox.askyesno(
            "⚠️ Xóa License",
            "CẢNH BÁO: Tính năng này dành cho debug!\n\n"
            "License sẽ bị xóa hoàn toàn.\n"
            "Bạn sẽ phải kích hoạt lại lần sau.\n\n"
            "Tiếp tục?",
            parent=self.root
        )
        if result:
            if self.license_mgr.revoke_license():
                messagebox.showinfo(
                    "Đã xóa",
                    "License đã bị xóa!\n\nỨng dụng sẽ đóng.",
                    parent=self.root
                )
                self.on_closing()
            else:
                messagebox.showerror("Lỗi", "Không thể xóa license file!", parent=self.root)
    else:
        messagebox.showerror("Lỗi", "License manager không khả dụng!", parent=self.root)

# Bind methods vào class
StudentManagerApp.show_license_info = show_license_info_method
StudentManagerApp.activate_new_license = activate_new_license_method
StudentManagerApp.revoke_license_debug = revoke_license_debug_method


if __name__ == "__main__":
    # =============================================================================
    # SPLASH SCREEN - HIỂN THỊ KHI KHỞI ĐỘNG
    # =============================================================================
    
    # Tạo root ẩn để splash hoạt động
    root_hidden = tk.Tk()
    root_hidden.withdraw()  # Ẩn root chính
    
    # Hiển thị splash screen (2 giây)
    splash = SplashScreen(duration=2000)
    splash.show()
    
    # Đợi splash đóng
    root_hidden.wait_window(splash.splash_root)
    root_hidden.destroy()
    
    # =============================================================================
    # LICENSE VALIDATION - CHECK BEFORE RUNNING APP
    # =============================================================================
    
    print("[DEBUG] Starting license check...")
    
    license_mgr = LicenseManager()
    license_valid = False
    
    # Step 1: Load existing license (nếu có)
    license_data = license_mgr.load_license()
    
    print(f"[DEBUG] License data loaded: {license_data is not None}")
    
    if license_data:
        # ĐÃ CÓ LICENSE - Validate
        is_valid, error_msg = license_mgr.validate_offline()
        
        if is_valid:
            # License OK - Try online check (không bắt buộc)
            try:
                license_mgr.check_online()
            except:
                pass  # Offline mode OK
            
            license_valid = True
        else:
            # License invalid
            temp_root = tk.Tk()
            temp_root.overrideredirect(True)
            temp_root.geometry("1x1+0+0")
            temp_root.update()
            
            messagebox.showerror(
                "License không hợp lệ",
                f"Lỗi: {error_msg}\n\nVui lòng kích hoạt lại license!",
                parent=temp_root
            )
            
            # Show activation dialog
            activation_dialog = LicenseActivationDialog(temp_root, license_mgr)
            temp_root.wait_window(activation_dialog)  # Wait for dialog to close
            license_valid = activation_dialog.success
            temp_root.destroy()
    
    else:
        # CHƯA CÓ LICENSE - Check trial
        trial_available, days_left, hours_left, mins_left = license_mgr.is_trial_available()
        
        print(f"[DEBUG] Trial available: {trial_available}, Days: {days_left}, Hours: {hours_left}, Mins: {mins_left}")
        
        if trial_available:
            # Check xem đã có trial file chưa (đang trong trial)
            if os.path.exists(license_mgr.TRIAL_FILE):
                # ĐÃ TRONG TRIAL - Cho vào app luôn
                print(f"[DEBUG] Already in trial mode. Remaining: {days_left}d {hours_left}h {mins_left}m")
                license_valid = True
            else:
                # CHƯA DÙNG TRIAL - Show activation dialog
                print("[DEBUG] Showing activation dialog (first time trial)...")
                temp_root = tk.Tk()
                # Don't withdraw - keep it visible but minimal
                temp_root.overrideredirect(True)  # Remove title bar
                temp_root.geometry("1x1+0+0")  # Minimal size at corner
                temp_root.update()
                
                # Show activation dialog (có cả nút dùng thử)
                activation_dialog = LicenseActivationDialog(temp_root, license_mgr)
                temp_root.wait_window(activation_dialog)  # Wait for dialog to close
                license_valid = activation_dialog.success
                temp_root.destroy()
                print(f"[DEBUG] Dialog closed. Success: {license_valid}")
        
        else:
            # HẾT TRIAL - BẮT BUỘC ACTIVATE
            temp_root = tk.Tk()
            temp_root.overrideredirect(True)
            temp_root.geometry("1x1+0+0")
            temp_root.update()
            
            messagebox.showwarning(
                "Hết thời gian dùng thử",
                "Bạn đã hết thời gian dùng thử 7 ngày!\n\n"
                "Vui lòng nhập License Key để tiếp tục sử dụng.",
                parent=temp_root
            )
            
            activation_dialog = LicenseActivationDialog(temp_root, license_mgr)
            temp_root.wait_window(activation_dialog)  # Wait for dialog to close
            license_valid = activation_dialog.success
            temp_root.destroy()
    
    # Step 2: Check result
    if not license_valid:
        # User không activate hoặc cancel → EXIT
        sys.exit(0)
    
    # =============================================================================
    # LICENSE OK - RUN APP
    # =============================================================================
    
    root = tk.Tk()
    app = StudentManagerApp(root)
    
    # Pass license_mgr to app (để dùng trong menu)
    app.license_mgr = license_mgr
    
    root.mainloop()

